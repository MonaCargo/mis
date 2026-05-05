# from typing import Any, Dict, List

# from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
# from sqlalchemy.orm import Session

# from app.db.models.exportOperation.export_uld_master import ExportUldMaster
# from app.db.session import get_db
# from app.utils.common.helperFunction import get_utc_now
# from app.utils.exportOperation.export_uld_master_cleaner import parse_uld_excel


# router = APIRouter(prefix="/export-uld", tags=["Export ULD Master"])






# # service function to upload uld data in db
# def upsert_uld_records(db: Session, records: List[Dict[str, str]]) -> Dict[str, Any]:
#     """Insert new ULD records; skip duplicates (uld_no is unique)."""
#     now = get_utc_now()
#     inserted = 0
#     skipped = 0

#     for record in records:
#         uld_no = record["uld_no"]
#         existing = db.query(ExportUldMaster).filter_by(uld_no=uld_no).first()

#         if existing:
#             skipped += 1
#             continue

#         db.add(ExportUldMaster(
#             uld_no=uld_no,
#             carrier=record["carrier"],
#             is_active=True,
#             created_at=now,
#             updated_at=now,
#         ))
#         inserted += 1

#     db.commit()
#     return {"inserted": inserted, "skipped": skipped, "total": len(records)}



# @router.post("/upload", summary="Upload Excel to bulk insert ULD master records")
# async def upload_uld_excel(
#     file: UploadFile = File(...),
#     db: Session = Depends(get_db),
# ):
#     # Validate file type
#     if not file.filename.endswith((".xlsx", ".xls")):
#         raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted.")

#     file_bytes = await file.read()

#     try:
#         records = parse_uld_excel(file_bytes)
#     except ValueError as e:
#         raise HTTPException(status_code=422, detail=str(e))

#     if not records:
#         raise HTTPException(status_code=400, detail="No valid rows found in the uploaded file.")

#     result = upsert_uld_records(db, records)

#     return {
#         "message": "Upload complete.",
#         "inserted": result["inserted"],
#         "skipped_duplicates": result["skipped"],
#         "total_rows_processed": result["total"],
#     }
















import io
import math

import pandas as pd
from sqlalchemy import select
from sqlalchemy.orm import aliased
from typing import Any, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, status
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.dialects.postgresql import insert

from app.core.dependency import verify_token_and_get_user
from app.db.models.exportOperation.export_fileupload_meta_log import ExportFileUploadMetaLog
from app.db.models.exportOperation.export_uld_master import ExportUldMaster
from app.db.models.user import User
from app.db.session import get_db
from app.schemas.domesticOperation.domestic_xray_report import PaginationMetadata
from app.schemas.exportOperation.car_message import UldInventoryRecord, UldStockRecord, UldStockSyncResponse
from app.schemas.exportOperation.location_master import CreateUldRequest
from app.services.exportOperation.uld_master_service import MultipleCarriersError, UldStockSyncService
from app.utils.common.helperFunction import get_utc_now
from app.utils.exportOperation.export_uld_inventory_all_carrier import SUPPORTED_EXTENSIONS, extract_all_carrier_uld_inventory
from app.utils.exportOperation.export_uld_master_cleaner import parse_uld_excel
from app.utils.exportOperation.extract_uld_inventry_pdf import extract_uld_stock

from app.db.session import engine
from fastapi.responses import StreamingResponse

from io             import BytesIO
from datetime       import datetime, timezone
from zoneinfo       import ZoneInfo
 
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils  import get_column_letter
 


router = APIRouter(prefix="/export-uld", tags=["Export ULD Master"])


async def upsert_uld_records(db: AsyncSession, records: List[Dict[str, str]]) -> Dict[str, Any]:
    now = get_utc_now()

    rows = [
        {
            "uld_no": r["uld_no"],
            "carrier": r["carrier"],
            "is_active": True,
            "created_at": now,
            "updated_at": now,
        }
        for r in records
    ]

    stmt = insert(ExportUldMaster).values(rows)
    stmt = stmt.on_conflict_do_nothing(index_elements=["uld_no"])  # skip duplicates at DB level

    result = await db.execute(stmt)
    await db.commit()

    inserted = result.rowcount
    skipped = len(rows) - inserted

    return {"inserted": inserted, "skipped": skipped, "total": len(rows)}


@router.post("/upload", summary="Upload Excel to bulk insert ULD master records")
async def upload_uld_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx or .xls files are accepted.")

    file_bytes = await file.read()

    try:
        records = parse_uld_excel(file_bytes)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))

    if not records:
        raise HTTPException(status_code=400, detail="No valid rows found in the uploaded file.")

    result = await upsert_uld_records(db, records)

    return {
        "message": "Upload complete.",
        "inserted": result["inserted"],
        "skipped_duplicates": result["skipped"],
        "total_rows_processed": result["total"],
    }



@router.post(
    "/new-uld/create",
    summary="Create a new ULD",
    status_code=201,
)
async def create_uld(
    payload: CreateUldRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),
):
    # check duplicate
    existing = await db.execute(
        select(ExportUldMaster).where(ExportUldMaster.uld_no == payload.uld_no.strip().upper())
    )
    if existing.scalar_one_or_none():
        raise HTTPException(status_code=400, detail=f"ULD '{payload.uld_no}' already exists")

    now = get_utc_now()
    uld = ExportUldMaster(
        uld_no=payload.uld_no.strip().upper(),
        carrier=payload.carrier.strip().upper(),
        uld_type=payload.uld_type.strip().upper() if payload.uld_type else None,
        is_active=True,
        created_at=now,
        updated_at=now,
        created_by=current_user.emp_id, 
    )
    db.add(uld)
    await db.commit()
    await db.refresh(uld)

    return {
        "success": True,
        "message": f"ULD '{uld.uld_no}' created successfully",
        "data": {
            "uld_id": uld.id,
            "uld_no": uld.uld_no,
            "carrier": uld.carrier,
        }
    }



# ======================== 🤢🤮

async def _write_uld_failure_log(filename, file_track_type, uploaded_by, now, error_message):
    try:
        async with AsyncSession(engine) as log_session:
            async with log_session.begin():
                log_session.add(ExportFileUploadMetaLog(
                    filename=filename,
                    file_type="pdf",
                    uploaded_by=uploaded_by,
                    uploaded_at=now,
                    file_track_type=file_track_type,
                    status="FAILED",
                    upload_meta={
                        "total_in_pdf": 0,
                        "inserted": 0,
                        "updated": 0,
                    },
                    error_message=str(error_message)[:500],
                    created_at=now,
                ))
    except Exception as log_err:
        print(f"⚠️ ULD log write failed: {log_err}")



MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB
@router.post(
"/upload-and-sync-by-inventry-pdf",
response_model=UldStockSyncResponse,
status_code=status.HTTP_200_OK,
summary="Upload ULD stock PDF and sync to database",
description=(
    "Accepts a ULD stock PDF, extracts all ULD records from it, "
    "then upserts them into export_uld_master. "
    "Existing ULDs are marked available and refreshed. "
    "New ULDs are created. ULDs not in this PDF are left untouched."
),
responses={
    200: {"description": "PDF processed and database synced successfully"},
    400: {"description": "Invalid file, empty PDF, or mixed carriers"},
    413: {"description": "File exceeds 10 MB limit"},
    500: {"description": "Unexpected error during extraction or sync"},
},
)
async def upload_and_sync_uld_stock(
    db: AsyncSession = Depends(get_db),
    file: UploadFile = File(...),
    current_user = Depends(verify_token_and_get_user),
) -> UldStockSyncResponse:
    
    now = get_utc_now()
    filename = file.filename or "unknown"

    # ✅ Capture primitive immediately — avoids lazy load on expired session
    emp_id = current_user.emp_id
    try:
        # ── 1. Validate file type ─────────────────────────────────────────────────
        if not (file.filename or "").lower().endswith(".pdf"):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only PDF files are allowed.",
            )

        # ── 2. Read & validate file size ──────────────────────────────────────────
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="File too large. Maximum allowed size is 10 MB.",
            )


        # ── 3. Extract records from PDF ───────────────────────────────────────────
        try:
            df = extract_uld_stock(io.BytesIO(contents))
        except Exception as exc:
        
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Failed to extract data from the PDF.",
            ) from exc

        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid ULD records found in the uploaded PDF.",
            )

        

        # ── 4. Convert DataFrame rows → UldStockRecord objects ────────────────────
        records: list[UldStockRecord] = []
        for row in df.where(pd.notnull(df), None).to_dict(orient="records"):
            # Normalize datetime/Timestamp fields to ISO strings for Pydantic
            normalized = {
                k: v.isoformat() if hasattr(v, "isoformat") else v
                for k, v in row.items()
            }
            records.append(UldStockRecord(**normalized))

        # ── 5. Sync into DB ───────────────────────────────────────────────────────
        service = UldStockSyncService(db=db)

        try:
            response = await service.sync(
                records=records,
                synced_by= emp_id,  # pass the employee ID of the authenticated user
            )
        except MultipleCarriersError as exc:
        
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc
        except Exception as exc:
        
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="An unexpected error occurred during ULD stock synchronization.",
            ) from exc
        # return response
        # ── 6. Write success log ──────────────────────────────────────────
        db.add(ExportFileUploadMetaLog(
            filename=filename,
            file_type="pdf",
            uploaded_by=emp_id,
            uploaded_at=now,
            file_track_type="ULD_INVENTRY_PDF",
            status="SUCCESS",
            upload_meta={
                 "carrier":        response.carrier,
                "total_in_pdf": response.total_received,
                "inserted":     response.total_created,
                "updated":      response.total_updated,
            },
            error_message=None,
            created_at=now,
        ))
        await db.commit()

        return response

    except HTTPException as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_uld_failure_log(
            filename=filename,
            file_track_type="ULD_INVENTRY_PDF",
            uploaded_by=emp_id,
            now=now,
            error_message=e.detail if isinstance(e.detail, str) else str(e.detail),
        )
        raise

    except Exception as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_uld_failure_log(
            filename=filename,
             file_track_type="ULD_INVENTRY_PDF",
            uploaded_by=emp_id,
            now=now,
            error_message=str(e),
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Processing failed: {str(e)}",
        )



# ========================  ULD iNVENTRU EXCEL UPLOAD FOR ALL INVENTRY =========================== 

@router.post(
    "/excel-upload-and-sync-uld-inventory-file",
    response_model=UldStockSyncResponse,
    status_code=status.HTTP_200_OK,
    summary="Upload all-carrier ULD inventory CSV/Excel and sync to DB",
    responses={
        200: {"description": "File processed and DB synced successfully"},
        400: {"description": "Invalid file type or no valid records found"},
        413: {"description": "File exceeds 10 MB limit"},
        500: {"description": "Unexpected error during extraction or sync"},
    },
)
async def upload_and_sync_uld_inventory(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
) -> UldStockSyncResponse:
 
    now      = get_utc_now()
    filename = file.filename
    emp_id   = current_user.emp_id
 
    try:
        # ── 1. Validate extension ─────────────────────────────────────────────
        if not any(filename.lower().endswith(ext) for ext in SUPPORTED_EXTENSIONS):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Unsupported file type. Allowed: {', '.join(sorted(SUPPORTED_EXTENSIONS))}",
            )
 
        # ── 2. Read & size-check ──────────────────────────────────────────────
        contents = await file.read()
        if len(contents) > MAX_FILE_SIZE:
            raise HTTPException(
                status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
                detail="File too large. Maximum allowed size is 10 MB.",
            )
 
        # ── 3. Clean → DataFrame from cleaner ────────────────────────────────
        try:
            df = extract_all_carrier_uld_inventory(io.BytesIO(contents), filename)
        except ValueError as exc:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=str(exc),
            ) from exc
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail=f"Failed to extract data from the file: {exc}",
            ) from exc
 
        if df.empty:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="No valid ULD records found in the uploaded file.",
            )
 
        
        # ── 4. DataFrame → Pydantic records ───────────────────────────────────────────
        records: list[UldInventoryRecord] = []
        for row in df.where(pd.notnull(df), None).to_dict(orient="records"):
            normalized = {
                k: v.isoformat() if hasattr(v, "isoformat") else v
                for k, v in row.items()
            }
            # carrier_code is None for rows like O30321** — skip those rows entirely
            if not normalized.get("carrier_code"):
                continue
            records.append(UldInventoryRecord(**normalized))
 
        # ── 5. Sync via service ───────────────────────────────────────────────
        service = UldStockSyncService(db=db)
        try:
            response = await service.sync_all_carrier_inventory_file(
                records=records,
                synced_by=emp_id,
            )
        except Exception as exc:
            raise HTTPException(
                status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
                detail="Unexpected error during ULD inventory synchronization.",
            ) from exc
 
        # ── 6. Audit log ──────────────────────────────────────────────────────
        db.add(ExportFileUploadMetaLog(
            filename        = filename,
            file_type       = filename.rsplit(".", 1)[-1],
            uploaded_by     = emp_id,
            uploaded_at     = now,
            file_track_type = "ULD_INVENTRY_ALL_CARRIER",
            status          = "SUCCESS",
            upload_meta     = {
                "total_in_file": response.total_received,
                "inserted":      response.total_created,
                "updated":       response.total_updated,
            },
            error_message   = None,
            created_at      = now,
        ))
        await db.commit()
        return response
 
    except HTTPException as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_uld_failure_log(
            filename=filename,
             file_track_type="ULD_INVENTRY_ALL_CARRIER",
            uploaded_by=emp_id,
            now=now,
            error_message=e.detail if isinstance(e.detail, str) else str(e.detail),
        )
        raise
 
    except Exception as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_uld_failure_log(
            filename=filename,
             file_track_type="ULD_INVENTRY_ALL_CARRIER",
            uploaded_by=emp_id,
            now=now,
            error_message=str(e),
        )
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Processing failed: {str(e)}",
        )
 

# ============== ULD MASTER DATA TO SHOW IN WEB TABLE WITH FILTER AND PAGINATION ==============


@router.get("/get-uld-master-list-with-pagination", summary="Get ULD master list with filters and pagination")
async def get_uld_master_list(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),

    # filters
    carrier:      Optional[str]  = Query(None, description="Carrier code e.g. AI, LH, EK | all"),
    is_available: Optional[bool] = Query(None, description="true | false"),
    is_active:    Optional[bool] = Query(None, description="true | false"),

    # pagination
    page:      int = Query(1,  ge=1,         description="Page number"),
    page_size: int = Query(20, ge=1, le=200, description="Records per page"),
):
    # ✅ call via class
    records, total = await UldStockSyncService.get_filtered_uld_master(
        db=db,
        carrier=carrier,
        is_available=is_available,
        is_active=is_active,
        page=page,
        page_size=page_size,
    )

    total_pages = math.ceil(total / page_size) if total > 0 else 1

    pagination = PaginationMetadata(
        current_page=page,
        page_size=page_size,
        total_records=total,
        total_pages=total_pages,
        has_previous=page > 1,
        has_next=page < total_pages,
        previous_page=page - 1 if page > 1 else None,
        next_page=page + 1 if page < total_pages else None,
    )

    data = [
        {
            "id":           r.id,
            "uld_no":       r.uld_no,
            "uld_type":     r.uld_type,
            "carrier":      r.carrier,
            "is_available": r.is_available,
            "is_active":    r.is_active,
            "created_by":   r.created_by,
            "updated_by":   r.updated_by,
            "created_at":   r.created_at.isoformat() if r.created_at else None,
            "updated_at":   r.updated_at.isoformat() if r.updated_at else None,
        }
        for r in records
    ]

    return {
        "success":    True,
        "message":    "ULD master records fetched successfully",
        "pagination": pagination,
        "data":       data,
    }


@router.get("/uld-master/carriers")
async def get_uld_carriers(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),
):
    """Returns distinct carrier codes — use this to populate the carrier dropdown."""
    carriers = await UldStockSyncService.get_distinct_carriers(db)
    return {
        "success":  True,
        "carriers": carriers,
    }



@router.get(
    "/search",
    summary="Search ULD by exact match"
)
async def search_in_uld_master(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),
    uld_no: str = Query(..., description="ULD number"),
):
    # ============================
    # Normalize Input
    # ============================
    # cleaned_uld = uld_no.strip().upper().replace(" ", "")
    cleaned_uld = uld_no.strip().upper()
    UserAlias = aliased(User)
    if not cleaned_uld:
        raise HTTPException(status_code=400, detail="ULD number is required")

    # ============================
    # Query (Exact Match)
    # ============================
    stmt = select(
        ExportUldMaster,
          UserAlias.name.label("created_by_emp_name")
                  ).outerjoin(
        UserAlias,
        UserAlias.emp_id == ExportUldMaster.created_by
    ).where(
        ExportUldMaster.uld_no == cleaned_uld,
        ExportUldMaster.is_active == True
    )

    result = await db.execute(stmt)
    row = result.first()

    # ============================
    # Response
    # ============================
    if not row:
        raise HTTPException(
            status_code=404,
            detail=f"ULD {cleaned_uld} not found"
        )
    
    uld = row[0]
    created_by_name = row.created_by_emp_name

    return {
        "status": "success",
        "data": {
            "id": uld.id,
            "uld_no": uld.uld_no,
            "uld_type": uld.uld_type,
            "carrier": uld.carrier,
            "is_available": uld.is_available,
            "is_active": uld.is_active,
            "created_by": uld.created_by,
            "created_by_emp_name": created_by_name,
            "updated_by": uld.updated_by,
            "created_at": uld.created_at,
            "updated_at": uld.updated_at,
        }
    }

# Expoer filterred data
@router.get(
    "/export-uld-master-data",
    summary="Export ULD master list as Excel based on filters",
    # response_class=StreamingResponse,
)
async def export_uld_master(
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),

    # filters (same as the list endpoint)
    carrier:      Optional[str]  = Query(None, description="Carrier code e.g. AI, LH, EK | all"),
    is_available: Optional[bool] = Query(None, description="true | false"),
    is_active:    Optional[bool] = Query(None, description="true | false"),
):
    # Step 1: get total count
    _, total = await UldStockSyncService.get_filtered_uld_master(
        db=db,
        carrier=carrier,
        is_available=is_available,
        is_active=is_active,
        page=1,
        page_size=1,
    )

    if total == 0:
        raise HTTPException(status_code=404, detail="No records found for the given filters.")

    # Step 2: fetch all matching records in one shot
    records, total = await UldStockSyncService.get_filtered_uld_master(
        db=db,
        carrier=carrier,
        is_available=is_available,
        is_active=is_active,
        page=1,
        page_size=total,
    )

    # ── Build Excel ──────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ULD Master"

    # ── Styles ───────────────────────────────────────────────────────────────
    header_font       = Font(bold=True, color="FFFFFF", size=11)
    header_fill       = PatternFill("solid", fgColor="1E3A5F")
    header_alignment  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    center_alignment  = Alignment(horizontal="center", vertical="center")
    left_alignment    = Alignment(horizontal="left",   vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    green_fill = PatternFill("solid", fgColor="C6EFCE")   # available
    red_fill   = PatternFill("solid", fgColor="FFC7CE")   # not available
    blue_fill  = PatternFill("solid", fgColor="BDD7EE")   # active
    gray_fill  = PatternFill("solid", fgColor="D9D9D9")   # inactive

    # ── Header row ────────────────────────────────────────────────────────────
    headers = [
        "S.No", "ULD No", "ULD Type", "Carrier",
        "Availability", "Status",
        # "Created By", "Updated By",
        "Created At (IST)", 
        # "Updated At (IST)",
    ]
    col_widths = [6, 18, 12, 10, 14, 10, 14, 14, 22, 22]

    ws.row_dimensions[1].height = 30
    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_alignment
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Helper: UTC → IST string ──────────────────────────────────────────────
    def to_ist(dt) -> str:
        if not dt:
            return "-"
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        ist = dt.astimezone(ZoneInfo("Asia/Kolkata"))
        return ist.strftime("%d-%b-%Y %H:%M")

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, r in enumerate(records, start=2):
        availability_text = "Available"     if r.is_available else "Not Available"
        active_text       = "Active"        if r.is_active    else "Inactive"
        avail_fill        = green_fill       if r.is_available else red_fill
        active_fill_cell  = blue_fill        if r.is_active    else gray_fill

        row_data = [
            row_idx - 1,          # S.No
            r.uld_no,
            r.uld_type,
            r.carrier,
            availability_text,
            active_text,
            # r.created_by or "-",
            # r.updated_by or "-",
            to_ist(r.created_at),
            # to_ist(r.updated_at),
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border    = thin_border
            cell.alignment = center_alignment if col_idx != 2 else left_alignment

            # Colored badge cells
            if col_idx == 5:   # Availability
                cell.fill = avail_fill
            elif col_idx == 6: # Status
                cell.fill = active_fill_cell

        ws.row_dimensions[row_idx].height = 18

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A2"

    # ── Summary row (totals) ──────────────────────────────────────────────────
    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total Records").font = Font(bold=True)
    ws.cell(row=summary_row, column=2, value=total).font = Font(bold=True)

    # ── Stream response ───────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    carrier_label = carrier if carrier and carrier != "all" else "ALL"
    filename = f"uld_master_{carrier_label}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# Toggle Availabilty of container like BT and PD

@router.get(
    "/toggle-availability/{uld_no}",
    summary="Toggle ULD availability (available ↔ unavailable)",
)
async def toggle_uld_availability(
    uld_no: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    cleaned = uld_no.strip().upper()
    result = await db.execute(
        select(ExportUldMaster).where(ExportUldMaster.uld_no == cleaned)
    )
    uld: Optional[ExportUldMaster] = result.scalar_one_or_none()

    if not uld:
        raise HTTPException(status_code=404, detail=f"ULD '{cleaned}' not found")

    uld.is_available = not uld.is_available
    uld.updated_at   = get_utc_now()
    uld.updated_by   = current_user.emp_id

    await db.commit()
    await db.refresh(uld)

    return {
        "success":      True,
        "uld_no":       uld.uld_no,
        "is_available": uld.is_available,
        "message":      f"ULD '{uld.uld_no}' marked as {'Available' if uld.is_available else 'Unavailable'}",
    }