



from datetime import date, datetime, timezone
from io import BytesIO
import math
from typing import Optional

from fastapi import APIRouter, HTTPException
from fastapi.params import Depends
from fastapi.responses import StreamingResponse
from fastapi.temp_pydantic_v1_params import Query
from sqlalchemy import and_, func, select
from sqlalchemy.ext.asyncio import AsyncSession
from app.api.routes.domesticOperation.domestic_xray_report import convert_ist_day_to_utc_range
from app.core.dependency import verify_token_and_get_user
from app.db.models.exportOperation.export_fileupload_meta_log import ExportFileUploadMetaLog
from app.db.models.user import User
from app.db.session import engine, get_db
from app.schemas.domesticOperation.domestic_xray_report import PaginationMetadata

router = APIRouter(prefix="/meta", tags=["Export File Upload Meta Logs"])




@router.get("/car-message")
async def get_upload_logs(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),

    # filters
    status:          Optional[str]  = Query(None, description="SUCCESS | FAILED | all"),
    file_track_type: Optional[str]  = Query(None, description="CAR_MESSAGE_AWB | WH_INVENTORY_PDF | ULD_STOCK_PDF | all"),
    start_date:      Optional[date] = Query(None, description="Filter from this date (YYYY-MM-DD)"),
    end_date:        Optional[date] = Query(None, description="Filter to this date (YYYY-MM-DD)"),

    # pagination
    page:      int = Query(1,  ge=1,  description="Page number"),
    page_size: int = Query(20, ge=1, le=100, description="Records per page"),
):
    # ── Convert IST dates → UTC datetimes ────────────────────────────────────
    # utc_start = None
    # utc_end   = None

    # if start_date:
    #     naive_start = datetime.combine(start_date, time.min)
    #     utc_start   = IST.localize(naive_start).astimezone(timezone.utc)

    # if end_date:
    #     naive_end = datetime.combine(end_date, time.max)
    #     utc_end   = IST.localize(naive_end).astimezone(timezone.utc)

    utc_start, _ = convert_ist_day_to_utc_range(start_date)
    _, utc_end = convert_ist_day_to_utc_range(end_date)

    # ── Fetch ─────────────────────────────────────────────────────────────────
    records, total = await get_filtered_upload_logs(
        db=db,
        status=status,
        file_track_type=file_track_type,
        start_date=utc_start,
        end_date=utc_end,
        page=page,
        page_size=page_size,
    )

    # ── Pagination ────────────────────────────────────────────────────────────
    total_pages = math.ceil(total / page_size) if total > 0 else 1

    pagination = PaginationMetadata(
        current_page=page,
        page_size=page_size,
        total_records=total,
        total_pages=total_pages,
        has_previous=page > 1,
        has_next=page < total_pages,
        previous_page=page - 1 if page > 1 else None,
        next_page=page + 1 if page < total_pages else None,
    )

    # ── Serialize ─────────────────────────────────────────────────────────────
    # data = [
    #     {
    #         "id":              r.id,
    #         "filename":        r.filename,
    #         "file_type":       r.file_type,
    #         "file_track_type": r.file_track_type,
    #         "uploaded_by":     r.uploaded_by,
    #         "uploaded_at":     r.uploaded_at,
    #         "status":          r.status,
    #         "upload_meta":     r.upload_meta,
    #         "error_message":   r.error_message,
    #         "created_at":      r.created_at,
    #     }
    #     for r in records
    # ]

    data = records

    return {
        "message":    "Upload logs fetched successfully",
        "success":    True,
        "pagination": pagination,
        "data":       data,
    }





@router.get("/car-message/file-upload-log/export")
async def export_upload_logs(
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),

    status: Optional[str] = Query(None),
    file_track_type: Optional[str] = Query(None),
    start_date: Optional[date] = Query(None),
    end_date: Optional[date] = Query(None),
):
    utc_start, _ = convert_ist_day_to_utc_range(start_date)
    _, utc_end = convert_ist_day_to_utc_range(end_date)

    records = await get_upload_logs_for_export(
        db=db,
        status=status,
        file_track_type=file_track_type,
        start_date=utc_start,
        end_date=utc_end,
    )

    file_stream = build_upload_logs_excel(records)

    return StreamingResponse(
        file_stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=upload_logs.xlsx"},
    )





# ==========


async def get_filtered_upload_logs(
    db: AsyncSession,
    status: Optional[str],           # "SUCCESS" | "FAILED" | None = all
    file_track_type: Optional[str],  # "CAR_MESSAGE_AWB" | "WH_INVENTORY_PDF" | "ULD_STOCK_PDF" | None = all
    start_date: Optional[date],
    end_date: Optional[date],
    page: int,
    page_size: int,
) -> tuple[list[ExportFileUploadMetaLog], int]:

    # query       = select(ExportFileUploadMetaLog)
    query = select(
        ExportFileUploadMetaLog,
        User.name.label("user_name")   # ✅ add user name
        ).outerjoin(
            User,
            User.emp_id == ExportFileUploadMetaLog.uploaded_by
    )

    count_query = select(func.count()).select_from(ExportFileUploadMetaLog)

    conditions = []

    if status and status != "all":
        if status not in ("SUCCESS", "FAILED"):
            raise HTTPException(status_code=400, detail=f"Invalid status: {status}. Use SUCCESS or FAILED.")
        conditions.append(ExportFileUploadMetaLog.status == status)

    if file_track_type and file_track_type != "all":
        conditions.append(ExportFileUploadMetaLog.file_track_type == file_track_type)

    if start_date:
        conditions.append(ExportFileUploadMetaLog.created_at >= start_date)

    if end_date:
        conditions.append(ExportFileUploadMetaLog.created_at <= end_date)

    if conditions:
        query       = query.where(and_(*conditions))
        count_query = count_query.where(and_(*conditions))

    total_result = await db.execute(count_query)
    total        = total_result.scalar() or 0

    offset = (page - 1) * page_size
    query  = query.order_by(ExportFileUploadMetaLog.created_at.desc()).offset(offset).limit(page_size)

    # result  = await db.execute(query)
    # records = result.scalars().all()

    # return list(records), total

    result = await db.execute(query)
    rows = result.mappings().all()

    records = []

    for row in rows:
        log = row[ExportFileUploadMetaLog]

        records.append({
            "id": log.id,
            "filename": log.filename,
            "file_type": log.file_type,
            "file_track_type": log.file_track_type,
            "uploaded_by": log.uploaded_by,
            "uploaded_by_name": row["user_name"],   # ✅ NEW FIELD
            "uploaded_at": log.uploaded_at,
            "status": log.status,
            "upload_meta": log.upload_meta,
            "error_message": log.error_message,
            "created_at": log.created_at,
        })

    return records, total





# =================== export file upload meta log service function ===================

async def get_upload_logs_for_export(
    db: AsyncSession,
    status: Optional[str],
    file_track_type: Optional[str],
    start_date: Optional[date],
    end_date: Optional[date],
) -> list[dict]:

    query = select(
        ExportFileUploadMetaLog,
        User.name.label("user_name")
    ).outerjoin(
        User,
        User.emp_id == ExportFileUploadMetaLog.uploaded_by
    )

    conditions = []

    if status and status != "all":
        conditions.append(ExportFileUploadMetaLog.status == status)

    if file_track_type and file_track_type != "all":
        conditions.append(ExportFileUploadMetaLog.file_track_type == file_track_type)

    if start_date:
        conditions.append(ExportFileUploadMetaLog.created_at >= start_date)

    if end_date:
        conditions.append(ExportFileUploadMetaLog.created_at <= end_date)

    if conditions:
        query = query.where(and_(*conditions))

    query = query.order_by(ExportFileUploadMetaLog.created_at.desc())

    result = await db.execute(query)
    rows = result.mappings().all()

    records = []

    for row in rows:
        log = row[ExportFileUploadMetaLog]

        records.append({
            "filename": log.filename,
            "file_type": log.file_type,
            "file_track_type": log.file_track_type,
            "uploaded_by": log.uploaded_by,
            "uploaded_by_name": row["user_name"],
            "uploaded_at": log.uploaded_at,
            "status": log.status,
            "error_message": log.error_message,
            "created_at": log.created_at,
        })

    return records


def build_upload_logs_excel(records: list[dict]) -> BytesIO:
    import pytz
    import openpyxl
    from openpyxl.styles import Font, Alignment

    IST = pytz.timezone("Asia/Kolkata")

    def to_ist(val):
        if not val:
            return ""
        if isinstance(val, datetime):
            if val.tzinfo is None:
                val = val.replace(tzinfo=timezone.utc)
            return val.astimezone(IST).strftime("%d-%b-%Y %H:%M")
        return str(val)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Upload Logs"

    headers = [
        "Filename",
        "File Type",
        "Track Type",
        "User Name",
        "Emp ID",
        "Uploaded At (IST)",
        "Status",
        "Error Message",
        "Created At (IST)",
    ]

    # Header style
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=r.get("filename"))
        ws.cell(row=row_idx, column=2, value=r.get("file_type"))
        ws.cell(row=row_idx, column=3, value=r.get("file_track_type"))
        ws.cell(row=row_idx, column=4, value=r.get("uploaded_by_name") or "-")
        ws.cell(row=row_idx, column=5, value=r.get("uploaded_by"))
        ws.cell(row=row_idx, column=6, value=to_ist(r.get("uploaded_at")))
        ws.cell(row=row_idx, column=7, value=r.get("status"))
        ws.cell(row=row_idx, column=8, value=r.get("error_message") or "")
        ws.cell(row=row_idx, column=9, value=to_ist(r.get("created_at")))

    # Auto width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf