from datetime import date, datetime, timedelta, timezone
from io import BytesIO
import io
import math
from typing import Any, Optional

from fastapi import APIRouter, Body, Form, HTTPException, Query, UploadFile, File, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.orm import Session
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.dependency import verify_token_and_get_user
from app.db.models.exportOperation.car_message import ExportAwbSkidItemSequence, ExportAwbSkidMapping, ExportCarMessageAwbMaster, ExportFlightBookingDetail, ExportFlightBookingHeader, ExportSkidLocationMapping
from app.db.models.exportOperation.export_fileupload_meta_log import ExportFileUploadMetaLog
from app.db.models.exportOperation.export_location_master import ExportLocationsMaster
from app.db.models.exportOperation.export_skid_master import ExportSkidMaster
from app.db.models.user import User
from app.db.session import get_db
from app.schemas.exportOperation.car_message import AvailableAwbForFlightBookingResponse, AvailableAwbForFlightBookingResponseList, AwbCheckerInfoResponse, AwbLookupError, AwbManualCreateRequest, AwbManualCreateResponse, CarMessageExcelExportRequest, CreateFlightBookingFromPdfResponse, CreateFlightBookingRequest, CreateFlightBookingResponse, CreateUldAssignmentRequest, DashboardStatsResponse, EditFlightBookingRequest, EditFlightBookingResponse, EditUldAssignmentRequest, FlightBookingAwbItem, FlightBookingByFlightResponse, FlightUldLoadingStatusResponse, PdfUpsertResponse, RetrieveSkidFromLocationRequest, ScanItemIntoUldRequest, ScanItemIntoUldResponse, UldAssignmentDataResponse, UldAssignmentResponse, UldMasterResponse, UldVerifyForLoadingResponse, UltraFastScanRequest, UpdateFlightDptDatetime
from app.schemas.exportOperation.uld_master import AddUldsRequest
from app.schemas.user import UserRead
from app.services.exportOperation.base_master import ultra_fast_scan_and_load
from app.services.exportOperation.car_message import CategoryLiteral, assign_ulds_to_flight_mobile, build_car_message_excel, close_per_uld__per_flight_service, create_flight_booking, create_manual_awb_service, create_uld_assignment, edit_flight_booking, edit_uld_assignment, enrich_awb_from_wh_inventory, extract_carrier_for_uld_filter, generate_flight_date_report, generate_loading_report_excel_shift_wise, get_available_awbs_for_flight_booking_dropdown, get_awb_checker_info, get_awb_data_filtered, get_awb_data_for_export, get_awb_missing_sequence_status, get_car_message_dashboard_stats, get_dashboard_drilldown_detail, get_dashboard_drilldown_detail_v2, get_dashboard_stats_v2, get_flight_awb_breakdown, get_flight_booking_by_flight_no_and_date, get_flight_full_detail, get_flight_history, get_flight_particular_flight_detail, get_flight_uld_loading_status, get_flights_by_date, get_latest_uld_assignment_service, get_loading_sheet_form_history_service, get_loading_sheet_form_service, get_loading_sheet_report_service, get_uld_assignment_by_flight, get_uld_master_list, get_uld_master_list_eligeble_for_assignment, get_uld_sequences_of_particular_flight, mark_awb_ultra_fast, retrieve_skid_from_location, save_export_car_message_awbs, save_loading_sheet_form_service, scan_item_into_uld, trace_sequence_full_info, unlock_per_uld_service, upsert_flight_booking_from_pdf, verify_uld_for_loading
from app.services.export_slot_file_upload_service import get_utc_now
from app.utils.exportOperation.car_message import clean_car_message
from app.utils.exportOperation.extract_flight_planning_data import extract_flight_planning
from app.utils.exportOperation.wh_inventry_pdf_data_extract import extract_export_inventory
from app.db.session import engine
from app.schemas.exportOperation.car_message import UpdateAwbPiecesRequest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from app.services.exportOperation.car_message_flow_audit_log import write_car_message_flow_audit
from app.utils.common.car_message_flow_audit_utils import CarMessageFlowModule, CarMessageFlowStep


router = APIRouter(
    prefix="/car-message-awb",
    tags=[]
)

async def _write_failure_log(filename, file_type, uploaded_by, now, meta, error_message):
    try:
        async with AsyncSession(engine) as log_session:
            async with log_session.begin():
                log_session.add(ExportFileUploadMetaLog(
                    filename=filename,
                    file_type=file_type,
                    uploaded_by=uploaded_by,
                    uploaded_at=now,
                    file_track_type="CAR_MESSAGE_AWB",
                    status="FAILED",
                    upload_meta=meta,
                    error_message=str(error_message)[:500],
                    created_at=now,
                ))
    except Exception as log_err:
        print(f"⚠️ Log write failed: {log_err}")
async def _write_inventory_failure_log(filename, uploaded_by, now, error_message):
    try:
        async with AsyncSession(engine) as log_session:
            async with log_session.begin():
                log_session.add(ExportFileUploadMetaLog(
                    filename=filename,
                    file_type="pdf",
                    uploaded_by=uploaded_by,
                    uploaded_at=now,
                    file_track_type="CAR_WH_INVENTORY_PDF",
                    status="FAILED",
                    upload_meta={
                        "total_in_pdf": 0,
                        "matched_updated": 0,
                        "not_found_count": 0,
                    },
                    error_message=str(error_message)[:500],
                    created_at=now,
                ))
    except Exception as log_err:
        print(f"⚠️ Inventory log write failed: {log_err}")

@router.post("/upload")
async def process_export_car_message_file(
    file: UploadFile = File(...),
    current_user = Depends(verify_token_and_get_user),
    db: AsyncSession = Depends(get_db)
):
    
    now = get_utc_now()
    filename = file.filename or "unknown"
    file_type = "unknown"
    emp_id = current_user.emp_id

    try:
        # ✅ Validate filename
        if not file.filename:
            raise HTTPException(status_code=400, detail="No file provided")

        # ✅ Detect file type
        file_extension = file.filename.split('.')[-1].lower()

        if file_extension in ("xlsx", "xls"):
            file_type = "excel"
        elif file_extension == "csv":
            file_type = "csv"
        else:
            raise HTTPException(
                status_code=400,
                detail="Unsupported file type. Allowed: .xlsx, .xls, .csv"
            )

        # ✅ Read file as bytes
        file_bytes = BytesIO(await file.read())

        # ✅ Clean file (returns cleaned + faulty)
        cleaned_df, faulty_df = clean_car_message(
            file_bytes,
            file_type
        )


        # ✅ Save using async ON CONFLICT service
        save_result = await save_export_car_message_awbs(
            db,
            cleaned_df,
            uploaded_by=emp_id,
        )

        # ✅ write log in SAME transaction
        db.add(ExportFileUploadMetaLog(
            filename=filename,
            file_type=file_type,
            uploaded_by=emp_id,
            uploaded_at=now,
            file_track_type="CAR_MESSAGE_AWB",
            status="SUCCESS",
            upload_meta={
                "total_received": save_result["total_received"],
                "inserted": save_result["inserted"],
                "updated": save_result["updated"],
                "already_present": save_result["already_present"],
                "faulty_rows": len(faulty_df),
            },
            error_message=None,
            created_at=now,
        ))

        # ✅ ONE commit — AWBs + log together
        await db.commit()

        return {
            "message": "File processed successfully",
            "success":True,
            **save_result,
            "faulty_rows_count": len(faulty_df),
        }

    except HTTPException as e:
        try:
            await db.rollback()
        except Exception:
            pass
        # ✅ CHANGED: use fresh session instead of broken db
        await _write_failure_log(
            filename=filename,
            file_type=file_type,
            uploaded_by=emp_id,
            now=now,
            meta={"total_received": 0, "inserted": 0,
                  "updated": 0, "already_present": 0, "faulty_rows": 0},
            error_message=e.detail if isinstance(e.detail, str) else str(e.detail),
        )
        raise

    except Exception as e:
        try:
            await db.rollback()
        except Exception:
            pass
        # ✅ CHANGED: use fresh session instead of broken db
        await _write_failure_log(
            filename=filename,
            file_type=file_type,
            uploaded_by=emp_id,
            now=now,
            meta={"total_received": 0, "inserted": 0,
                  "updated": 0, "already_present": 0, "faulty_rows": 0},
            error_message=str(e),
        )
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")

# @router.get("/search/by-awb", summary="Search AWB by substring")
# async def search_awb(
#     q: str = Query(..., min_length=1, description="Substring to search in AWB no"),
#     db: AsyncSession = Depends(get_db),
# ):
    
    
#     async def search_awb_by_substring(
#         db: AsyncSession,
#         awb_substring: str,
#     ) -> list:
#         stmt = (
#             select(ExportCarMessageAwbMaster)
#             .where(ExportCarMessageAwbMaster.awb_no.ilike(f"%{awb_substring}%"))
#             .order_by(ExportCarMessageAwbMaster.created_at.desc())
#         )
#         result = await db.execute(stmt)
#         rows = result.scalars().all()
#         return rows

#     rows = await search_awb_by_substring(db, q.strip())

#     if not rows:
#         return {"message": "No records found.","success":True, "data": []}

#     return {
#         "message": f"{len(rows)} record(s) found.",
#         "success":True,
#         "data": [
#             {
#                 "id": r.id,
#                 "awb_no": r.awb_no,
#                 "origin": r.origin,
#                 "destination": r.destination,
#                 "sb_no": r.sb_no,
#                 "sb_date": r.sb_date,
#                 "pcs": r.pcs,
#                 "gross_wt": r.gross_wt,
#                 "chg_wt": r.chg_wt,
#                 "nog": r.nog,
#                 "car_msg_date": r.car_msg_date,
#                 "car_msg_time": r.car_msg_time,
#                 "created_at": r.created_at,
#             }
#             for r in rows
#         ],
#     }


@router.get("/search/by-awb", summary="Search AWB by exact match")
async def search_awb(
    q: str = Query(..., min_length=1, description="AWB no to search"),
    db: AsyncSession = Depends(get_db)
):
    from collections import defaultdict

    # ── 1. AWB + scanned count ────────────────────────────────────
    stmt = (
        select(
            ExportCarMessageAwbMaster,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .outerjoin(
            ExportAwbSkidItemSequence,
            ExportAwbSkidItemSequence.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(ExportCarMessageAwbMaster.awb_no == q.strip())
        .group_by(ExportCarMessageAwbMaster.id)
    )

    result = await db.execute(stmt)
    row = result.first()

    if not row:
        return {"success": False, "message": "No records found.", "data": None}
    

    awb = row.ExportCarMessageAwbMaster
    scanned_pcs = row.scanned_pcs

    # ── 2. Fetch all sequences for this AWB ───────────────────────
    seq_stmt = (
        select(ExportAwbSkidItemSequence)
        .where(ExportAwbSkidItemSequence.awb_master_id == awb.id)
        .order_by(
            ExportAwbSkidItemSequence.mapping_id,
            ExportAwbSkidItemSequence.sequence_no,
        )
    )
    seq_result = await db.execute(seq_stmt)
    sequences = seq_result.scalars().all()

    # Group sequences by mapping_id
    seq_by_mapping = defaultdict(list)
    for s in sequences:
        seq_by_mapping[s.mapping_id].append({
            "id": s.id,
            "sequence_no": s.sequence_no,
            "sequence_date_time": s.sequence_date_time,
            "scan_by_device": s.scan_by_device,
            "scanned_by": s.scanned_by,
        })

    # ── 3. Fetch all skids mapped to this AWB ─────────────────────
    skid_stmt = (
        select(ExportAwbSkidMapping, ExportSkidMaster)
        .join(
            ExportSkidMaster,
            ExportSkidMaster.id == ExportAwbSkidMapping.skid_id,
        )
        .where(ExportAwbSkidMapping.awb_master_id == awb.id)
        .order_by(ExportAwbSkidMapping.created_at.asc())
    )
    skid_result = await db.execute(skid_stmt)
    skid_rows = skid_result.all()

    # ── 4. Fetch current locations for all skids of this AWB ──────
    skid_ids = [sr.ExportSkidMaster.id for sr in skid_rows]

    current_location_map = {}
    if skid_ids:
        loc_stmt = (
            select(ExportSkidLocationMapping, ExportLocationsMaster)
            .join(
                ExportLocationsMaster,
                ExportLocationsMaster.id == ExportSkidLocationMapping.location_id,
            )
            .where(
                ExportSkidLocationMapping.skid_id.in_(skid_ids),
                ExportSkidLocationMapping.is_current == True,
            )
        )
        loc_result = await db.execute(loc_stmt)
        loc_rows = loc_result.all()

        for lr in loc_rows:
            current_location_map[lr.ExportSkidLocationMapping.skid_id] = lr.ExportLocationsMaster.loc

    return {
        "success": True,
        "message": "Record found.",
        "data": {
            # ── AWB master info ───────────────────────────────────
            "id": awb.id,
            "awb_no": awb.awb_no,
            "origin": awb.origin,
            "destination": awb.destination,
            "is_ultra_fast": getattr(awb, "is_ultra_fast", False),
            "is_manually_created": getattr(awb, "is_manually_created", False),
            "sb_no": awb.sb_no,
            "sb_date": awb.sb_date,
            "hwb_no": awb.hwb_no,
            "gross_wt": awb.gross_wt,
            "volumetric_wt": awb.volumetric_wt,
            "chg_wt": awb.chg_wt,
            "nog": awb.nog,
            "shc": awb.shc,
            "remarks": awb.remarks,
            "manual_creation_remarks": awb.manual_creation_remarks,
            "car_msg_date": awb.car_msg_date,
            "car_msg_time": awb.car_msg_time,
            "car_message_datetime_combo": awb.car_message_datetime_combo,
            "uploaded_by": awb.uploaded_by,
            "created_at": awb.created_at,
            "updated_at": awb.updated_at,

            # ── pcs summary ───────────────────────────────────────
            "total_pcs": awb.pcs,
            "scanned_pcs": scanned_pcs,
            "remaining_pcs": (awb.pcs - scanned_pcs) if awb.pcs is not None else None,
            "is_complete": (scanned_pcs >= awb.pcs) if awb.pcs is not None else False,

            # ── skids with sequences nested ───────────────────────
            "skids": [
                {
                    "mapping_id": sr.ExportAwbSkidMapping.id,
                    "skid_id": sr.ExportSkidMaster.id,
                    "skid_no": sr.ExportSkidMaster.skid_no,
                    "skid_type": sr.ExportSkidMaster.skid_type,
                    "is_virtual": sr.ExportAwbSkidMapping.is_virtual,
                    "virtual_skid_no": sr.ExportAwbSkidMapping.virtual_skid_no,
                    "is_locked": sr.ExportSkidMaster.is_locked,
                    "mapped_at": sr.ExportAwbSkidMapping.created_at,
                    "scanned_count": len(seq_by_mapping[sr.ExportAwbSkidMapping.id]),
                    "sequences": seq_by_mapping[sr.ExportAwbSkidMapping.id],
                    "current_location": current_location_map.get(sr.ExportSkidMaster.id, None),
                    
                }
                for sr in skid_rows
            ],
        },
    } 


@router.get("/all-awb-data", summary="Get all AWB master records")
async def get_all_awb(
    db: AsyncSession = Depends(get_db),
):
    stmt = (
        select(ExportCarMessageAwbMaster)
        .order_by(ExportCarMessageAwbMaster.created_at.desc())
    )

    result = await db.execute(stmt)
    rows = result.scalars().all()

    if not rows:
        return {
            "success": True,
            "message": "No records found.",
            "data": [],
        }

    return {
        "success": True,
        "message": f"{len(rows)} record(s) found.",
        "data": [
            {
                "id": r.id,
                "awb_no": r.awb_no,
                "origin": r.origin,
                "destination": r.destination,
                "sb_no": r.sb_no,
                "sb_date": r.sb_date,
                "pcs": r.pcs,
                "gross_wt": r.gross_wt,
                "chg_wt": r.chg_wt,
                "nog": r.nog,
                "car_msg_date": r.car_msg_date,
                "car_msg_time": r.car_msg_time,
                "created_at": r.created_at,
            }
            for r in rows
        ],
    }


# 🤢 Get all awb of car message based on date filter status feature ....
@router.get("/awb-data-with-paginatoin", summary="Get AWB master records with filters and pagination")
async def get_awb_data(
    # startDate: str = Query(..., example="2026-01-24"),
    # endDate: str = Query(..., example="2026-01-24"),
    startDate: Optional[date],
    endDate: Optional[date],
    status: str = Query("all", example="all"),  # all | rcs | not_rcs
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=401),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    """
    Get AWB master records filtered by date range and status (paginated).
    status options: all | rcs | not_rcs
    """
    records, total = await get_awb_data_filtered(
        db=db,
        start_date=startDate,
        end_date=endDate,
        status=status,
        page=page,
        page_size=page_size,
    )

    total_pages = math.ceil(total / page_size) if total else 1

    pagination = {
        "current_page": page,
        "page_size": page_size,
        "total_records": total,
        "total_pages": total_pages,
        "has_previous": page > 1,
        "has_next": page < total_pages,
        "previous_page": page - 1 if page > 1 else None,
        "next_page": page + 1 if page < total_pages else None,
    }

    return {
        "success": True,
        "message": f"{total} record(s) found.",
        "pagination": pagination,
        "data": records,
    }


@router.post("/awb-data-export-by-filters")
async def export_awb_data(
    payload: CarMessageExcelExportRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),
):
    records = await get_awb_data_for_export(
        db=db,
        start_date=payload.startDate,
        end_date=payload.endDate,
        status=payload.status,
    )

    if not records:
        raise HTTPException(status_code=404, detail="No records found for the selected filters.")

    if len(records) > 50000:
        raise HTTPException(status_code=400, detail="Too many records. Please narrow your date range.")

    buf = build_car_message_excel(records)

    filename = f"car_message_awb_{payload.startDate}_{payload.endDate}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )






# =========== ✌️ UPLOAD CAR message pdf wh inventry for fligh booking page ===================

@router.post("/wh-inventry-pdf/upload-extract-and-save")
async def upload_inventory_pdf(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    now = get_utc_now()
    filename = file.filename or "unknown"
     # ✅ Capture primitive immediately — avoid lazy load issues in except block
    emp_id = current_user.emp_id  # ← remove this line if no auth on this route

    try:
        if not file.filename.lower().endswith(".pdf"):
            raise HTTPException(status_code=400, detail="Only PDF files allowed")

        contents = await file.read()  # ← read once here

        if len(contents) > 10 * 1024 * 1024:
            raise HTTPException(status_code=400, detail="File too large (Max 10MB)")

        # ✅ Wrap already-read bytes in BytesIO — do NOT use file.file again
        import io
        df = extract_export_inventory(io.BytesIO(contents))

        if df.empty:
            raise HTTPException(status_code=400, detail="No valid records found in pdf")

        print(df.head(3))

        result = await enrich_awb_from_wh_inventory(db, df)

         # ✅ Write success log
        db.add(ExportFileUploadMetaLog(
            filename=filename,
            file_type="pdf",
            uploaded_by=emp_id,
            uploaded_at=now,
            file_track_type="CAR_WH_INVENTORY_PDF",
            status="SUCCESS",
            upload_meta={
                "total_in_pdf":    result["total_in_pdf"],
                "matched_updated": result["matched"],
                "not_found_count": result["not_found_count"],
            },
            error_message=None,
            created_at=now,
        ))
        await db.commit()


        return {
            "success": True,
            "message": "PDF processed successfully.",
            "total_in_pdf":    result["total_in_pdf"],
            "matched_updated": result["matched"],
            "not_found_count": result["not_found_count"],
            "not_found_awbs":  result["not_found"],
        }
    except HTTPException as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_inventory_failure_log(
            filename=filename,
            uploaded_by=emp_id,
            now=now,
            error_message=e.detail if isinstance(e.detail, str) else str(e.detail),
        )
        raise

    except Exception as e:
        try:
            await db.rollback()
        except Exception:
            pass
        await _write_inventory_failure_log(
            filename=filename,
            uploaded_by=emp_id,
            now=now,
            error_message=str(e),
        )
        raise HTTPException(status_code=500, detail=f"Processing failed: {str(e)}")




# ======✌️ Get those awb which is allowed to select in flight booking screen dropdown 
@router.get(
    "/flight-booking/available-awbs",
    response_model=AvailableAwbForFlightBookingResponseList,
    summary="AWBs available for flight booking",
)
async def get_available_awbs(
    db: AsyncSession = Depends(get_db),
):
    try:
        data =  await get_available_awbs_for_flight_booking_dropdown(
            db=db,
        )

        return AvailableAwbForFlightBookingResponseList(
            success=True,
            message="Get all valid and availble awb for flight booking",
            data = data
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    


# ✌️======Create new flight booking =============== 
@router.post(
    "/flight-booking/create",
    response_model=CreateFlightBookingResponse,
    summary="Create a new flight booking",
    status_code=201
)
async def create_booking(
    payload: CreateFlightBookingRequest,
    db: AsyncSession = Depends(get_db),
    current_user:UserRead= Depends(verify_token_and_get_user),
):
    
    return await create_flight_booking(db=db, payload=payload, booked_by=current_user.emp_id)




# Edit flight booking 
@router.get(
    "/flight-booking/by-flt-no-and-date",
    response_model=FlightBookingByFlightResponse,
    summary="Get flight booking by flight number and date",
)
async def get_booking_by_flight(
    flight_no: str = Query(..., description="e.g. AI101"),
    flight_date: date = Query(..., description="e.g. 2026-03-16"),
    db: AsyncSession = Depends(get_db),
):
    return await get_flight_booking_by_flight_no_and_date(
        db=db,
        flight_no=flight_no,
        flight_date=flight_date,
    )


# ✌️ Edit flight booking--------=====
@router.patch(
    "/flight-booking/{header_id}/edit",
    response_model=EditFlightBookingResponse,
    summary="Edit an existing flight booking",
)
async def edit_booking(
    header_id: int,
    payload: EditFlightBookingRequest,
    db: AsyncSession = Depends(get_db),
     current_user:UserRead= Depends(verify_token_and_get_user),
   
):
    return await edit_flight_booking(
        db=db,
        header_id=header_id,
        payload=payload,
        edited_by=current_user.emp_id,
    )

# =================== ✌️ ULD ASSIGNMENT ✌️ ========================================

@router.get("/uld-assignment/uld-master", response_model=list[UldMasterResponse])
async def get_uld_master(db: AsyncSession = Depends(get_db)):
    return await get_uld_master_list(db=db)


@router.get("/uld-assignment/uld-master/eligible-for-uld-assignment", response_model=list[UldMasterResponse])
async def get_uld_master(
     flight_no: str | None = Query(None),
    db: AsyncSession = Depends(get_db)
    ):


    carriers = None

    if flight_no:
        carrier_code = await extract_carrier_for_uld_filter(db=db, flight_no=flight_no)
        if carrier_code:
            carriers = [carrier_code]
    print("FLIGHT NO:", flight_no)
    print("EXTRACTED CARRIER:", carrier_code)
    return await get_uld_master_list_eligeble_for_assignment(db=db,carriers=carriers)


@router.get("/uld-assignment/by-flight", response_model=UldAssignmentDataResponse | None, description="GET CREATED ULD ASSIGNMENT BY FLIGHT NO. AND FLIGHT DATE")
async def get_assignment_by_flight(
    flight_no: str = Query(...),
    flight_date: date = Query(...),
    db: AsyncSession = Depends(get_db),
):
    return await get_uld_assignment_by_flight(
        db=db,
        flight_no=flight_no,
        flight_date=flight_date,
    )


@router.post("/uld-assignment/create", response_model=UldAssignmentResponse, status_code=201)
async def create_assignment(
    payload: CreateUldAssignmentRequest,
    db: AsyncSession = Depends(get_db),
    current_user:UserRead= Depends(verify_token_and_get_user),
):
    
    return await create_uld_assignment(db=db, payload=payload, assigned_by=current_user.emp_id)


@router.patch("/uld-assignment/{assignment_id}/edit", response_model=UldAssignmentResponse)
async def edit_assignment(
    assignment_id: int,
    payload: EditUldAssignmentRequest,
    db: AsyncSession = Depends(get_db),
    current_user:UserRead= Depends(verify_token_and_get_user),
):
  
    return await edit_uld_assignment(
        db=db,
        assignment_id=assignment_id,
        payload=payload,
        edited_by=current_user.emp_id,
    )


# ========================= ✌️ Skid retrival from location ==============================


@router.get(
"/flight-booking/get-all-flight/by-date",
summary="Get all booked flights on a particular date",
)
async def get_flights_by_date_route(
    flight_date: date = Query(..., description="e.g. 2026-03-16"),
    db: AsyncSession = Depends(get_db),
):
    return await get_flights_by_date(db=db, flight_date=flight_date)


@router.get(
    "/flight-booking/{header_id}/full-flight-detail",
    summary="Get full flight detail — AWBs + skids + sequences + ULDs",
)
async def get_flight_full_detail_route(
    header_id: int,
    db: AsyncSession = Depends(get_db),
):
    return await get_flight_full_detail(db=db, header_id=header_id)

@router.patch(
    "/skid/retrieve",
    summary="Retrieve skid from its current location by mapping id",
)
async def retrieve_skid(
    payload: RetrieveSkidFromLocationRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user),
):
    return await retrieve_skid_from_location(
        db=db,
        mapping_id=payload.mapping_id,
        retrieved_by=current_user.emp_id,
    )





# ===============👌👌 EXPORT ULD/PALLET LOADING BY SCANNING SEQUENCE [LAST STEP OF PROCESS]====================

# ── 1. Verify ULD ──────────────────────────────────────────
@router.get(
    "/flight/{flight_header_id}/uld-loading/verify-uld",
    response_model=UldVerifyForLoadingResponse,
    summary="Verify ULD belongs to flight before scanning",
)
async def verify_uld_for_loading_route(
    flight_header_id: int,
    uld_no: str = Query(...),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await verify_uld_for_loading(db=db, flight_header_id=flight_header_id, uld_no=uld_no)


# ── 2. Scan item into ULD ──────────────────────────────────
@router.post(
    "/flight/{flight_header_id}/uld-loading/scan-item",
    response_model=ScanItemIntoUldResponse,
    summary="Scan item barcode into selected ULD",
    status_code=201,
)
async def scan_item_into_uld_route(
    flight_header_id: int,
    payload: ScanItemIntoUldRequest,
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    print(f"Received scan request: flight_header_id={flight_header_id}, payload={payload}") 
    return await scan_item_into_uld(
        db=db,
        flight_header_id=flight_header_id,
        payload=payload,
        loaded_by=current_user.emp_id,
    )


# ── 3. Get loading status ──────────────────────────────────
@router.get(
    "/flight/{flight_header_id}/uld-loading/status",
    response_model=FlightUldLoadingStatusResponse,
    summary="Get ULD loading status for flight",
)
async def get_loading_status_route(
    flight_header_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await get_flight_uld_loading_status(db=db, flight_header_id=flight_header_id)









# ======================================= car message report related ====================


@router.get(
    "/flight-date-report",
    summary="Download full flight report by date as Excel",
)
async def download_flight_date_report(
    # flight_date: date = Query(..., description="e.g. 2026-03-20"),
    from_date: date = Query(..., description="e.g. 2026-03-20"),
    to_date: date = Query(..., description="e.g. 2026-03-25"),
    db: AsyncSession = Depends(get_db),
):
    return await generate_flight_date_report(db=db, from_date=from_date,to_date=to_date)





@router.get(
    "/dashboard/stats",
    # response_model=DashboardStatsResponse,
    summary="Get dashboard stats for a selected IST date",
)
async def get_dashboard_stats_route(
    report_date: date = Query(..., description="IST date e.g. 2026-03-20"),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await get_car_message_dashboard_stats(db=db, report_date=report_date)

@router.get(
    "/dashboard/statsv2/v2",
    summary="Get dashboard stats for a selected IST date, filtered by AWB source category",
    description=(
        "Returns AWB, scanning, skid/location, and flight stats for the given IST date.\n\n"
        "**category** options:\n"
        "- `all` — no source filter, all AWBs for that date\n"
        "- `car_message` — only AWBs where source IS NULL (uploaded via CAR message)\n"
        "- `export_tp` — only AWBs where source = `EXPORT_TP_XRAY`\n"
        "- `import_tp` — only AWBs where source = `IMPORT_TP_XRAY`\n"
        "- `tp_combine` — combined Export + Import TP (source IN `EXPORT_TP_XRAY`, `IMPORT_TP_XRAY`)\n\n"
        "Flight stats are identical across all categories (not source-filtered)."
    ),
)
async def get_dashboard_stats_v2_route(
    report_date: date = Query(..., description="IST date e.g. 2026-03-20"),
    category: CategoryLiteral = Query(
        ...,
        description="AWB source category: all | car_message | export_tp | import_tp | tp_combine",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await get_dashboard_stats_v2(
        db=db,
        report_date=report_date,
        category=category,
    )

@router.get(
    "/dashboard/drilldown/detail",
    summary="Get drilldown detail for a dashboard stat box",
)
async def get_dashboard_detail_route(
    report_date: date = Query(..., description="IST date e.g. 2026-03-20"),
    detail_type: str = Query(
        ...,
        description="all_awbs | rcs_awbs | non_rcs_awbs | scanned_awbs | used_skids | others_awbs"
    ),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await get_dashboard_drilldown_detail(
        db=db,
        report_date=report_date,
        detail_type=detail_type,
    )




@router.get(
    "/dashboard/drilldown/detailv2/v2",
    summary="Get drilldown detail for a v2 dashboard stat box",
    description=(
        "Returns detailed records for a specific stat box on the v2 dashboard.\n\n"
        "**detail_type** options:\n"
        "- `all_awbs` — all AWBs for the date, filtered by category\n"
        "- `scanned_awbs` — AWBs with at least one scan, with scanner details\n"
        "- `used_skids` — skids used for the date AWBs, with location and base info\n"
        "- `flight_bookings` — all flights for the date (not source-filtered)\n\n"
        "- `others_awbs` — all those awb scanned today but have carmessage date different \n\n"
        "**category** options:\n"
        "- `all` — no source filter\n"
        "- `car_message` — source IS NULL\n"
        "- `export_tp` — source = EXPORT_TP_XRAY\n"
        "- `import_tp` — source = IMPORT_TP_XRAY\n"
        "- `tp_combine` — source IN (EXPORT_TP_XRAY, IMPORT_TP_XRAY)\n\n"
        "Note: `flight_bookings` ignores the category param and always returns all flights."
    ),
)
async def get_dashboard_drilldown_detail_v2_route(
    report_date: date = Query(..., description="IST date e.g. 2026-03-20"),
    detail_type: str = Query(
        ...,
        description="all_awbs | scanned_awbs | used_skids | flight_bookings | others_awbs",
    ),
    category: CategoryLiteral = Query(
        ...,
        description="all | car_message | export_tp | import_tp | tp_combine",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await get_dashboard_drilldown_detail_v2(
        db=db,
        report_date=report_date,
        detail_type=detail_type,
        category=category,
    )











# ========================================= Flight booking by pdf ==========================
# @router.post(
#     "/create-flight-from-pdf-upload",
#     response_model=CreateFlightBookingFromPdfResponse,
#     summary="Create flight booking by uploading the planning PDF",
#     status_code=201,
# )
# async def create_booking_from_pdf_upload(
#     pdf_file: UploadFile = File(..., description="Export Planning Report PDF"),
#     flight_dpt_datetime: datetime = Form(
#         ...,
#         description="Departure datetime in IST. e.g. 2026-04-02T22:30:00",
#     ),
#     db: AsyncSession = Depends(get_db),
#     current_user: UserRead = Depends(verify_token_and_get_user),
# ):
#     # ── Step 1: validate file type ────────────────────────────────────────
#     if pdf_file.content_type not in ("application/pdf", "application/octet-stream"):
#         raise HTTPException(
#             status_code=400,
#             detail="Uploaded file must be a PDF.",
#         )
 
#     # ── Step 2: read bytes and run OCR extraction ─────────────────────────
#     pdf_bytes = await pdf_file.read()
#     if not pdf_bytes:
#         raise HTTPException(status_code=400, detail="Uploaded PDF is empty.")
 
#     try:
#         df = extract_flight_planning(io.BytesIO(pdf_bytes))
#     except Exception as e:
#         raise HTTPException(
#             status_code=422,
#             detail=f"PDF extraction failed: {str(e)}",
#         )
 
#     if df.empty:
#         raise HTTPException(
#             status_code=422,
#             detail="No AWB records could be extracted from the PDF. "
#                    "Check that the file is a valid Export Planning Report.",
#         )
 
#     # ── Step 3: pull flight-level fields (same value on every row) ────────
#     flight_no   = df["FLIGHT_NUM"].iloc[0]
#     flight_date = df["FLIGHT_DATE"].iloc[0]     # already a date object after extraction
 
#     if not flight_no:
#         raise HTTPException(status_code=422, detail="Could not extract Flight Number from PDF.")
#     if not flight_date:
#         raise HTTPException(status_code=422, detail="Could not extract Flight Date from PDF.")
 
#     # ── Step 4: deduplicate AWB numbers ───────────────────────────────────
#     # The same AWB appears on multiple rows in the PDF (one per location/ULD).
#     # We collapse to unique AWB numbers — pcs come from the DB, not the PDF.
#     unique_awb_nos: list[str] = (
#         df["AWB_NUM"]
#         .dropna()
#         .drop_duplicates()
#         .tolist()
#     )
 
#     if not unique_awb_nos:
#         raise HTTPException(
#             status_code=422,
#             detail="No valid AWB numbers found in the extracted PDF data.",
#         )
 
#     # ── Step 5: bulk-fetch AWB master records ─────────────────────────────
#     result = await db.execute(
#         select(
#             ExportCarMessageAwbMaster.id,
#             ExportCarMessageAwbMaster.awb_no,
#             ExportCarMessageAwbMaster.pcs,
#         ).where(ExportCarMessageAwbMaster.awb_no.in_(unique_awb_nos))
#     )
#     db_awb_map = {row.awb_no: row for row in result.mappings().all()}
 
#     # ── Step 6: resolve awb_no → AwbBookingItem (book ALL pcs) ───────────
#     not_found: list[AwbLookupError] = []
#     awb_items: list[FlightBookingAwbItem] = []
 
#     for awb_no in unique_awb_nos:
#         db_row = db_awb_map.get(awb_no)
 
#         if not db_row:
#             not_found.append(AwbLookupError(
#                 awb_no=awb_no,
#                 reason="Not found in export_car_message_awb_master",
#             ))
#             continue
 
#         if not db_row.pcs:
#             not_found.append(AwbLookupError(
#                 awb_no=awb_no,
#                 reason="AWB has no pcs recorded — cannot book",
#             ))
#             continue
 
#         awb_items.append(FlightBookingAwbItem(
#             awb_master_id=db_row.id,
#             booked_pcs=db_row.pcs,      # always book full pcs from DB
#         ))
 
#     if not awb_items:
#         raise HTTPException(
#             status_code=400,
#             detail={
#                 "message": "None of the AWBs from the PDF exist in the database.",
#                 "not_found_awbs": [e.model_dump() for e in not_found],
#             },
#         )
 
#     # ── Step 7: build standard request and delegate to existing service ───
#     booking_request = CreateFlightBookingRequest(
#         flight_no=flight_no,
#         flight_date=flight_date,
#         flight_dpt_datetime=flight_dpt_datetime,
#         awbs=awb_items,
#     )

#     print("Booking request constructed from PDF:", booking_request)

#     # return {
#     #     "success": True,
#     #     "message": "Flight booking created from PDF. See 'booking' for details. "
#     #                "Any AWBs that could not be processed are listed in 'not_found_awbs'.",
#     # }
 
#     booking_result = await create_flight_booking(
#         db=db,
#         payload=booking_request,
#         booked_by=current_user.emp_id,
#     )
 
#     return CreateFlightBookingFromPdfResponse(
#           success=True,                                          # ← add
#     message="Flight booking created successfully from PDF.",  # ← add
#         booking=booking_result,
#         not_found_awbs=not_found,
#     )
 




@router.post(
    "/create-flight-from-pdf-upload",
    response_model=PdfUpsertResponse,
    summary="Create or update flight booking by uploading the planning PDF",
    status_code=200,
)
async def create_booking_from_pdf_upload(
    pdf_file: UploadFile = File(...),
    # flight_dpt_datetime: datetime = Form(...),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    if pdf_file.content_type not in ("application/pdf", "application/octet-stream"):
        raise HTTPException(status_code=400, detail="Uploaded file must be a PDF.")

    pdf_bytes = await pdf_file.read()
    if not pdf_bytes:
        raise HTTPException(status_code=400, detail="Uploaded PDF is empty.")

    try:
        df = extract_flight_planning(io.BytesIO(pdf_bytes))
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"PDF extraction failed: {str(e)}")

    if df.empty:
        raise HTTPException(
            status_code=422,
            detail="No AWB records could be extracted from the PDF.",
        )

    return await upsert_flight_booking_from_pdf(
        db=db,
        df=df,
        # flight_dpt_datetime=flight_dpt_datetime,
        booked_by=current_user.emp_id,
    )






# ========= ultra fast process 🤢 ============

# ── Mark AWB ultra-fast ────────────────────────────────────────
@router.patch(
    "/awb/{awb_master_id}/mark-ultra-fast",
    summary="Mark or unmark AWB as ultra-fast",
)
async def mark_awb_ultra_fast_route(
    awb_master_id: int,
    is_ultra_fast: bool = Query(...),
    remarks: str | None = Query(
        None,
        description="Remarks for marking ultra-fast (required if is_ultra_fast=true)",
    ),
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    # print(f"mark_awb_ultra_fast_route called with: awb_master_id={awb_master_id}, is_ultra_fast={is_ultra_fast}, remarks={remarks}")
    return await mark_awb_ultra_fast(
        db=db,
        awb_master_id=awb_master_id,
        is_ultra_fast=is_ultra_fast,
        marked_by=current_user.emp_id,
        remarks=remarks 
    )


# ── Ultra-fast ULD scan ────────────────────────────────────────
@router.post(
    "/uld-loading/{flight_header_id}/ultra-fast-scan",
    summary="Scan barcodes directly into ULD for ultra-fast AWB",
)
async def ultra_fast_scan_route(
    flight_header_id: int,
    payload: UltraFastScanRequest,
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    return await ultra_fast_scan_and_load(
        db=db,
        flight_header_id=flight_header_id,
        uld_assignment_detail_id=payload.uld_assignment_detail_id,
        awb_master_id=payload.awb_master_id,
        sequence_nos=payload.sequence_nos,
        loaded_by=current_user.emp_id,
    )



# ====================== Manual awb creation ===========
@router.post(
    "/awb/manual-create",
    response_model=AwbManualCreateResponse,
    status_code=201,
    summary="Create AWB manually",
)
async def create_manual_awb(
    payload: AwbManualCreateRequest,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    return await create_manual_awb_service(
        db=db,
        data=payload,
        emp_id=current_user.emp_id,
    )





# =============== Search awb across all table FOR web table not mobile app ================

from sqlalchemy import select, func
from sqlalchemy import inspect

def model_to_dict(obj):
    return {
        c.key: getattr(obj, c.key)
        for c in inspect(obj).mapper.column_attrs
    }
@router.get(
    "/car-message/awb-search-for-web",
    summary="Search AWB by exact match across all records for web table",
)
async def search_awb_across_all(
    db: AsyncSession = Depends(get_db),
    awb_no: str = Query(...),
):
    awb_no = awb_no.strip()
    if not awb_no:
        raise HTTPException(status_code=400, detail="AWB number is required.")

    # ── Normalize ─────────────────────────────────────────────────────────────
    cleaned_awb = awb_no.replace("-", "").replace(" ", "")
    if len(cleaned_awb) == 10:
        cleaned_awb = "0" + cleaned_awb
    if len(cleaned_awb) != 11:
        raise HTTPException(status_code=400, detail=f"Invalid AWB number: {awb_no}")

    # ── 1. Fetch AWB master ────────────────────────────────────────────────────
    awb_record = (
        await db.execute(
            select(ExportCarMessageAwbMaster)
            .where(ExportCarMessageAwbMaster.awb_no == cleaned_awb)
        )
    ).scalars().first()

    if not awb_record:
        raise HTTPException(status_code=404, detail=f"AWB {cleaned_awb} not found.")

    # ── 2. Fetch flight bookings + header in ONE join query ────────────────────
    # No selectinload needed — explicit join, full control
    flight_rows = (
        await db.execute(
            select(ExportFlightBookingDetail, ExportFlightBookingHeader)
            .join(
                ExportFlightBookingHeader,
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id
            )
            .where(ExportFlightBookingDetail.awb_master_id == awb_record.id)
        )
    ).all()  # returns list of (detail_row, header_row) tuples

    # ── 3. Count scanned pcs ──────────────────────────────────────────────────
    scanned_pcs: int = (
        await db.execute(
            select(func.count())
            .where(ExportAwbSkidItemSequence.awb_master_id == awb_record.id)
        )
    ).scalar_one()

    # ── 3b. Resolve user names for ultrafast / manual ──────────────────────────
    emp_ids = [
        e for e in (awb_record.is_ultra_fast_marked_by, awb_record.manual_created_by)
        if e
    ]
    name_by_emp: dict[str, str] = {}
    if emp_ids:
        user_rows = (
            await db.execute(
                select(User.emp_id, User.name).where(User.emp_id.in_(emp_ids))
            )
        ).all()
        name_by_emp = {emp_id: name for emp_id, name in user_rows}

    # ── 4. Shape response ─────────────────────────────────────────────────────
    flight_bookings = [
        {
            "flight_no":           header.flight_no,
            "flight_date":         header.flight_date.isoformat(),
            "flight_dpt_datetime": header.flight_dpt_datetime.isoformat(),
            "flight_is_active":    header.is_active,
            "booked_pcs":          detail.booked_pcs,
        }
        for detail, header in flight_rows
    ]

    total_booked_pcs = sum(d["booked_pcs"] for d in flight_bookings)

    return {
        "status": "success",
        "data": {
        # ✅ Spreads all AWB columns flat — exactly like before
        **model_to_dict(awb_record),

        # ✅ New fields just sit alongside
        "scanned_pcs":      scanned_pcs,
        "total_booked_pcs": total_booked_pcs,
        "flight_bookings":  flight_bookings,

        # resolved names
            "ultra_fast_by_name": (
                name_by_emp.get(awb_record.is_ultra_fast_marked_by)
                if awb_record.is_ultra_fast_marked_by else None
            ),
            "manual_created_by_name": (
                name_by_emp.get(awb_record.manual_created_by)
                if awb_record.manual_created_by else None
            ),
    },
    }

# @router.get(
#     "/car-message/awb-search-for-web",
#     summary="Search AWB by exact match across all records for web table",
# )
# async def search_awb_across_all(
#     db: AsyncSession = Depends(get_db),   # ✅ FIX
#     awb_no: str = Query(...),             # ✅ also better
# ):

#     awb_no = awb_no.strip()

#     if not awb_no:
#         raise HTTPException(status_code=400, detail="AWB number is required.")

#     # ── Normalize: strip non-digits, pad to 11 ────────────────────────────────
#     cleaned_awb = awb_no.replace("-", "").replace(" ", "")
#     if len(cleaned_awb) == 10:
#         cleaned_awb = "0" + cleaned_awb
#     if len(cleaned_awb) != 11:
#         raise HTTPException(status_code=400, detail=f"Invalid AWB number: {awb_no}")
    
#     # 2. Database Query
#     # Since awb_no is indexed and has a UniqueConstraint, this will be very fast.
#     stmt = select(ExportCarMessageAwbMaster).where(
#         ExportCarMessageAwbMaster.awb_no == cleaned_awb
#     )
#     result = await db.execute(stmt)
#     awb_record = result.scalars().first()

#     # 3. Return Response
#     if not awb_record:
#         raise HTTPException(
#             status_code=404, 
#             detail=f"AWB {cleaned_awb} not found."
#         )

#     return {
#         "status": "success",
#         "data": awb_record
#     }



#  =============== 🫥 INDIVIDUAL ULD CLOSING or Unclosing  PER FLIGHTS PER DATE ====================

@router.post("/uld/{uld_detail_id}/close")
async def close_uld(
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    return await close_per_uld__per_flight_service(
        db=db,
        uld_assignment_detail_id=uld_detail_id,
        closed_by=current_user.emp_id,
    )

@router.post("/uld/{uld_detail_id}/open")
async def unlock_uld(
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    return await unlock_per_uld_service(
        db=db,
        uld_assignment_detail_id=uld_detail_id,
        unlocked_by=current_user.emp_id,
    )

@router.get("/uld/{uld_no}/get-latest-uld-assignment")
async def get_latest_close_uld_assignment(
    uld_no: str,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):
    return await get_latest_uld_assignment_service(
        db=db,
        uld_no=uld_no,
    )

# ========================== FIGHT HISTORY RELATED ROUTES =======================================

@router.get("/flight/by-date-for-history", 
            summary="Get all flights on a particular date for flight history screen"
            )
async def get_flights(date: date, db: AsyncSession = Depends(get_db)):
    return await get_flight_history(db, date)

@router.get("/flight/{flight_id}/awb-breakdown")
async def flight_awb_breakdown(
    flight_id: int,
    db: AsyncSession = Depends(get_db),
):
    return await get_flight_awb_breakdown(db, flight_id)


@router.get("/flight/{flight_id}/partial-flight-detail",summary="Get partial flight detail for flight history screen (for main listing)")
async def get_flight_detail(flight_id: int, db: AsyncSession = Depends(get_db)):
    return await get_flight_particular_flight_detail(db, flight_id)

@router.get("/flight/{flight_id}/uld/{uld_detail_id}/history-sequences",summary="Get sequences of a particular ULD in a particular flight for flight history screen")
async def get_uld_sequences(
    flight_id: int,
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db)
):
    return await get_uld_sequences_of_particular_flight(db, flight_id, uld_detail_id)


# ============================= LOADING SHEET rEPORT ==============================================
@router.get("/flight/{flight_id}/uld/{uld_detail_id}/loading-sheet")
async def get_loading_sheet(
    flight_id: int,
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db),
):
    return await get_loading_sheet_report_service(
        db,
        flight_id,
        uld_detail_id
    )






# ─────────────────────────────────────────────────────────────────────────────
#  SAVE — POST /loading-sheet-form/save
# ─────────────────────────────────────────────────────────────────────────────
 
@router.post("/loading-sheet-form/save")
async def save_loading_sheet_form(
    flight_id: int = Body(...),
    uld_detail_id: int = Body(...),
    form_data: dict = Body(...),
    
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user)
):
    """
    Save (or update) the loading sheet form for a Flight + ULD combo.
 
    Request body:
    ```json
    {
        "flight_id": 14,
        "uld_detail_id": 19,
        "saved_by": "523520",
        "form_data": {
            "operational": { ... },
            "user_input": { ... }
        }
    }
    ```
 
    - First save → creates new row + first history entry
    - Subsequent saves → updates live row + appends new history entry
    - API snapshot is auto-captured at save time for audit trail
    """
    result = await save_loading_sheet_form_service(
        db=db,
        flight_id=flight_id,
        uld_detail_id=uld_detail_id,
        form_data=form_data,
        saved_by=current_user.emp_id,
    )
    return result
 
 
# ─────────────────────────────────────────────────────────────────────────────
#  LOAD — GET /get-saved/loading-sheet-form-data/{flight_id}/{uld_detail_id}
# ─────────────────────────────────────────────────────────────────────────────
 
@router.get("/get-saved/loading-sheet-form-data/{flight_id}/{uld_detail_id}")
async def get_loading_sheet_form(
    flight_id: int,
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Load the saved loading sheet form for a Flight + ULD combo.
 
    Returns:
    ```json
    {
        "is_saved": true,
        "form_data": { "operational": {...}, "user_input": {...} },
        "last_saved_at": "2026-04-29T16:30:00+05:30",
        "last_saved_by": "523520"
    }
    ```
 
    If no form was saved yet:
    ```json
    {
        "is_saved": false,
        "form_data": null,
        "last_saved_at": null,
        "last_saved_by": null
    }
    ```
    """
    result = await get_loading_sheet_form_service(
        db=db,
        flight_id=flight_id,
        uld_detail_id=uld_detail_id,
    )
    return result
 
 
# ─────────────────────────────────────────────────────────────────────────────
#  HISTORY — GET /loading-sheet-form/{flight_id}/{uld_detail_id}/history
# ─────────────────────────────────────────────────────────────────────────────
 
@router.get("/loading-sheet-form/{flight_id}/{uld_detail_id}/history")
async def get_loading_sheet_form_history(
    flight_id: int,
    uld_detail_id: int,
    db: AsyncSession = Depends(get_db),
):
    """
    Get the complete save history for a loading sheet form.
    Returns all snapshots, newest first.
 
    Each entry contains:
    - `form_data`: what the form looked like at that save
    - `api_snapshot`: what the API returned at that moment
    - `saved_at`: when the save happened
    - `saved_by`: who saved it
    - `save_type`: "first" | "manual" | "auto"
 
    Example response:
    ```json
    [
        {
            "id": 12,
            "form_data": { "operational": {...}, "user_input": {...} },
            "api_snapshot": { "flight_info": {...}, "awb_list": [...], ... },
            "saved_at": "2026-04-29T16:45:00+05:30",
            "saved_by": "523520",
            "save_type": "manual"
        },
        {
            "id": 8,
            "form_data": { ... },
            "api_snapshot": { ... },
            "saved_at": "2026-04-29T16:30:00+05:30",
            "saved_by": "523520",
            "save_type": "first"
        }
    ]
    ```
    """
    result = await get_loading_sheet_form_history_service(
        db=db,
        flight_id=flight_id,
        uld_detail_id=uld_detail_id,
    )
    return result
 



@router.post(
    "/flight/{flight_header_id}/uld-assignment/assign-ulds/mobile",
    response_model=UldAssignmentResponse,
)
async def assign_ulds_to_flight_mobile_route(
    flight_header_id: int,
    payload: AddUldsRequest,           # { uld_ids: List[int] }
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    
    print(f"assign_ulds_to_flight_mobile_route called with: flight_header_id={flight_header_id}, payload={payload}")
    return await assign_ulds_to_flight_mobile(
        db=db,
        flight_header_id=flight_header_id,
        uld_ids=payload.uld_ids,
        uld_detail_ids_to_remove=payload.uld_detail_ids_to_remove,  # NEW: list of uld_detail_ids to unassign
        performed_by=current_user.emp_id,
    )



@router.get("/awb/{awb_no}/missing-sequence-status")
async def fetch_awb_sequence_status_route(
    awb_no: str,
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    """
    Get sequence number status for an AWB:
      - loaded:       scanned and loaded on a ULD
      - scanned_only: scanned but not yet loaded
      - missing:      in valid pcs range but never scanned
      - extra:        scanned but outside the valid pcs range (data sanity)
    """
    print(f"fetch_awb_sequence_status_route called with: awb_no={awb_no}")
    return await get_awb_missing_sequence_status(
        db=db,
        awb_no=awb_no.strip().upper(),
    )


@router.get("/trace-skid-flight-by-seq/{sequence_no}")
async def trace_sequence_route(
    sequence_no: str,
    db: AsyncSession = Depends(get_db),
    current_user: UserRead = Depends(verify_token_and_get_user),
):
    """
    Trace a sequence_no:
      - Which skid it was scanned onto
      - Which flight it was actually loaded onto (if loaded on ULD)
      - All booked/candidate flights for the AWB (AWB may split across flights)
    Each flight includes its departure datetime (flight_dpt_datetime in UTC).
    """
    print(f"trace_sequence_route called with: sequence_no={sequence_no}")
    return await trace_sequence_full_info(
        db=db,
        sequence_no=sequence_no.strip().upper(),
    )



# ========= Uld loading report shift wise ===============
@router.get("/loading-report/shift-wise")
async def loading_report(
    from_dt: datetime = Query(..., description="Start datetime (ISO 8601), e.g. 2024-06-01T00:00:00"),
    to_dt: datetime   = Query(..., description="End datetime   (ISO 8601), e.g. 2024-06-01T23:59:59"),
    db: AsyncSession  = Depends(get_db),
):
    """
    Returns a streamed Excel (.xlsx) file containing all ULD loading records
    whose loaded_at timestamp falls within [from_dt, to_dt].
 
    Columns: Flight No | Flight Date | AWB No | Sequence No | Scanned By |
             ULD No | Loaded By | Loaded At (IST)
    """
    filename = (
        f"loading_report_{from_dt.strftime('%Y%m%d_%H%M')}_{to_dt.strftime('%Y%m%d_%H%M')}.xlsx"
    )
 
    stream = await generate_loading_report_excel_shift_wise(db, from_dt, to_dt)
 
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )






# =============================✌️ AWB Checker INFO on Flight booking AWB selection time =======================

@router.get(
    "/check-individual-awb-for-flight-booking",
    response_model=AwbCheckerInfoResponse,
    summary="AWB status, piece counts and per-flight booking breakdown",
)
async def check_awb_availability(
    awb_no: str = Query(
        ...,
        min_length=11,
        max_length=11,
        pattern=r"^\d{11}$",
        description="11-digit numeric AWB number",
        example="09830280051",
    ),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
) -> AwbCheckerInfoResponse:
    return await get_awb_checker_info(awb_no=awb_no, db=db)








@router.put("/flight-booking/update-departure-datetime/{header_id}")
async def update_flight_dpt_datetime(
    header_id: int,
    payload: UpdateFlightDptDatetime,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(verify_token_and_get_user),
):

    def to_utc(dt: datetime) -> datetime:
        if dt.tzinfo is not None:
            return dt.astimezone(timezone.utc)
        ist_offset = timedelta(hours=5, minutes=30)
        return (dt - ist_offset).replace(tzinfo=timezone.utc)

    result = await db.execute(
        select(ExportFlightBookingHeader).where(
            ExportFlightBookingHeader.id == header_id,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    header = result.scalar_one_or_none()

    if not header:
        raise HTTPException(status_code=404, detail="Flight booking not found")

    now = get_utc_now()

    # Always normalize input to aware UTC (handles both naive-IST and aware input)
    new_dpt_utc = to_utc(payload.flight_dpt_datetime)

    # Block editing if already departed
    # if header.flight_dpt_datetime <= now:
    #     raise HTTPException(
    #         status_code=400,
    #         detail="Flight has already departed — departure time cannot be edited",
    #     )

    # New departure cannot be in the past
    if new_dpt_utc <= now:
        raise HTTPException(
            status_code=400,
            detail="New departure time cannot be in the past",
        )

    # ✅ Must be within 3 days of booking time (compare UTC vs UTC)
    max_allowed = header.booked_at + timedelta(days=3)
    if new_dpt_utc > max_allowed:
        raise HTTPException(
            status_code=400,
            detail=f"Departure must be within 3 days of booking (by {max_allowed.isoformat()})",
        )

    header.flight_dpt_datetime = new_dpt_utc
    header.updated_at = now

    await db.commit()
    await db.refresh(header)

    return {
        "success": True,
        "message": "Departure datetime updated successfully",
        "flight_dpt_datetime": header.flight_dpt_datetime.isoformat(),
    }



# ========== report of booked flights ==============================

@router.get("/flight-booking/report-download")
async def download_flight_booking_report(
    start_date: date = Query(...),
    end_date: date = Query(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Excel report for all flights with flight_date between start_date and end_date
    (inclusive). Flights ordered newest date first, then by departure time.
    Flight details repeat on every AWB row.
    """
    IST = timezone(timedelta(hours=5, minutes=30))

    def _to_ist(dt):
        if dt is None:
            return None
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(IST)

    if end_date < start_date:
        raise HTTPException(status_code=400, detail="end_date cannot be before start_date")

    stmt = (
        select(
            ExportFlightBookingHeader.id.label("header_id"),
            ExportFlightBookingHeader.flight_no,
            ExportFlightBookingHeader.flight_date,
            ExportFlightBookingHeader.flight_dpt_datetime,
            ExportFlightBookingHeader.is_active,
            ExportFlightBookingHeader.booked_by,
            ExportFlightBookingHeader.booked_at,
            User.name.label("booked_by_name"),
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("awb_total_pcs"),
            ExportCarMessageAwbMaster.gross_wt,
        )
        .outerjoin(User, User.emp_id == ExportFlightBookingHeader.booked_by)
        .outerjoin(
            ExportFlightBookingDetail,
            ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
        )
        .outerjoin(
            ExportCarMessageAwbMaster,
            ExportCarMessageAwbMaster.id == ExportFlightBookingDetail.awb_master_id,
        )
        .where(ExportFlightBookingHeader.flight_date >= start_date)
        .where(ExportFlightBookingHeader.flight_date <= end_date)
        .where(ExportFlightBookingHeader.is_active.is_(True))
        .order_by(
            ExportFlightBookingHeader.flight_date.desc(),
            ExportFlightBookingHeader.flight_dpt_datetime.desc(),
            ExportCarMessageAwbMaster.awb_no.asc(),
        )
    )

    result = await db.execute(stmt)
    rows = result.all()

    if not rows:
        raise HTTPException(status_code=404, detail="No flights found for this date range")

    # ── Group rows by flight header, preserving the query order ────────
    grouped = []          # list of (header_row, [awb_rows...])
    seen = {}             # header_id -> index in grouped
    for r in rows:
        if r.header_id not in seen:
            seen[r.header_id] = len(grouped)
            grouped.append((r, []))
        if r.awb_no:      # skip the placeholder AWB cols from an empty outerjoin
            grouped[seen[r.header_id]][1].append(r)

    # ── Build workbook ────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Flight Booking Report"

    NCOLS = 8  # A..H

    title_font = Font(name="Arial", bold=True, size=14)
    range_font = Font(name="Arial", bold=True, size=11)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="1F4E78")
    cell_font = Font(name="Arial")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    last_col_letter = ws.cell(row=1, column=NCOLS).column_letter

    # Row 1: title (merged)
    ws.merge_cells(f"A1:{last_col_letter}1")
    c = ws["A1"]
    c.value = "Flight Booking Report"
    c.font = title_font
    c.alignment = center

    # Row 2: date range (merged)
    ws.merge_cells(f"A2:{last_col_letter}2")
    c = ws["A2"]
    c.value = f"{start_date.strftime('%d-%b-%Y')}  to  {end_date.strftime('%d-%b-%Y')}"
    c.font = range_font
    c.alignment = center

    # Row 3: blank spacer

    # Row 4: header
    headers = [
        "Flight No", "Flight Date", "Departure (IST)",
        "Booked By (Emp ID)", "Booked By (Name)", "Booked At",
        "AWB No", "Total Pcs", "Booked Pcs",
    ]
    head_row = 4
    for col, h in enumerate(headers, start=1):
        cc = ws.cell(row=head_row, column=col, value=h)
        cc.font = header_font
        cc.fill = header_fill
        cc.alignment = center
        cc.border = border

    # ── Data ───────────────────────────────────────────────────────────
    DATE_FMT = "dd-mmm-yyyy"
    DATETIME_FMT = "dd-mmm-yyyy hh:mm"

    cur = head_row + 1
    for header_r, awb_list in grouped:
        dep_ist = _to_ist(header_r.flight_dpt_datetime)
        dep_val = dep_ist.replace(tzinfo=None) if dep_ist else None  # openpyxl: no tz-aware
        flight_date_val = header_r.flight_date  # already a date object

        booked_at_ist = _to_ist(header_r.booked_at)
        booked_at_val = booked_at_ist.replace(tzinfo=None) if booked_at_ist else None

        def write_flight_row(awb=None):
            # flight columns — repeated on every row
            ws.cell(row=cur, column=1, value=header_r.flight_no).font = cell_font

            c2 = ws.cell(row=cur, column=2, value=flight_date_val)
            c2.font = cell_font
            if flight_date_val:
                c2.number_format = DATE_FMT

            c3 = ws.cell(row=cur, column=3, value=dep_val)
            c3.font = cell_font
            if dep_val:
                c3.number_format = DATETIME_FMT

            ws.cell(row=cur, column=4, value=header_r.booked_by or "").font = cell_font
            ws.cell(row=cur, column=5, value=header_r.booked_by_name or "").font = cell_font

            c6 = ws.cell(row=cur, column=6, value=booked_at_val)
            c6.font = cell_font
            if booked_at_val:
                c6.number_format = DATETIME_FMT

            # AWB columns
            ws.cell(row=cur, column=7, value=(awb.awb_no if awb else "") or "").font = cell_font

            total_pcs_cell = ws.cell(row=cur, column=8)
            if awb and awb.awb_total_pcs is not None:
                total_pcs_cell.value = int(awb.awb_total_pcs)
                total_pcs_cell.number_format = "0"
            total_pcs_cell.font = cell_font

            pcs_cell = ws.cell(row=cur, column=9)
            if awb and awb.booked_pcs is not None:
                pcs_cell.value = int(awb.booked_pcs)
                pcs_cell.number_format = "0"
            pcs_cell.font = cell_font

            # borders + alignment for all 8 columns
            for col in range(1, NCOLS + 1):
                cell = ws.cell(row=cur, column=col)
                cell.border = border
                cell.alignment = left if col <= 5 else center

        if not awb_list:
            write_flight_row(None)
            cur += 1
            continue

        for a in awb_list:
            write_flight_row(a)
            cur += 1

    # ── Column widths + freeze ─────────────────────────────────────────
    widths = [12, 13, 18, 18, 22, 18, 14, 11, 11]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=head_row, column=i).column_letter].width = w
    # ws.freeze_panes = "A5"

    # ── Stream out ─────────────────────────────────────────────────────
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"flight_booking_report_{start_date.isoformat()}_to_{end_date.isoformat()}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )



@router.patch("/{awb_no}/update-pieces")
async def update_awb_pieces(
    awb_no: str,
    payload: UpdateAwbPiecesRequest,
    db: AsyncSession = Depends(get_db),
    current_user = Depends(verify_token_and_get_user)
):
    # 1. Fetch AWB record
    stmt = select(ExportCarMessageAwbMaster).where(ExportCarMessageAwbMaster.awb_no == awb_no)
    result = await db.execute(stmt)
    awb_record = result.scalars().first()

    if not awb_record:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"AWB {awb_no} not found"
        )

    # 2. Rule check
    if awb_record.source == "CAR_MESSAGE" or awb_record.source is None:
        raise HTTPException(
            status_code=400,
            detail="Shipments with source file as 'CAR_MESSAGE' (or missing source) cannot be edited."
        )

    # 3. Validation: new pieces >= current pieces
    current_pcs = awb_record.pcs if awb_record.pcs is not None else 0
    if payload.pcs < current_pcs:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"New pieces ({payload.pcs}) cannot be less than current pieces ({current_pcs})."
        )

    # ── capture OLD value BEFORE mutating ──────────────────────
    old_pcs = current_pcs   # already holds the pre-update value

    # 4. Apply update
    awb_record.pcs = payload.pcs
    awb_record.updated_at = datetime.now(timezone.utc)
    awb_record.updated_pieces_by = current_user.emp_id
    awb_record.updated_pcs_at = get_utc_now()

   
    # ── write audit log (individual AWB op → no flight id) ─────
    await write_car_message_flow_audit(
        db=db,
        awb_reference_id=awb_record.id,        # ← PK is `id`
        flight_reference_id=None,              # individual op, no flight
        module=CarMessageFlowModule.AWB_MASTER,
        flow_step=CarMessageFlowStep.AWB_MASTER,
        record_id=awb_record.id,               # ← same `id`
        action="UPDATE_AWB_PIECES",
        performed_by=current_user.emp_id,
        changes={
            "pcs": {"old": old_pcs, "new": payload.pcs}
        },
        note=f"Pieces updated for AWB {awb_no}",
    )

    # 5. Commit (record change + audit log together)
    try:
        await db.commit()
        await db.refresh(awb_record)
    except Exception as e:
        await db.rollback()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Database error while updating pieces."
        )

    return {
        "status": "success",
        "message": f"Successfully updated pieces for AWB {awb_no}",
        "data": {
            "awb_no": awb_record.awb_no,
            "pcs": awb_record.pcs
        }
    }