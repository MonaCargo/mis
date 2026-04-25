# services/export_car_message_awb_service.py

from collections import defaultdict
from datetime import date, datetime, time, timedelta, timezone
import re
from typing import Optional
from fastapi import status
from zoneinfo import ZoneInfo

from fastapi import HTTPException
import openpyxl
import pytz
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import and_, case, distinct, func, or_, select, text, update
from sqlalchemy.dialects.postgresql import insert
from app.api.routes.domesticOperation.domestic_xray_report import convert_ist_day_to_utc_range
from app.db.models.exportOperation.car_message import (
    ExportAwbSkidItemSequence,
    ExportAwbSkidMapping,
    ExportCarMessageAwbMaster,
    ExportFlightBookingDetail,
    ExportFlightBookingHeader,
    ExportSequenceItemUldLoading,
    ExportSkidBaseMapping,
    ExportSkidLocationMapping,
    ExportUldAssignment,
    ExportUldAssignmentDetail
)
from app.db.models.exportOperation.export_base_master import ExportBaseMaster
from app.db.models.exportOperation.export_carrier_master import ExportCarrierMaster
from app.db.models.exportOperation.export_location_master import ExportLocationsMaster
from app.db.models.exportOperation.export_skid_master import ExportSkidMaster
from app.db.models.exportOperation.export_uld_master import ExportUldMaster
from app.db.models.user import User
from app.schemas.exportOperation.car_message import AvailableAwbForFlightBookingResponse, AwbChangeRecord, AwbDaySummary, AwbLoadingStatusItem, AwbLookupError, AwbManualCreateRequest, CreateFlightBookingRequest, CreateFlightBookingResponse, CreateUldAssignmentRequest, DashboardStatsResponse, EditFlightBookingRequest, EditFlightBookingResponse, EditUldAssignmentRequest, FlightBookingAwbItem, FlightBookingByFlightResponse, FlightBookingDetailResponse, FlightBookingDetailWithAwbResponse, FlightUldLoadingStatusResponse, PdfUpsertResponse, ScanItemIntoUldRequest, ScanItemIntoUldResponse, ScanItemResult, ScanningDaySummary, SkidDaySummary, UldAssignmentDataResponse, UldAssignmentDetailResponse, UldAssignmentResponse, UldLoadingStatusItem, UldMasterResponse, UldVerifyForLoadingResponse
from app.services.exportOperation.car_message_flow_audit_log import write_car_message_flow_audit
from app.utils.common.car_message_flow_audit_utils import CarMessageFlowModule, CarMessageFlowStep
from app.utils.common.helperFunction import get_utc_now


# ── Constants ─────────────────────────────────────────────────────────────────
FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING = {"RCS", "TFD", "RCT"} 

# ── 😎 reusable booked_pcs subquery }───────────────────────────────
def _booked_pcs_subquery():
    return (
        select(
            ExportFlightBookingDetail.awb_master_id,
            func.sum(ExportFlightBookingDetail.booked_pcs).label("booked_pcs"),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
            ),
        )
        .group_by(ExportFlightBookingDetail.awb_master_id)
        .subquery()
    )


def to_utc(dt: datetime) -> datetime:
    if dt.tzinfo is not None:
        return dt.astimezone(timezone.utc)
    ist_offset = timedelta(hours=5, minutes=30)
    return (dt - ist_offset).replace(tzinfo=timezone.utc)


# def _build_skid_activity_log(
#     skid: dict,
#     mapping_row,
#     location_history: list,
#     base_drop: dict | None = None, 
# ) -> list[dict]:

#     activity = []

#     # ── 1. Skid assigned to AWB ────────────────────────────────
#     activity.append({
#         "action": "SKID_ASSIGNED",
#         "label": "Skid assigned to AWB",
#         "performed_by": skid.get("mapped_by"),
#         "timestamp": mapping_row.mapping_created_at,
#         "detail": {
#             "skid_no": skid.get("skid_no"),
#             "virtual_skid_no": skid.get("virtual_skid_no"),
#             "is_virtual": skid.get("is_virtual"),
#         },
#     })

#     # ── 2. Retrieved — most recent is_current=False + picked_at set ──
#     retrieved = [
#         loc for loc in location_history
#         if not loc.is_current and
#         loc.picked_at and loc.picked_by
#          and not loc.is_relocation
      
#     ]

#     if retrieved:
#         # ✅ most recent retrieval only
#         most_recent = max(retrieved, key=lambda x: x.picked_at)
#         activity.append({
#             "action": "RETRIEVED_FROM_LOCATION",
#             "label": f"Retrieved from {most_recent.area_code} — {most_recent.loc}",
#             "performed_by": most_recent.picked_by,
#             "timestamp": most_recent.picked_at,
#             "detail": {
#                 "location_code": most_recent.area_code,
#                 "location_name": most_recent.loc,
#             },
#         })

#        # ── 3. Dropped at base ─────────────────────────────────────
#     # if base_drop:
#     #     activity.append({
#     #         "action": "DROPPED_AT_BASE",
#     #         "label": f"Dropped at base — {base_drop['base_name']}",
#     #         "performed_by": base_drop["dropped_by"],
#     #         "timestamp": base_drop["dropped_at"],
#     #         "detail": {
#     #             "base_id": base_drop["base_id"],
#     #             "base_name": base_drop["base_name"],
#     #         },
#     #     })

#     # ── BASE DROP EVENTS (all cycles) ─────────────────────────
#     # base_drop is now a list (multiple cycles possible)
#     if base_drop:
#         base_drops = base_drop if isinstance(base_drop, list) else [base_drop]
#         for drop in base_drops:
#             activity.append({
#                 "action": "DROPPED_AT_BASE",
#                 "label": f"Dropped at base — {drop.get('base_name')} (cycle {drop.get('cycle_no', 1)})",
#                 "performed_by": drop.get("dropped_by"),
#                 "timestamp": drop.get("dropped_at"),
#                 "detail": {
#                     "base_id": drop.get("base_id"),
#                     "base_name": drop.get("base_name"),
#                     "cycle_no": drop.get("cycle_no", 1),   # ✅ shows cycle
#                 },
#             })


#     # ── sort by timestamp ──────────────────────────────────────
#     activity.sort(
#         key=lambda x: x["timestamp"] if x["timestamp"]
#         else datetime.min.replace(tzinfo=timezone.utc)
#     )

#     return activity

def _build_skid_activity_log(
    skid: dict,
    mapping_row,
    location_history: list,
    base_drop: list | dict | None = None,
) -> list[dict]:

    activity = []

    # ── 1. Skid assigned ──────────────────────────────────────
    activity.append({
        "action": "SKID_ASSIGNED",
        "label": "Skid assigned to AWB",
        "performed_by": skid.get("mapped_by"),
        "timestamp": mapping_row.mapping_created_at,
        "detail": {
            "skid_no": skid.get("skid_no"),
            "virtual_skid_no": skid.get("virtual_skid_no"),
            "is_virtual": skid.get("is_virtual"),
        },
    })

    # ── 2. ALL location events in chronological order ──────────
    # Each location row gives us:
    #   assigned_at → PLACED / RELOCATED event
    #   picked_at   → RETRIEVED event (if set)
    for loc in location_history:

        # placed or relocated
        activity.append({
            "action": "RELOCATED_TO_LOCATION" if loc.is_relocation else "PLACED_AT_LOCATION",
            "label": (
                f"Relocated to {loc.area_code} — {loc.loc}"
                if loc.is_relocation
                else f"Placed at {loc.area_code} — {loc.loc}"
            ),
            "performed_by": loc.assigned_by,
            "timestamp": loc.assigned_at,
            "detail": {
                "location_code": loc.area_code,
                "location_name": loc.loc,
                "is_relocation": loc.is_relocation,
            },
        })

        # retrieved from this location (if picked)
        if loc.picked_at and loc.picked_by:
            activity.append({
                "action": "RETRIEVED_FROM_LOCATION",
                "label": f"Retrieved from {loc.area_code} — {loc.loc}",
                "performed_by": loc.picked_by,
                "timestamp": loc.picked_at,
                "detail": {
                    "location_code": loc.area_code,
                    "location_name": loc.loc,
                },
            })

    # ── 3. ALL base drop events (multi-cycle) ─────────────────
    if base_drop:
        base_drops = base_drop if isinstance(base_drop, list) else [base_drop]
        for drop in base_drops:
            activity.append({
                "action": "DROPPED_AT_BASE",
                "label": f"Dropped at base — {drop.get('base_name')} (cycle {drop.get('cycle_no', 1)})",
                "performed_by": drop.get("dropped_by"),
                "timestamp": drop.get("dropped_at"),
                "detail": {
                    "base_id": drop.get("base_id"),
                    "base_name": drop.get("base_name"),
                    "cycle_no": drop.get("cycle_no", 1),
                },
            })

    # ── sort all events by timestamp ───────────────────────────
    activity.sort(
        key=lambda x: x["timestamp"] if x["timestamp"]
        else datetime.min.replace(tzinfo=timezone.utc)
    )

    return activity

# def _get_skid_retrieval_status(
#     location_history: list,
#     base_drop: dict | None,
# ) -> str:
#     if base_drop:
#         return "AT_BASE"
#     if location_history:
#         most_recent = max(location_history, key=lambda x: x.assigned_at)
#         if not most_recent.is_current and most_recent.picked_at:
#             return "RETRIEVED"
#     return "PENDING"

# 🤢
def _get_skid_retrieval_status(location_history, base_drop) -> str:
    if not location_history:
        return "PENDING"

    base_drops = base_drop if isinstance(base_drop, list) else ([base_drop] if base_drop else [])

    last_loc = sorted(location_history, key=lambda x: x.assigned_at)[-1]

    if base_drops:
        last_drop = sorted(base_drops, key=lambda x: x.get("dropped_at"))[-1]
        last_drop_at = last_drop.get("dropped_at")
        last_picked_at = last_loc.picked_at

        if last_picked_at and last_drop_at and last_drop_at >= last_picked_at:
            return "AT_BASE"

    if last_loc.picked_at:
        return "RETRIEVED"

    if last_loc.is_current:
        return "AT_LOCATION"

    return "PENDING"

# COMMON PRIVATE FUN END ---------------------------------------------------




async def save_export_car_message_awbs(db: AsyncSession, df, uploaded_by: str = None, ):

    if df.empty:
        return {
            "total_received": 0,
            "inserted": 0,
            "already_present": 0,
        }

    records = df.to_dict(orient="records")

    # 🔥 Add timestamps
    now = get_utc_now()

    for record in records:
        record["created_at"] = now
        record["updated_at"] = now
        record["uploaded_by"] = uploaded_by 

    stmt = insert(ExportCarMessageAwbMaster).values(records)

    # stmt = stmt.on_conflict_do_nothing(
    #     constraint="uq_awb_car_msg"
    # )

    # 🤢🤮
    stmt = stmt.on_conflict_do_update(
    constraint="uq_awb_car_msg",
    set_={
        "pcs": stmt.excluded.pcs,
        "gross_wt": stmt.excluded.gross_wt,
        "hwb_no": stmt.excluded.hwb_no,
        "volumetric_wt": stmt.excluded.volumetric_wt,
        "chg_wt": stmt.excluded.chg_wt,

        "nog":                      stmt.excluded.nog,   # ✅ was missing
        "shc":                      stmt.excluded.shc,   # ✅ was missing

            "car_msg_date": stmt.excluded.car_msg_date,
        "car_msg_time": stmt.excluded.car_msg_time,

        "origin": stmt.excluded.origin,
        "destination": stmt.excluded.destination,

        "sb_no": stmt.excluded.sb_no,
        "sb_date": stmt.excluded.sb_date,

    
        "car_message_datetime_combo": stmt.excluded.car_message_datetime_combo,
        # "rcs_datetime": stmt.excluded.rcs_datetime,

        

        # "status": stmt.excluded.status,
        # "agent": stmt.excluded.agent,

        "updated_at": now,
    },
    where=(ExportCarMessageAwbMaster.is_manually_created == True)
    ).returning(
    text("(xmax = 0) AS is_inserted")   # ✅ True = inserted, False = updated
)


    result = await db.execute(stmt)
    # await db.commit()
    await db.flush()               # ✅ flush only — gets rowcount without committing

    rows = result.fetchall()

    # inserted_count = result.rowcount or 0
    # total_received = len(records)
    # already_present = total_received - inserted_count

    inserted_count = sum(1 for r in rows if r[0] is True)
    updated_count = sum(1 for r in rows if r[0] is False)
    total_received = len(records)
    skipped_count = total_received - len(rows)  # where clause was False → not updated

    return {
        "total_received": total_received,
        "inserted": inserted_count,
        "updated": updated_count,
        "already_present": skipped_count,
    }



# ✌️======== Extract awb data from give export wh_inventry pdf for car message =================================
async def enrich_awb_from_wh_inventory(db: AsyncSession, df) -> dict:
    """
    Match extracted PDF rows against ExportCarMessageAwbMaster by awb_no
    and update: status, rcs_datetime, agent, vol_mc.

    Rules:
        - If master.status is already RCS → skip status, rcs_datetime, vol_mc
        - If not RCS yet → always update status, rcs_datetime, vol_mc
        - agent → update if currently NULL (always, regardless of status)
    """
    if df.empty:
        return {
            "total_in_pdf": 0,
            "matched": 0,
            "not_found": [],
            "not_found_count": 0,
        }

    # ── 1. Deduplicate — one row per AWB (latest datetime wins) ──────────────
    df = df.copy()
    df_unique = (
        df.sort_values("DATETIME")
          .drop_duplicates(subset="AWB", keep="last")
          .set_index("AWB")
    )
    pdf_awb_set = set(df_unique.index.tolist())

    # ── 2. Fetch matching master rows — single query ───────────────────────── 
    stmt = select(ExportCarMessageAwbMaster).where(
        ExportCarMessageAwbMaster.awb_no.in_(pdf_awb_set)
    )
    result  = await db.execute(stmt)
    masters = result.scalars().all()

    # # ── Constants ─────────────────────────────────────────────────────────────────
    # FINAL_STATUSES = {"RCS", "TFD", "RCT"} 

    # ── 3. Update fields ──────────────────────────────────────────────────────
    matched: set[str] = set()
    now = get_utc_now()

    for master in masters:
        if master.awb_no not in df_unique.index:
            continue

        row = df_unique.loc[master.awb_no]

        # agent — always update if currently NULL
        if master.agent is None:
            master.agent = _val(row.get("AGENT"))

        # If already RCS — final status, don't touch anything else
        # if master.status == "RCS":
        if master.status in FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING:
            pass

        else:
            # Not yet RCS — update status, datetime, vol_mc
            master.status       = _val(row.get("STATUS"))
            master.rcs_datetime = _to_datetime(row.get("DATETIME"))
            master.vol_mc       = _to_float(row.get("VOL_MC"))

        master.updated_at = now
        matched.add(master.awb_no)

    # # ── 4. Commit ─────────────────────────────────────────────────────────────
    # await db.commit()

     # ── 4. Flush only — route will commit ──────────────────────────────────
    await db.flush()   # ✅ replaces await db.commit()

    not_found = sorted(pdf_awb_set - matched)

    return {
        "total_in_pdf":    len(df_unique),
        "matched":         len(matched),
        "not_found":       not_found,
        "not_found_count": len(not_found),
    }
# ── Small helpers (private) ───────────────────────────────────────────────────
 
def _val(val):
    """pandas NA / empty string → None, else stripped string."""
    import pandas as pd
    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    return s or None
 
 
def _to_datetime(val):
    """Convert PDF datetime (IST naive) → UTC aware datetime for DB."""
    import pandas as pd
    import pytz
    from datetime import datetime

    IST = pytz.timezone("Asia/Kolkata")
    UTC = pytz.utc

    if val is None:
        return None
    try:
        if pd.isna(val):
            return None
    except (TypeError, ValueError):
        pass

    if isinstance(val, pd.Timestamp):
        dt = val.to_pydatetime().replace(tzinfo=None)
    elif isinstance(val, datetime):
        dt = val.replace(tzinfo=None)
    else:
        return None

    # Treat as IST → convert to UTC
    return IST.localize(dt).astimezone(UTC)

 
 
def _to_float(val):
    """Safe float conversion, None on failure."""
    try:
        return float(val)
    except (TypeError, ValueError):
        return None
    
#  =====------------ End of extract and save awb data from pdf whexport-inventry -------------------   


#✌️=======Get allowed awb for fligh booking screen dropdown =======================

# async def get_available_awbs_for_flight_booking_dropdown(
#     db: AsyncSession,
#     origin: Optional[str] = None,
#     destination: Optional[str] = None,
# ) -> list[AvailableAwbForFlightBookingResponse]:

#     # ── Subquery 1: booked pcs ─────────────────────────────
#     booked_subq = _booked_pcs_subquery()

#     # ── Subquery 2: scanned pcs per AWB ───────────────────
#     scanned_subq = (
#         select(
#             ExportAwbSkidItemSequence.awb_master_id,
#             func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
#         )
#         .group_by(ExportAwbSkidItemSequence.awb_master_id)
#         .subquery()
#     )

#     # ── Subquery 3: total distinct skids per AWB ───────────
#     total_skids_subq = (
#         select(
#             ExportAwbSkidMapping.awb_master_id,
#             func.count(ExportAwbSkidMapping.skid_id.distinct()).label("total_skids"),
#         )
#         .group_by(ExportAwbSkidMapping.awb_master_id)
#         .subquery()
#     )

#     # ── Subquery 4: located skids scoped to same AWB session
#     ever_located_subq = (
#         select(
#             ExportAwbSkidMapping.awb_master_id,
#             func.count(ExportAwbSkidMapping.skid_id.distinct()).label("ever_located_skids"),
#         )
#         .join(
#             ExportSkidLocationMapping,
#             and_(
#                 ExportSkidLocationMapping.skid_id == ExportAwbSkidMapping.skid_id,
#                 ExportSkidLocationMapping.awb_master_id == ExportAwbSkidMapping.awb_master_id,  # ✅ scope
#             ),
#         )
#         .group_by(ExportAwbSkidMapping.awb_master_id)
#         .subquery()
#     )

#     # ── Main query ─────────────────────────────────────────
#     remaining_pcs_expr = (
#         ExportCarMessageAwbMaster.pcs
#         - func.coalesce(booked_subq.c.booked_pcs, 0)
#     )

#     stmt = (
#         select(
#             ExportCarMessageAwbMaster.id.label("awb_master_id"),
#             ExportCarMessageAwbMaster.awb_no,
#             ExportCarMessageAwbMaster.origin,
#             ExportCarMessageAwbMaster.destination,
#             ExportCarMessageAwbMaster.pcs.label("total_pcs"),
#             func.coalesce(booked_subq.c.booked_pcs, 0).label("booked_pcs"),
#             remaining_pcs_expr.label("remaining_pcs"),
#             ExportCarMessageAwbMaster.agent,
#             ExportCarMessageAwbMaster.rcs_datetime,
#         )
#         .outerjoin(booked_subq, ExportCarMessageAwbMaster.id == booked_subq.c.awb_master_id)
#         .join(scanned_subq, ExportCarMessageAwbMaster.id == scanned_subq.c.awb_master_id)
#         .join(total_skids_subq, ExportCarMessageAwbMaster.id == total_skids_subq.c.awb_master_id)
#         .join(ever_located_subq, ExportCarMessageAwbMaster.id == ever_located_subq.c.awb_master_id)
#         .where(
#             ExportCarMessageAwbMaster.status == "RCS",
#             ExportCarMessageAwbMaster.pcs.isnot(None),
#             remaining_pcs_expr > 0,
#             # ✅ all pcs scanned
#             scanned_subq.c.scanned_pcs >= ExportCarMessageAwbMaster.pcs,
#             # ✅ ALL skids located at least once in this AWB session
#             ever_located_subq.c.ever_located_skids == total_skids_subq.c.total_skids,
#         )
#         .order_by(ExportCarMessageAwbMaster.rcs_datetime.desc())
#     )

#     if origin:
#         stmt = stmt.where(
#             ExportCarMessageAwbMaster.origin == origin.strip().upper()
#         )
#     if destination:
#         stmt = stmt.where(
#             ExportCarMessageAwbMaster.destination == destination.strip().upper()
#         )

#     result = await db.execute(stmt)
#     rows = result.mappings().all()

#     return [AvailableAwbForFlightBookingResponse(**row) for row in rows]

# 🤢
async def get_available_awbs_for_flight_booking_dropdown(
    db: AsyncSession,
    origin: Optional[str] = None,
    destination: Optional[str] = None,
) -> list[AvailableAwbForFlightBookingResponse]:

    # ── Subquery 1: booked pcs per AWB across active flights ───────
    booked_subq = _booked_pcs_subquery()

    # ── Subquery 2: scanned pcs per AWB ───────────────────────────
    scanned_subq = (
        select(
            ExportAwbSkidItemSequence.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .group_by(ExportAwbSkidItemSequence.awb_master_id)
        .subquery()
    )

    # ── remaining — scanned-based for normal, total_pcs-based for ultra-fast ──
    remaining_pcs_expr = case(
        (ExportCarMessageAwbMaster.is_ultra_fast == True,
         ExportCarMessageAwbMaster.pcs - func.coalesce(booked_subq.c.booked_pcs, 0)),
        else_=(
            func.coalesce(scanned_subq.c.scanned_pcs, 0)
            - func.coalesce(booked_subq.c.booked_pcs, 0)
        )
    )

    stmt = (
        select(
            ExportCarMessageAwbMaster.id.label("awb_master_id"),
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
            func.coalesce(scanned_subq.c.scanned_pcs, 0).label("scanned_pcs"),   # ✅ expose for frontend
            func.coalesce(booked_subq.c.booked_pcs, 0).label("booked_pcs"),
            remaining_pcs_expr.label("remaining_pcs"),
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportCarMessageAwbMaster.is_ultra_fast,
        )
        .outerjoin(booked_subq, ExportCarMessageAwbMaster.id == booked_subq.c.awb_master_id)
        .outerjoin(scanned_subq, ExportCarMessageAwbMaster.id == scanned_subq.c.awb_master_id)  # ✅ outerjoin so AWBs with 0 scanned still appear (filtered below)
        .where(
            # ExportCarMessageAwbMaster.status == "RCS",
            ExportCarMessageAwbMaster.status.in_(FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING),            ExportCarMessageAwbMaster.pcs.isnot(None),

             # ✅ ultra-fast bypasses scanned gate {ULTRA_FAST OR HAVE AT LEAST SCANNED ONE PCS}
            or_(
                ExportCarMessageAwbMaster.is_ultra_fast == True,
                func.coalesce(scanned_subq.c.scanned_pcs, 0) > 0,
            ),
            # func.coalesce(scanned_subq.c.scanned_pcs, 0) > 0,   # ✅ must have at least 1 scanned pc

            remaining_pcs_expr > 0,                              # ✅ scanned - booked > 0
        )
        .order_by(
            ExportCarMessageAwbMaster.is_ultra_fast.desc(),
            ExportCarMessageAwbMaster.rcs_datetime.desc(),
        )
    )

    if origin:
        stmt = stmt.where(
            ExportCarMessageAwbMaster.origin == origin.strip().upper()
        )
    if destination:
        stmt = stmt.where(
            ExportCarMessageAwbMaster.destination == destination.strip().upper()
        )

    result = await db.execute(stmt)
    rows = result.mappings().all()

    return [AvailableAwbForFlightBookingResponse(**row) for row in rows]

# ========= ✌️✌️  CREATE new flight booking ──────────────────────────────────────=========
async def create_flight_booking(
    db: AsyncSession,
    payload: CreateFlightBookingRequest,
    booked_by: str,  # emp_id from auth
) -> CreateFlightBookingResponse:

    now = get_utc_now()
    awb_ids = [item.awb_master_id for item in payload.awbs]

   
    # ── Check 1: duplicate flight on same date ─────────────────
    existing_header = await db.execute(
        select(ExportFlightBookingHeader.id).where(
            ExportFlightBookingHeader.flight_no == payload.flight_no,
            ExportFlightBookingHeader.flight_date == payload.flight_date,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    if existing_header.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"Flight {payload.flight_no} already booked on {payload.flight_date}"
        )

     # ✅ ADD HERE — departure in past check
    if to_utc(payload.flight_dpt_datetime) <= now:
        raise HTTPException(
            status_code=400,
            detail=f"Flight {payload.flight_no} departure datetime is in the past — cannot create booking",
        )


    # ── Check 2: fetch all AWBs in one query ───────────────────
    awb_result = await db.execute(
        select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.pcs,
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.is_ultra_fast,   # ← ADD
        ).where(ExportCarMessageAwbMaster.id.in_(awb_ids))
    )
    awb_map = {row.id: row for row in awb_result.mappings().all()}

    # ── Check 3: fetch already booked pcs for these AWBs in one query ──
    booked_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            func.sum(ExportFlightBookingDetail.booked_pcs).label("booked_pcs"),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
            ),
        )
        .where(ExportFlightBookingDetail.awb_master_id.in_(awb_ids))
        .group_by(ExportFlightBookingDetail.awb_master_id)
    )
    booked_map = {row.awb_master_id: row.booked_pcs for row in booked_result.mappings().all()}

    # ── Fetch scanned pcs for these AWBs ──────────────────────
    scanned_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .where(ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids))
        .group_by(ExportAwbSkidItemSequence.awb_master_id)
    )
    scanned_map = {
        row.awb_master_id: row.scanned_pcs
        for row in scanned_result.mappings().all()
    }
    

    # ── Check 4: validate each AWB ─────────────────────────────
    errors = []
    for item in payload.awbs:
        awb = awb_map.get(item.awb_master_id)

        if not awb:
            errors.append(f"AWB id {item.awb_master_id} not found")
            continue

        # if awb.status != "RCS":
        #     errors.append(f"AWB {awb.awb_no} is not in RCS status")
        #     continue

        if awb.status not in FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING:   # ✅ CHANGE — was != "RCS"
            errors.append(
                f"AWB {awb.awb_no} status '{awb.status}' is not eligible for flight booking"
            )
            continue

        # already_booked = booked_map.get(item.awb_master_id, 0)
        # remaining = awb.pcs - already_booked

        # if remaining <= 0:
        #     errors.append(f"AWB {awb.awb_no} is fully booked — no pcs remaining")
        #     continue

        # if item.booked_pcs > remaining:
        #     errors.append(
        #         f"AWB {awb.awb_no}: requested {item.booked_pcs} pcs "
        #         f"but only {remaining} remaining"
        #     )

        already_booked = booked_map.get(item.awb_master_id, 0)
        scanned_pcs = scanned_map.get(item.awb_master_id, 0)

        # ✅ available = scanned - already booked elsewhere
        # available = scanned_pcs - already_booked
        # 🤮🤮
        is_ultra_fast = awb.is_ultra_fast  # ← need this in awb_map query

        if is_ultra_fast:
            # ✅ ultra-fast: use total_pcs instead of scanned_pcs
            available = (awb.pcs or 0) - already_booked
        else:
            available = scanned_pcs - already_booked
        # ------

        if available <= 0:
            # errors.append(
            #     f"AWB {awb.awb_no}: no available scanned pcs "
            #     f"(scanned={scanned_pcs}, already_booked={already_booked})"
            # )
            errors.append(
                f"AWB {awb.awb_no}: no pcs available "
                f"({'total' if is_ultra_fast else 'scanned'}={awb.pcs if is_ultra_fast else scanned_pcs}, booked={already_booked})"
            )
            continue

        if item.booked_pcs > available:
            errors.append(
                f"AWB {awb.awb_no}: requested {item.booked_pcs} pcs "
                f"but only {available} available "
                f"(scanned={scanned_pcs}, already_booked={already_booked})"
            )
            continue

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    # ── All checks passed — insert header + details ────────────
    header = ExportFlightBookingHeader(
        flight_no=payload.flight_no,
        flight_date=payload.flight_date,
        flight_dpt_datetime=to_utc(payload.flight_dpt_datetime),
        booked_by=booked_by,
        booked_at=now,
        is_active=True,
        created_at=now,
        updated_at=now,
    )
    db.add(header)
    await db.flush()  # get header.id without committing

    details = [
        ExportFlightBookingDetail(
             flight_header_id=header.id,
            awb_master_id=item.awb_master_id,
            booked_pcs=item.booked_pcs,
        )
        for item in payload.awbs
    ]
    db.add_all(details)

    await db.flush()  # get detail ids


     # ── Audit log — one entry per AWB ─────────────────────────
    for item in payload.awbs:
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=item.awb_master_id,
            flight_reference_id=header.id,
            module=CarMessageFlowModule.FLIGHT_BOOKING,
            flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
            record_id=header.id,
            action="CREATE",
            performed_by=booked_by,
            changes={
                "event": "FLIGHT_BOOKING_CREATED",
                "flight_no": header.flight_no,
                "flight_date": str(header.flight_date),
                "flight_dpt_datetime": str(header.flight_dpt_datetime),
                "booked_pcs": item.booked_pcs,
            },
        )
        
    await db.commit()   # commits header + details + logs together
    await db.refresh(header)

    # ── Build response ─────────────────────────────────────────
    return CreateFlightBookingResponse(
        success=True,
        message="Successfully Created flight booking.",
        header_id=header.id,
        flight_no=header.flight_no,
        flight_date=header.flight_date,
        flight_dpt_datetime=header.flight_dpt_datetime,
        total_awbs=len(details),
        total_pcs=sum(item.booked_pcs for item in payload.awbs),
        details=[
            FlightBookingDetailResponse(
                awb_master_id=item.awb_master_id,
                awb_no=awb_map[item.awb_master_id].awb_no,
                booked_pcs=item.booked_pcs,
                total_pcs=awb_map[item.awb_master_id].pcs,  # ✅ add this
            )
            for item in payload.awbs
        ],
    )







async def get_flight_booking_by_flight_no_and_date(
    db: AsyncSession,
    flight_no: str,
    flight_date: date,
) -> FlightBookingByFlightResponse:

    # ── Fetch header ───────────────────────────────────────────
    header_result = await db.execute(
        select(ExportFlightBookingHeader).where(
            ExportFlightBookingHeader.flight_no == flight_no.strip().upper(),
            ExportFlightBookingHeader.flight_date == flight_date,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    header = header_result.scalar_one_or_none()
    if not header:
        raise HTTPException(
            status_code=404,
            detail=f"No active booking found for flight {flight_no} on {flight_date}",
        )

    # ── Fetch details with AWB info ────────────────────────────
    details_result = await db.execute(
        select(
            ExportFlightBookingDetail.id.label("detail_id"),
            ExportFlightBookingDetail.awb_master_id,
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
             ExportCarMessageAwbMaster.is_ultra_fast,
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportCarMessageAwbMaster.is_ultra_fast,
            ExportCarMessageAwbMaster.is_manually_created,
            
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
        )
      
        .where(ExportFlightBookingDetail.flight_header_id == header.id)
    )
    details = details_result.mappings().all()

    if not details:
        raise HTTPException(
            status_code=404,
            detail="Flight booking has no AWB details",
        )

    awb_ids = [d.awb_master_id for d in details]

    # ✅🤢 ADD — scanned pcs + is_ultra_fast per AWB
    scanned_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .where(ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids))
        .group_by(ExportAwbSkidItemSequence.awb_master_id)
    )
    scanned_map = {row.awb_master_id: row.scanned_pcs for row in scanned_result.mappings().all()}

    # ── Booked in OTHER flights for these AWBs (single query) ──
    booked_elsewhere_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            func.sum(ExportFlightBookingDetail.booked_pcs).label("booked_pcs"),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
              
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
                ExportFlightBookingHeader.id != header.id,   # exclude current
            ),
        )
        .where(ExportFlightBookingDetail.awb_master_id.in_(awb_ids))
        .group_by(ExportFlightBookingDetail.awb_master_id)
    )
    booked_elsewhere = {
        row.awb_master_id: row.booked_pcs
        for row in booked_elsewhere_result.mappings().all()
    }

    # ── Build detail response ──────────────────────────────────
    detail_list = []
    for d in details:
        booked_in_others = booked_elsewhere.get(d.awb_master_id, 0)
        remaining = d.total_pcs - booked_in_others - d.booked_pcs

        detail_list.append(
            FlightBookingDetailWithAwbResponse(
                detail_id=d.detail_id,
                awb_master_id=d.awb_master_id,
                awb_no=d.awb_no,
                origin=d.origin,
                destination=d.destination,
                total_pcs=d.total_pcs,
                booked_pcs=d.booked_pcs,
                scanned_pcs=scanned_map.get(d.awb_master_id, 0),   # ✅ ADD
                is_ultra_fast=d.is_ultra_fast,        #✅ ADD
                booked_in_other_flights=booked_in_others,
                remaining_pcs=remaining,
                agent=d.agent,
                rcs_datetime=d.rcs_datetime,
            )
        )

    return FlightBookingByFlightResponse(
        header_id=header.id,
        flight_no=header.flight_no,
        flight_date=header.flight_date,
        flight_dpt_datetime=header.flight_dpt_datetime,
        booked_by=header.booked_by,
        booked_at=header.booked_at,
        total_awbs=len(detail_list),
        total_pcs=sum(d.booked_pcs for d in detail_list),
        details=detail_list,
    )



async def edit_flight_booking(
    db: AsyncSession,
    header_id: int,
    payload: EditFlightBookingRequest,
    edited_by: str,
) -> EditFlightBookingResponse:

    now = get_utc_now()

    # ── Fetch header ───────────────────────────────────────────
    header = await db.get(ExportFlightBookingHeader, header_id)
    if not header or not header.is_active:
        raise HTTPException(
            status_code=404,
            detail="Flight booking not found"
        )

    if header.flight_dpt_datetime <= now:
        raise HTTPException(
            status_code=400,
            detail=f"Flight {header.flight_no} has already departed — editing is not allowed"
        )

    # ── Fetch existing details for this header ─────────────────
    existing_result = await db.execute(
        select(ExportFlightBookingDetail).where(
            ExportFlightBookingDetail.flight_header_id == header_id
        )
    )
    existing_details = {
        d.id: d for d in existing_result.scalars().all()
    }

    # ✅ ADD THIS — snapshot before any mutation {which used in log}
    old_pcs_map = {
        d_id: d.booked_pcs
        for d_id, d in existing_details.items()
    }

        # ✅ ADD HERE — block AWB removal + pcs reduction if items already loaded
    # fetch loaded counts per AWB on this flight in one query
    loaded_per_awb_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.awb_master_id,
            func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
            ExportCarMessageAwbMaster.awb_no,
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportSequenceItemUldLoading.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(
            ExportSequenceItemUldLoading.flight_header_id == header_id,
        )
        .group_by(
            ExportSequenceItemUldLoading.awb_master_id,
            ExportCarMessageAwbMaster.awb_no,
        )
    )
    loaded_per_awb = {
        row.awb_master_id: row
        for row in loaded_per_awb_result.mappings().all()
    }

    block_errors = []

    # block removal of AWBs that have loaded items
    for did in payload.removed_detail_ids:
        detail = existing_details.get(did)
        if not detail:
            continue
        loaded = loaded_per_awb.get(detail.awb_master_id)
        if loaded and loaded.loaded_count > 0:
            block_errors.append(
                f"AWB {loaded.awb_no} cannot be removed — "
                f"{loaded.loaded_count} item(s) already loaded into ULD on this flight"
            )

    # block pcs reduction below already loaded count
    for item in payload.awbs:
        if not item.detail_id or item.detail_id not in existing_details:
            continue
        loaded = loaded_per_awb.get(item.awb_master_id)
        if loaded and loaded.loaded_count > 0:
            if item.booked_pcs < loaded.loaded_count:
                awb = awb_map.get(item.awb_master_id) if 'awb_map' in dir() else None
                # awb_map not built yet at this point — fetch awb_no from loaded row
                block_errors.append(
                    f"AWB {loaded.awb_no}: cannot reduce pcs to {item.booked_pcs} — "
                    f"{loaded.loaded_count} item(s) already loaded into ULD"
                )

    if block_errors:
        raise HTTPException(status_code=400, detail=block_errors)
    
    # ── All AWB ids in payload ─────────────────────────────────
    awb_ids = [item.awb_master_id for item in payload.awbs]

    # ── Fetch AWB master data in one query ─────────────────────
    awb_result = await db.execute(
        select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.pcs,
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.is_ultra_fast,   # ← need this for validation
        ).where(ExportCarMessageAwbMaster.id.in_(awb_ids))
    )
    awb_map = {row.id: row for row in awb_result.mappings().all()}

    # ── Booked pcs in OTHER active flights (exclude current) ───
    booked_elsewhere_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            func.sum(ExportFlightBookingDetail.booked_pcs).label("booked_pcs"),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
                ExportFlightBookingHeader.id != header_id,  # ← exclude current flight
            ),
        )
        .where(ExportFlightBookingDetail.awb_master_id.in_(awb_ids))
        .group_by(ExportFlightBookingDetail.awb_master_id)
    )
    booked_elsewhere = {
        row.awb_master_id: row.booked_pcs
        for row in booked_elsewhere_result.mappings().all()
    }

    # ── Fetch scanned pcs ──────────────────────────────────────
    scanned_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .where(ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids))
        .group_by(ExportAwbSkidItemSequence.awb_master_id)
    )
    scanned_map = {
        row.awb_master_id: row.scanned_pcs
        for row in scanned_result.mappings().all()
    }

    # ── Validate each AWB ──────────────────────────────────────
    errors = []
    for item in payload.awbs:
        awb = awb_map.get(item.awb_master_id)

        if not awb:
            errors.append(f"AWB id {item.awb_master_id} not found")
            continue

        # if awb.status != "RCS":
        #     errors.append(f"AWB {awb.awb_no} is not in RCS status")

        if awb.status not in FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING:   # ✅ CHANGE — was != "RCS"
            errors.append(
                f"AWB {awb.awb_no} status '{awb.status}' is not eligible for flight booking"
            )
            continue

        # booked_in_others = booked_elsewhere.get(item.awb_master_id, 0)
        # available = awb.pcs - booked_in_others  # excludes current flight

        # if item.booked_pcs > available:
        #     errors.append(
        #         f"AWB {awb.awb_no}: max {available} pcs available "
        #         f"({awb.pcs} total − {booked_in_others} in other flights)"
        #     )

        booked_in_others = booked_elsewhere.get(item.awb_master_id, 0)
        scanned_pcs = scanned_map.get(item.awb_master_id, 0)
        # available = scanned_pcs - booked_in_others   # ✅ scanned based cap
        # 🔥 ADD THIS
        is_ultra_fast = awb.is_ultra_fast or False

        if is_ultra_fast:
            available = (awb.pcs or 0) - booked_in_others
        else:
            available = scanned_pcs - booked_in_others

        # ✅ NEW (add this)
        if item.booked_pcs <= 0:
            errors.append(
                f"AWB {awb.awb_no}: booked pcs must be greater than 0"
            )
            continue

        if available <= 0:
            errors.append(
                f"AWB {awb.awb_no}: no available scanned pcs "
                f"(scanned={scanned_pcs}, booked_elsewhere={booked_in_others})"
                #  f"AWB {awb.awb_no}: no pcs available for booking"
            )
            continue

        if item.booked_pcs > available:
            errors.append(
                f"AWB {awb.awb_no}: max {available} pcs available "
                f"(scanned={scanned_pcs}, booked_elsewhere={booked_in_others})"
                f"===="
                 f"AWB {awb.awb_no}: requested {item.booked_pcs} pcs, "
                f"but only {available} pcs available"
            )
            continue

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    # ── Delete removed AWBs ────────────────────────────────────
    for detail_id in payload.removed_detail_ids:
        detail = existing_details.get(detail_id)
        if detail:
            await db.delete(detail)

    # ── Update existing / insert new ───────────────────────────
    for item in payload.awbs:
        if item.detail_id and item.detail_id in existing_details:
            # ✅ update existing row
            existing_details[item.detail_id].booked_pcs = item.booked_pcs
        else:
            # ✅ new AWB — insert
            db.add(
                ExportFlightBookingDetail(
                    flight_header_id=header_id,
                    awb_master_id=item.awb_master_id,
                    booked_pcs=item.booked_pcs,
                )
            )

    # ── Update header timestamp ────────────────────────────────
    header.updated_at = now



    # ── 😎 Audit log — track what actually changed ────────────────
    added_awbs = [
        item for item in payload.awbs
        if not item.detail_id or item.detail_id not in existing_details
    ]
    updated_awbs = [
        item for item in payload.awbs
        if item.detail_id and item.detail_id in existing_details
        and old_pcs_map[item.detail_id] != item.booked_pcs  # ✅ uses snapshot
    ]

    # log for newly added AWBs
    for item in added_awbs:
        awb = awb_map.get(item.awb_master_id)  # ✅ ADD THIS
        if not awb:
            continue
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=item.awb_master_id,
            flight_reference_id=header_id,
            module=CarMessageFlowModule.FLIGHT_BOOKING,
            flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
            record_id=header_id,
            action="UPDATE",
            performed_by=edited_by,
            changes={
                "event": "AWB_ADDED",
                "flight_no": header.flight_no,
                "awb_no": awb.awb_no,
                "booked_pcs": item.booked_pcs,
                "summary": (
                f"AWB {awb.awb_no} added to flight {header.flight_no} "
                f"with {item.booked_pcs} pcs"
            ),
            },
        )

    # log for pcs changed AWBs
    for item in updated_awbs:
        awb = awb_map.get(item.awb_master_id)  # ✅ ADD THIS
        if not awb:
            continue
        old_pcs = old_pcs_map[item.detail_id]
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=item.awb_master_id,
            flight_reference_id=header_id,
            module=CarMessageFlowModule.FLIGHT_BOOKING,
            flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
            record_id=header_id,
            action="UPDATE",
            performed_by=edited_by,
            changes={
                "event": "PCS_UPDATED",
                "flight_no": header.flight_no,
                "booked_pcs_before": old_pcs,
                "awb_no": awb.awb_no,
                "booked_pcs_after": item.booked_pcs,
                "summary": (
                f"AWB {awb.awb_no} pcs updated from {old_pcs} "
                f"to {item.booked_pcs} on flight {header.flight_no} and "
                f"flight date {str(header.flight_date)}"
            ),
            },
        )

    # log for removed AWBs — need awb_master_id from existing details
    # if payload.removed_detail_ids:
    #     removed_awb_ids = {
    #         existing_details[did].awb_master_id
    #         for did in payload.removed_detail_ids
    #         if did in existing_details
    #     }
    #     for awb_id in removed_awb_ids:
    #         await write_car_message_flow_audit(
    #             db=db,
    #             awb_reference_id=awb_id,
    #             flight_reference_id=header_id,
    #             module=CarMessageFlowModule.FLIGHT_BOOKING,
    #             flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
    #             record_id=header_id,
    #             action="UPDATE",
    #             performed_by=edited_by,
    #             changes={
    #                 "event": "AWB_REMOVED",
    #                 "flight_no": header.flight_no,
    #                 "awb_no": awb.awb_no if awb else "UNKNOWN",
    #                 "summary": (
    #             f"AWB {awb.awb_no if awb else 'UNKNOWN'} removed from "
    #             f"flight {header.flight_no}"
    #             f"flight date {str(header.flight_date)}"
    #             f"({detail.booked_pcs} pcs freed)"
    #         ),
    #             },
    #         )

    for did in payload.removed_detail_ids:
        detail = existing_details.get(did)
        if not detail:
            continue
        awb = awb_map.get(detail.awb_master_id)  # ✅ now awb is defined
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=detail.awb_master_id,
            flight_reference_id=header_id,
            module=CarMessageFlowModule.FLIGHT_BOOKING,
            flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
            record_id=header_id,
            action="UPDATE",
            performed_by=edited_by,
            changes={
                "event": "AWB_REMOVED",
                "flight_no": header.flight_no,
                "flight_date": str(header.flight_date),
                "awb_no": awb.awb_no if awb else "UNKNOWN",
                "booked_pcs_removed": old_pcs_map.get(did),  # ✅ from snapshot
                "summary": (
                    f"AWB {awb.awb_no if awb else 'UNKNOWN'} removed from "
                    f"flight {header.flight_no} ({header.flight_date}) — "
                    f"{old_pcs_map.get(did)} pcs freed"
                ),
            },
        )
# ------
    await db.commit()

    await db.refresh(header)

    # ── Return updated booking ─────────────────────────────────
    updated = await get_flight_booking_by_flight_no_and_date(
        db=db,
        flight_no=header.flight_no,
        flight_date=header.flight_date,
    )

    return EditFlightBookingResponse(
        success=True,
        message=f"Flight {header.flight_no} booking updated successfully",
        data=updated,
    )


# ======================== ✌️ ULD ASSIGNMENT CREATE AND  EDIT GET ...✌️ ======================================
# ============================-----------------------------------==============================================



# ── Helper: build assignment data response ─────────────────────
async def _build_assignment_response(
    db: AsyncSession,
    assignment: ExportUldAssignment,
    header: ExportFlightBookingHeader,
) -> UldAssignmentDataResponse:

    details_result = await db.execute(
        select(
            ExportUldAssignmentDetail.id.label("detail_id"),
            ExportUldAssignmentDetail.uld_id,
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        )
        .join(ExportUldMaster, ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
        .where(ExportUldAssignmentDetail.assignment_id == assignment.id)
        .order_by(ExportUldMaster.uld_no)
    )
    details = details_result.mappings().all()

    return UldAssignmentDataResponse(
        assignment_id=assignment.id,
        flight_header_id=header.id,
        flight_no=header.flight_no,
        flight_date=str(header.flight_date),
        flight_dpt_datetime=header.flight_dpt_datetime,
        total_ulds=len(details),
        assigned_by=assignment.assigned_by,
        assigned_at=assignment.assigned_at,
        ulds=[UldAssignmentDetailResponse(**d) for d in details],
    )


# ── Helper: departure check ────────────────────────────────────
def _check_not_departed(header: ExportFlightBookingHeader, flight_no: str):
    if header.flight_dpt_datetime <= datetime.now(timezone.utc):
        raise HTTPException(
            status_code=400,
            detail=f"Flight {flight_no} has already departed — operation not allowed",
        )

# ── reusable helper ────────────────────────────────────────────
# this help to check that already assignd uld not reused utilf that flight departure
async def _check_ulds_not_active_on_another_flight(
    db: AsyncSession,
    uld_ids: list[int],
    exclude_assignment_id: int | None = None,  # for edit — exclude current
) -> list[str]:
    """
    Returns list of error messages for ULDs already assigned
    to a flight that has not yet departed.
    """
    now = datetime.now(timezone.utc)

    # Find ULDs already assigned to another active non-departed flight
    stmt = (
        select(
            ExportUldMaster.uld_no,
            ExportFlightBookingHeader.flight_no,
            ExportFlightBookingHeader.flight_dpt_datetime,
        )
        .join(
            ExportUldAssignmentDetail,
            ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
        )
        .join(
            ExportUldAssignment,
            and_(
                ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id,
                ExportUldAssignment.is_active == True,
            ),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
                ExportUldAssignment.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
                ExportFlightBookingHeader.flight_dpt_datetime > now,  # not yet departed
            ),
        )
        .where(ExportUldMaster.id.in_(uld_ids))
    )

    # exclude current assignment in edit mode
    if exclude_assignment_id:
        stmt = stmt.where(
            ExportUldAssignment.id != exclude_assignment_id
        )

    result = await db.execute(stmt)
    rows = result.mappings().all()

    return [
        f"ULD {row.uld_no} is already assigned to flight {row.flight_no} "
        f"which departs at {row.flight_dpt_datetime.strftime('%d %b %Y %H:%M')} UTC — "
        f"cannot assign until that flight departs"
        for row in rows
    ]


# ── GET all active ULDs ────────────────────────────────────────
async def get_uld_master_list(db: AsyncSession) -> list[UldMasterResponse]:
    result = await db.execute(
        select(
            ExportUldMaster.id.label("uld_id"),
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        )
        .where(ExportUldMaster.is_active == True)
        .order_by(ExportUldMaster.uld_no)
    )

    return [UldMasterResponse(**row) for row in result.mappings().all()]

# ──🤮🤢 Get ULD which is assign to a flight and that flight  depart. If flighrt not depart now then those are not include 
# async def get_uld_master_list_eligeble_for_assignment(db: AsyncSession) -> list[UldMasterResponse]:

#     now = get_utc_now()

#     # ── Subquery: ULD ids already assigned to non-departed active flights ──
#     assigned_uld_subq = (
#         select(ExportUldAssignmentDetail.uld_id)
#         .join(
#             ExportUldAssignment,
#             ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id,
#         )
#         .join(
#             ExportFlightBookingHeader,
#             ExportUldAssignment.flight_header_id == ExportFlightBookingHeader.id,
#         )
#         .where(
#             ExportUldAssignment.is_active == True,
#             ExportFlightBookingHeader.is_active == True,
#             ExportFlightBookingHeader.flight_dpt_datetime > now,  # ✅ not yet departed
#         )
#         .subquery()
#     )

#     result = await db.execute(
#         select(
#             ExportUldMaster.id.label("uld_id"),
#             ExportUldMaster.uld_no,
#             ExportUldMaster.carrier,
#         )
#         .where(
#             ExportUldMaster.is_active == True,
#             ExportUldMaster.id.notin_(select(assigned_uld_subq)),  # ✅ exclude assigned
#         )
#         .order_by(ExportUldMaster.uld_no)
#     )
#     rows = result.mappings().all()
#     print(len(rows))

#     return [UldMasterResponse(**row) for row in rows]

async def get_uld_master_list_eligeble_for_assignment(
    db: AsyncSession,
    carriers: list[str] | None = None,   # ✅ ADD
) -> list[UldMasterResponse]:

    stmt  = (
        select(
            ExportUldMaster.id.label("uld_id"),
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        )
        .where(
            ExportUldMaster.is_active == True,
            ExportUldMaster.is_available == True,   # ✅ ADD THIS
        )
        .order_by(ExportUldMaster.uld_no)
    )
        # ✅ ADD
    if carriers:
        stmt = stmt.where(ExportUldMaster.carrier.in_(carriers))

    result = await db.execute(stmt)

    rows = result.mappings().all()
    print(len(rows))

    return [UldMasterResponse(**row) for row in rows]


# ── GET assignment by flight ───────────────────────────────────
async def get_uld_assignment_by_flight(
    db: AsyncSession,
    flight_no: str,
    flight_date: date,
) -> UldAssignmentDataResponse | None:

    # fetch flight header
    header_result = await db.execute(
        select(ExportFlightBookingHeader).where(
            ExportFlightBookingHeader.flight_no == flight_no.strip().upper(),
            ExportFlightBookingHeader.flight_date == flight_date,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    header = header_result.scalar_one_or_none()
    if not header:
        raise HTTPException(
            status_code=404,
            detail=f"No active flight booking found for {flight_no} on {flight_date}",
        )

    # fetch assignment
    assignment_result = await db.execute(
        select(ExportUldAssignment).where(
            ExportUldAssignment.flight_header_id == header.id,
            ExportUldAssignment.is_active == True,
        )
    )
    assignment = assignment_result.scalar_one_or_none()

    if not assignment:
        return None  # no assignment yet — frontend shows create form

    return await _build_assignment_response(db, assignment, header)


# ── CREATE assignment ──────────────────────────────────────────
# async def create_uld_assignment(
#     db: AsyncSession,
#     payload: CreateUldAssignmentRequest,
#     assigned_by: str,
# ) -> UldAssignmentResponse:

#     now = get_utc_now()

#     # fetch header
#     header = await db.get(ExportFlightBookingHeader, payload.flight_header_id)
#     if not header or not header.is_active:
#         raise HTTPException(status_code=404, detail="Flight booking not found")

#     # departure check
#     _check_not_departed(header, header.flight_no)

#     # check no existing assignment
#     existing = await db.execute(
#         select(ExportUldAssignment.id).where(
#             ExportUldAssignment.flight_header_id == payload.flight_header_id,
#             ExportUldAssignment.is_active == True,
#         )
#     )
#     if existing.scalar_one_or_none():
#         raise HTTPException(
#             status_code=400,
#             detail=f"ULD assignment already exists for flight {header.flight_no} — use edit instead",
#         )

#     # validate ULDs exist and are active — one query
#     uld_result = await db.execute(
#         select(ExportUldMaster.id).where(
#             ExportUldMaster.id.in_(payload.uld_ids),
#             ExportUldMaster.is_active == True,
#         )
#     )
#     valid_uld_ids = {row.id for row in uld_result.all()}
#     invalid = set(payload.uld_ids) - valid_uld_ids
#     if invalid:
#         raise HTTPException(
#             status_code=400,
#             detail=f"ULD ids not found or inactive: {sorted(invalid)}",
#         )
    
#     # ── Check ULDs not active on another non-departed flight ───
#     conflicts = await _check_ulds_not_active_on_another_flight(
#         db=db,
#         uld_ids=payload.uld_ids,
#     )
#     if conflicts:
#         raise HTTPException(status_code=400, detail=conflicts)

#     # insert assignment
#     assignment = ExportUldAssignment(
#         flight_header_id=payload.flight_header_id,
#         assigned_by=assigned_by,
#         assigned_at=now,
#         is_active=True,
#         created_at=now,
#         updated_at=now,
#     )
#     db.add(assignment)
#     await db.flush()



#     # insert details
#     db.add_all([
#         ExportUldAssignmentDetail(
#             assignment_id=assignment.id,
#             uld_id=uld_id,
#             created_at=now,
#         )
#         for uld_id in payload.uld_ids
#     ])

#     # ✅ fetch ULD info for readable 😎 Log
#     uld_info_result = await db.execute(
#         select(
#             ExportUldMaster.id,
#             ExportUldMaster.uld_no,
#             ExportUldMaster.carrier,
#         ).where(ExportUldMaster.id.in_(payload.uld_ids))
#     )
#     uld_info_map = {row.id: row for row in uld_info_result.mappings().all()}

#     # ✅ fetch all AWBs on this flight for per-AWB logging
#     awb_ids_result = await db.execute(
#         select(ExportFlightBookingDetail.awb_master_id).where(
#             ExportFlightBookingDetail.flight_header_id == payload.flight_header_id
#         )
#     )
#     awb_ids_on_flight = [row.awb_master_id for row in awb_ids_result.all()]

#     # ✅ audit log — one entry per AWB
#     for awb_id in awb_ids_on_flight:
#         await write_car_message_flow_audit(
#             db=db,
#             awb_reference_id=awb_id,
#             flight_reference_id=payload.flight_header_id,
#             module=CarMessageFlowModule.ULD_ASSIGNMENT,
#             flow_step=CarMessageFlowStep.ULD_ASSIGNMENT,
#             record_id=assignment.id,
#             action="CREATE",
#             performed_by=assigned_by,
#             changes={
#                 "event": "ULD_ASSIGNMENT_CREATED",
#                 "flight_no": header.flight_no,
#                 "flight_date": str(header.flight_date),
#                 "uld_count": len(payload.uld_ids),
#                 "ulds": [
#                     {
#                         "uld_id": uid,
#                         "uld_no": uld_info_map[uid].uld_no,
#                         "carrier": uld_info_map[uid].carrier,
#                     }
#                     for uid in payload.uld_ids
#                     if uid in uld_info_map
#                 ],
#                 "summary": (
#                     f"{len(payload.uld_ids)} ULD(s) assigned to flight "
#                     f"{header.flight_no} ({header.flight_date}): "
#                     f"{', '.join(uld_info_map[uid].uld_no for uid in payload.uld_ids if uid in uld_info_map)}"
#                 ),
#             },
#         )


# # ------
#     await db.commit()
#     await db.refresh(assignment)

#     data = await _build_assignment_response(db, assignment, header)
#     return UldAssignmentResponse(
#         success=True,
#         message=f"ULD assignment created for flight {header.flight_no} — {len(payload.uld_ids)} ULDs assigned",
#         data=data,
#     )

# 🤢🤮
from sqlalchemy import update

# ── CREATE assignment ──────────────────────────────────────────
async def create_uld_assignment(
    db: AsyncSession,
    payload: CreateUldAssignmentRequest,
    assigned_by: str,
) -> UldAssignmentResponse:

    now = get_utc_now()

    # fetch header
    header = await db.get(ExportFlightBookingHeader, payload.flight_header_id)
    if not header or not header.is_active:
        raise HTTPException(status_code=404, detail="Flight booking not found")

    # ❌ REMOVED departure check
    # _check_not_departed(header, header.flight_no)

    # check no existing assignment
    existing = await db.execute(
        select(ExportUldAssignment.id).where(
            ExportUldAssignment.flight_header_id == payload.flight_header_id,
            ExportUldAssignment.is_active == True,
        )
    )
    if existing.scalar_one_or_none():
        raise HTTPException(
            status_code=400,
            detail=f"ULD assignment already exists for flight {header.flight_no} — use edit instead",
        )

    # validate ULDs exist and are active
    uld_result = await db.execute(
        select(ExportUldMaster.id).where(
            ExportUldMaster.id.in_(payload.uld_ids),
            ExportUldMaster.is_active == True,
        )
    )
    valid_uld_ids = {row.id for row in uld_result.all()}
    invalid = set(payload.uld_ids) - valid_uld_ids
    if invalid:
        raise HTTPException(
            status_code=400,
            detail=f"ULD ids not found or inactive: {sorted(invalid)}",
        )

    # ✅ NEW: check availability
    available_result = await db.execute(
        select(ExportUldMaster.id).where(
            ExportUldMaster.id.in_(payload.uld_ids),
            ExportUldMaster.is_available == True,
        )
    )
    available_ids = {row.id for row in available_result.all()}
    not_available = set(payload.uld_ids) - available_ids

    if not_available:
        raise HTTPException(
            status_code=400,
            detail=f"ULD not available: {sorted(not_available)}",
        )

    # insert assignment
    assignment = ExportUldAssignment(
        flight_header_id=payload.flight_header_id,
        assigned_by=assigned_by,
        assigned_at=now,
        is_active=True,
        created_at=now,
        updated_at=now,
    )
    db.add(assignment)
    await db.flush()

    # insert details
    db.add_all([
        ExportUldAssignmentDetail(
            assignment_id=assignment.id,
            uld_id=uld_id,
            created_at=now,
        )
        for uld_id in payload.uld_ids
    ])

    # ✅ mark assigned ULDs as unavailable
    await db.execute(
        update(ExportUldMaster)
        .where(ExportUldMaster.id.in_(payload.uld_ids))
        .values(is_available=False)
    )

    # ✅ fetch ULD info for readable 😎 Log
    uld_info_result = await db.execute(
        select(
            ExportUldMaster.id,
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        ).where(ExportUldMaster.id.in_(payload.uld_ids))
    )
    uld_info_map = {row.id: row for row in uld_info_result.mappings().all()}

    # ✅ fetch all AWBs on this flight for per-AWB logging
    awb_ids_result = await db.execute(
        select(ExportFlightBookingDetail.awb_master_id).where(
            ExportFlightBookingDetail.flight_header_id == payload.flight_header_id
        )
    )
    awb_ids_on_flight = [row.awb_master_id for row in awb_ids_result.all()]

    # ✅ audit log — one entry per AWB
    for awb_id in awb_ids_on_flight:
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=awb_id,
            flight_reference_id=payload.flight_header_id,
            module=CarMessageFlowModule.ULD_ASSIGNMENT,
            flow_step=CarMessageFlowStep.ULD_ASSIGNMENT,
            record_id=assignment.id,
            action="CREATE",
            performed_by=assigned_by,
            changes={
                "event": "ULD_ASSIGNMENT_CREATED",
                "flight_no": header.flight_no,
                "flight_date": str(header.flight_date),
                "uld_count": len(payload.uld_ids),
                "ulds": [
                    {
                        "uld_id": uid,
                        "uld_no": uld_info_map[uid].uld_no,
                        "carrier": uld_info_map[uid].carrier,
                    }
                    for uid in payload.uld_ids
                    if uid in uld_info_map
                ],
                "summary": (
                    f"{len(payload.uld_ids)} ULD(s) assigned to flight "
                    f"{header.flight_no} ({header.flight_date}): "
                    f"{', '.join(uld_info_map[uid].uld_no for uid in payload.uld_ids if uid in uld_info_map)}"
                ),
            },
        )

    await db.commit()
    await db.refresh(assignment)

    data = await _build_assignment_response(db, assignment, header)

    return UldAssignmentResponse(
        success=True,
        message=f"ULD assignment created for flight {header.flight_no} — {len(payload.uld_ids)} ULDs assigned",
        data=data,
    )


# ── EDIT assignment ────────────────────────────────────────────
# async def edit_uld_assignment(
#     db: AsyncSession,
#     assignment_id: int,
#     payload: EditUldAssignmentRequest,
#     edited_by: str,
# ) -> UldAssignmentResponse:

#     now = get_utc_now()

#     # fetch assignment
#     assignment = await db.get(ExportUldAssignment, assignment_id)
#     if not assignment or not assignment.is_active:
#         raise HTTPException(status_code=404, detail="ULD assignment not found")

#     # fetch header for departure check
#     header = await db.get(ExportFlightBookingHeader, assignment.flight_header_id)
#     _check_not_departed(header, header.flight_no)

#     # fetch existing details
#     existing_result = await db.execute(
#         select(ExportUldAssignmentDetail).where(
#             ExportUldAssignmentDetail.assignment_id == assignment_id
#         )
#     )
#     existing_details = {d.id: d for d in existing_result.scalars().all()}
#     existing_uld_ids = {d.uld_id for d in existing_details.values()}

#     errors = []

#     # ✅ ADD HERE — block ULD removal if items already loaded
#     if payload.uld_detail_ids_to_remove:
#         loaded_result = await db.execute(
#             select(
#                 ExportSequenceItemUldLoading.uld_assignment_detail_id,
#                 func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
#                 ExportUldMaster.uld_no,
#             )
#             .join(
#                 ExportUldAssignmentDetail,
#                 ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id,
#             )
#             .join(
#                 ExportUldMaster,
#                 ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
#             )
#             .where(
#                 ExportSequenceItemUldLoading.uld_assignment_detail_id.in_(
#                     payload.uld_detail_ids_to_remove
#                 )
#             )
#             .group_by(
#                 ExportSequenceItemUldLoading.uld_assignment_detail_id,
#                 ExportUldMaster.uld_no,
#             )
#         )
#         loaded_rows = loaded_result.mappings().all()

#         block_errors = [
#             f"ULD {row.uld_no} cannot be removed — "
#             f"{row.loaded_count} item(s) already loaded into it"
#             for row in loaded_rows
#             if row.loaded_count > 0
#         ]

#         if block_errors:
#             raise HTTPException(status_code=400, detail=block_errors)

#     # ── validate new ULDs to add ───────────────────────────────
#     if payload.uld_ids_to_add:
#         # check duplicates against already assigned
#         already_assigned = set(payload.uld_ids_to_add) & existing_uld_ids
#         if already_assigned:
#             # get uld_nos for readable error
#             dup_result = await db.execute(
#                 select(ExportUldMaster.uld_no).where(
#                     ExportUldMaster.id.in_(already_assigned)
#                 )
#             )
#             dup_nos = [r.uld_no for r in dup_result.all()]
#             errors.append(f"ULDs already assigned to this flight: {', '.join(dup_nos)}")

#         # check active in master
#         uld_result = await db.execute(
#             select(ExportUldMaster.id).where(
#                 ExportUldMaster.id.in_(payload.uld_ids_to_add),
#                 ExportUldMaster.is_active == True,
#             )
#         )
#         valid_ids = {row.id for row in uld_result.all()}
#         invalid = set(payload.uld_ids_to_add) - valid_ids
#         if invalid:
#             errors.append(f"ULD ids not found or inactive: {sorted(invalid)}")

#         # ✅ only check flight conflicts if ULDs are valid so far
#         # no point querying flight conflicts for invalid/inactive ULDs
#         if not errors:
#             conflicts = await _check_ulds_not_active_on_another_flight(
#                 db=db,
#                 uld_ids=payload.uld_ids_to_add,
#                 exclude_assignment_id=assignment_id,
#             )
#             if conflicts:
#                 errors.extend(conflicts)


#     # ── validate remove ids belong to this assignment ──────────
#     if payload.uld_detail_ids_to_remove:
#         invalid_removes = set(payload.uld_detail_ids_to_remove) - set(existing_details.keys())
#         if invalid_removes:
#             errors.append(f"Detail ids do not belong to this assignment: {sorted(invalid_removes)}")

#         # block if removing all and not adding any
#         remaining_after_remove = len(existing_details) - len(payload.uld_detail_ids_to_remove)
#         net_adds = len(payload.uld_ids_to_add)
#         if remaining_after_remove + net_adds < 1:
#             errors.append("At least one ULD must remain in the assignment")

#     if errors:
#         raise HTTPException(status_code=400, detail=errors)

#     # ── delete removed details ─────────────────────────────────
#     for detail_id in payload.uld_detail_ids_to_remove:
#         await db.delete(existing_details[detail_id])

#     # ── insert new ULDs ────────────────────────────────────────
#     db.add_all([
#         ExportUldAssignmentDetail(
#             assignment_id=assignment_id,
#             uld_id=uld_id,
#             created_at=now,
#         )
#         for uld_id in payload.uld_ids_to_add
#     ])

#     assignment.updated_at = now
    
# # ====== Log related used 😎
#     # ✅ snapshot existing uld_id per detail_id before any delete
#     existing_uld_id_map = {
#         d_id: d.uld_id
#         for d_id, d in existing_details.items()
#     }

#     # ✅ fetch ULD info for both added and removed ULDs
#     all_relevant_uld_ids = list(
#         set(payload.uld_ids_to_add) |
#         {existing_uld_id_map[did] for did in payload.uld_detail_ids_to_remove if did in existing_uld_id_map}
#     )

#     uld_info_map = {}
#     if all_relevant_uld_ids:
#         uld_info_result = await db.execute(
#             select(
#                 ExportUldMaster.id,
#                 ExportUldMaster.uld_no,
#                 ExportUldMaster.carrier,
#             ).where(ExportUldMaster.id.in_(all_relevant_uld_ids))
#         )
#         uld_info_map = {row.id: row for row in uld_info_result.mappings().all()}

#     # ✅ fetch AWBs on this flight
#     awb_ids_result = await db.execute(
#         select(ExportFlightBookingDetail.awb_master_id).where(
#             ExportFlightBookingDetail.flight_header_id == assignment.flight_header_id
#         )
#     )
#     awb_ids_on_flight = [row.awb_master_id for row in awb_ids_result.all()]

#     # ✅ build readable added/removed uld lists
#     added_ulds = [
#         {
#             "uld_id": uid,
#             "uld_no": uld_info_map[uid].uld_no,
#             "carrier": uld_info_map[uid].carrier,
#         }
#         for uid in payload.uld_ids_to_add
#         if uid in uld_info_map
#     ]

#     removed_ulds = [
#         {
#             "uld_id": existing_uld_id_map[did],
#             "uld_no": uld_info_map[existing_uld_id_map[did]].uld_no,
#         }
#         for did in payload.uld_detail_ids_to_remove
#         if did in existing_uld_id_map
#         and existing_uld_id_map[did] in uld_info_map
#     ]

#     added_uld_nos = [u["uld_no"] for u in added_ulds]
#     removed_uld_nos = [u["uld_no"] for u in removed_ulds]

#     # ✅ audit log — one entry per AWB
#     for awb_id in awb_ids_on_flight:
#         await write_car_message_flow_audit(
#             db=db,
#             awb_reference_id=awb_id,
#             flight_reference_id=assignment.flight_header_id,
#             module=CarMessageFlowModule.ULD_ASSIGNMENT,
#             flow_step=CarMessageFlowStep.ULD_ASSIGNMENT,
#             record_id=assignment_id,
#             action="UPDATE",
#             performed_by=edited_by,
#             changes={
#                 "event": "ULD_ASSIGNMENT_UPDATED",
#                 "flight_no": header.flight_no,
#                 "flight_date": str(header.flight_date),
#                 "added_ulds": added_ulds,
#                 "removed_ulds": removed_ulds,
#                 "summary": (
#                     f"Flight {header.flight_no} ({header.flight_date}) "
#                     "ULD assignment updated — "
#                     + (f"Added: {', '.join(added_uld_nos)}. " if added_uld_nos else "No ULDs added. ")
#                     + (f"Removed: {', '.join(removed_uld_nos)}." if removed_uld_nos else "No ULDs removed.")
#                 ),
#             },
#         )

#     # --------------
#     await db.commit()
#     await db.refresh(assignment)

#     data = await _build_assignment_response(db, assignment, header)
#     return UldAssignmentResponse(
#         success=True,
#         message=f"ULD assignment updated for flight {header.flight_no}",
#         data=data,
#     )


# 🤢🤮
async def edit_uld_assignment(
    db: AsyncSession,
    assignment_id: int,
    payload: EditUldAssignmentRequest,
    edited_by: str,
) -> UldAssignmentResponse:

    now = get_utc_now()

    # fetch assignment
    assignment = await db.get(ExportUldAssignment, assignment_id)
    if not assignment or not assignment.is_active:
        raise HTTPException(status_code=404, detail="ULD assignment not found")

    # fetch header for departure check
    header = await db.get(ExportFlightBookingHeader, assignment.flight_header_id)
    # _check_not_departed(header, header.flight_no)

    # fetch existing details
    existing_result = await db.execute(
        select(ExportUldAssignmentDetail).where(
            ExportUldAssignmentDetail.assignment_id == assignment_id
        )
    )
    existing_details = {d.id: d for d in existing_result.scalars().all()}
    existing_uld_ids = {d.uld_id for d in existing_details.values()}

    errors = []

    # ✅ block ULD removal if items already loaded
    if payload.uld_detail_ids_to_remove:
        loaded_result = await db.execute(
            select(
                ExportSequenceItemUldLoading.uld_assignment_detail_id,
                func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
                ExportUldMaster.uld_no,
            )
            .join(
                ExportUldAssignmentDetail,
                ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id,
            )
            .join(
                ExportUldMaster,
                ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
            )
            .where(
                ExportSequenceItemUldLoading.uld_assignment_detail_id.in_(
                    payload.uld_detail_ids_to_remove
                )
            )
            .group_by(
                ExportSequenceItemUldLoading.uld_assignment_detail_id,
                ExportUldMaster.uld_no,
            )
        )
        loaded_rows = loaded_result.mappings().all()

        block_errors = [
            f"ULD {row.uld_no} cannot be removed — "
            f"{row.loaded_count} item(s) already loaded into it"
            for row in loaded_rows
            if row.loaded_count > 0
        ]

        if block_errors:
            raise HTTPException(status_code=400, detail=block_errors)

    # ── validate new ULDs to add ───────────────────────────────
    if payload.uld_ids_to_add:
        already_assigned = set(payload.uld_ids_to_add) & existing_uld_ids
        if already_assigned:
            dup_result = await db.execute(
                select(ExportUldMaster.uld_no).where(
                    ExportUldMaster.id.in_(already_assigned)
                )
            )
            dup_nos = [r.uld_no for r in dup_result.all()]
            errors.append(f"ULDs already assigned to this flight: {', '.join(dup_nos)}")

        uld_result = await db.execute(
            select(ExportUldMaster.id).where(
                ExportUldMaster.id.in_(payload.uld_ids_to_add),
                ExportUldMaster.is_active == True,
            )
        )
        valid_ids = {row.id for row in uld_result.all()}
        invalid = set(payload.uld_ids_to_add) - valid_ids
        if invalid:
            errors.append(f"ULD ids not found or inactive: {sorted(invalid)}")

        # if not errors:
        #     conflicts = await _check_ulds_not_active_on_another_flight(
        #         db=db,
        #         uld_ids=payload.uld_ids_to_add,
        #         exclude_assignment_id=assignment_id,
        #     )
        #     if conflicts:
        #         errors.extend(conflicts)

    available_result = await db.execute(
    select(ExportUldMaster.id).where(
        ExportUldMaster.id.in_(payload.uld_ids_to_add),
        ExportUldMaster.is_available == True,
    )
    )

    available_ids = {row.id for row in available_result.all()}
    not_available = set(payload.uld_ids_to_add) - available_ids

    if not_available:
        errors.append(f"ULD not available: {sorted(not_available)}")

    # ── validate remove ids belong to this assignment ──────────
    if payload.uld_detail_ids_to_remove:
        invalid_removes = set(payload.uld_detail_ids_to_remove) - set(existing_details.keys())
        if invalid_removes:
            errors.append(f"Detail ids do not belong to this assignment: {sorted(invalid_removes)}")

        remaining_after_remove = len(existing_details) - len(payload.uld_detail_ids_to_remove)
        net_adds = len(payload.uld_ids_to_add)
        if remaining_after_remove + net_adds < 1:
            errors.append("At least one ULD must remain in the assignment")

    if errors:
        raise HTTPException(status_code=400, detail=errors)

    # ── delete removed details ─────────────────────────────────
    removed_uld_ids = []
    for detail_id in payload.uld_detail_ids_to_remove:
        if detail_id in existing_details:
            removed_uld_ids.append(existing_details[detail_id].uld_id)
            await db.delete(existing_details[detail_id])

    # ── insert new ULDs ────────────────────────────────────────
    db.add_all([
        ExportUldAssignmentDetail(
            assignment_id=assignment_id,
            uld_id=uld_id,
            created_at=now,
        )
        for uld_id in payload.uld_ids_to_add
    ])

    # ✅ mark added ULDs as unavailable
    if payload.uld_ids_to_add:
        await db.execute(
            update(ExportUldMaster)
            .where(ExportUldMaster.id.in_(payload.uld_ids_to_add))
            .values(is_available=False)
        )

    # ✅ mark removed ULDs as available (MOVED HERE)
    if removed_uld_ids:
        await db.execute(
            update(ExportUldMaster)
            .where(ExportUldMaster.id.in_(removed_uld_ids))
            .values(is_available=True)
        )

    assignment.updated_at = now

    # ====== (NO CHANGE BELOW — your logs remain same 😎) ======

    existing_uld_id_map = {
        d_id: d.uld_id
        for d_id, d in existing_details.items()
    }

    all_relevant_uld_ids = list(
        set(payload.uld_ids_to_add) |
        {existing_uld_id_map[did] for did in payload.uld_detail_ids_to_remove if did in existing_uld_id_map}
    )

    uld_info_map = {}
    if all_relevant_uld_ids:
        uld_info_result = await db.execute(
            select(
                ExportUldMaster.id,
                ExportUldMaster.uld_no,
                ExportUldMaster.carrier,
            ).where(ExportUldMaster.id.in_(all_relevant_uld_ids))
        )
        uld_info_map = {row.id: row for row in uld_info_result.mappings().all()}

    awb_ids_result = await db.execute(
        select(ExportFlightBookingDetail.awb_master_id).where(
            ExportFlightBookingDetail.flight_header_id == assignment.flight_header_id
        )
    )
    awb_ids_on_flight = [row.awb_master_id for row in awb_ids_result.all()]

    added_ulds = [
        {
            "uld_id": uid,
            "uld_no": uld_info_map[uid].uld_no,
            "carrier": uld_info_map[uid].carrier,
        }
        for uid in payload.uld_ids_to_add
        if uid in uld_info_map
    ]

    removed_ulds = [
        {
            "uld_id": existing_uld_id_map[did],
            "uld_no": uld_info_map[existing_uld_id_map[did]].uld_no,
        }
        for did in payload.uld_detail_ids_to_remove
        if did in existing_uld_id_map
        and existing_uld_id_map[did] in uld_info_map
    ]

    added_uld_nos = [u["uld_no"] for u in added_ulds]
    removed_uld_nos = [u["uld_no"] for u in removed_ulds]

    for awb_id in awb_ids_on_flight:
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=awb_id,
            flight_reference_id=assignment.flight_header_id,
            module=CarMessageFlowModule.ULD_ASSIGNMENT,
            flow_step=CarMessageFlowStep.ULD_ASSIGNMENT,
            record_id=assignment_id,
            action="UPDATE",
            performed_by=edited_by,
            changes={
                "event": "ULD_ASSIGNMENT_UPDATED",
                "flight_no": header.flight_no,
                "flight_date": str(header.flight_date),
                "added_ulds": added_ulds,
                "removed_ulds": removed_ulds,
                "summary": (
                    f"Flight {header.flight_no} ({header.flight_date}) "
                    "ULD assignment updated — "
                    + (f"Added: {', '.join(added_uld_nos)}. " if added_uld_nos else "No ULDs added. ")
                    + (f"Removed: {', '.join(removed_uld_nos)}." if removed_uld_nos else "No ULDs removed.")
                ),
            },
        )

    await db.commit()
    await db.refresh(assignment)

    data = await _build_assignment_response(db, assignment, header)

    return UldAssignmentResponse(
        success=True,
        message=f"ULD assignment updated for flight {header.flight_no}",
        data=data,
    )
# ============ END ULD =========================


# ======================✌️Skid Retrival from location ===================
async def get_flights_by_date(
    db: AsyncSession,
    flight_date: date,
) -> list[dict]:

    now = get_utc_now()
    result = await db.execute(
        select(
            ExportFlightBookingHeader.id.label("header_id"),
            ExportFlightBookingHeader.flight_no,
            ExportFlightBookingHeader.flight_date,
            ExportFlightBookingHeader.flight_dpt_datetime,
            ExportFlightBookingHeader.booked_by,
            ExportFlightBookingHeader.booked_at,
            func.count(ExportFlightBookingDetail.id).label("total_awbs"),
            func.sum(ExportFlightBookingDetail.booked_pcs).label("total_pcs"),
        )
        .join(
            ExportFlightBookingDetail,
            ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
        )
        .where(
            ExportFlightBookingHeader.flight_date == flight_date,
            ExportFlightBookingHeader.is_active == True,
            ExportFlightBookingHeader.flight_dpt_datetime > now,  # ✅ not yet departed
        )
        .group_by(ExportFlightBookingHeader.id)
        .order_by(ExportFlightBookingHeader.flight_dpt_datetime.asc())
    )
    rows = result.mappings().all()

    return [dict(row) for row in rows]


async def get_flight_full_detail(
    db: AsyncSession,
    header_id: int,
) -> dict:

    # ── Fetch header ───────────────────────────────────────────
    header = await db.get(ExportFlightBookingHeader, header_id)
    if not header or not header.is_active:
        raise HTTPException(status_code=404, detail="Flight booking not found")

    # ── Fetch AWBs booked on this flight ───────────────────────
    awb_result = await db.execute(
        select(
            ExportFlightBookingDetail.id.label("detail_id"),
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.id.label("awb_master_id"),
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
            ExportCarMessageAwbMaster.gross_wt,
            ExportCarMessageAwbMaster.chg_wt,
            ExportCarMessageAwbMaster.nog,
            ExportCarMessageAwbMaster.shc,
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.rcs_datetime,
            # ✅ Add the ultra-fast flag here
            ExportCarMessageAwbMaster.is_ultra_fast,
            ExportCarMessageAwbMaster.is_manually_created,
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(ExportFlightBookingDetail.flight_header_id == header_id)
        .order_by(ExportCarMessageAwbMaster.awb_no)
    )
    awb_rows = awb_result.mappings().all()

    if not awb_rows:
        raise HTTPException(status_code=404, detail="No AWBs found for this flight")

    awb_master_ids = [row.awb_master_id for row in awb_rows]



    # ── Fetch skid mappings for all AWBs in one query ──────────
    skid_result = await db.execute(
        select(
            ExportAwbSkidMapping.id.label("mapping_id"),
            ExportAwbSkidMapping.awb_master_id,
            ExportAwbSkidMapping.skid_id,
            ExportAwbSkidMapping.virtual_skid_no,
            ExportAwbSkidMapping.is_virtual,
            ExportAwbSkidMapping.mapped_by,
            ExportAwbSkidMapping.mapped_at,
            ExportAwbSkidMapping.is_skid_used_complete,
             ExportAwbSkidMapping.created_at.label("mapping_created_at"),
            ExportSkidMaster.skid_no,
        )
        .outerjoin(
            ExportSkidMaster,
            ExportAwbSkidMapping.skid_id == ExportSkidMaster.id,
        )
        .where(ExportAwbSkidMapping.awb_master_id.in_(awb_master_ids))
        .order_by(ExportAwbSkidMapping.awb_master_id)
    )
    skid_rows = skid_result.mappings().all()

    mapping_ids = [row.mapping_id for row in skid_rows]

    # ── Fetch sequences for all mappings in one query ──────────
    seq_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.id.label("sequence_id"),
            ExportAwbSkidItemSequence.mapping_id,
            ExportAwbSkidItemSequence.awb_master_id,
            ExportAwbSkidItemSequence.sequence_no,
            ExportAwbSkidItemSequence.sequence_date_time,
            ExportAwbSkidItemSequence.scanned_by,
            ExportAwbSkidItemSequence.scan_by_device,
        )
        .where(ExportAwbSkidItemSequence.mapping_id.in_(mapping_ids))
        .order_by(
            ExportAwbSkidItemSequence.mapping_id,
            ExportAwbSkidItemSequence.sequence_date_time.asc(),
        )
    ) if mapping_ids else None

    seq_rows = seq_result.mappings().all() if seq_result else []

    # ── Fetch current locations for all skids in one query ────
    location_result = await db.execute(
        select(
            ExportSkidLocationMapping.skid_id,
            ExportSkidLocationMapping.assigned_at,
            ExportSkidLocationMapping.assigned_by,
            ExportSkidLocationMapping.picked_at,
            ExportSkidLocationMapping.picked_by,
            ExportSkidLocationMapping.is_relocation,
            ExportLocationsMaster.id.label("location_id"),
            ExportLocationsMaster.loc,
            ExportLocationsMaster.area_code,
        )
        .join(
            ExportLocationsMaster,
            ExportSkidLocationMapping.location_id == ExportLocationsMaster.id,
        )
        .where(
            ExportSkidLocationMapping.skid_id.in_(
                [row.skid_id for row in skid_rows if row.skid_id]
            ),
            ExportSkidLocationMapping.mapping_id.in_(mapping_ids),# 😂😂
            ExportSkidLocationMapping.is_current == True,   # only current location
        )
    ) if any(row.skid_id for row in skid_rows) else None

    location_rows = location_result.mappings().all() if location_result else []

     # ──🤢 loaded pcs per AWB for this flight======== ────────────────────
    loaded_per_awb_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.awb_master_id,
            func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
        )
        .where(ExportSequenceItemUldLoading.flight_header_id == header_id)
        .group_by(ExportSequenceItemUldLoading.awb_master_id)
    ) if mapping_ids else None

    loaded_per_awb: dict[int, int] = {
        row.awb_master_id: row.loaded_pcs
        for row in (loaded_per_awb_result.mappings().all() if loaded_per_awb_result else [])
    }
    # -------

# =========== activity log releted ---
    # ── Fetch full location history for all skids ──────────────
    location_history_result = await db.execute(
        select(
            ExportSkidLocationMapping.skid_id,
            ExportSkidLocationMapping.assigned_at,
             ExportSkidLocationMapping.mapping_id,
            ExportSkidLocationMapping.assigned_by,
            ExportSkidLocationMapping.picked_at,
            ExportSkidLocationMapping.picked_by,
            ExportSkidLocationMapping.is_relocation,
            ExportSkidLocationMapping.is_current,
            ExportLocationsMaster.loc,
            ExportLocationsMaster.area_code,
        )
        .join(
            ExportLocationsMaster,
            ExportSkidLocationMapping.location_id == ExportLocationsMaster.id,
        )
        .where(
            ExportSkidLocationMapping.skid_id.in_(
                [row.skid_id for row in skid_rows if row.skid_id]
            ),
            ExportSkidLocationMapping.mapping_id.in_(mapping_ids),  
        )
        .order_by(
            ExportSkidLocationMapping.skid_id,
            ExportSkidLocationMapping.assigned_at.asc(),  # asc for activity timeline
        )
    ) if any(row.skid_id for row in skid_rows) else None

    location_history_rows = location_history_result.mappings().all() if location_history_result else []

    location_history_by_skid: dict[int, list] = {}
    for loc in location_history_rows:
        location_history_by_skid.setdefault(loc.skid_id, []).append(loc)
# ==========-------------------

#===================✌️  GET THE DATA DROPED AT BASE 
    base_drop_result = await db.execute(
        select(
            ExportSkidBaseMapping.mapping_id,
            ExportSkidBaseMapping.skid_id,
            ExportSkidBaseMapping.dropped_by,
             ExportSkidBaseMapping.cycle_no,
            ExportSkidBaseMapping.dropped_at,
            ExportBaseMaster.base_name,
            ExportBaseMaster.id.label("base_id"),
        )
        .join(ExportBaseMaster, ExportSkidBaseMapping.base_id == ExportBaseMaster.id)
        .where(ExportSkidBaseMapping.mapping_id.in_(mapping_ids))
        .order_by(ExportSkidBaseMapping.cycle_no.asc())   # ✅ ADD — chronological
    ) if mapping_ids else None

    base_drop_rows = base_drop_result.mappings().all() if base_drop_result else []

    # keyed by mapping_id
    # base_drop_by_mapping: dict[int, dict] = {
    #     row.mapping_id: dict(row)
    #     for row in base_drop_rows
    # }

    # ✅ new — list per mapping_id (keeps all cycles)
    base_drop_by_mapping: dict[int, list] = {}
    for row in base_drop_rows:
        base_drop_by_mapping.setdefault(row.mapping_id, []).append(dict(row))
# ------------

# ── 🤢 ─────────────────
    # ── loaded pcs per mapping — GLOBAL (across all flights) ──────
    # loaded_result = await db.execute(
    #     select(
    #         ExportSequenceItemUldLoading.mapping_id,
    #         func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
    #     )
    #     .where(ExportSequenceItemUldLoading.mapping_id.in_(mapping_ids))
    #     .group_by(ExportSequenceItemUldLoading.mapping_id)
    # ) if mapping_ids else None

    loaded_result = await db.execute(
    select(
        ExportAwbSkidItemSequence.mapping_id,
        func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
    )
    .join(
        ExportSequenceItemUldLoading,
        ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
    )
    .where(
        ExportAwbSkidItemSequence.mapping_id.in_(mapping_ids)
    )
    .group_by(ExportAwbSkidItemSequence.mapping_id)
)
    
    loaded_by_mapping_global: dict[int, int] = {
        row.mapping_id: row.loaded_pcs
        for row in (loaded_result.mappings().all() if loaded_result else [])
    }

    # ── loaded pcs per mapping — THIS FLIGHT ONLY ─────────────────
    # loaded_this_flight_result = await db.execute(
    #     select(
    #         ExportSequenceItemUldLoading.mapping_id,
    #         func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
    #     )
    #     .where(
    #         ExportSequenceItemUldLoading.flight_header_id == header_id,
    #         ExportSequenceItemUldLoading.mapping_id.in_(mapping_ids),
    #     )
    #     .group_by(ExportSequenceItemUldLoading.mapping_id)
    # ) if mapping_ids else None

    loaded_this_flight_result = await db.execute(
    select(
        ExportAwbSkidItemSequence.mapping_id,
        func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
    )
    .join(
        ExportSequenceItemUldLoading,
        ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
    )
    .where(
        ExportSequenceItemUldLoading.flight_header_id == header_id,
        ExportAwbSkidItemSequence.mapping_id.in_(mapping_ids),
    )
    .group_by(ExportAwbSkidItemSequence.mapping_id)
)
    
    loaded_by_mapping_this_flight: dict[int, int] = {
        row.mapping_id: row.loaded_pcs
        for row in (loaded_this_flight_result.mappings().all() if loaded_this_flight_result else [])
    }


# ------->
    # location keyed by skid_id — one current location per skid
    location_by_skid: dict[int, dict] = {
        row.skid_id: {
            "location_id": row.location_id,
            "location_name": row.loc,
            "location_code": row.area_code,
            "assigned_at": row.assigned_at,
            "assigned_by": row.assigned_by,
            "picked_at": row.picked_at,
            "picked_by": row.picked_by,
            "is_relocation": row.is_relocation,
        }
        for row in location_rows
    }


    # ── ULD assignment for this flight ─────────────────────────
    uld_result = await db.execute(
        select(
            ExportUldAssignment.id.label("assignment_id"),
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
            ExportUldAssignmentDetail.uld_id,
        )
        .join(
            ExportUldAssignmentDetail,
            ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id,
        )
        .join(
            ExportUldMaster,
            ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
        )
        .where(
            ExportUldAssignment.flight_header_id == header_id,
            ExportUldAssignment.is_active == True,
        )
        .order_by(ExportUldMaster.uld_no)
    )
    uld_rows = uld_result.mappings().all()

    # ── Build nested structure in memory ───────────────────────

    # sequences grouped by mapping_id
    seq_by_mapping: dict[int, list] = {}
    for seq in seq_rows:
        seq_by_mapping.setdefault(seq.mapping_id, []).append({
            "sequence_id": seq.sequence_id,
            "sequence_no": seq.sequence_no,
            "sequence_date_time": seq.sequence_date_time,
            "scanned_by": seq.scanned_by,
            "scan_by_device": seq.scan_by_device,
        })

    # skids grouped by awb_master_id
    skids_by_awb: dict[int, list] = {}
    for skid in skid_rows:

        location_history = location_history_by_skid.get(skid.skid_id, []) if skid.skid_id else []  # ← ADD
        base_drop = base_drop_by_mapping.get(skid.mapping_id) 

        skid_scanned_pcs = len(seq_by_mapping.get(skid.mapping_id, []))
        skid_loaded_pcs_global = loaded_by_mapping_global.get(skid.mapping_id, 0)       # ✅ all flights
        skid_loaded_pcs_this_flight = loaded_by_mapping_this_flight.get(skid.mapping_id, 0)  # ✅ this flight
        skid_remaining_pcs = skid_scanned_pcs - skid_loaded_pcs_global   
        
        skids_by_awb.setdefault(skid.awb_master_id, []).append({
            "mapping_id": skid.mapping_id,
            "skid_id": skid.skid_id,
            "skid_no": skid.skid_no,
            "virtual_skid_no": skid.virtual_skid_no,
            "is_virtual": skid.is_virtual,
            "is_skid_used_complete": skid.is_skid_used_complete,
            "sequences": seq_by_mapping.get(skid.mapping_id, []),
            "scanned_pcs": len(seq_by_mapping.get(skid.mapping_id, [])),
            "current_location": location_by_skid.get(skid.skid_id),

            # 🤢 clear naming
            "loaded_pcs_this_flight": skid_loaded_pcs_this_flight,   # how many loaded into THIS flight ULD
            "loaded_pcs_total": skid_loaded_pcs_global,               # how many loaded across ALL flights
            "remaining_pcs": skid_remaining_pcs,                      # scanned - total loaded = truly remaining
            "can_relocate": (
                not skid.is_skid_used_complete
                and skid_remaining_pcs > 0
                and skid_loaded_pcs_global > 0    # ✅ at least something loaded globally
            ),


            # ✅ only 3 events now
            "activity_log": _build_skid_activity_log(
                skid=dict(skid),
                mapping_row=skid,
                location_history=location_history_by_skid.get(skid.skid_id, []) if skid.skid_id else [],
                base_drop=base_drop,        
            ),
           
             "retrieval_status": _get_skid_retrieval_status(    # ← ADD
                location_history=location_history,
                base_drop=base_drop,
    ),
        })

    # assemble AWBs
    awbs = []
    for awb in awb_rows:
        skids = skids_by_awb.get(awb.awb_master_id, [])
        total_scanned = sum(s["scanned_pcs"] for s in skids)
        loaded_for_flight = loaded_per_awb.get(awb.awb_master_id, 0)
        is_awb_fully_loaded_for_flight = loaded_for_flight >= awb.booked_pcs
        awbs.append({
            "detail_id": awb.detail_id,
            "awb_master_id": awb.awb_master_id,
            "awb_no": awb.awb_no,
            "origin": awb.origin,
            "destination": awb.destination,
            "total_pcs": awb.total_pcs,
            "booked_pcs": awb.booked_pcs,
            "gross_wt": awb.gross_wt,
            "chg_wt": awb.chg_wt,
            "nog": awb.nog,
            "shc": awb.shc,
            "agent": awb.agent,
            "status": awb.status,
            "rcs_datetime": awb.rcs_datetime,
            "total_scanned_pcs": total_scanned,
            "is_ultra_fast": awb.is_ultra_fast,
            "is_manually_created": awb.is_manually_created,

            "loaded_pcs_this_flight": loaded_for_flight,           # ✅ ADD
            "is_fully_loaded_for_flight": is_awb_fully_loaded_for_flight,  # ✅ ADD — frontend uses this to hide retrieve button

            "skids": skids,
        })

    return {
        "header_id": header.id,
        "flight_no": header.flight_no,
        "flight_date": str(header.flight_date),
        "flight_dpt_datetime": header.flight_dpt_datetime,
        "booked_by": header.booked_by,
        "booked_at": header.booked_at,
        "total_awbs": len(awbs),
        "total_pcs": sum(a["booked_pcs"] for a in awbs),
        "total_scanned_pcs": sum(a["total_scanned_pcs"] for a in awbs),
        "ulds": [dict(u) for u in uld_rows],
        "awbs": awbs,
    }



async def retrieve_skid_from_location(
    db: AsyncSession,
    mapping_id: int,
    retrieved_by: str,
) -> dict:

    now = get_utc_now()

    # ── Fetch mapping with skid info ───────────────────────────
    mapping_result = await db.execute(
        select(
            ExportAwbSkidMapping.id.label("mapping_id"),
            ExportAwbSkidMapping.skid_id,
            ExportAwbSkidMapping.awb_master_id,
            ExportAwbSkidMapping.is_virtual,
            ExportAwbSkidMapping.virtual_skid_no,
            ExportSkidMaster.skid_no,
        )
        .join(
            ExportSkidMaster,
            ExportAwbSkidMapping.skid_id == ExportSkidMaster.id,
        )
        .where(ExportAwbSkidMapping.id == mapping_id)
    )
    mapping = mapping_result.mappings().one_or_none()

    if not mapping:
        raise HTTPException(
            status_code=404,
            detail=f"Mapping id {mapping_id} not found",
        )

    # ── Fetch most recent active location for this skid ────────
    location_result = await db.execute(
        select(ExportSkidLocationMapping)
        .where(
            ExportSkidLocationMapping.skid_id == mapping.skid_id,
            # 🤢
                ExportSkidLocationMapping.mapping_id == mapping_id,  # ensure location belongs to this mapping
            ExportSkidLocationMapping.is_current == True,
        )
        .order_by(ExportSkidLocationMapping.assigned_at.desc())
        .limit(1)
    )
    location = location_result.scalar_one_or_none()

    if not location:
        raise HTTPException(
            status_code=404,
            detail=(
                f"Skid {mapping.skid_no} has no active location — "
                "already retrieved or never placed at a location"
            ),
        )

    # ── Fetch location master for readable info ────────────────
    location_master_result = await db.execute(
        select(ExportLocationsMaster).where(
            ExportLocationsMaster.id == location.location_id
        )
    )
    location_master = location_master_result.scalar_one_or_none()

    # ── Mark as retrieved ──────────────────────────────────────
    location.is_current = False
    location.picked_at = now
    location.picked_by = retrieved_by

    # ── Audit log ──────────────────────────────────────────────
    await write_car_message_flow_audit(
        db=db,
        awb_reference_id=location.awb_master_id,
        flight_reference_id=None,
        module=CarMessageFlowModule.SKID_RETRIEVAL,      # ← was LOCATION_MAPPING
        flow_step=CarMessageFlowStep.SKID_RETRIEVAL,     # ← was STEP_LOCATION_MAPPING
        record_id=location.id,
        action="UPDATE",
        performed_by=retrieved_by,
        changes={
            "event": "SKID_RETRIEVED",
            "mapping_id": mapping_id,
            "skid_id": mapping.skid_id,
            "skid_no": mapping.skid_no,
            "is_virtual": mapping.is_virtual,
            "virtual_skid_no": mapping.virtual_skid_no,
            "location_id": location.location_id,
            "location_code": location_master.area_code if location_master else None,
            "location_name": location_master.loc if location_master else None,
            "summary": (
                f"Skid {mapping.skid_no} retrieved from "
                f"{location_master.area_code if location_master else location.location_id} "
                f"by {retrieved_by}"
            ),
        },
    )

    await db.commit()

    return {
        "success": True,
        "message": (
            f"Skid {mapping.skid_no} successfully retrieved from "
            f"{location_master.area_code if location_master else 'location'}"
        ),
        "data": {
            "mapping_id": mapping_id,
            "skid_id": mapping.skid_id,
            "skid_no": mapping.skid_no,
            "is_virtual": mapping.is_virtual,
            "virtual_skid_no": mapping.virtual_skid_no,
            "awb_master_id": mapping.awb_master_id,
            "location_id": location.location_id,
            "location_code": location_master.area_code if location_master else None,
            "location_name": location_master.loc if location_master else None,
            "retrieved_by": retrieved_by,
            "retrieved_at": now,
        },
    }



# ================== 👌👌✌️ EXPORT ULD/PALLET LOADING BY SCANNING SEQUENCE [LAST STEP OF PROCESS]================= 


# ── 1. Verify ULD belongs to flight ───────────────────────
async def verify_uld_for_loading(
    db: AsyncSession,
    flight_header_id: int,
    uld_no: str,
) -> UldVerifyForLoadingResponse:

    uld_no = uld_no.strip().upper()

    # verify ULD exists and belongs to this flight
    result = await db.execute(
        select(
            ExportUldAssignmentDetail.id.label("uld_assignment_detail_id"),
            ExportUldMaster.id.label("uld_id"),
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        )
        .join(ExportUldAssignment,
              ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id)
        .join(ExportUldMaster,
              ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
        .where(
            ExportUldAssignment.flight_header_id == flight_header_id,
            ExportUldAssignment.is_active == True,
            ExportUldMaster.uld_no == uld_no,
        )
    )
    row = result.mappings().first()

    if not row:
        raise HTTPException(
            status_code=404,
            detail=f"ULD '{uld_no}' is not assigned to this flight",
        )

    # count already loaded items for this ULD
    loaded_result = await db.execute(
        select(func.count(ExportSequenceItemUldLoading.id)).where(
            ExportSequenceItemUldLoading.uld_assignment_detail_id == row.uld_assignment_detail_id,
        )
    )
    already_loaded = loaded_result.scalar() or 0

    return UldVerifyForLoadingResponse(
        success=True,
        message=f"ULD '{uld_no}' verified — {already_loaded} items already loaded",
        uld_assignment_detail_id=row.uld_assignment_detail_id,
        uld_id=row.uld_id,
        uld_no=row.uld_no,
        carrier=row.carrier,
        already_loaded=already_loaded,
    )


# ── 2. 🤢 Scan item into ULD (comment on 10 APR 6:52) ────────────────────────────────── 
# async def scan_item_into_uld(
#     db: AsyncSession,
#     flight_header_id: int,
#     payload: ScanItemIntoUldRequest,
#     loaded_by: str,
# ) -> ScanItemIntoUldResponse:

#     now = get_utc_now()

#     # ── verify flight ──────────────────────────────────────
#     flight = await db.get(ExportFlightBookingHeader, flight_header_id)
#     if not flight or not flight.is_active:
#         raise HTTPException(status_code=404, detail="Flight not found")

#     if flight.flight_dpt_datetime <= now:
#         raise HTTPException(status_code=400, detail="Flight has already departed")

#     # ── verify ULD belongs to this flight ─────────────────
#     uld_detail_result = await db.execute(
#         select(
#             ExportUldAssignmentDetail.id,
#             ExportUldMaster.uld_no,
#         )
#         .join(ExportUldAssignment,
#               ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id)
#         .join(ExportUldMaster,
#               ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
#         .where(
#             ExportUldAssignmentDetail.id == payload.uld_assignment_detail_id,
#             ExportUldAssignment.flight_header_id == flight_header_id,
#             ExportUldAssignment.is_active == True,
#         )
#     )
#     uld_detail = uld_detail_result.mappings().first()
#     if not uld_detail:
#         raise HTTPException(status_code=400, detail="ULD does not belong to this flight")

#     # ── fetch all sequences in one query ───────────────────
#     seq_result = await db.execute(
#         select(
#             ExportAwbSkidItemSequence.id.label("sequence_id"),
#             ExportAwbSkidItemSequence.mapping_id,
#             ExportAwbSkidItemSequence.awb_master_id,
#             ExportAwbSkidItemSequence.sequence_no,
#         )
#         .where(ExportAwbSkidItemSequence.sequence_no.in_(payload.sequence_nos))
#     )
#     seq_map = {row.sequence_no: row for row in seq_result.mappings().all()}

#     # ── fetch flight AWB ids in one query ──────────────────
#     flight_awb_result = await db.execute(
#         select(
#             ExportFlightBookingDetail.awb_master_id,
#             ExportCarMessageAwbMaster.awb_no,
#             ExportFlightBookingDetail.booked_pcs,    # ✅ ADD booked_pcs
#         )
#         .join(ExportCarMessageAwbMaster,
#               ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id)
#         .where(ExportFlightBookingDetail.flight_header_id == flight_header_id)
#     )
#     flight_awb_rows = flight_awb_result.mappings().all()
#     flight_awb_map = {row.awb_master_id: row.awb_no for row in flight_awb_rows}

#     # ✅ ADD — booked_pcs per AWB for this flight
#     booked_pcs_map = {row.awb_master_id: row.booked_pcs for row in flight_awb_rows}

#     # ── fetch already loaded sequences in one query ────────
#     already_loaded_result = await db.execute(
#         select(
#             ExportSequenceItemUldLoading.sequence_id,
#             ExportUldMaster.uld_no.label("loaded_uld_no"),
#         )
#         .join(ExportUldAssignmentDetail,
#               ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id)
#         .join(ExportUldMaster,
#               ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
#         .where(
#             ExportSequenceItemUldLoading.sequence_id.in_(
#                 [s.sequence_id for s in seq_map.values()]
#             )
#         )
#     )
#     already_loaded_map = {
#         row.sequence_id: row.loaded_uld_no
#         for row in already_loaded_result.mappings().all()
#     }

#     # ✅ ADD — loaded count per AWB for this flight (for cap check)
#     loaded_count_result = await db.execute(
#         select(
#             ExportSequenceItemUldLoading.awb_master_id,
#             func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
#         )
#         .where(ExportSequenceItemUldLoading.flight_header_id == flight_header_id)
#         .group_by(ExportSequenceItemUldLoading.awb_master_id)
#     )
#     loaded_count_map = {
#         row.awb_master_id: row.loaded_count
#         for row in loaded_count_result.mappings().all()
#     }

#     # ── fetch base drop status — CHANGED to cycle-aware check ─
#     mapping_ids_in_seq = list({s.mapping_id for s in seq_map.values()})

#     # ✅ CHANGED — base drop must be AFTER last retrieval (cycle-aware)
#     base_dropped_mapping_ids: set[int] = set()
#     if mapping_ids_in_seq:
#         for mid in mapping_ids_in_seq:
#             # get last retrieval picked_at for this mapping's skid
#             last_retrieval_result = await db.execute(
#                 select(ExportSkidLocationMapping.picked_at)
#                 .join(
#                     ExportAwbSkidMapping,
#                     ExportSkidLocationMapping.skid_id == ExportAwbSkidMapping.skid_id,
#                 )
#                 .where(
#                     ExportAwbSkidMapping.id == mid,
#                     ExportSkidLocationMapping.picked_at.isnot(None),
#                 )
#                 .order_by(ExportSkidLocationMapping.picked_at.desc())
#                 .limit(1)
#             )
#             last_retrieval_at = last_retrieval_result.scalar_one_or_none()

#             if not last_retrieval_at:
#                 continue  # never retrieved → not at base

#             # check base drop exists AFTER that retrieval
#             base_check = await db.execute(
#                 select(ExportSkidBaseMapping.id).where(
#                     ExportSkidBaseMapping.mapping_id == mid,
#                     ExportSkidBaseMapping.dropped_at >= last_retrieval_at,
#                 )
#             )
#             if base_check.scalar_one_or_none():
#                 base_dropped_mapping_ids.add(mid)

#     # ── process each sequence_no ───────────────────────────
#     to_insert = []
#     results = []

#     for seq_no in payload.sequence_nos:
#         seq = seq_map.get(seq_no)

#         # ── not found in system
#         if not seq:
#             results.append(ScanItemResult(
#                 sequence_no=seq_no,
#                 awb_no="—",
#                 success=False,
#                 message=f"Item '{seq_no}' not found in system",
#             ))
#             continue

#         # ── not part of this flight
#         awb_no = flight_awb_map.get(seq.awb_master_id)
#         if not awb_no:
#             results.append(ScanItemResult(
#                 sequence_no=seq_no,
#                 awb_no="—",
#                 success=False,
#                 message=f"Item '{seq_no}' does not belong to any AWB on this flight",
#             ))
#             continue

#         # ── skid not dropped at base yet (cycle-aware)
#         if seq.mapping_id not in base_dropped_mapping_ids:
#             results.append(ScanItemResult(
#                 sequence_no=seq_no,
#                 awb_no=awb_no,
#                 success=False,
#                 message=f"Item '{seq_no}' cannot be loaded — skid has not been dropped at base yet",
#             ))
#             continue

#         # ── already loaded in a ULD
#         loaded_uld = already_loaded_map.get(seq.sequence_id)
#         if loaded_uld:
#             results.append(ScanItemResult(
#                 sequence_no=seq_no,
#                 awb_no=awb_no,
#                 success=False,
#                 message=f"Already loaded into ULD {loaded_uld}",
#             ))
#             continue

#         # ✅ ADD — booked_pcs cap check per AWB per flight
#         booked_pcs = booked_pcs_map.get(seq.awb_master_id, 0)
#         already_loaded_count = loaded_count_map.get(seq.awb_master_id, 0)

#         if already_loaded_count >= booked_pcs:
#             results.append(ScanItemResult(
#                 sequence_no=seq_no,
#                 awb_no=awb_no,
#                 success=False,
#                 message=(
#                     f"AWB {awb_no} fully loaded for this flight — "
#                     f"{already_loaded_count}/{booked_pcs} pcs done. "
#                     "Remaining pcs go to next flight."
#                 ),
#             ))
#             continue

#         # ── all good — add to insert list
#         to_insert.append(
#             ExportSequenceItemUldLoading(
#                 flight_header_id=flight_header_id,
#                 uld_assignment_detail_id=payload.uld_assignment_detail_id,
#                 sequence_id=seq.sequence_id,
#                 awb_master_id=seq.awb_master_id,
#                 mapping_id=seq.mapping_id,
#                 loaded_by=loaded_by,
#                 loaded_at=now,
#                 created_at=now,
#             )
#         )
#         results.append(ScanItemResult(
#             sequence_no=seq_no,
#             awb_no=awb_no,
#             success=True,
#             message=f"Loaded into ULD {uld_detail.uld_no}",
#         ))
#         # ✅ ADD — increment in-memory so next seq in same batch is checked correctly
#         loaded_count_map[seq.awb_master_id] = already_loaded_count + 1

#     # ── bulk insert all valid items ────────────────────────
#     if to_insert:
#         db.add_all(to_insert)
#         await db.flush()

#         # ✅ ADD — auto mark skid complete if all its sequences loaded into any ULD
#         inserted_mapping_ids = {item.mapping_id for item in to_insert}
#         for mid in inserted_mapping_ids:
#             total_seq_result = await db.execute(
#                 select(func.count(ExportAwbSkidItemSequence.id)).where(
#                     ExportAwbSkidItemSequence.mapping_id == mid
#                 )
#             )
#             total_seqs = total_seq_result.scalar() or 0

#             loaded_seq_result = await db.execute(
#                 select(func.count(ExportSequenceItemUldLoading.id))
#                 .join(
#                     ExportAwbSkidItemSequence,
#                     ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
#                 )
#                 .where(ExportAwbSkidItemSequence.mapping_id == mid)
#             )
#             loaded_seqs = loaded_seq_result.scalar() or 0

#             if total_seqs > 0 and loaded_seqs >= total_seqs:
#                 # all sequences loaded — mark complete and unlock skid
#                 await db.execute(
#                     update(ExportAwbSkidMapping)
#                     .where(ExportAwbSkidMapping.id == mid)
#                     .values(is_skid_used_complete=True)
#                 )
#                 # fetch skid_id for unlock
#                 skid_id_result = await db.execute(
#                     select(ExportAwbSkidMapping.skid_id)
#                     .where(ExportAwbSkidMapping.id == mid)
#                 )
#                 skid_id = skid_id_result.scalar_one_or_none()
#                 if skid_id:
#                     await db.execute(
#                         update(ExportSkidMaster)
#                         .where(ExportSkidMaster.id == skid_id)
#                         .values(
#                             is_locked=False,
#                             locked_at=None,
#                             locked_by_user_id=None,
#                             updated_at=now,
#                         )
#                     )

#         await db.commit()

#     total_loaded = sum(1 for r in results if r.success)
#     total_failed = sum(1 for r in results if not r.success)

#     return ScanItemIntoUldResponse(
#         success=True,
#         message=f"{total_loaded} loaded, {total_failed} failed",
#         uld_no=uld_detail.uld_no,
#         total_submitted=len(payload.sequence_nos),
#         total_loaded=total_loaded,
#         total_failed=total_failed,
#         results=results,
#     )


# 🤢🤮New scan_item_into_uld with both handling case narmal and utrafast. (Add on on 10 APR 6:52 by commenting above)
async def scan_item_into_uld(
    db: AsyncSession,
    flight_header_id: int,
    payload: ScanItemIntoUldRequest,
    loaded_by: str,
) -> ScanItemIntoUldResponse:

    now = get_utc_now()

    # ── verify flight ──────────────────────────────────────
    flight = await db.get(ExportFlightBookingHeader, flight_header_id)
    if not flight or not flight.is_active:
        raise HTTPException(status_code=404, detail="Flight not found")
    if flight.flight_dpt_datetime <= now:
        raise HTTPException(status_code=400, detail="Flight has already departed")

    # ── verify ULD ─────────────────────────────────────────
    uld_detail_result = await db.execute(
        select(ExportUldAssignmentDetail.id, ExportUldMaster.uld_no, ExportUldAssignmentDetail.is_closed, )
        .join(ExportUldAssignment, ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id)
        .join(ExportUldMaster, ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
        .where(
            ExportUldAssignmentDetail.id == payload.uld_assignment_detail_id,
            ExportUldAssignment.flight_header_id == flight_header_id,
            ExportUldAssignment.is_active == True,
        )
    )
    uld_detail = uld_detail_result.mappings().first()
    if not uld_detail:
        raise HTTPException(status_code=400, detail="ULD does not belong to this flight")
    
    if uld_detail.is_closed:
        raise HTTPException(
            status_code=400,
            detail=f"ULD {uld_detail.uld_no} is already closed"
        )

    # ── fetch flight AWBs ──────────────────────────────────
    flight_awb_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.is_ultra_fast,
        )
        .join(ExportCarMessageAwbMaster,
              ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id)
        .where(ExportFlightBookingDetail.flight_header_id == flight_header_id)
    )
    flight_awb_rows = flight_awb_result.mappings().all()
    flight_awb_map    = {row.awb_master_id: row.awb_no for row in flight_awb_rows}
    booked_pcs_map    = {row.awb_master_id: row.booked_pcs for row in flight_awb_rows}
    ultra_fast_map    = {row.awb_master_id: row.is_ultra_fast for row in flight_awb_rows}
    # ✅ prefix map for ultra-fast barcode AWB detection
    awb_no_to_id_map  = {row.awb_no: row.awb_master_id for row in flight_awb_rows}

    # ── fetch existing sequences in ONE query ──────────────
    seq_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.id.label("sequence_id"),
            ExportAwbSkidItemSequence.mapping_id,
            ExportAwbSkidItemSequence.awb_master_id,
            ExportAwbSkidItemSequence.sequence_no,
        )
        .where(ExportAwbSkidItemSequence.sequence_no.in_(payload.sequence_nos))
    )
    seq_map = {row.sequence_no: row for row in seq_result.mappings().all()}

    # ── split barcodes: normal vs ultra-fast ──────────────
    normal_seq_nos: list[str] = []
    ultra_fast_by_awb: dict[int, list[str]] = {}

    for seq_no in payload.sequence_nos:
        if seq_no in seq_map:
            # normal_seq_nos.append(seq_no)
            seq = seq_map[seq_no]
            # ✅ check if this seq belongs to ultra-fast AWB
            if ultra_fast_map.get(seq.awb_master_id):
                # already created in previous call — handle in ultra-fast path
                ultra_fast_by_awb.setdefault(seq.awb_master_id, []).append(seq_no)
            else:
                normal_seq_nos.append(seq_no)
        else:
            # not in DB — check via AWB prefix
            awb_prefix = seq_no[:11]
            uf_awb_id = awb_no_to_id_map.get(awb_prefix)
            if uf_awb_id and ultra_fast_map.get(uf_awb_id):
                ultra_fast_by_awb.setdefault(uf_awb_id, []).append(seq_no)
            else:
                # unknown barcode — will be caught in normal processing loop
                normal_seq_nos.append(seq_no)

    # ══════════════════════════════════════════════════════
    # NORMAL PATH PROCESSING
    # ══════════════════════════════════════════════════════

    # ── already loaded sequences ───────────────────────────
    normal_seq_ids = [seq_map[s].sequence_id for s in normal_seq_nos if s in seq_map]
    already_loaded_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.sequence_id,
            ExportUldMaster.uld_no.label("loaded_uld_no"),
        )
        .join(ExportUldAssignmentDetail,
              ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id)
        .join(ExportUldMaster, ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
        .where(ExportSequenceItemUldLoading.sequence_id.in_(normal_seq_ids))
    ) if normal_seq_ids else None
    already_loaded_map = {
        row.sequence_id: row.loaded_uld_no
        for row in (already_loaded_result.mappings().all() if already_loaded_result else [])
    }

    # ── loaded count per AWB this flight ───────────────────
    loaded_count_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.awb_master_id,
            func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
        )
        .where(ExportSequenceItemUldLoading.flight_header_id == flight_header_id)
        .group_by(ExportSequenceItemUldLoading.awb_master_id)
    )
    loaded_count_map = {
        row.awb_master_id: row.loaded_count
        for row in loaded_count_result.mappings().all()
    }

    # ── cycle-aware base drop check — ONE batch query ──────
    # get last picked_at per mapping in ONE query
    mapping_ids_in_seq = list({seq_map[s].mapping_id for s in normal_seq_nos if s in seq_map})

    base_dropped_mapping_ids: set[int] = set()
    if mapping_ids_in_seq:
        # get last retrieval per mapping in one query using window function
        last_retrieval_subq = (
            select(
                ExportSkidLocationMapping.mapping_id,
                func.max(ExportSkidLocationMapping.picked_at).label("last_picked_at"),
            )
            .where(
                ExportSkidLocationMapping.mapping_id.in_(mapping_ids_in_seq),
                ExportSkidLocationMapping.picked_at.isnot(None),
            )
            .group_by(ExportSkidLocationMapping.mapping_id)
            .subquery()
        )

        # check base drop after last retrieval in ONE query
        valid_base_result = await db.execute(
            select(ExportSkidBaseMapping.mapping_id)
            .join(
                last_retrieval_subq,
                and_(
                    ExportSkidBaseMapping.mapping_id == last_retrieval_subq.c.mapping_id,
                    ExportSkidBaseMapping.dropped_at >= last_retrieval_subq.c.last_picked_at,
                )
            )
            .where(ExportSkidBaseMapping.mapping_id.in_(mapping_ids_in_seq))
        )
        base_dropped_mapping_ids = {row.mapping_id for row in valid_base_result.all()}

    # ── process normal sequences ───────────────────────────
    to_insert = []
    results: list[ScanItemResult] = []

    for seq_no in normal_seq_nos:
        seq = seq_map.get(seq_no)

        if not seq:
            results.append(ScanItemResult(
                sequence_no=seq_no, awb_no="—",
                success=False, message=f"Item '{seq_no}' not found in system",
            ))
            continue

        awb_no = flight_awb_map.get(seq.awb_master_id)
        if not awb_no:
            results.append(ScanItemResult(
                sequence_no=seq_no, awb_no="—",
                success=False, message=f"Item '{seq_no}' does not belong to any AWB on this flight",
            ))
            continue

        if seq.mapping_id not in base_dropped_mapping_ids:
            results.append(ScanItemResult(
                sequence_no=seq_no, awb_no=awb_no,
                success=False, message=f"Item '{seq_no}' cannot be loaded — skid not dropped at base yet",
            ))
            continue

        loaded_uld = already_loaded_map.get(seq.sequence_id)
        if loaded_uld:
            results.append(ScanItemResult(
                sequence_no=seq_no, awb_no=awb_no,
                success=False, message=f"Already loaded into ULD {loaded_uld}",
            ))
            continue

        booked_pcs = booked_pcs_map.get(seq.awb_master_id, 0)
        already_loaded_count = loaded_count_map.get(seq.awb_master_id, 0)
        if already_loaded_count >= booked_pcs:
            results.append(ScanItemResult(
                sequence_no=seq_no, awb_no=awb_no,
                success=False,
                message=f"AWB {awb_no} fully loaded — {already_loaded_count}/{booked_pcs} pcs done",
            ))
            continue

        to_insert.append(ExportSequenceItemUldLoading(
            flight_header_id=flight_header_id,
            uld_assignment_detail_id=payload.uld_assignment_detail_id,
            sequence_id=seq.sequence_id,
            awb_master_id=seq.awb_master_id,
            mapping_id=seq.mapping_id,
            loaded_by=loaded_by,
            loaded_at=now,
            created_at=now,
        ))
        results.append(ScanItemResult(
            sequence_no=seq_no, awb_no=awb_no,
            success=True, message=f"Loaded into ULD {uld_detail.uld_no}",
        ))
        loaded_count_map[seq.awb_master_id] = already_loaded_count + 1

    # ── bulk insert normal ─────────────────────────────────
    if to_insert:
        db.add_all(to_insert)
        await db.flush()

        # auto complete check
        inserted_mapping_ids = {item.mapping_id for item in to_insert}
        for mid in inserted_mapping_ids:
            total_seqs = await db.scalar(
                select(func.count(ExportAwbSkidItemSequence.id)).where(
                    ExportAwbSkidItemSequence.mapping_id == mid)
            ) or 0
            loaded_seqs = await db.scalar(
                select(func.count(ExportSequenceItemUldLoading.id))
                .join(ExportAwbSkidItemSequence,
                      ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id)
                .where(ExportAwbSkidItemSequence.mapping_id == mid)
            ) or 0
            if total_seqs > 0 and loaded_seqs >= total_seqs:
                await db.execute(
                    update(ExportAwbSkidMapping)
                    .where(ExportAwbSkidMapping.id == mid)
                    .values(is_skid_used_complete=True)
                )
                skid_id = await db.scalar(
                    select(ExportAwbSkidMapping.skid_id)
                    .where(ExportAwbSkidMapping.id == mid)
                )
                if skid_id:
                    await db.execute(
                        update(ExportSkidMaster)
                        .where(ExportSkidMaster.id == skid_id)
                        .values(is_locked=False, locked_at=None,
                                locked_by_user_id=None, updated_at=now)
                    )

    # ══════════════════════════════════════════════════════
    # ULTRA-FAST PATH — batch per AWB in ONE call each
    # ══════════════════════════════════════════════════════
    if ultra_fast_by_awb:
        # get system skid + base once
        system_skid_id = await db.scalar(
            select(ExportSkidMaster.id).where(
                ExportSkidMaster.skid_no == "ULTRAFAST-SYSTEM-SKID")
        )
        system_base_id = await db.scalar(
            select(ExportBaseMaster.id).where(
                ExportBaseMaster.base_name == "ULTRAFAST-AUTO-BASE")
        )
        if not system_skid_id or not system_base_id:
            raise HTTPException(status_code=500,
                detail="Ultra-fast system skid or base not seeded in DB")

        for uf_awb_id, uf_seq_nos in ultra_fast_by_awb.items():
            awb_no = flight_awb_map.get(uf_awb_id, "—")
            booked_pcs = booked_pcs_map.get(uf_awb_id, 0)

            # get or create mapping
            mapping_id = await db.scalar(
                select(ExportAwbSkidMapping.id).where(
                    ExportAwbSkidMapping.awb_master_id == uf_awb_id,
                    ExportAwbSkidMapping.skid_id == system_skid_id,
                    ExportAwbSkidMapping.is_virtual == True,
                )
            )
            if not mapping_id:
                new_mapping = ExportAwbSkidMapping(
                    awb_master_id=uf_awb_id,
                    skid_id=system_skid_id,
                    is_virtual=True,
                    virtual_skid_no="ULTRAFAST",
                    mapped_by=loaded_by,
                    mapped_at=now,
                    created_at=now,
                )
                db.add(new_mapping)
                await db.flush()
                mapping_id = new_mapping.id

            # get or create base drop
            has_base = await db.scalar(
                select(ExportSkidBaseMapping.id).where(
                    ExportSkidBaseMapping.mapping_id == mapping_id)
            )
            if not has_base:
                db.add(ExportSkidBaseMapping(
                    mapping_id=mapping_id,
                    skid_id=system_skid_id,
                    awb_master_id=uf_awb_id,
                    base_id=system_base_id,
                    cycle_no=1,
                    dropped_by="ULTRAFAST-AUTO",
                    dropped_at=now,
                    created_at=now,
                ))
                await db.flush()

            # already loaded count for this AWB this flight
            already_loaded = loaded_count_map.get(uf_awb_id, 0)

            # check existing sequences in batch
            existing_uf_result = await db.execute(
                select(
                    ExportAwbSkidItemSequence.sequence_no,
                    ExportAwbSkidItemSequence.id.label("sequence_id"),
                ).where(ExportAwbSkidItemSequence.sequence_no.in_(uf_seq_nos))
            )
            existing_uf_map = {
                row.sequence_no: row.sequence_id
                for row in existing_uf_result.mappings().all()
            }

            already_loaded_uf_ids: set[int] = set()
            if existing_uf_map:
                loaded_uf_result = await db.execute(
                    select(ExportSequenceItemUldLoading.sequence_id).where(
                        ExportSequenceItemUldLoading.sequence_id.in_(
                            list(existing_uf_map.values()))
                    )
                )
                already_loaded_uf_ids = {row.sequence_id for row in loaded_uf_result.all()}

            uf_to_insert_seqs = []
            seen: set[str] = set()
            in_memory = already_loaded

            for seq_no in uf_seq_nos:
                if seq_no in seen:
                    results.append(ScanItemResult(
                        sequence_no=seq_no, awb_no=awb_no,
                        success=False, message="Duplicate in batch",
                    ))
                    continue
                seen.add(seq_no)

                if seq_no in existing_uf_map:
                    seq_id = existing_uf_map[seq_no]
                    msg = "Already loaded into ULD" if seq_id in already_loaded_uf_ids \
                        else "Sequence already scanned — contact supervisor"
                    results.append(ScanItemResult(
                        sequence_no=seq_no, awb_no=awb_no,
                        success=False, message=msg,
                    ))
                    continue

                if in_memory >= booked_pcs:
                    results.append(ScanItemResult(
                        sequence_no=seq_no, awb_no=awb_no,
                        success=False,
                        message=f"AWB fully loaded — {in_memory}/{booked_pcs} pcs done",
                    ))
                    continue

                uf_to_insert_seqs.append(ExportAwbSkidItemSequence(
                    awb_master_id=uf_awb_id,
                    mapping_id=mapping_id,
                    sequence_no=seq_no,
                    sequence_date_time=now,
                    scanned_by=loaded_by,
                    scan_by_device="ULTRAFAST-ULD-GATE",
                ))
                results.append(ScanItemResult(
                    sequence_no=seq_no, awb_no=awb_no,
                    success=True, message=f"Loaded into ULD {uld_detail.uld_no}",
                ))
                in_memory += 1

            if uf_to_insert_seqs:
                db.add_all(uf_to_insert_seqs)
                await db.flush()

                db.add_all([
                    ExportSequenceItemUldLoading(
                        flight_header_id=flight_header_id,
                        uld_assignment_detail_id=payload.uld_assignment_detail_id,
                        sequence_id=s.id,
                        awb_master_id=uf_awb_id,
                        mapping_id=mapping_id,
                        loaded_by=loaded_by,
                        loaded_at=now,
                        created_at=now,
                    )
                    for s in uf_to_insert_seqs
                ])
                await db.flush()

                if in_memory >= booked_pcs:
                    await db.execute(
                        update(ExportAwbSkidMapping)
                        .where(ExportAwbSkidMapping.id == mapping_id)
                        .values(is_skid_used_complete=True)
                    )

                # sync loaded_count_map for any subsequent normal seq of same AWB
                loaded_count_map[uf_awb_id] = in_memory

    # ── single commit ──────────────────────────────────────
    await db.commit()

    total_loaded = sum(1 for r in results if r.success)
    total_failed = len(results) - total_loaded

    return ScanItemIntoUldResponse(
        success=True,
        message=f"{total_loaded} loaded, {total_failed} failed",
        uld_no=uld_detail.uld_no,
        total_submitted=len(payload.sequence_nos),
        total_loaded=total_loaded,
        total_failed=total_failed,
        results=results,
    )

# ── 3. Get loading status ──────────────────────────────────
async def get_flight_uld_loading_status(
    db: AsyncSession,
    flight_header_id: int,
) -> FlightUldLoadingStatusResponse:

    # ── fetch flight ───────────────────────────────────────
    flight = await db.get(ExportFlightBookingHeader, flight_header_id)
    if not flight or not flight.is_active:
        raise HTTPException(status_code=404, detail="Flight not found")

    # ── fetch all AWBs with booked pcs ─────────────────────
    awb_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.awb_no,
                # ExportCarMessageAwbMaster.origin,
                # ExportCarMessageAwbMaster.destination,
                ExportCarMessageAwbMaster.pcs,
                # ExportCarMessageAwbMaster.gross_wt,
                # ExportCarMessageAwbMaster.chg_wt,
                # ExportCarMessageAwbMaster.nog,
                # ExportCarMessageAwbMaster.shc,
                # ExportCarMessageAwbMaster.agent,
                ExportCarMessageAwbMaster.status,
                # ExportCarMessageAwbMaster.rcs_datetime,
                ExportCarMessageAwbMaster.is_manually_created,
                ExportCarMessageAwbMaster.is_ultra_fast,
        )
        .join(ExportCarMessageAwbMaster,
              ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id)
        .where(ExportFlightBookingDetail.flight_header_id == flight_header_id)
    )
    awb_rows = awb_result.mappings().all()

    awb_ids = [r.awb_master_id for r in awb_rows]
    total_to_load = sum(r.booked_pcs for r in awb_rows)

    # ✅ ADD HERE — after awb_ids, before loaded_per_awb query
    # ── fetch all mapping_ids for AWBs on this flight ──────────
    mapping_ids_result = await db.execute(
        select(ExportAwbSkidMapping.id.label("mapping_id")).where(
            ExportAwbSkidMapping.awb_master_id.in_(awb_ids)
        )
    ) if awb_ids else None

    all_mapping_ids = [
        row.mapping_id
        for row in (mapping_ids_result.all() if mapping_ids_result else [])
    ]

    # ── which of those mappings have been dropped at base ──────
    base_drop_result = await db.execute(
        select(ExportSkidBaseMapping.mapping_id).where(
            ExportSkidBaseMapping.mapping_id.in_(all_mapping_ids)
        )
    ) if all_mapping_ids else None

    base_dropped_mapping_ids = {
        row.mapping_id
        for row in (base_drop_result.all() if base_drop_result else [])
    }

    # 🤢🤮

    # ── fetch mappings where skid still at location (not retrieved yet) ──
    at_location_result = await db.execute(
        select(ExportSkidLocationMapping.mapping_id).where(
            ExportSkidLocationMapping.mapping_id.in_(all_mapping_ids),
            ExportSkidLocationMapping.is_current == True,  # still at location
            ExportSkidLocationMapping.picked_at == None,   # never retrieved
        )
    ) if all_mapping_ids else None

    at_location_mapping_ids = {
        row.mapping_id
        for row in (at_location_result.all() if at_location_result else [])
    }

    # ── loaded pcs per AWB ─────────────────────────────────
    loaded_per_awb_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.awb_master_id,
            func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
        )
        .where(
            ExportSequenceItemUldLoading.flight_header_id == flight_header_id,
            ExportSequenceItemUldLoading.awb_master_id.in_(awb_ids),
        )
        .group_by(ExportSequenceItemUldLoading.awb_master_id)
    )
    loaded_per_awb = {
        r.awb_master_id: r.loaded_pcs
        for r in loaded_per_awb_result.mappings().all()
    }

    # ── loaded pcs per ULD ─────────────────────────────────
    uld_result = await db.execute(
        select(
            ExportUldAssignmentDetail.id.label("uld_assignment_detail_id"),
            ExportUldAssignmentDetail.is_closed,
            ExportUldAssignmentDetail.closed_at,
            ExportUldAssignmentDetail.closed_by,
            ExportUldMaster.id.label("uld_id"),
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
            func.count(ExportSequenceItemUldLoading.id).label("loaded_count"),
        )
        .join(ExportUldAssignment,
              ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id)
        .join(ExportUldMaster,
              ExportUldAssignmentDetail.uld_id == ExportUldMaster.id)
        .outerjoin(ExportSequenceItemUldLoading,
              ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id)
        .where(
            ExportUldAssignment.flight_header_id == flight_header_id,
            ExportUldAssignment.is_active == True,
        )
        .group_by(
            ExportUldAssignmentDetail.id,
            ExportUldMaster.id,
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier,
        )
    )
    uld_rows = uld_result.mappings().all()

    total_loaded = sum(r.loaded_count for r in uld_rows)


    # ── fetch all loaded sequences per ULD to show in screen which are scanned already ────────────────────
    uld_detail_ids = [r.uld_assignment_detail_id for r in uld_rows]

    sequences_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.uld_assignment_detail_id,
            ExportSequenceItemUldLoading.sequence_id,
            ExportSequenceItemUldLoading.awb_master_id,
            ExportSequenceItemUldLoading.loaded_by,
            ExportSequenceItemUldLoading.loaded_at,
            ExportAwbSkidItemSequence.sequence_no,
            ExportCarMessageAwbMaster.awb_no,
        )
        .join(
            ExportAwbSkidItemSequence,
            ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportSequenceItemUldLoading.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(
            ExportSequenceItemUldLoading.flight_header_id == flight_header_id,
            ExportSequenceItemUldLoading.uld_assignment_detail_id.in_(uld_detail_ids),
        )
        .order_by(ExportSequenceItemUldLoading.loaded_at.asc())
    ) if uld_detail_ids else None

    sequences_rows = sequences_result.mappings().all() if sequences_result else []

    # group sequences by uld_assignment_detail_id
    sequences_by_uld: dict[int, list] = {}
    for seq in sequences_rows:
        sequences_by_uld.setdefault(seq.uld_assignment_detail_id, []).append({
            "sequence_id": seq.sequence_id,
            "sequence_no": seq.sequence_no,
            "awb_master_id": seq.awb_master_id,
            "awb_no": seq.awb_no,
            "loaded_by": seq.loaded_by,
            "loaded_at": seq.loaded_at,
        })


    # ── fetch ALL scanned sequences for this flight ────────────
    # all sequences belonging to AWBs booked on this flight
    # regardless of ULD loading status
    all_sequences_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.id.label("sequence_id"),
            ExportAwbSkidItemSequence.sequence_no,
            ExportAwbSkidItemSequence.awb_master_id,
            ExportAwbSkidItemSequence.mapping_id,
            ExportAwbSkidItemSequence.sequence_date_time,
            ExportAwbSkidItemSequence.scanned_by,
            ExportAwbSkidItemSequence.scan_by_device,
            ExportCarMessageAwbMaster.awb_no,
            # ✅ is it loaded into ULD or not
            ExportSequenceItemUldLoading.id.label("loading_id"),
            ExportSequenceItemUldLoading.uld_assignment_detail_id,
            ExportSequenceItemUldLoading.loaded_by,
            ExportSequenceItemUldLoading.loaded_at,
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportAwbSkidItemSequence.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .outerjoin(                                      # ← outerjoin — include unloaded too
            ExportSequenceItemUldLoading,
            and_(
                ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
                ExportSequenceItemUldLoading.flight_header_id == flight_header_id,
            ),
        )
        .where(
            ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids),  # scoped to flight AWBs
        )
        .order_by(ExportAwbSkidItemSequence.sequence_date_time.asc())
    ) if awb_ids else None

    all_sequences_rows = all_sequences_result.mappings().all() if all_sequences_result else []

    # ======🤢🤮
    # # ✅ NEW — globally loaded sequence ids
    # globally_loaded_result = await db.execute(
    #     select(ExportSequenceItemUldLoading.sequence_id).where(
    #         ExportSequenceItemUldLoading.sequence_id.in_(
    #             [row.sequence_id for row in all_sequences_rows]
    #         )
    #     )
    # ) if all_sequences_rows else None

    # globally_loaded_sequence_ids = {
    #     row.sequence_id
    #     for row in (globally_loaded_result.all() if globally_loaded_result else [])
    # }-------------

    # ====== ✅ NEW GLOBAL LOADING MAP (WITH DETAILS)
    global_loading_result = await db.execute(
        select(
            ExportSequenceItemUldLoading.sequence_id,
            ExportFlightBookingHeader.flight_no,
            ExportFlightBookingHeader.flight_date,
            ExportUldMaster.uld_no,
        )
        .join(
            ExportFlightBookingHeader,
            ExportSequenceItemUldLoading.flight_header_id == ExportFlightBookingHeader.id,
        )
        .join(
            ExportUldAssignmentDetail,
            ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id,
        )
        .join(
            ExportUldMaster,
            ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
        )
        .where(
            ExportSequenceItemUldLoading.sequence_id.in_(
                [row.sequence_id for row in all_sequences_rows]
            )
        )
    ) if all_sequences_rows else None


    global_loading_map = {
        row.sequence_id: {
            "flight_no": row.flight_no,
            "flight_date": row.flight_date,
            "uld_no": row.uld_no,
        }
        for row in (global_loading_result.all() if global_loading_result else [])
    }
    # ======
    # ======🤢🤮

    # build flat list with loading status per sequence
    all_sequences = [
        {
            "sequence_id": row.sequence_id,
            "sequence_no": row.sequence_no,
            "awb_master_id": row.awb_master_id,
            "awb_no": row.awb_no,
            "mapping_id": row.mapping_id,
            "sequence_date_time": row.sequence_date_time,
            "scanned_by": row.scanned_by,
            "scan_by_device": row.scan_by_device,
            "is_loaded": row.loading_id is not None,   # ← True/False

             "is_loaded_globally": row.sequence_id in global_loading_map, # ✅ ADD

            "uld_assignment_detail_id": row.uld_assignment_detail_id,         # ← None if not loaded
            "loaded_by": row.loaded_by,
            "loaded_at": row.loaded_at,
            # ✅ ADD — frontend uses this to block scan attempt
           "is_eligible_to_load": (
    row.loading_id is None

     and row.sequence_id not in global_loading_map   # ✅ FIX

    # 🤢🤮
    and row.mapping_id not in at_location_mapping_ids
    
    and row.mapping_id in base_dropped_mapping_ids
),
"ineligible_reason": (
    # "Already loaded into ULD"
    # if row.loading_id is not None
    # else "Skid not dropped at base yet"
    # if row.mapping_id not in base_dropped_mapping_ids
    # else None    # ← eligible — no reason needed
     "Already loaded into ULD"
    if row.loading_id is not None
#     else (
#     f"Already loaded in Flight {global_loading_map[row.sequence_id]['flight_no']} "
#     f"({global_loading_map[row.sequence_id]['flight_date']}) "
#     f"in ULD {global_loading_map[row.sequence_id]['uld_no']}"
# )
else (
        f"Already loaded in Flight {global_loading_map[row.sequence_id]['flight_no']} "
        f"({global_loading_map[row.sequence_id]['flight_date'].strftime('%d-%b-%Y')}) "
        f"in ULD {global_loading_map[row.sequence_id]['uld_no']}"
    )
if row.sequence_id in global_loading_map
    else "Skid is still at location — retrieve first"   # ← ADD 🤮🤢
    if row.mapping_id in at_location_mapping_ids
    else "Skid not dropped at base yet"
    if row.mapping_id not in base_dropped_mapping_ids
    else None
),
        }
        for row in all_sequences_rows
    ]

    total_scanned = len(all_sequences)
    total_loaded_sequences = sum(1 for s in all_sequences if s["is_loaded"])
    total_pending_sequences = total_scanned - total_loaded_sequences

    # ── build response ─────────────────────────────────────
    awb_status = [
        AwbLoadingStatusItem(
            awb_master_id=r.awb_master_id,
            is_ultra_fast=r.is_ultra_fast,
            is_manually_created=r.is_manually_created,
            awb_no=r.awb_no,
            total_pcs=r.pcs,
            booked_pcs=r.booked_pcs,
            loaded_pcs=loaded_per_awb.get(r.awb_master_id, 0),
            pending_pcs=r.booked_pcs - loaded_per_awb.get(r.awb_master_id, 0),
        )
        for r in awb_rows
    ]

    uld_status = [
        UldLoadingStatusItem(
            uld_assignment_detail_id=r.uld_assignment_detail_id,
            uld_id=r.uld_id,
            uld_no=r.uld_no,
            carrier=r.carrier,
            loaded_count=r.loaded_count,
            sequences=sequences_by_uld.get(r.uld_assignment_detail_id, []),  # ← ADD
             # ✅ ADD THESE
        is_closed=r.is_closed,
        closed_by=r.closed_by,
        closed_at=r.closed_at,
        )
        for r in uld_rows
    ]

    return FlightUldLoadingStatusResponse(
        success=True,
        message=f"{total_loaded}/{total_to_load} items loaded",
        flight_header_id=flight.id,
        flight_no=flight.flight_no,
        flight_date=flight.flight_date,
        flight_dpt_datetime=flight.flight_dpt_datetime,
        total_to_load=total_to_load,
        total_loaded=total_loaded,
        total_pending=total_to_load - total_loaded,

         total_scanned=total_scanned,                        # ← ADD
        total_loaded_sequences=total_loaded_sequences,      # ← ADD
        total_pending_sequences=total_pending_sequences,    # ← ADD

        is_fully_loaded=total_loaded >= total_to_load,
        ulds=uld_status,
        awbs=awb_status,
        all_sequences=all_sequences,                        # ← ADD
    )







# ============================================= report car-message =====================


from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import date, timezone
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, and_, outerjoin
from fastapi.responses import StreamingResponse

async def generate_flight_date_report(
    db: AsyncSession,
    # flight_date: date,
    from_date:date,
    to_date:date,
) -> StreamingResponse:

    from sqlalchemy.orm import aliased

    # ── Latest location per mapping subquery ───────────────────
    latest_loc_subq = (
        select(
            ExportSkidLocationMapping.mapping_id,
            func.max(ExportSkidLocationMapping.assigned_at).label("max_assigned_at"),
        )
        .group_by(ExportSkidLocationMapping.mapping_id)
        .subquery()
    )

    LocMapping = aliased(ExportSkidLocationMapping)

    latest_base_subq = (
        select(
            ExportSkidBaseMapping.mapping_id,
            func.max(ExportSkidBaseMapping.dropped_at).label("max_dropped_at"),
        )
        .group_by(ExportSkidBaseMapping.mapping_id)
        .subquery()
    )

    BaseDrop = aliased(ExportSkidBaseMapping)   # ← ADD

    # ── One big query — sequence level with all joins ──────────
    result = await db.execute(
        select(
            # Flight info
            ExportFlightBookingHeader.flight_no,
            ExportFlightBookingHeader.flight_date,
            ExportFlightBookingHeader.flight_dpt_datetime,
            ExportFlightBookingHeader.booked_by.label("flight_booked_by"),
            ExportFlightBookingHeader.booked_at.label("flight_booked_at"),

            # AWB info
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("awb_total_pcs"),
            ExportCarMessageAwbMaster.gross_wt,
            ExportCarMessageAwbMaster.chg_wt,
            ExportCarMessageAwbMaster.nog,
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.status.label("awb_status"),
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportFlightBookingDetail.booked_pcs,

            # Skid info
            ExportSkidMaster.skid_no,
            ExportAwbSkidMapping.is_virtual,
            ExportAwbSkidMapping.virtual_skid_no,
            ExportAwbSkidMapping.is_skid_used_complete,
            ExportAwbSkidMapping.mapped_by,
            ExportAwbSkidMapping.mapped_at,

            # Sequence info
            ExportAwbSkidItemSequence.id.label("sequence_id"),
            ExportAwbSkidItemSequence.sequence_no,
            ExportAwbSkidItemSequence.sequence_date_time,
            ExportAwbSkidItemSequence.scanned_by,
            ExportAwbSkidItemSequence.scan_by_device,

            # Location info
            ExportLocationsMaster.area_code.label("location_code"),
            ExportLocationsMaster.loc.label("location_name"),
            LocMapping.assigned_at.label("location_assigned_at"),
            LocMapping.assigned_by.label("location_assigned_by"),
            LocMapping.picked_at.label("location_picked_at"),
            LocMapping.picked_by.label("location_picked_by"),
            LocMapping.is_current.label("is_at_location"),

            # Base drop info
            ExportBaseMaster.base_name,
            # ExportSkidBaseMapping.dropped_at,
            # ExportSkidBaseMapping.dropped_by,

            BaseDrop.dropped_at,    # ✅ was ExportSkidBaseMapping.dropped_at
            BaseDrop.dropped_by,    # ✅ was ExportSkidBaseMapping.dropped_by

            # ULD info
            ExportUldMaster.uld_no,
            ExportUldMaster.carrier.label("uld_carrier"),
            ExportSequenceItemUldLoading.loaded_by.label("uld_loaded_by"),
            ExportSequenceItemUldLoading.loaded_at.label("uld_loaded_at"),
        )
        .join(
            ExportFlightBookingDetail,
            ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        # .join(
        #     ExportAwbSkidMapping,
        #     # ExportAwbSkidMapping.awb_master_id == ExportCarMessageAwbMaster.id,
        #     ExportAwbSkidMapping.id == ExportAwbSkidItemSequence.mapping_id,

        # )
         .join(
            ExportSequenceItemUldLoading,
            ExportSequenceItemUldLoading.flight_header_id == ExportFlightBookingHeader.id,
        )

      
        # .join(
        #     ExportAwbSkidItemSequence,
        #     ExportAwbSkidItemSequence.mapping_id == ExportAwbSkidMapping.id,
        # )
       

        .join(
            ExportAwbSkidItemSequence,
            ExportAwbSkidItemSequence.id == ExportSequenceItemUldLoading.sequence_id,
        )
        # 3. THEN join mapping (now sequence exists ✅)
        .join(
            ExportAwbSkidMapping,
            ExportAwbSkidMapping.id == ExportAwbSkidItemSequence.mapping_id,
        )
        .join(
            ExportSkidMaster,
            ExportAwbSkidMapping.skid_id == ExportSkidMaster.id,
        )
        # ── Location — most recent only via two-step join ──────
        .outerjoin(
            latest_loc_subq,
            latest_loc_subq.c.mapping_id == ExportAwbSkidMapping.id,
        )
        .outerjoin(
            LocMapping,
            and_(
                LocMapping.mapping_id == ExportAwbSkidMapping.id,
                LocMapping.assigned_at == latest_loc_subq.c.max_assigned_at,
            ),
        )
        .outerjoin(
            ExportLocationsMaster,
            LocMapping.location_id == ExportLocationsMaster.id,
        )
        # ── Base drop ──────────────────────────────────────────
        # .outerjoin(
        #     ExportSkidBaseMapping,
        #     ExportSkidBaseMapping.mapping_id == ExportAwbSkidMapping.id,
        # )
        # .outerjoin(
        #     ExportBaseMaster,
        #     ExportSkidBaseMapping.base_id == ExportBaseMaster.id,
        # )

        # ── Base drop — latest only ────────────────────────────
        .outerjoin(
            latest_base_subq,
            latest_base_subq.c.mapping_id == ExportAwbSkidMapping.id,
        )
        .outerjoin(
            BaseDrop,
            and_(
                BaseDrop.mapping_id == ExportAwbSkidMapping.id,
                BaseDrop.dropped_at == latest_base_subq.c.max_dropped_at,
            ),
        )
        .outerjoin(
            ExportBaseMaster,
            BaseDrop.base_id == ExportBaseMaster.id,
        )



        # ── ULD loading ────────────────────────────────────────
    #     .outerjoin(
    #         ExportSequenceItemUldLoading,
    #         # ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
    #          and_(
    #     ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
    #     ExportSequenceItemUldLoading.flight_header_id == ExportFlightBookingHeader.id,
    # )
           
    #     )
        .outerjoin(
            ExportUldAssignmentDetail,
            ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id,
        )
        .outerjoin(
            ExportUldMaster,
            ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
        )
        .where(
            # ExportFlightBookingHeader.flight_date == flight_date,
            ExportFlightBookingHeader.flight_date >= from_date,
            ExportFlightBookingHeader.flight_date <= to_date,

            ExportFlightBookingHeader.is_active == True,
  ExportSequenceItemUldLoading.awb_master_id == ExportFlightBookingDetail.awb_master_id,
             # 🔥 ADD THIS — restrict sequence to AWB of this flight
    # ExportFlightBookingDetail.awb_master_id == ExportAwbSkidItemSequence.awb_master_id,

    # 🤢
    # 🔥 ADD THIS — avoid sequences from other flights
    # or_(
    #     ExportSequenceItemUldLoading.flight_header_id == ExportFlightBookingHeader.id,
    #     ExportSequenceItemUldLoading.id.is_(None)
    # )
        )
        .order_by(
            ExportFlightBookingHeader.flight_date.asc(), 
            ExportFlightBookingHeader.flight_no,
            ExportCarMessageAwbMaster.awb_no,
            ExportAwbSkidMapping.id,
            ExportAwbSkidItemSequence.sequence_date_time.asc(),
        )
    )

    rows = result.mappings().all()

    # ── Build Excel ────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Report {from_date} to {to_date}"

    # ── Styles ─────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    section_fills = {
        "flight":   PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
        "awb":      PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid"),
        "skid":     PatternFill(start_color="FEF9E7", end_color="FEF9E7", fill_type="solid"),
        "sequence": PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid"),
        "location": PatternFill(start_color="F4ECF7", end_color="F4ECF7", fill_type="solid"),
        "base":     PatternFill(start_color="FDFEFE", end_color="FDFEFE", fill_type="solid"),
        "uld":      PatternFill(start_color="FEF0CD", end_color="FEF0CD", fill_type="solid"),
        "audit":    PatternFill(start_color="EAECEE", end_color="EAECEE", fill_type="solid"),
    }

    # ── Column definitions ─────────────────────────────────────
    columns = [
        # (header label, field key, width, section)
        ("Flight No",           "flight_no",                14, "flight"),
        ("Flight Date",         "flight_date",              14, "flight"),
        ("Departure (IST)",     "flight_dpt_datetime",      22, "flight"),

        ("AWB No",              "awb_no",                   16, "awb"),
        ("Origin",              "origin",                   10, "awb"),
        ("Destination",         "destination",              12, "awb"),
        ("AWB Total Pcs",       "awb_total_pcs",            14, "awb"),
        ("Booked Pcs",          "booked_pcs",               12, "awb"),
        ("Gross Wt",            "gross_wt",                 12, "awb"),
        ("Chg Wt",              "chg_wt",                   12, "awb"),
        ("NOG",                 "nog",                      22, "awb"),
        ("Agent",               "agent",                    14, "awb"),
        ("AWB Status",          "awb_status",               12, "awb"),
        ("RCS DateTime (IST)",  "rcs_datetime",             22, "awb"),

        ("Skid No",             "skid_no",                  16, "skid"),
        # ("Is Virtual",          "is_virtual",               12, "skid"),
        # ("Virtual Skid No",     "virtual_skid_no",          16, "skid"),
        # ("Scan Complete",       "is_skid_used_complete",    14, "skid"),
        ("Mapped By",           "mapped_by",                14, "skid"),
        ("Mapped At (IST)",     "mapped_at",                22, "skid"),

        ("Sequence No",         "sequence_no",              24, "sequence"),
        ("Scan DateTime (IST)", "sequence_date_time",       22, "sequence"),
        ("Scanned By",          "scanned_by",               14, "sequence"),
        # ("Scan Device",         "scan_by_device",           20, "sequence"),

        # ("Location Code",       "location_code",            14, "location"),
        ("Location Name",       "location_name",            22, "location"),
        ("Placed At (IST)",     "location_assigned_at",     22, "location"),
        ("Placed By",           "location_assigned_by",     14, "location"),
        ("Retrieved At (IST)",  "location_picked_at",       22, "location"),
        ("Retrieved By",        "location_picked_by",       14, "location"),
        # ("Still At Location",   "is_at_location",           16, "location"),

        ("Base Name",           "base_name",                14, "base"),
        ("Dropped At (IST)",    "dropped_at",               22, "base"),
        ("Dropped By",          "dropped_by",               14, "base"),

        ("ULD No",              "uld_no",                   16, "uld"),
        # ("ULD Carrier",         "uld_carrier",              12, "uld"),
        ("Loaded At (IST)",     "uld_loaded_at",            22, "uld"),
        ("Loaded By",           "uld_loaded_by",            14, "uld"),

        ("Flight Booked By",    "flight_booked_by",         16, "audit"),
        ("Flight Booked At",    "flight_booked_at",         22, "audit"),
    ]

    # ── Section header row (row 1) ─────────────────────────────
    section_labels = {
        "flight":   "✈ FLIGHT",
        "awb":      "📦 AWB",
        "skid":     "🔲 SKID",
        "sequence": "🔍 SEQUENCE",
        "location": "📍 LOCATION",
        "base":     "🏠 BASE",
        "uld":      "📤 ULD",
        "audit":    "📋 AUDIT",
    }

    section_start: dict[str, int] = {}
    section_end: dict[str, int] = {}
    for col_idx, (_, _, _, section) in enumerate(columns, 1):
        if section not in section_start:
            section_start[section] = col_idx
        section_end[section] = col_idx

    for section, label in section_labels.items():
        if section not in section_start:
            continue
        start = section_start[section]
        end = section_end[section]
        if start != end:
            ws.merge_cells(
                start_row=1, start_column=start,
                end_row=1, end_column=end,
            )
        cell = ws.cell(row=1, column=start, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # ── Column header row (row 2) ──────────────────────────────
    for col_idx, (label, _, width, section) in enumerate(columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.font = Font(bold=True, size=9)
        cell.fill = section_fills[section]
        cell.alignment = center_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 32
    ws.freeze_panes = "A3"

    # ── Format helper ──────────────────────────────────────────
    def fmt(val):
        if val is None:
            return "—"
        if isinstance(val, bool):
            return "Yes" if val else "No"
        if hasattr(val, "tzinfo") and val.tzinfo:
            # from datetime import timedelta
            # ist = val.astimezone(timezone.utc) + timedelta(hours=5, minutes=30)
            # return ist.strftime("%d-%m-%Y %H:%M")
            from zoneinfo import ZoneInfo
            ist = val.astimezone(ZoneInfo("Asia/Kolkata"))
            return ist.strftime("%d-%m-%Y %H:%M")
        if isinstance(val, date) and not hasattr(val, "hour"):
            return val.strftime("%d-%m-%Y")
        return val

    alt_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")

    # ── Data rows ──────────────────────────────────────────────
    for row_idx, row in enumerate(rows, 3):
        for col_idx, (_, field, _, _) in enumerate(columns, 1):
            value = fmt(row.get(field))
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = left_align
            cell.border = thin_border
            cell.font = Font(size=9)
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    # ── Summary row ────────────────────────────────────────────
    if rows:
        summary_row = len(rows) + 4
        ws.cell(
            row=summary_row, column=1,
            value=f"Total Sequences: {len(rows)}"
        ).font = Font(bold=True, size=10)
        ws.cell(
            row=summary_row, column=4,
            value=f"Total AWBs: {len(set(r['awb_no'] for r in rows))}"
        ).font = Font(bold=True, size=10)
        ws.cell(
            row=summary_row, column=len(columns) - 3,
            value=f"Loaded into ULD: {sum(1 for r in rows if r.get('uld_no'))}"
        ).font = Font(bold=True, size=10)

    # ── Stream ─────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    # filename = f"flight_report_{flight_date}.xlsx"
    filename = f"flight_report_{from_date}_to_{to_date}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )

    # ==========================✌️✌️✌️✌️✌️✌️  REPORT DASHBOARD ========================

def ist_date_to_utc_range(ist_date: date) -> tuple[datetime, datetime]:
    """
    Convert IST date to UTC datetime range.
    IST = UTC+5:30
    e.g. 2026-03-20 IST → 2026-03-19 18:30:00 UTC to 2026-03-20 18:29:59 UTC
    """
    ist_tz = ZoneInfo("Asia/Kolkata")
    utc_tz = ZoneInfo("UTC")

    start_ist = datetime(ist_date.year, ist_date.month, ist_date.day, 0, 0, 0, tzinfo=ist_tz)
    end_ist = datetime(ist_date.year, ist_date.month, ist_date.day, 23, 59, 59, 999999, tzinfo=ist_tz)

    return start_ist.astimezone(utc_tz), end_ist.astimezone(utc_tz)


async def get_car_message_dashboard_stats(
    db: AsyncSession,
    report_date: date,
) -> dict:

    # ── Convert IST date to UTC range ──────────────────────────
    day_start_utc, day_end_utc = ist_date_to_utc_range(report_date)  # ← first line

    # ══════════════════════════════════════════════════════════
    # SECTION 1 — AWB MASTER STATS (by car_message_datetime_combo)
    # ══════════════════════════════════════════════════════════
    awb_stats_result = await db.execute(
        select(
            func.count(ExportCarMessageAwbMaster.id).label("total_awb"),
            func.count(ExportCarMessageAwbMaster.id).filter(
                ExportCarMessageAwbMaster.status == "RCS"
            ).label("total_rcs"),
            func.count(ExportCarMessageAwbMaster.id).filter(
                ExportCarMessageAwbMaster.status != "RCS",
                ExportCarMessageAwbMaster.status.isnot(None)
            ).label("total_non_rcs"),
            func.count(ExportCarMessageAwbMaster.id).filter(
                ExportCarMessageAwbMaster.status.is_(None)
            ).label("total_no_status"),
            func.coalesce(
                func.sum(ExportCarMessageAwbMaster.pcs), 0
            ).label("total_pcs"),
        )
        .where(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= day_start_utc,
            ExportCarMessageAwbMaster.car_message_datetime_combo <= day_end_utc,
        )
    )
    awb_row = awb_stats_result.mappings().one()

    # get AWB ids for this date
    awb_ids_result = await db.execute(
        select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.pcs,
        )
        .where(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= day_start_utc,
            ExportCarMessageAwbMaster.car_message_datetime_combo <= day_end_utc,
        )
    )
    awb_rows = awb_ids_result.mappings().all()
    awb_ids = [r.id for r in awb_rows]
    awb_pcs_map = {r.id: (r.pcs or 0) for r in awb_rows}
    total_awb_pcs = sum(awb_pcs_map.values())

    # ══════════════════════════════════════════════════════════
    # SECTION 2 — SCANNING STATS
    # ══════════════════════════════════════════════════════════

    if awb_ids:
        # ── All scanned sequences for these AWBs ───────────────
        scanning_result = await db.execute(
            select(
                ExportAwbSkidItemSequence.awb_master_id,
                func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
            )
            .where(ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids))
            .group_by(ExportAwbSkidItemSequence.awb_master_id)
        )
        scanned_by_awb = {
            r.awb_master_id: r.scanned_pcs
            for r in scanning_result.mappings().all()
        }

        scanned_awbs = len(scanned_by_awb)  # only AWBs with scans appear in dict

        total_scanned_for_date_awbs = sum(scanned_by_awb.values())
        total_pending_pcs = max(0, total_awb_pcs - total_scanned_for_date_awbs)
        scan_pct = round(
            (total_scanned_for_date_awbs / total_awb_pcs * 100), 1
        ) if total_awb_pcs > 0 else 0

        # ── Others scanned — sequences for AWBs NOT in this date
        # i.e. scanning happened today but AWB belongs to different date
        others_result = await db.execute(
            select(
                func.count(ExportAwbSkidItemSequence.id).label("others_scanned_pcs"),
                func.count(
                    distinct(ExportAwbSkidItemSequence.awb_master_id)
                ).label("others_awb_count"),
            )
            .join(
                ExportCarMessageAwbMaster,
                ExportAwbSkidItemSequence.awb_master_id == ExportCarMessageAwbMaster.id,
            )
            .where(
                # sequences scanned on selected date (IST)
                ExportAwbSkidItemSequence.sequence_date_time >= day_start_utc,
                ExportAwbSkidItemSequence.sequence_date_time <= day_end_utc,
                # but AWB does NOT belong to this date
                ExportCarMessageAwbMaster.id.notin_(awb_ids),
            )
        )
        others_row = others_result.mappings().one()

        # -------
                # =====================================================
        # 🆕 NEW: OTHER AWB TOTAL PCS
        # =====================================================
        other_awb_totals_result = await db.execute(
            select(
                ExportCarMessageAwbMaster.id,
                ExportCarMessageAwbMaster.pcs.label("total_pcs"),
            )
            .join(
                ExportAwbSkidItemSequence,
                ExportAwbSkidItemSequence.awb_master_id
                == ExportCarMessageAwbMaster.id,
            )
            .where(
                ExportAwbSkidItemSequence.sequence_date_time >= day_start_utc,
                ExportAwbSkidItemSequence.sequence_date_time <= day_end_utc,
                ExportCarMessageAwbMaster.id.notin_(awb_ids),
            )
            .group_by(
                ExportCarMessageAwbMaster.id,
                ExportCarMessageAwbMaster.pcs,
            )
        )

        other_awb_rows = other_awb_totals_result.mappings().all()

        # 🆕 NEW
        other_awb_ids = [r.id for r in other_awb_rows]

        # 🆕 NEW
        other_awb_total_pcs = sum(r.total_pcs or 0 for r in other_awb_rows)

        # =====================================================
        # 🆕 NEW: SCANNED EVER FOR OTHER AWBs
        # =====================================================
        if other_awb_ids:
            other_scanned_ever_result = await db.execute(
                select(
                    ExportAwbSkidItemSequence.awb_master_id,
                    func.count(ExportAwbSkidItemSequence.id).label("scanned_ever"),
                )
                .where(ExportAwbSkidItemSequence.awb_master_id.in_(other_awb_ids))
                .group_by(ExportAwbSkidItemSequence.awb_master_id)
            )

            # 🆕 NEW
            other_scanned_ever_map = {
                r.awb_master_id: r.scanned_ever
                for r in other_scanned_ever_result.mappings().all()
            }
        else:
            # 🆕 NEW
            other_scanned_ever_map = {}

        # 🆕 NEW
        other_awbs_scanned_ever = sum(other_scanned_ever_map.values())

        # 🆕 NEW
        other_awbs_pending = max(
            0, other_awb_total_pcs - other_awbs_scanned_ever
        )

        # =====================================================
        # 🆕 NEW: % CALCULATIONS
        # =====================================================
        others_scanned_today = others_row.get("others_scanned_pcs") or 0

        # 🆕 NEW
        other_date_pct = round(
            (others_scanned_today / other_awb_total_pcs * 100), 1
        ) if other_awb_total_pcs > 0 else 0

        # 🆕 NEW
        merged_scanned = total_scanned_for_date_awbs + others_scanned_today

        # 🆕 NEW
        merged_total_pcs = total_awb_pcs + other_awb_total_pcs

        # 🆕 NEW
        merged_pct = round(
            (merged_scanned / merged_total_pcs * 100), 1
        ) if merged_total_pcs > 0 else 0

        # --------

    else:
        total_scanned_for_date_awbs = 0
        total_pending_pcs = 0
        scanned_awbs = 0    # ✅ ADD
        scan_pct = 0
        others_row = {"others_scanned_pcs": 0, "others_awb_count": 0}

         # 🆕 NEW DEFAULTS
        other_awb_total_pcs = 0
        other_awbs_pending = 0
        other_date_pct = 0
        merged_scanned = 0
        merged_total_pcs = 0
        merged_pct = 0

    # ══════════════════════════════════════════════════════════
    # SECTION 3 — SKID & LOCATION STATS (only date AWBs, no others)
    # ══════════════════════════════════════════════════════════

    if awb_ids:
        # all mappings for date AWBs
        mapping_result = await db.execute(
            select(
                ExportAwbSkidMapping.id.label("mapping_id"),
                ExportAwbSkidMapping.skid_id,
                ExportAwbSkidMapping.is_skid_used_complete,
            )
            .where(ExportAwbSkidMapping.awb_master_id.in_(awb_ids))
        )
        mapping_rows = mapping_result.mappings().all()
        mapping_ids = [r.mapping_id for r in mapping_rows]
        total_skids_used = len({r.skid_id for r in mapping_rows if r.skid_id})
        total_skids_complete = sum(
            1 for r in mapping_rows if r.is_skid_used_complete
        )

        # location stats — scoped to date AWBs only
        if mapping_ids:

            # ── Latest location row per skid ──────────────────────────
            latest_loc_subq = (
                select(
                    ExportSkidLocationMapping.skid_id,
                    func.max(ExportSkidLocationMapping.id).label("latest_id"),  # ✅ latest row id
                )
                .where(ExportSkidLocationMapping.mapping_id.in_(mapping_ids))
                .group_by(ExportSkidLocationMapping.skid_id)
                .subquery()
            )

            loc_result = await db.execute(
                select(
                    func.count(
                        distinct(ExportSkidLocationMapping.skid_id)
                    ).label("total_ever_located"),
                    func.count(
                        distinct(ExportSkidLocationMapping.skid_id)
                    ).filter(
                        ExportSkidLocationMapping.is_current == True
                    ).label("skids_at_location"),
                    func.count(
                        distinct(ExportSkidLocationMapping.skid_id)
                    ).filter(
                        ExportSkidLocationMapping.is_current == False,
                        ExportSkidLocationMapping.picked_at.isnot(None),
                    ).label("skids_retrieved"),
                )
                .join(
        latest_loc_subq,
        ExportSkidLocationMapping.id == latest_loc_subq.c.latest_id,  # ✅ only latest row
    )
                .where(
                    ExportSkidLocationMapping.mapping_id.in_(mapping_ids)
                )
            )
            loc_row = loc_result.mappings().one()
            total_ever_located = loc_row.total_ever_located or 0
            skids_at_location = loc_row.skids_at_location or 0
            skids_retrieved = loc_row.skids_retrieved or 0
        else:
            total_ever_located = skids_at_location = skids_retrieved = 0

        skids_not_located = max(0, total_skids_used - total_ever_located)

    else:
        total_skids_used = total_skids_complete = 0
        total_ever_located = skids_at_location = skids_retrieved = skids_not_located = 0

    # ══════════════════════════════════════════════════════════
    # SECTION 4 — FLIGHT BOOKING STATS (by flight_date = IST date)
    # ══════════════════════════════════════════════════════════
    flight_stats_result = await db.execute(
        select(
            func.count(ExportFlightBookingHeader.id).label("total_flights"),
            func.count(ExportFlightBookingHeader.id).filter(
                ExportFlightBookingHeader.flight_dpt_datetime <= day_end_utc,
                ExportFlightBookingHeader.flight_dpt_datetime >= day_start_utc,
            ).label("departing_today"),
        )
        .where(
            ExportFlightBookingHeader.flight_date == report_date,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    flight_row = flight_stats_result.mappings().one()

    # ── Build response ─────────────────────────────────────────
    return {
        "report_date": str(report_date),

        # ── Section 1: AWB Master ──────────────────────────────
        "awb": {
            "total_awb": awb_row.total_awb or 0,
            "total_rcs": awb_row.total_rcs or 0,
            "total_non_rcs": awb_row.total_non_rcs or 0,
            "total_no_status": awb_row.total_no_status or 0,
            "total_pcs": total_awb_pcs,
        },

        # ── Section 2: Scanning ────────────────────────────────
        "scanning": {
            # for AWBs of this date
            "total_awb_pcs": total_awb_pcs,
            "scanned_awbs": scanned_awbs,  
            "scanned_pcs": total_scanned_for_date_awbs,
            "pending_pcs": total_pending_pcs,
            "scan_completion_pct": scan_pct,

            # other AWBs scanned on this date (different car message date)
            "others": {
                "others_scanned_pcs": others_row.get("others_scanned_pcs") or 0,
                "others_awb_count": others_row.get("others_awb_count") or 0,
                "note": "Scanning done on this date for AWBs from other car message dates",

                  # 🆕 NEW
                "other_awb_total_pcs": other_awb_total_pcs,
                "other_awbs_pending": other_awbs_pending,
                "other_date_pct": other_date_pct,
            },

             # 🆕 NEW BLOCK
            "merged": {
                "merged_scanned_pcs": merged_scanned,
                "merged_total_pcs": merged_total_pcs,
                "merged_pct": merged_pct,
            },

            # combined view
            "total_scanned_pcs_on_this_date": (
                total_scanned_for_date_awbs
                + (others_row.get("others_scanned_pcs") or 0)
            ),
        },

        # ── Section 3: Skid & Location (date AWBs only) ────────
        "skid_and_location": {
            "total_skids_used": total_skids_used,
            "total_skids_complete": total_skids_complete,
            "total_skids_incomplete": max(
                0, total_skids_used - total_skids_complete
            ),
            "total_ever_located": total_ever_located,
            "skids_at_location": skids_at_location,
            "skids_retrieved": skids_retrieved,
            "skids_not_located_yet": skids_not_located,
        },

        "flight": {
            "total_flights": flight_row.total_flights or 0,
            "departing_today": flight_row.departing_today or 0,
        },
    }





# ✌️====== Get detaild drilldown for dashboard stats
async def get_dashboard_drilldown_detail(
    db: AsyncSession,
    report_date: date,
    detail_type: str,  # "all_awbs" | "rcs_awbs" | "non_rcs_awbs" | "scanned_awbs" | "used_skids"
) -> dict:

    day_start_utc, day_end_utc = ist_date_to_utc_range(report_date)
    IST = ZoneInfo("Asia/Kolkata")

    def to_ist_str(dt) -> str | None:
        if dt is None:
            return None
        return dt.astimezone(IST).strftime("%d-%m-%Y %H:%M")

    # ── Get AWB ids for this date ──────────────────────────────
    awb_ids_result = await db.execute(
        select(ExportCarMessageAwbMaster.id).where(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= day_start_utc,
            ExportCarMessageAwbMaster.car_message_datetime_combo <= day_end_utc,
        )
    )
    awb_ids = [r.id for r in awb_ids_result.all()]

    # ══════════════════════════════════════════════════════════
    # AWB LIST TYPES — all / rcs / non_rcs
    # ══════════════════════════════════════════════════════════
    if detail_type in ("all_awbs", "rcs_awbs", "non_rcs_awbs"):

        stmt = select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportCarMessageAwbMaster.car_message_datetime_combo,
        ).where(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= day_start_utc,
            ExportCarMessageAwbMaster.car_message_datetime_combo <= day_end_utc,
        )

        if detail_type == "rcs_awbs":
            stmt = stmt.where(ExportCarMessageAwbMaster.status == "RCS")
        elif detail_type == "non_rcs_awbs":
            stmt = stmt.where(
                ExportCarMessageAwbMaster.status != "RCS",
                ExportCarMessageAwbMaster.status.isnot(None),
            )

        stmt = stmt.order_by(ExportCarMessageAwbMaster.awb_no)
        result = await db.execute(stmt)
        rows = result.mappings().all()

        return {
            "detail_type": detail_type,
            "report_date": str(report_date),
            "total": len(rows),
            "items": [
                {
                    "awb_no": r.awb_no,
                    "origin": r.origin,
                    "destination": r.destination,
                    "total_pcs": r.total_pcs,
                    "status": r.status or "—",
                    "agent": r.agent or "—",
                    "rcs_datetime": to_ist_str(r.rcs_datetime),
                    "car_message_datetime": to_ist_str(r.car_message_datetime_combo),
                }
                for r in rows
            ],
        }

    # ══════════════════════════════════════════════════════════
    # SCANNED AWBs — awb level summary with scanned pcs count
    # same AWBs used in stats (awb_ids of this date that have scans)
    # ══════════════════════════════════════════════════════════
    elif detail_type == "scanned_awbs":

        if not awb_ids:
            return {"detail_type": detail_type, "report_date": str(report_date), "total": 0, "items": []}

        result = await db.execute(
            select(
                ExportCarMessageAwbMaster.id.label("awb_master_id"),
                ExportCarMessageAwbMaster.awb_no,
                ExportCarMessageAwbMaster.origin,
                ExportCarMessageAwbMaster.destination,
                ExportCarMessageAwbMaster.pcs.label("total_pcs"),
                ExportCarMessageAwbMaster.status,
                ExportCarMessageAwbMaster.agent,
                func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
            )
            .join(
                ExportAwbSkidItemSequence,
                ExportAwbSkidItemSequence.awb_master_id == ExportCarMessageAwbMaster.id,
            )
            .where(ExportCarMessageAwbMaster.id.in_(awb_ids))
            .group_by(
                ExportCarMessageAwbMaster.id,
                ExportCarMessageAwbMaster.awb_no,
                ExportCarMessageAwbMaster.origin,
                ExportCarMessageAwbMaster.destination,
                ExportCarMessageAwbMaster.pcs,
                ExportCarMessageAwbMaster.status,
                ExportCarMessageAwbMaster.agent,
            )
            .order_by(ExportCarMessageAwbMaster.awb_no)
        )

        # ── scanner🫥🫥 details per AWB (which persons scanned those awb) ────────────────────────────────
        scanner_result = await db.execute(
            select(
                ExportAwbSkidItemSequence.awb_master_id,
                ExportAwbSkidItemSequence.mapping_id,
                ExportAwbSkidItemSequence.scanned_by,
                User.name.label("scanned_by_name"), 
                ExportAwbSkidItemSequence.scan_by_device,
                ExportSkidMaster.skid_no,
                func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
                func.max(ExportAwbSkidItemSequence.sequence_date_time).label("last_scanned_at"),
            )
            .join(
                ExportAwbSkidMapping,
                ExportAwbSkidItemSequence.mapping_id == ExportAwbSkidMapping.id,
            )
            .join(
                ExportSkidMaster,
                ExportAwbSkidMapping.skid_id == ExportSkidMaster.id,
            ).join(
                User,   # 👈 join users table
                User.emp_id == ExportAwbSkidItemSequence.scanned_by,
                isouter=True   # optional: allows null if no matching user
            )
            .where(ExportAwbSkidItemSequence.awb_master_id.in_(awb_ids))
            .group_by(
                ExportAwbSkidItemSequence.awb_master_id,
                ExportAwbSkidItemSequence.mapping_id,
                ExportAwbSkidItemSequence.scanned_by,
                User.name,   # 👈 must group by since it's selected
                ExportAwbSkidItemSequence.scan_by_device,
                ExportSkidMaster.skid_no,
            )
            .order_by(ExportAwbSkidItemSequence.awb_master_id)
        )
        scanner_rows = scanner_result.mappings().all()

        rows = result.mappings().all()
# 🫥🫥
        scanners_by_awb: dict[int, list] = defaultdict(list) 

        for r in scanner_rows:
            scanners_by_awb[r.awb_master_id].append({
                "emp_id": r.scanned_by,
                "name": r.scanned_by_name or "", 
                "scan_by_device": r.scan_by_device,
                "skid_no": r.skid_no,
                "scanned_pcs": r.scanned_pcs,
                "last_scanned_at": to_ist_str(r.last_scanned_at),
            })

        return {
            "detail_type": detail_type,
            "report_date": str(report_date),
            "total": len(rows),
            "items": [
                {
                    "awb_no": r.awb_no,
                    "origin": r.origin,
                    "destination": r.destination,
                    "total_pcs": r.total_pcs or 0,
                    "scanned_pcs": r.scanned_pcs,
                    "pending_pcs": max(0, (r.total_pcs or 0) - r.scanned_pcs),
                    "status": r.status or "—",
                    "agent": r.agent or "—",
                    "scan_pct": round(
                        r.scanned_pcs / r.total_pcs * 100, 1
                    ) if r.total_pcs else 0,
                     "scanners": scanners_by_awb.get(r.awb_master_id, []),
                }
                for r in rows
            ],
        }

    # ══════════════════════════════════════════════════════════
    # USED SKIDS — skid_no, awb, location, base drop, retrieved
    # scoped strictly to awb_ids of this date
    # ══════════════════════════════════════════════════════════
    elif detail_type == "used_skids":

        if not awb_ids:
            return {"detail_type": detail_type, "report_date": str(report_date), "total": 0, "items": []}

        # ── get all mappings for date AWBs ─────────────────────
        mapping_result = await db.execute(
            select(
                ExportAwbSkidMapping.id.label("mapping_id"),
                ExportAwbSkidMapping.skid_id,
                ExportAwbSkidMapping.awb_master_id,
                ExportAwbSkidMapping.is_skid_used_complete,
                ExportSkidMaster.skid_no,
                ExportSkidMaster.skid_type,
                ExportCarMessageAwbMaster.awb_no,
            )
            .join(ExportSkidMaster,
                  ExportAwbSkidMapping.skid_id == ExportSkidMaster.id)
            .join(ExportCarMessageAwbMaster,
                  ExportAwbSkidMapping.awb_master_id == ExportCarMessageAwbMaster.id)
            .where(ExportAwbSkidMapping.awb_master_id.in_(awb_ids))
            .order_by(ExportSkidMaster.skid_no)
        )
        mapping_rows = mapping_result.mappings().all()

        if not mapping_rows:
            return {"detail_type": detail_type, "report_date": str(report_date), "total": 0, "items": []}

        mapping_ids = [r.mapping_id for r in mapping_rows]

        # ── scanned pcs per mapping ────────────────────────────────
        seq_count_result = await db.execute(
            select(
                ExportAwbSkidItemSequence.mapping_id,
                func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
            )
            .where(ExportAwbSkidItemSequence.mapping_id.in_(mapping_ids))
            .group_by(ExportAwbSkidItemSequence.mapping_id)
        )
        seq_count_map = {
            r.mapping_id: r.scanned_pcs
            for r in seq_count_result.mappings().all()
        }

        # ── current location per mapping ───────────────────────
        loc_result = await db.execute(
            select(
                ExportSkidLocationMapping.mapping_id,
                ExportLocationsMaster.loc.label("location_name"),
                ExportLocationsMaster.area_code,
                ExportSkidLocationMapping.is_current,
                ExportSkidLocationMapping.picked_at,
            )
            .join(ExportLocationsMaster,
                  ExportSkidLocationMapping.location_id == ExportLocationsMaster.id)
            .where(
                ExportSkidLocationMapping.mapping_id.in_(mapping_ids),
            )
            .order_by(ExportSkidLocationMapping.assigned_at.desc())
        )
        loc_rows = loc_result.mappings().all()

        # build location map — latest location per mapping
        loc_map: dict[int, dict] = {}
        for row in loc_rows:
            if row.mapping_id not in loc_map:
                loc_map[row.mapping_id] = {
                    "location_name": row.location_name,
                    "area_code": row.area_code,
                    "is_current": row.is_current,
                    "is_retrieved": row.picked_at is not None,
                    "picked_at": to_ist_str(row.picked_at),
                }

        # ── base drop per mapping ──────────────────────────────
        base_result = await db.execute(
            select(
                ExportSkidBaseMapping.mapping_id,
                ExportBaseMaster.base_name,
                ExportSkidBaseMapping.dropped_at,
            )
            .join(ExportBaseMaster,
                  ExportSkidBaseMapping.base_id == ExportBaseMaster.id)
            .where(ExportSkidBaseMapping.mapping_id.in_(mapping_ids))
        )
        base_map = {
            r.mapping_id: {
                "base_name": r.base_name,
                "dropped_at": to_ist_str(r.dropped_at),
            }
            for r in base_result.mappings().all()
        }

    # Get scanner info of this skid
        scanner_result = await db.execute(
            select(
                ExportAwbSkidItemSequence.mapping_id,
                ExportAwbSkidItemSequence.scanned_by,
                # ExportAwbSkidItemSequence.sequence_date_time,
                User.name.label("user_name")   # ✅ NEW
            )
            .join(
                User,
                ExportAwbSkidItemSequence.scanned_by == User.emp_id,  # 🔥 KEY JOIN
                isouter=True  # ✅ important (in case user missing)
            )
            .where(
                ExportAwbSkidItemSequence.mapping_id.in_(mapping_ids)
            ) .distinct(
        ExportAwbSkidItemSequence.mapping_id,
        ExportAwbSkidItemSequence.scanned_by
    )
        )
        scanner_rows = scanner_result.mappings().all()

        scanner_map: dict[int, list] = {}

        for row in scanner_rows:
            if row.mapping_id not in scanner_map:
                scanner_map[row.mapping_id] = []

            scanner_map[row.mapping_id].append({
                "emp_id": row.scanned_by,
                "name": row.user_name,  # ✅ NEW
                # "scanned_at": to_ist_str(row.sequence_date_time),
            })

        # ── build response ─────────────────────────────────────
        items = []
        for r in mapping_rows:
            loc = loc_map.get(r.mapping_id)
            base = base_map.get(r.mapping_id)
            scanned_pcs = seq_count_map.get(r.mapping_id, 0)  # ✅ get scanned pcs

            items.append({
                "skid_no": r.skid_no,
                "skid_type": r.skid_type,
                "awb_no": r.awb_no,
                "scan_complete": r.is_skid_used_complete,

                "scanned_pcs": scanned_pcs,
                "scanners": scanner_map.get(r.mapping_id, []),


                # location
                "location_name": loc["location_name"] if loc else None,
                "area_code": loc["area_code"] if loc else None,
                "is_at_location": loc["is_current"] if loc else False,
                "is_retrieved": loc["is_retrieved"] if loc else False,
                "retrieved_at": loc["picked_at"] if loc else None,
                "ever_located": loc is not None,

                # base drop
                "base_dropped": base is not None,
                "base_name": base["base_name"] if base else None,
                "base_dropped_at": base["dropped_at"] if base else None,
            })

        return {
            "detail_type": detail_type,
            "report_date": str(report_date),
            "total": len(items),
            "items": items,
        }
    
    # ══════════════════════════════════════════════════════════
    # FLIGHT BOOKINGS DETAIL
    # ══════════════════════════════════════════════════════════
    elif detail_type == "flight_bookings":

        # fetch all flights on this date
        flights_result = await db.execute(
            select(
                ExportFlightBookingHeader.id.label("header_id"),
                ExportFlightBookingHeader.flight_no,
                ExportFlightBookingHeader.flight_date,
                ExportFlightBookingHeader.flight_dpt_datetime,
                ExportFlightBookingHeader.booked_by,
                ExportFlightBookingHeader.booked_at,
            )
            .where(
                ExportFlightBookingHeader.flight_date == report_date,
                ExportFlightBookingHeader.is_active == True,
            )
            .order_by(ExportFlightBookingHeader.flight_dpt_datetime)
        )
        flight_rows = flights_result.mappings().all()

        if not flight_rows:
            return {
                "detail_type": detail_type,
                "report_date": str(report_date),
                "total": 0,
                "items": [],
            }

        flight_header_ids = [r.header_id for r in flight_rows]

        # ── AWBs per flight ────────────────────────────────────
        awb_detail_result = await db.execute(
            select(
                ExportFlightBookingDetail.flight_header_id,
                ExportFlightBookingDetail.booked_pcs,
                ExportCarMessageAwbMaster.awb_no,
                ExportCarMessageAwbMaster.origin,
                ExportCarMessageAwbMaster.destination,
                ExportCarMessageAwbMaster.id.label("awb_master_id"),
            )
            .join(
                ExportCarMessageAwbMaster,
                ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
            )
            .where(
                ExportFlightBookingDetail.flight_header_id.in_(flight_header_ids)
            )
        )
        awb_detail_rows = awb_detail_result.mappings().all()

        # group AWBs by flight
        awbs_by_flight: dict[int, list] = {}
        awb_ids_by_flight: dict[int, list[int]] = {}
        for r in awb_detail_rows:
            awbs_by_flight.setdefault(r.flight_header_id, []).append({
                "awb_no": r.awb_no,
                "origin": r.origin,
                "destination": r.destination,
                "booked_pcs": r.booked_pcs,
                "awb_master_id": r.awb_master_id,
            })
            awb_ids_by_flight.setdefault(r.flight_header_id, []).append(r.awb_master_id)

        # ── Skids per flight — via sequences loaded ────────────
        skid_result = await db.execute(
            select(
                ExportSequenceItemUldLoading.flight_header_id,
                ExportAwbSkidMapping.skid_id,
                ExportSkidMaster.skid_no,
                ExportSkidMaster.skid_type,
                ExportCarMessageAwbMaster.awb_no,
                func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
            )
            .join(
                ExportAwbSkidItemSequence,
                ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id,
            )
            .join(
                ExportAwbSkidMapping,
                ExportAwbSkidItemSequence.mapping_id == ExportAwbSkidMapping.id,
            )
            .join(
                ExportSkidMaster,
                ExportAwbSkidMapping.skid_id == ExportSkidMaster.id,
            )
            .join(
                ExportCarMessageAwbMaster,
                ExportSequenceItemUldLoading.awb_master_id == ExportCarMessageAwbMaster.id,
            )
            .where(
                ExportSequenceItemUldLoading.flight_header_id.in_(flight_header_ids)
            )
            .group_by(
                ExportSequenceItemUldLoading.flight_header_id,
                ExportAwbSkidMapping.skid_id,
                ExportSkidMaster.skid_no,
                ExportSkidMaster.skid_type,
                ExportCarMessageAwbMaster.awb_no,
            )
        )
        skid_rows = skid_result.mappings().all()

        skids_by_flight: dict[int, list] = {}
        for r in skid_rows:
            skids_by_flight.setdefault(r.flight_header_id, []).append({
                "skid_no": r.skid_no,
                "skid_type": r.skid_type,
                "awb_no": r.awb_no,
                "loaded_pcs": r.loaded_pcs,
            })

        # ── ULDs per flight ────────────────────────────────────
        uld_result = await db.execute(
            select(
                ExportUldAssignment.flight_header_id,
                ExportUldMaster.uld_no,
                ExportUldMaster.carrier,
                func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
            )
            .join(
                ExportUldAssignmentDetail,
                ExportUldAssignmentDetail.assignment_id == ExportUldAssignment.id,
            )
            .join(
                ExportUldMaster,
                ExportUldAssignmentDetail.uld_id == ExportUldMaster.id,
            )
            .outerjoin(
                ExportSequenceItemUldLoading,
                ExportSequenceItemUldLoading.uld_assignment_detail_id == ExportUldAssignmentDetail.id,
            )
            .where(
                ExportUldAssignment.flight_header_id.in_(flight_header_ids),
                ExportUldAssignment.is_active == True,
            )
            .group_by(
                ExportUldAssignment.flight_header_id,
                ExportUldMaster.uld_no,
                ExportUldMaster.carrier,
            )
        )
        uld_rows = uld_result.mappings().all()

        ulds_by_flight: dict[int, list] = {}
        for r in uld_rows:
            ulds_by_flight.setdefault(r.flight_header_id, []).append({
                "uld_no": r.uld_no,
                "carrier": r.carrier,
                "loaded_pcs": r.loaded_pcs,
            })

        # ── loaded pcs per flight ──────────────────────────────
        loaded_result = await db.execute(
            select(
                ExportSequenceItemUldLoading.flight_header_id,
                func.count(ExportSequenceItemUldLoading.id).label("loaded_pcs"),
            )
            .where(
                ExportSequenceItemUldLoading.flight_header_id.in_(flight_header_ids)
            )
            .group_by(ExportSequenceItemUldLoading.flight_header_id)
        )
        loaded_by_flight = {
            r.flight_header_id: r.loaded_pcs
            for r in loaded_result.mappings().all()
        }

        # ── build items ────────────────────────────────────────
        items = []
        for f in flight_rows:
            fid = f.header_id
            total_booked = sum(
                a["booked_pcs"] for a in awbs_by_flight.get(fid, [])
            )
            loaded_pcs = loaded_by_flight.get(fid, 0)

            items.append({
                "flight_no": f.flight_no,
                "flight_date": str(f.flight_date),
                "flight_dpt_datetime": to_ist_str(f.flight_dpt_datetime),
                "booked_by": f.booked_by,
                "booked_at": to_ist_str(f.booked_at),
                "total_booked_pcs": total_booked,
                "total_loaded_pcs": loaded_pcs,
                "pending_pcs": max(0, total_booked - loaded_pcs),
                "awbs": awbs_by_flight.get(fid, []),
                "skids": skids_by_flight.get(fid, []),
                "ulds": ulds_by_flight.get(fid, []),
            })

        return {
            "detail_type": detail_type,
            "report_date": str(report_date),
            "total": len(items),
            "items": items,
        }

    raise HTTPException(
        status_code=400,
        detail=f"Invalid detail_type '{detail_type}'. "
               "Use: all_awbs | rcs_awbs | non_rcs_awbs | scanned_awbs | used_skids" | "flight_bookings",
    )

# =================================== ✌️flight create by pdf service ======================


async def upsert_flight_booking_from_pdf(
    db: AsyncSession,
    df,                          # extracted DataFrame
    # flight_dpt_datetime: datetime,
    booked_by: str,
) -> PdfUpsertResponse:

    now = get_utc_now()

    # ── Pull flight info from PDF ───────────────────────────────
    flight_no   = str(df["FLIGHT_NUM"].iloc[0]).strip().upper()
    flight_date = df["FLIGHT_DATE"].iloc[0]       # already date object

    # ── Departure check ────────────────────────────────────────
    # Convert IST departure input to UTC for comparison
    from zoneinfo import ZoneInfo
    ist = ZoneInfo("Asia/Kolkata")
    utc = ZoneInfo("UTC")


    # ✅ END OF NEXT DAY (23:59:59 IST)
    dpt_ist = datetime.combine(
        flight_date + timedelta(days=1),
        time(23, 59, 59)
    ).replace(tzinfo=ist)

    # ✅ Convert to UTC for DB
    dpt_utc = dpt_ist.astimezone(utc)

    # if flight_dpt_datetime.tzinfo is None:
    #     # assume IST if no tz given
    #     dpt_ist = flight_dpt_datetime.replace(tzinfo=ist)
    # else:
    #     dpt_ist = flight_dpt_datetime

    # dpt_utc = dpt_ist.astimezone(utc)

    # if dpt_utc <= now:
    #     raise HTTPException(
    #         status_code=400,
    #         detail=f"Flight {flight_no} has already departed — upload rejected.",
    #     )
    print(f"DEBUG: Flight {flight_no} departure IST: {dpt_ist}, departure UTC: {dpt_utc}, now UTC: {now}")
  
    # ── Unique AWB nos from PDF ─────────────────────────────────
    pdf_awb_nos: list[str] = (
        df["AWB_NUM"].dropna().drop_duplicates().tolist()
    )

    # ── NEW🫥: Sum LOC_PCS per AWB from PDF (part-shipment aggregation) ──
    # pdf_awb_pcs: dict[str, int] = (
    #     df[df["AWB_NUM"].notna()]
    #     .groupby("AWB_NUM")["LOC_PCS"]
    #     .sum()
    #     .astype(int)
    #     .to_dict()
    # )

# 🤢
    pdf_awb_pcs: dict[str, int] = (
    df[df["AWB_NUM"].notna()]
    .groupby("AWB_NUM")["LOC_PCS"]
    .apply(lambda x: int(x.dropna().sum()))   # ← drop NA before summing
    .to_dict()
    )

        # ── Check if flight already exists ─────────────────────────
    existing_header_result = await db.execute(
        select(ExportFlightBookingHeader).where(
            ExportFlightBookingHeader.flight_no == flight_no,
            ExportFlightBookingHeader.flight_date == flight_date,
            ExportFlightBookingHeader.is_active == True,
        )
    )
    existing_header = existing_header_result.scalar_one_or_none()
    is_new_flight = existing_header is None

    # 🚨 NEW CHECK — block updates (EXTING FLIGHT) if flight already departed
    if existing_header:
        if existing_header.flight_dpt_datetime <= now:
            # raise HTTPException(
            #     status_code=400,
            #     detail=f"Flight {flight_no} already departed — cannot modify",
            # )
            return PdfUpsertResponse(
            success=False,
            message=f"Flight {flight_no} already departed — no changes applied.",
            flight_no=flight_no,
            flight_date=str(flight_date),
            is_new_flight=False,
            total_awbs_in_pdf=len(df),

            added=[],
            removed=[],
            updated=[],
            skipped=[],
            unchanged=[],

            not_found_in_db=[],

            added_count=0,
            removed_count=0,
            updated_count=0,
            skipped_count=0,
            unchanged_count=0,
            not_found_count=0,
        )
    else:
        if dpt_utc <= now:
            return PdfUpsertResponse(
                success=False,
                message=f"Flight {flight_no} has already departed (Have old flight date from current in pdf) — upload rejected.",
                flight_no=flight_no,
                flight_date=str(flight_date),
                is_new_flight=True,
                total_awbs_in_pdf=len(df),

                added=[],
                removed=[],
                updated=[],
                skipped=[],
                unchanged=[],

                not_found_in_db=[],

                added_count=0,
                removed_count=0,
                updated_count=0,
                skipped_count=0,
                unchanged_count=0,
                not_found_count=0,
            )



    # ── Fetch AWB master records for PDF AWBs ──────────────────
    awb_master_result = await db.execute(
        select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.pcs,
            ExportCarMessageAwbMaster.status,   # ← ADD THIS
            ExportCarMessageAwbMaster.is_ultra_fast,  # ✅ ADD THIS
        ).where(ExportCarMessageAwbMaster.awb_no.in_(pdf_awb_nos))
    )
    db_awb_map = {
        row.awb_no: row
        for row in awb_master_result.mappings().all()
    }

    # AWBs in PDF but not in DB
    not_found_in_db = [
        AwbLookupError(awb_no=a, reason="Not found in AWB master")
        for a in pdf_awb_nos if a not in db_awb_map
    ]

    # # Valid AWBs from PDF that exist in DB (Not in used in new way)
    # valid_pdf_awb_ids = {
    #     db_awb_map[a].id
    #     for a in pdf_awb_nos
    #     if a in db_awb_map and db_awb_map[a].pcs
    # }

    # # ── Check if flight already exists ─────────────────────────
    # existing_header_result = await db.execute(
    #     select(ExportFlightBookingHeader).where(
    #         ExportFlightBookingHeader.flight_no == flight_no,
    #         ExportFlightBookingHeader.flight_date == flight_date,
    #         ExportFlightBookingHeader.is_active == True,
    #     )
    # )
    # existing_header = existing_header_result.scalar_one_or_none()
    # is_new_flight = existing_header is None

    # # 🚨 NEW CHECK — block updates (EXTING FLIGHT) if flight already departed
    # if existing_header:
    #     if existing_header.flight_dpt_datetime <= now:
    #         raise HTTPException(
    #             status_code=400,
    #             detail=f"Flight {flight_no} already departed — cannot modify",
    #         )

    # ══════════════════════════════════════════════════════════
    # CASE A — NEW FLIGHT → delegate to create service
    # ══════════════════════════════════════════════════════════
    if is_new_flight:
        awb_items = []
        skipped: list[AwbChangeRecord] = []
        added: list[AwbChangeRecord] = []

        for awb_no in pdf_awb_nos:
            row = db_awb_map.get(awb_no)
            if not row:
                continue

            # ✅🤢 ADD — skip non-RCS instead of letting create_flight_booking raise
            # if row.status != "RCS":
            #     skipped.append(AwbChangeRecord(
            #         awb_no=awb_no,
            #         action="SKIPPED",
            #         reason=f"AWB not in RCS status (current: {row.status})",
            #     ))

            if row.status not in FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    reason=f"AWB not eligible for booking (current status: {row.status})",
                ))
                continue

            pdf_pcs = pdf_awb_pcs.get(awb_no, 0)
            if not pdf_pcs:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    reason="AWB has no pcs pdf",
                ))
                continue

            # ✅ ADD THIS BLOCK HERE
            # scanned_result = await db.execute(
            #     select(
            #         func.count(ExportAwbSkidItemSequence.id)
            #     ).where(ExportAwbSkidItemSequence.awb_master_id == row.id)
            # )
            # scanned = scanned_result.scalar() or 0

            # if pdf_pcs > scanned:
            #     skipped.append(AwbChangeRecord(
            #         awb_no=awb_no,
            #         action="SKIPPED",
            #         reason=f"PDF pcs ({pdf_pcs}) exceeds scanned pcs ({scanned})",
            #     ))
            #     continue

            # ✅ NEW LOGIC (correct + ultra-fast support)

            # ── 1🤮. Get booked elsewhere ─────────────────────
            booked_elsewhere_result = await db.execute(
                select(
                    func.coalesce(func.sum(ExportFlightBookingDetail.booked_pcs), 0)
                ).where(
                    ExportFlightBookingDetail.awb_master_id == row.id
                )
            )
            booked_elsewhere = booked_elsewhere_result.scalar() or 0


            # ── 2. Compute available ────────────────────────
            if row.is_ultra_fast:
                available = (row.pcs or 0) - booked_elsewhere
            else:
                scanned_result = await db.execute(
                    select(func.count(ExportAwbSkidItemSequence.id))
                    .where(ExportAwbSkidItemSequence.awb_master_id == row.id)
                )
                scanned = scanned_result.scalar() or 0

                available = scanned - booked_elsewhere


            # ── 3. Validate ────────────────────────────────
            # if available <= 0:
            #     skipped.append(AwbChangeRecord(
            #         awb_no=awb_no,
            #         action="SKIPPED",
            #         reason=f"No available pcs (available={available})",
            #     ))
            #     continue

            if available <= 0:
                if row.is_ultra_fast:
                    skipped.append(AwbChangeRecord(
                        awb_no=awb_no,
                        action="SKIPPED",
                        reason=(
                            f"AWB fully booked: total pcs ({row.pcs or 0}) already allocated to flights"
                        ),
                    ))
                else:
                    skipped.append(AwbChangeRecord(
                        awb_no=awb_no,
                        action="SKIPPED",
                        reason=(
                            f"AWB has no scanned pcs available "
                            f"(scanned={scanned}, booked_elsewhere={booked_elsewhere})"
                        ),
                    ))
                continue

            if pdf_pcs > available:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    reason=f"PDF pcs ({pdf_pcs}) exceeds available ({available})",
                ))
                continue

            # ------------

            awb_items.append(FlightBookingAwbItem(
                awb_master_id=row.id,
                booked_pcs=pdf_pcs, #🫥
            ))
            added.append(AwbChangeRecord(
                awb_no=awb_no,
                action="ADDED",
                reason="New flight created from PDF",
                new_pcs=pdf_pcs, #🫥
            ))

        # if not awb_items:
        #     raise HTTPException(
        #         status_code=400,
        #         detail="No valid AWBs with pcs found in PDF.",
        #     )
        if not awb_items:
            return PdfUpsertResponse(
                success=False,
                # message="No valid AWBs found in PDF",
                message=(
    f"Flight {flight_no} not created: "
    f"{len(pdf_awb_nos)} AWBs in PDF, "
    f"{len(not_found_in_db)} not found in system"
),
                flight_no=flight_no,
                flight_date=str(flight_date),
                is_new_flight=True,
                total_awbs_in_pdf=len(pdf_awb_nos),

                added=[],
                removed=[],
                updated=[],
                skipped=skipped,
                unchanged=[],
                not_found_in_db=not_found_in_db,

                added_count=0,
                removed_count=0,
                updated_count=0,
                skipped_count=len(skipped),
                unchanged_count=0,
                not_found_count=len(not_found_in_db),
            )

        booking_request = CreateFlightBookingRequest(
            flight_no=flight_no,
            flight_date=flight_date,
            flight_dpt_datetime=dpt_utc,
            awbs=awb_items,
        )
        await create_flight_booking(
            db=db,
            payload=booking_request,
            booked_by=booked_by,
        )

        return PdfUpsertResponse(
            success=True,
            message=f"New flight {flight_no} created with {len(added)} AWBs.",
            flight_no=flight_no,
            flight_date=str(flight_date),
            is_new_flight=True,
            total_awbs_in_pdf=len(pdf_awb_nos),
            added=added,
            removed=[],
            updated=[],
            skipped=skipped,
            unchanged=[],
            not_found_in_db=not_found_in_db,
            added_count=len(added),
            removed_count=0,
            updated_count=0,
            skipped_count=len(skipped),
            unchanged_count=0,
            not_found_count=len(not_found_in_db),
        )

    # ══════════════════════════════════════════════════════════
    # CASE B — EXISTING FLIGHT → upsert
    # ══════════════════════════════════════════════════════════
    header_id = existing_header.id

    # ── Fetch existing booking details ─────────────────────────
    existing_details_result = await db.execute(
        select(
            ExportFlightBookingDetail.id.label("detail_id"),
            ExportFlightBookingDetail.awb_master_id,
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(ExportFlightBookingDetail.flight_header_id == header_id)
    )
    existing_details = {
        row.awb_master_id: row
        for row in existing_details_result.mappings().all()
    }

    existing_awb_ids = set(existing_details.keys())

    # ── Check scanning status for existing AWBs ────────────────
    # scanning_started = skid mapping exists AND has at least 1 sequence
    scanning_result = await db.execute(
        select(
            ExportAwbSkidMapping.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("seq_count"),
        )
        .join(
            ExportAwbSkidItemSequence,
            ExportAwbSkidItemSequence.mapping_id == ExportAwbSkidMapping.id,
        )
        .where(
            ExportAwbSkidMapping.awb_master_id.in_(existing_awb_ids),
        )
        .group_by(ExportAwbSkidMapping.awb_master_id)
        .having(func.count(ExportAwbSkidItemSequence.id) > 0)
    )
    scanning_started_ids = {
        row.awb_master_id
        for row in scanning_result.mappings().all()
    }

    # ── Fetch scanned pcs for existing AWBs ───────────────────
        # ── Fetch scanned pcs — ALL AWBs (existing + new from PDF) ─
    all_awb_ids = existing_awb_ids.union({
        db_awb_map[a].id for a in pdf_awb_nos if a in db_awb_map
    })

    scanned_result = await db.execute(
        select(
            ExportAwbSkidItemSequence.awb_master_id,
            func.count(ExportAwbSkidItemSequence.id).label("scanned_pcs"),
        )
        .where(ExportAwbSkidItemSequence.awb_master_id.in_(all_awb_ids))
        .group_by(ExportAwbSkidItemSequence.awb_master_id)
    )
    scanned_map = {
        row.awb_master_id: row.scanned_pcs
        for row in scanned_result.mappings().all()
    }

    #-----🤢
    booked_elsewhere_result = await db.execute(
        select(
            ExportFlightBookingDetail.awb_master_id,
            func.coalesce(func.sum(ExportFlightBookingDetail.booked_pcs), 0).label("booked_pcs"),
        )
        .join(
            ExportFlightBookingHeader,
            and_(
                ExportFlightBookingDetail.flight_header_id == ExportFlightBookingHeader.id,
                ExportFlightBookingHeader.is_active == True,
                ExportFlightBookingHeader.id != header_id,
            )
        )
        .where(ExportFlightBookingDetail.awb_master_id.in_(all_awb_ids))
        .group_by(ExportFlightBookingDetail.awb_master_id)
   )

    booked_elsewhere_map = {
        row.awb_master_id: row.booked_pcs
        for row in booked_elsewhere_result.mappings().all()
    }

    # ── Process changes ────────────────────────────────────────
    added: list[AwbChangeRecord] = []
    removed: list[AwbChangeRecord] = []
    updated: list[AwbChangeRecord] = []
    skipped: list[AwbChangeRecord] = []
    unchanged: list[AwbChangeRecord] = []

    # ── 1. AWBs in PDF but NOT in existing booking → ADD ───────
    awbs_to_add: list[FlightBookingAwbItem] = []
    for awb_no in pdf_awb_nos:
        row = db_awb_map.get(awb_no)
        if not row:
            continue  # already in not_found_in_db

        # ✅ ADD
        awb_full = db_awb_map.get(awb_no)
        # if awb_full and awb_full.status != "RCS":  # need status in db_awb_map query
        #     skipped.append(AwbChangeRecord(
        #         awb_no=awb_no,
        #         action="SKIPPED",
        #         reason="AWB not in RCS status",
        #     ))

        if awb_full and awb_full.status not in FINAL_STATUSES_FROM_WH_INVENTRY_FLT_BOOKING:
            skipped.append(AwbChangeRecord(
                awb_no=awb_no,
                action="SKIPPED",
                reason=f"AWB not eligible for booking (current status: {awb_full.status})",
            ))
            continue
       
        # CORRECT ✅ — check PDF pcs first, consistent with Case A
        pdf_pcs = pdf_awb_pcs.get(awb_no, 0)
        if not pdf_pcs:
            skipped.append(AwbChangeRecord(
                awb_no=awb_no,
                action="SKIPPED",
                reason="AWB has no pcs in PDF",
            ))
            continue
        
        if row.id not in existing_awb_ids:
            # scanned = scanned_map.get(row.id, 0)
            # booked_el = booked_elsewhere_map.get(row.id, 0)

            # available = scanned - booked_el
            booked_el = booked_elsewhere_map.get(row.id, 0)
            if row.is_ultra_fast:
                available = (row.pcs or 0) - booked_elsewhere_map.get(row.id, 0)
            else:
                scanned = scanned_map.get(row.id, 0)
                # booked_el = booked_elsewhere_map.get(row.id, 0)
                available = scanned - booked_el

            if available <= 0:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    # reason=f"No available scanned pcs (scanned={scanned}, booked_elsewhere={booked_el})",
                    reason=f"No available pcs. (available={available})",
                    new_pcs=pdf_pcs,
                ))
                continue

            if pdf_pcs > available:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    reason=f"PDF pcs ({pdf_pcs}) exceeds available scanned pcs ({available})",
                    new_pcs=pdf_pcs,
                ))
                continue

            awbs_to_add.append(FlightBookingAwbItem(
                awb_master_id=row.id,
                booked_pcs=pdf_pcs,
            ))
            added.append(AwbChangeRecord(
                awb_no=awb_no,
                action="ADDED",
                reason="AWB present in PDF but not in existing booking",
                new_pcs=pdf_pcs,
            ))

    # ── 2. AWBs in existing booking but NOT in PDF ─────────────
    #       Remove if scanning NOT started, skip if started
    awb_ids_in_pdf = {
        db_awb_map[a].id for a in pdf_awb_nos if a in db_awb_map
    }
    detail_ids_to_remove: list[int] = []

    for awb_id, detail_row in existing_details.items():
        if awb_id not in awb_ids_in_pdf:
            scanning = awb_id in scanning_started_ids
            if scanning:
                skipped.append(AwbChangeRecord(
                    awb_no=detail_row.awb_no,
                    action="SKIPPED",
                    reason="AWB missing from new PDF but scanning has started — cannot remove",
                    old_pcs=detail_row.booked_pcs,
                    scanning_started=True,
                ))
            else:
                detail_ids_to_remove.append(detail_row.detail_id)
                removed.append(AwbChangeRecord(
                    awb_no=detail_row.awb_no,
                    action="REMOVED",
                    reason="AWB not present in new PDF and scanning has not started",
                    old_pcs=detail_row.booked_pcs,
                    scanning_started=False,
                ))

    # ── 3. AWBs in both — check pcs change ────────────────────
    pcs_updates: list[tuple[int, int]] = []  # (detail_id, new_pcs)

    for awb_no in pdf_awb_nos:
        row = db_awb_map.get(awb_no)
        if not row or row.id not in existing_awb_ids:
            continue  # handled above

        detail_row = existing_details[row.id]
        # db_pcs = row.pcs or 0
        db_pcs = pdf_awb_pcs.get(awb_no, 0)
        booked_pcs = detail_row.booked_pcs
        scanning = row.id in scanning_started_ids

        if db_pcs == booked_pcs:
            unchanged.append(AwbChangeRecord(
                awb_no=awb_no,
                action="UNCHANGED",
                reason="Pcs unchanged",
                old_pcs=booked_pcs,
                new_pcs=db_pcs,
                scanning_started=scanning,
            ))
        elif db_pcs != booked_pcs:
            # ✅🤢 ADD — check remaining pcs across other flights
            booked_elsewhere = booked_elsewhere_map.get(row.id, 0)
            # get scanned pcs first
            # scanned_pcs = scanned_map.get(row.id, 0)
            # # ✅ FIX — use scanned pcs (already fetched above)
            # remaining_for_this_flight = scanned_pcs - booked_elsewhere

            # if db_pcs > remaining_for_this_flight:
            #     skipped.append(AwbChangeRecord(
            #         awb_no=awb_no,
            #         action="SKIPPED",
            #         reason=(
            #             f"PDF pcs ({db_pcs}) exceeds available remaining "
            #             f"({remaining_for_this_flight}) — keeping existing {booked_pcs} pcs"
            #         ),
            #         old_pcs=booked_pcs,
            #         new_pcs=db_pcs,
            #         scanning_started=scanning,
            #     ))
                # continue  # skip the update --------------->
            # get scanned pcs first 
            scanned_pcs = scanned_map.get(row.id, 0)
            booked_elsewhere = booked_elsewhere_map.get(row.id, 0)

            if row.is_ultra_fast:
                remaining_for_this_flight = (row.pcs or 0) - booked_elsewhere
            else:
                if db_pcs > scanned_pcs:
                    skipped.append(AwbChangeRecord(
                        awb_no=awb_no,
                        action="SKIPPED",
                        reason=f"PDF pcs ({db_pcs}) exceeds scanned pcs ({scanned_pcs})",
                        old_pcs=booked_pcs,
                        new_pcs=db_pcs,
                        scanning_started=scanning,
                    ))
                    continue

                remaining_for_this_flight = scanned_pcs - booked_elsewhere

            # ── ❗ CRITICAL CHECK: prevent overbooking ───────────
            if db_pcs > remaining_for_this_flight:
                skipped.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="SKIPPED",
                    reason=(
                        f"PDF pcs ({db_pcs}) exceeds available remaining "
                        f"({remaining_for_this_flight})"
                    ),
                    old_pcs=booked_pcs,
                    new_pcs=db_pcs,
                    scanning_started=scanning,
                ))
                continue

            if scanning:
                # check loaded count — cannot reduce below loaded
                loaded_result = await db.execute(
                    select(
                        func.count(ExportSequenceItemUldLoading.id).label("loaded")
                    ).where(
                        ExportSequenceItemUldLoading.flight_header_id == header_id,
                        ExportSequenceItemUldLoading.awb_master_id == row.id,
                    )
                )
                loaded_count = loaded_result.scalar() or 0

                if db_pcs < loaded_count:
                    skipped.append(AwbChangeRecord(
                        awb_no=awb_no,
                        action="SKIPPED",
                        reason=f"Cannot reduce pcs to {db_pcs} — {loaded_count} items already loaded into ULD",
                        old_pcs=booked_pcs,
                        new_pcs=db_pcs,
                        scanning_started=True,
                    ))
                else:
                    pcs_updates.append((detail_row.detail_id, db_pcs))
                    updated.append(AwbChangeRecord(
                        awb_no=awb_no,
                        action="PCS_UPDATED",
                        reason=f"Pcs updated from {booked_pcs} to {db_pcs} (scanning started but update is safe)",
                        old_pcs=booked_pcs,
                        new_pcs=db_pcs,
                        scanning_started=True,
                    ))
            else:
                pcs_updates.append((detail_row.detail_id, db_pcs))
                updated.append(AwbChangeRecord(
                    awb_no=awb_no,
                    action="PCS_UPDATED",
                    reason=f"Pcs updated from {booked_pcs} to {db_pcs}",
                    old_pcs=booked_pcs,
                    new_pcs=db_pcs,
                    scanning_started=False,
                ))

    # ── Apply all DB changes ───────────────────────────────────

    # Remove details
    if detail_ids_to_remove:
        await db.execute(
            ExportFlightBookingDetail.__table__.delete().where(
                ExportFlightBookingDetail.id.in_(detail_ids_to_remove)
            )
        )

    # Update pcs
    for detail_id, new_pcs in pcs_updates:
        await db.execute(
            ExportFlightBookingDetail.__table__.update()
            .where(ExportFlightBookingDetail.id == detail_id)
            .values(booked_pcs=new_pcs)
        )

    # Insert new AWBs
    if awbs_to_add:
        db.add_all([
            ExportFlightBookingDetail(
                flight_header_id=header_id,
                awb_master_id=item.awb_master_id,
                booked_pcs=item.booked_pcs,
            )
            for item in awbs_to_add
        ])

    # Update header timestamp
    existing_header.updated_at = now

    # Write audit logs for all changes
    all_changes = added + removed + updated
    for change in all_changes:
        awb_row = next(
            (r for r in db_awb_map.values() if r.awb_no == change.awb_no), None
        )
        if not awb_row:
            continue
        await write_car_message_flow_audit(
            db=db,
            awb_reference_id=awb_row.id,
            flight_reference_id=header_id,
            module=CarMessageFlowModule.FLIGHT_BOOKING,
            flow_step=CarMessageFlowStep.FLIGHT_BOOKING,
            record_id=header_id,
            action="UPDATE",
            performed_by=booked_by,
            changes={
                "event": f"PDF_UPSERT_{change.action}",
                "flight_no": flight_no,
                "flight_date": str(flight_date),
                "awb_no": change.awb_no,
                "reason": change.reason,
                "old_pcs": change.old_pcs,
                "new_pcs": change.new_pcs,
                "scanning_started": change.scanning_started,
            },
        )

    await db.commit()

    # ── Build summary message ──────────────────────────────────
    parts = []
    if added:    parts.append(f"{len(added)} added")
    if removed:  parts.append(f"{len(removed)} removed")
    if updated:  parts.append(f"{len(updated)} pcs updated")
    if skipped:  parts.append(f"{len(skipped)} skipped")
    if unchanged: parts.append(f"{len(unchanged)} unchanged")
    summary = f"Flight {flight_no} updated — " + ", ".join(parts) + "."

    return PdfUpsertResponse(
        success=True,
        message=summary,
        flight_no=flight_no,
        flight_date=str(flight_date),
        is_new_flight=False,
        total_awbs_in_pdf=len(pdf_awb_nos),
        added=added,
        removed=removed,
        updated=updated,
        skipped=skipped,
        unchanged=unchanged,
        not_found_in_db=not_found_in_db,
        added_count=len(added),
        removed_count=len(removed),
        updated_count=len(updated),
        skipped_count=len(skipped),
        unchanged_count=len(unchanged),
        not_found_count=len(not_found_in_db),
    )














async def mark_awb_ultra_fast(
    db: AsyncSession,
    awb_master_id: int,
    is_ultra_fast: bool,
    marked_by: str,
    remarks: str | None = None,
) -> dict:

    awb = await db.get(ExportCarMessageAwbMaster, awb_master_id)
    if not awb:
        raise HTTPException(status_code=404, detail="AWB not found")
    
    

    # ── Prevent re-marking or unmarking ─────────────────────
    if awb.is_ultra_fast and is_ultra_fast:
        raise HTTPException(
            status_code=400,
            detail=f"AWB {awb.awb_no} is already marked ultra-fast — cannot mark again",
        )

    if awb.is_ultra_fast and not is_ultra_fast:
        raise HTTPException(
            status_code=400,
            detail=f"AWB {awb.awb_no} is already ultra-fast — cannot revert to normal"
        )

    # ── Block if scanning already started ─────────────────────
    if is_ultra_fast:
        scan_check = await db.execute(
            select(ExportAwbSkidItemSequence.id)
            .join(
                ExportAwbSkidMapping,
                ExportAwbSkidItemSequence.mapping_id == ExportAwbSkidMapping.id,
            )
            .where(ExportAwbSkidMapping.awb_master_id == awb_master_id)
            .limit(1)
        )
        if scan_check.scalar_one_or_none():
            raise HTTPException(
                status_code=400,
                detail=(
                    f"AWB {awb.awb_no} already has scanning started — "
                    "cannot mark as ultra-fast"
                ),
            )

    old_value = awb.is_ultra_fast
    awb.is_ultra_fast = is_ultra_fast

        # ✅ Track who marked and when

    awb.is_ultra_fast_marked_by = marked_by if is_ultra_fast else None
    awb.is_ultra_fast_marked_at = get_utc_now() if is_ultra_fast else None

    # ✅ Add remarks if provided
    if remarks:
        awb.remarks = remarks

    await write_car_message_flow_audit(
        db=db,
        awb_reference_id=awb_master_id,
        flight_reference_id=None,
        module=CarMessageFlowModule.AWB_MASTER,
        flow_step=CarMessageFlowStep.AWB_MASTER,
        record_id=awb_master_id,
        action="UPDATE",
        performed_by=marked_by,
        changes={
            "event": "ULTRA_FAST_MARKED",
            "awb_no": awb.awb_no,
            "is_ultra_fast_before": old_value,
            "is_ultra_fast_after": is_ultra_fast,
             "remarks": remarks,
        },
    )

    await db.commit()

    return {
        "success": True,
        "message": (
            f"AWB {awb.awb_no} marked as ultra-fast"
            if is_ultra_fast
            else f"AWB {awb.awb_no} ultra-fast mode removed"
        ),
        "awb_no": awb.awb_no,
        "remarks": remarks,
        "is_ultra_fast": is_ultra_fast,
    }












# 🤢------- Get All awb of car message table to show in table in frontend with pagination -------------------
async def get_awb_data_filtered(
    db: AsyncSession,
    start_date: Optional[date],
    end_date: Optional[date],
    status: str,
    page: int,
    page_size: int,
):
    # ============================
    # Date Parsing
    # ============================

    # ============================
    # Date Conversion (IST → UTC)
    # ============================
    start_dt, _ = convert_ist_day_to_utc_range(start_date)
    _, end_dt = convert_ist_day_to_utc_range(end_date)


    conditions = []

    if start_dt:
        conditions.append(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= start_dt
        )

    if end_dt:
        conditions.append(
            ExportCarMessageAwbMaster.car_message_datetime_combo <= end_dt
        )

    # ============================
    # Base Conditions
    # ============================
    # conditions = [
    #     ExportCarMessageAwbMaster.car_message_datetime_combo >= start_dt,
    #     ExportCarMessageAwbMaster.car_message_datetime_combo <= end_dt,
    # ]

    # ============================
    # Status Filter
    # ============================
    # if status != "all":
    #     if status == "rcs":
    #         conditions.append(
    #             ExportCarMessageAwbMaster.status == "RCS"
    #         )
    #     elif status == "not_rcs":
    #         conditions.append(
    #             ExportCarMessageAwbMaster.status != "RCS"
    #         )
    #     else:
    #         raise HTTPException(
    #             status_code=400,
    #             detail=f"Invalid status '{status}'. Allowed: all | rcs | not_rcs"
    #         )

    if status != "all":
        if status == "rcs":
            # Must have status 'RCS' AND a timestamp
            conditions.append(ExportCarMessageAwbMaster.status == "RCS")
            conditions.append(ExportCarMessageAwbMaster.rcs_datetime.isnot(None))
            
        elif status == "not_rcs":
            # (Status is not RCS) OR (Status is RCS but timestamp is missing)
            conditions.append(
                or_(
                    ExportCarMessageAwbMaster.status != "RCS",
                    and_(
                        ExportCarMessageAwbMaster.status == "RCS",
                        ExportCarMessageAwbMaster.rcs_datetime.is_(None)
                    )
                )
            )
        else:
            raise HTTPException(
                status_code=400,
                detail=f"Invalid status '{status}'. Allowed: all | rcs | not_rcs"
            )

    # ============================
    # Base Query (selected columns)
    # ============================
    base_query = (
        select(
            ExportCarMessageAwbMaster.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.sb_no,
            ExportCarMessageAwbMaster.sb_date,
            ExportCarMessageAwbMaster.pcs,
            ExportCarMessageAwbMaster.gross_wt,
            ExportCarMessageAwbMaster.chg_wt,
            ExportCarMessageAwbMaster.nog,
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.car_msg_date,
            ExportCarMessageAwbMaster.car_msg_time,
            ExportCarMessageAwbMaster.car_message_datetime_combo,
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportCarMessageAwbMaster.is_ultra_fast,
            ExportCarMessageAwbMaster.created_at,
            ExportCarMessageAwbMaster.is_manually_created,
                ExportCarMessageAwbMaster.manual_created_by,
                ExportCarMessageAwbMaster.manual_creation_remarks,
                ExportCarMessageAwbMaster.manual_pcs,

        )
        .where(and_(*conditions))
    )

    # ============================
    # Count Query
    # ============================
    count_stmt = (
        select(func.count())
        .select_from(ExportCarMessageAwbMaster)
        .where(and_(*conditions))
    )

    total_result = await db.execute(count_stmt)
    total = total_result.scalar() or 0

    # ============================
    # Pagination
    # ============================
    offset = (page - 1) * page_size

    data_stmt = (
        base_query
        .order_by(ExportCarMessageAwbMaster.created_at.desc())
        .offset(offset)
        .limit(page_size)
    )

    result = await db.execute(data_stmt)
    rows = result.mappings().all()

    records = [dict(row) for row in rows]

    return records, total


async def get_awb_data_for_export(
    db: AsyncSession,
   start_date: Optional[date],
    end_date: Optional[date],
    status: str,
) -> list[dict]:


        # ============================
    # Convert IST → UTC
    # ============================
    start_dt, _ = convert_ist_day_to_utc_range(start_date)
    _, end_dt = convert_ist_day_to_utc_range(end_date)

    # ============================
    # Conditions
    # ============================
    conditions = []

    if start_dt:
        conditions.append(
            ExportCarMessageAwbMaster.car_message_datetime_combo >= start_dt
        )

    if end_dt:
        conditions.append(
            ExportCarMessageAwbMaster.car_message_datetime_combo <= end_dt
        )

    if status != "all":
        if status == "rcs":
            conditions.append(ExportCarMessageAwbMaster.status == "RCS")
            conditions.append(ExportCarMessageAwbMaster.rcs_datetime.isnot(None))
        elif status == "not_rcs":
            conditions.append(
                or_(
                    ExportCarMessageAwbMaster.status != "RCS",
                    and_(
                        ExportCarMessageAwbMaster.status == "RCS",
                        ExportCarMessageAwbMaster.rcs_datetime.is_(None)
                    )
                )
            )
        else:
            raise HTTPException(status_code=400, detail=f"Invalid status '{status}'.")

    stmt = (
        select(
            ExportCarMessageAwbMaster.awb_no,
            ExportCarMessageAwbMaster.origin,
            ExportCarMessageAwbMaster.destination,
            ExportCarMessageAwbMaster.sb_no,
            ExportCarMessageAwbMaster.sb_date,
            ExportCarMessageAwbMaster.pcs,
            ExportCarMessageAwbMaster.gross_wt,
            ExportCarMessageAwbMaster.chg_wt,
            ExportCarMessageAwbMaster.nog,
            ExportCarMessageAwbMaster.status,
            ExportCarMessageAwbMaster.agent,
            ExportCarMessageAwbMaster.car_msg_date,
            ExportCarMessageAwbMaster.car_msg_time,
            ExportCarMessageAwbMaster.car_message_datetime_combo,
            ExportCarMessageAwbMaster.rcs_datetime,
            ExportCarMessageAwbMaster.is_ultra_fast,
            ExportCarMessageAwbMaster.is_manually_created,
            ExportCarMessageAwbMaster.manual_pcs,
            ExportCarMessageAwbMaster.created_at,
        )
        .where(and_(*conditions))
        .order_by(ExportCarMessageAwbMaster.created_at.desc())
    )

    result = await db.execute(stmt)
    rows = result.mappings().all()
    return [dict(row) for row in rows]


def build_car_message_excel(records: list[dict]) -> BytesIO:
    IST = pytz.timezone("Asia/Kolkata")

    def format_status(row: dict) -> str:
        status = row.get("status")
        rcs_dt = row.get("rcs_datetime")

        if status == "RCS":
            return "RCS" if rcs_dt else ""
        return status or ""

    def to_ist(val):
        if not val:
            return ""
        if isinstance(val, datetime):
            if val.tzinfo is None:
                val = val.replace(tzinfo=timezone.utc)
            return val.astimezone(IST).strftime("%d-%b-%Y %H:%M")
        return str(val)

    def fmt_date(val):
        if not val:
            return ""
        return val.strftime("%d-%b-%Y") if hasattr(val, "strftime") else str(val)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Car Message AWB"

    headers = [
        "AWB No", "Origin", "Destination", "SB No", "SB Date",
        "Pcs", "Gross Wt", "Chg Wt", "NOG", "Status", "Agent",
         "Car Msg Datetime (IST)",
        "RCS Datetime (IST)", "Ultra Fast", "Manually Created",
        "Manual Pcs", "Created At (IST)",
    ]

    # Header row style
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill("solid", fgColor="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(records, 2):
        ws.cell(row=row_idx, column=1,  value=r.get("awb_no"))
        ws.cell(row=row_idx, column=2,  value=r.get("origin"))
        ws.cell(row=row_idx, column=3,  value=r.get("destination"))
        ws.cell(row=row_idx, column=4,  value=r.get("sb_no"))
        ws.cell(row=row_idx, column=5,  value=fmt_date(r.get("sb_date")))
        ws.cell(row=row_idx, column=6,  value=r.get("pcs"))
        ws.cell(row=row_idx, column=7,  value=r.get("gross_wt"))
        ws.cell(row=row_idx, column=8,  value=r.get("chg_wt"))
        ws.cell(row=row_idx, column=9,  value=r.get("nog"))
        # ws.cell(row=row_idx, column=10, value=r.get("status"))
        ws.cell(row=row_idx, column=10, value=format_status(r))
        ws.cell(row=row_idx, column=11, value=r.get("agent"))
        # ws.cell(row=row_idx, column=12, value=fmt_date(r.get("car_msg_date")))
        # ws.cell(row=row_idx, column=13, value=r.get("car_msg_time"))
        ws.cell(row=row_idx, column=12, value=to_ist(r.get("car_message_datetime_combo")))
        ws.cell(row=row_idx, column=13, value=to_ist(r.get("rcs_datetime")))
        ws.cell(row=row_idx, column=14, value="YES" if r.get("is_ultra_fast") else "NO")
        ws.cell(row=row_idx, column=15, value="YES" if r.get("is_manually_created") else "NO")
        ws.cell(row=row_idx, column=16, value=r.get("manual_pcs"))
        ws.cell(row=row_idx, column=17, value=to_ist(r.get("created_at")))

    # Auto column width
    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf







# 🤢------- Create manual AWB (for cases where PDF upload is not possible in early stage) -------------------
async def create_manual_awb_service(
    db: AsyncSession,
    data: AwbManualCreateRequest,
    emp_id: str,
):
    # ── 1. Check duplicate ─────────────────────
    result = await db.execute(
        select(ExportCarMessageAwbMaster).where(
            ExportCarMessageAwbMaster.awb_no == data.awb_no
        )
    )
    existing = result.scalar_one_or_none()

    if existing:
        raise HTTPException(
            status_code=status.HTTP_409_CONFLICT,
            detail=f"AWB '{data.awb_no}' already exists",
        )

    now = get_utc_now()
    now_ist = datetime.now(ZoneInfo("Asia/Kolkata"))

    # ─────────────────────────────────────────────
    # IST → UTC Builder
    # ─────────────────────────────────────────────
    IST = pytz.timezone("Asia/Kolkata")
    UTC = pytz.utc
    def build_utc_combo(date_value, time_value):
        if not date_value:
            return None

        if not time_value:
            time_value = "00:00:00"

        try:
            dt = datetime.combine(
                date_value,
                datetime.strptime(time_value, "%H:%M:%S").time()
            )

            ist_dt = IST.localize(dt)
            return ist_dt.astimezone(UTC)

        except Exception:
            return None
    
    # ── 2. Create AWB ──────────────────────────
    new_awb = ExportCarMessageAwbMaster(
        awb_no=data.awb_no,
        pcs=data.pcs,


        # ✅ manual flags
        is_manually_created=True,
        manual_created_by=emp_id,
        manual_creation_remarks=data.manual_creation_remarks,
        manual_pcs=data.pcs,

        # Manual also at ultra fast by default, since manual creation is typically for urgent cases without PDF
        is_ultra_fast=True,
        is_ultra_fast_marked_by=emp_id,
        is_ultra_fast_marked_at=now,

        # optional defaults
        status="RCS",   # if your flow requires
        # rcs_datetime=now,

        # ✅ car message date time and combo for easier querying and filtering when manual create
        car_msg_date=now_ist.date(),
        car_msg_time=now_ist.strftime("%H:%M:%S"),
        car_message_datetime_combo=build_utc_combo(now_ist.date(), now_ist.strftime("%H:%M:%S")),

        created_at=now,
        updated_at=now,
    )

    db.add(new_awb)
    await db.commit()
    await db.refresh(new_awb)

    return {
        "success": True,
        "message": "AWB created successfully",
        "awb_id": new_awb.id,
        "awb_no": new_awb.awb_no,
        "is_manually_created": new_awb.is_manually_created,
        "is_ultra_fast" : new_awb.is_ultra_fast,
        "manual_pcs": new_awb.manual_pcs,
    }










# ==========

async def extract_carrier_for_uld_filter(
    db: AsyncSession,
    flight_no: str,
) -> str | None:
    """
    Extracts carrier code from flight_no for ULD filtering.
    Returns carrier_code string if found, None otherwise.
    ================
    Extract carrier code from flight number.

    Logic:
    1. Validate flight_no format (2 or 3 letters + digits)
    2. Try to match carrier in DB
    3. Return carrier if found
    4. Otherwise return None with debug reason
    """
    flight_no = flight_no.strip().upper()

    # ── Try 2-char first, then 3-char ─────────────────────────
    match2 = re.match(r"^([A-Z0-9]{2})(\d{1,4})$", flight_no)
    match3 = re.match(r"^([A-Z0-9]{3})(\d{1,4})$", flight_no)

    if not match2 and not match3:
        print(f"❌ Invalid flight_no format: {flight_no}")
        return None

    # ── 2-char lookup first ────────────────────────────────────
    if match2:
        result = await db.execute(
            select(ExportCarrierMaster.carrier_code).where(
                ExportCarrierMaster.carrier_code == match2.group(1),
                ExportCarrierMaster.is_active == True,
            )
        )
        code = result.scalar_one_or_none()
        if code:
            return code

    # ── 3-char lookup fallback ─────────────────────────────────
    if match3:
        result = await db.execute(
            select(ExportCarrierMaster.carrier_code).where(
                ExportCarrierMaster.carrier_code == match3.group(1),
                ExportCarrierMaster.is_active == True,
            )
        )
        code = result.scalar_one_or_none()
        if code:
            return code
    # ── Final fallback ─────────────────────────────
    print(f"❌ Carrier extraction failed for flight_no: {flight_no}")
    return None





#  =============== 🫥 INDIVIDUAL ULD CLOSING  PER FLIGHTS PER DATE ====================

async def close_per_uld__per_flight_service(
    db: AsyncSession,
    uld_assignment_detail_id: int,
    closed_by: str,
):
    now = get_utc_now()

    # ── get ULD detail ─────────────────────
    uld = await db.get(ExportUldAssignmentDetail, uld_assignment_detail_id)
    print(f"Debug: Closing ULD Assignment Detail ID {uld_assignment_detail_id} by {closed_by} : uld- {uld}")

    if not uld:
        raise HTTPException(404, "ULD not found")

    if uld.is_closed:
        raise HTTPException(400, "ULD already closed")

    # ── get flight id ──────────────────────
    assignment = await db.get(ExportUldAssignment, uld.assignment_id)
    flight_id = assignment.flight_header_id

    # ── get sequences loaded in THIS ULD ───
    loaded_count = await db.scalar(
        select(func.count(ExportSequenceItemUldLoading.id))
        .where(
            ExportSequenceItemUldLoading.uld_assignment_detail_id == uld_assignment_detail_id
        )
    ) or 0

    if loaded_count == 0:
        raise HTTPException(400, "Cannot close empty ULD")

    # ✅ OPTIONAL VALIDATION
    # check if any pending sequences still expected in this ULD
    # (usually not required unless you assign capacity per ULD)

    # ── mark closed ─────────────────────────
    uld.is_closed = True
    uld.closed_by = closed_by
    uld.closed_at = now

    await db.commit()

    return {
        "success": True,
        "message": "ULD closed successfully",
        "uld_id": uld_assignment_detail_id,
        "loaded_count": loaded_count,
    }






# ============================== ✈️✈️FIGHT HISTORY RELATED ROUTES======================

async def get_flight_history(db, date):
    flights = await db.execute(
        select(ExportFlightBookingHeader)
        .where(
            ExportFlightBookingHeader.flight_date == date,
            ExportFlightBookingHeader.is_active == True
        )
    )

    flights = flights.scalars().all()

    result = []

    for f in flights:
        # AWB + pcs
        details = await db.execute(
            select(ExportFlightBookingDetail)
            .where(ExportFlightBookingDetail.flight_header_id == f.id)
        )

        details = details.scalars().all()

        total_awbs = len(details)
        total_pcs = sum(d.booked_pcs for d in details)

        result.append({
            "flight_id": f.id,
            "flight_no": f.flight_no,
            "flight_date": f.flight_date,
            "departure": f.flight_dpt_datetime,
            "booked_by": f.booked_by,
            "booked_at": f.booked_at,
            "total_awbs": total_awbs,
            "total_pcs": total_pcs,
        })

    return result


# This give which awb is booked in which flight with how many pcs and how many total pcs in that awb and pending pcs in that awb for a particular flight
async def get_flight_awb_breakdown(db: AsyncSession, flight_id: int):
    result = await db.execute(
        select(
            ExportFlightBookingDetail.id,
            ExportCarMessageAwbMaster.awb_no,
            ExportFlightBookingDetail.booked_pcs,
            ExportCarMessageAwbMaster.pcs.label("total_pcs"),
        )
        .join(
            ExportCarMessageAwbMaster,
            ExportFlightBookingDetail.awb_master_id == ExportCarMessageAwbMaster.id,
        )
        .where(
            ExportFlightBookingDetail.flight_header_id == flight_id
        )
    )

    rows = result.mappings().all()

    awbs = []
    total_booked = 0
    total_pcs = 0

    for r in rows:
        total_booked += r.booked_pcs or 0
        total_pcs += r.total_pcs or 0

        awbs.append({
            "awb_no": r.awb_no,
            "booked_pcs": r.booked_pcs or 0,
            "total_pcs": r.total_pcs or 0,
            "pending_pcs": (r.total_pcs or 0) - (r.booked_pcs or 0),
        })

    return {
        "flight_id": flight_id,
        "total_awbs": len(awbs),
        "total_booked_pcs": total_booked,
        "total_pcs": total_pcs,
        "awbs": awbs,
    }



async def get_flight_particular_flight_detail(db, flight_id):

    # 1. Assignment
    assignment = await db.execute(
        select(ExportUldAssignment)
        .where(
            ExportUldAssignment.flight_header_id == flight_id,
            ExportUldAssignment.is_active == True
        )
    )
    assignment = assignment.scalar_one_or_none()

    if not assignment:
        return []

    # 2. ULD + ULD MASTER (FIXED JOIN)
    ulds = await db.execute(
        select(
            ExportUldAssignmentDetail,
            ExportUldMaster
        )
        .join(
            ExportUldMaster,
            ExportUldAssignmentDetail.uld_id == ExportUldMaster.id
        )
        .where(
            ExportUldAssignmentDetail.assignment_id == assignment.id
        )
    )

    rows = ulds.all()

    result = []

    for u, uld in rows:

        # 3. sequences loaded in this ULD
        seqs = await db.execute(
            select(ExportSequenceItemUldLoading)
            .where(
                ExportSequenceItemUldLoading.uld_assignment_detail_id == u.id
            )
        )
        seqs = seqs.scalars().all()

        total_pcs = len(seqs)

        first_scan = min([s.loaded_at for s in seqs], default=None)
        last_scan = max([s.loaded_at for s in seqs], default=None)

        result.append({
            "uld_detail_id": u.id,
            "uld_no": uld.uld_no,   # ✅ FIXED
            "total_loaded_pcs": total_pcs,
            "first_scan": first_scan,
            "last_scan": last_scan,
            "is_closed": u.is_closed,
            "closed_by": u.closed_by,
            "closed_at": u.closed_at,
        })

    return result

async def get_uld_sequences_of_particular_flight(db, flight_id, uld_detail_id):
    rows = await db.execute(
        select(
            ExportSequenceItemUldLoading,
            ExportAwbSkidItemSequence,
            ExportCarMessageAwbMaster.awb_no  # ✅ CORRECT SOURCE
        )
        .join(
            ExportAwbSkidItemSequence,
            ExportSequenceItemUldLoading.sequence_id == ExportAwbSkidItemSequence.id
        ).join(  # ✅ IMPORTANT JOIN
            ExportCarMessageAwbMaster,
            ExportAwbSkidItemSequence.awb_master_id == ExportCarMessageAwbMaster.id
        )
        .where(
            ExportSequenceItemUldLoading.uld_assignment_detail_id == uld_detail_id
        )
    )

    rows = rows.all()

    result = []

    for load, seq, awb_no in rows:
        result.append({
            "sequence_no": seq.sequence_no,
            "awb_id": seq.awb_master_id,
            "awb_no": awb_no,
            "loaded_by": load.loaded_by,
            "loaded_at": load.loaded_at,
            "scanned_by": seq.scanned_by,
            "scan_time": seq.sequence_date_time,
        })

    return result