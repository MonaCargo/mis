


"""
utils/digital_reports/import_dept/seg_report_builder.py
Exact layout match to target image.

Metrics per date group (6): Flight | MAWB | HAWB | Piece | Gross MT | Chg MT
AIR INDIA is an expandable parent: parent summary row (deduped flight count)
followed by its indented children — Air India (Delhi) and Air India (TP).

show_zeros flag:
  True  → 0 / 0.000 shown (default, current behaviour)
  False → cells with no cargo are left BLANK (Point 7)
"""

import io
from datetime import date, datetime, timezone
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Colours (ARGB) ────────────────────────────────────────────────────────────
_C_BLUE       = "FFD9E1F2"   # light blue  — odd date columns
_C_GRAY       = "FFD6DCE4"   # light gray  — even date columns
_C_YELLOW     = "FFFFFF00"   # yellow      — Total column + Total row

_IST = ZoneInfo("Asia/Kolkata")


def _to_ist(iso_str: str) -> str:
    """ISO datetime string → IST display string 'DD/MM/YYYY HH:MM'."""
    try:
        dt = datetime.fromisoformat(iso_str.replace("Z", "+00:00"))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(_IST).strftime("%d/%m/%Y %H:%M")
    except Exception:
        return iso_str

_C_YELLOW_HDR = "FFFFD966"   # darker yellow — row-1 note bar
_C_GREEN      = "FFE2EFDA"   # light green — PAX group header
_C_PEACH      = "FFFCE4D6"   # light peach — CAO group header
_C_WHITE      = "FFFFFFFF"
_C_HDR_BLUE   = "FFD9E1F2"   # row-1 From/To date area

# ── Border ────────────────────────────────────────────────────────────────────
_THIN   = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_FONT  = "Arial"

# 6 metrics now — MAWB (was AWB) + HAWB (new)
_METRICS = [
    "Flight\nCount", "MAWB\nCount", "HAWB\nCount",
    "Piece\nCount", "Gross\nWgt (MT)", "Chg\nWgt (MT)",
]
_METRIC_KEYS = [
    "flight_count", "mawb_count", "hawb_count",
    "pcs", "gross_wgt_mt", "chg_wgt_mt",
]
_FMT_INT     = "#,##0"
_FMT_FLOAT   = "#,##0.000"

LABEL_COLS = 2   # A = Airline Name, B = Airline Code
METRIC_N   = 6   # cols per date group


def _s(cell, bold=False, bg=None, h_align="center", wrap=False,
        num_fmt=None, size=9, color="FF000000"):
    cell.font      = Font(name=_FONT, bold=bold, size=size, color=color)
    cell.alignment = Alignment(horizontal=h_align, vertical="center",
                                wrap_text=wrap)
    cell.border    = _BORDER
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)
    if num_fmt:
        cell.number_format = num_fmt


def _merge_write(ws, r, c1, c2, value, **skw):
    if c2 > c1:
        ws.merge_cells(start_row=r, start_column=c1, end_row=r, end_column=c2)
    cell = ws.cell(row=r, column=c1, value=value)
    _s(cell, **skw)
    for c in range(c1 + 1, c2 + 1):
        _s(ws.cell(row=r, column=c), bg=skw.get("bg"))


def _date_bg(i: int) -> str:
    return _C_BLUE if i % 2 == 0 else _C_GRAY


def _empty_m() -> dict:
    return {
        "flight_count": 0, "mawb_count": 0, "hawb_count": 0,
        "pcs": 0, "gross_wgt_mt": 0.0, "chg_wgt_mt": 0.0,
    }


def _write_metrics(ws, row, col, metrics, bg, bold=False, show_zeros=True):
    """
    Write the 6 metric cells. When show_zeros is False, a 0 value is written
    as an empty cell (blank) instead of 0 / 0.000.
    """
    fmts = [_FMT_INT, _FMT_INT, _FMT_INT, _FMT_INT, _FMT_FLOAT, _FMT_FLOAT]
    for j, (key, fmt) in enumerate(zip(_METRIC_KEYS, fmts)):
        val = metrics.get(key, 0)
        if not show_zeros and (val == 0 or val is None):
            # blank cell — keep styling/border, no value or number format
            c = ws.cell(row=row, column=col + j, value=None)
            _s(c, bold=bold, bg=bg)
        else:
            c = ws.cell(row=row, column=col + j, value=val)
            _s(c, bold=bold, bg=bg, num_fmt=fmt)


# ═════════════════════════════════════════════════════════════════════════════
# SIMPLE (compact) FORMAT
# ═════════════════════════════════════════════════════════════════════════════

_C_AI_PARENT   = "FFBDD7EE"   # medium blue — AI parent summary row
_C_AI_CHILD    = "FFDEEBF7"   # pale blue   — AI Delhi / TP child rows


def build_excel(report: dict, show_zeros: bool = True) -> bytes:
    wb  = Workbook()
    ws  = wb.active
    ws.title = "Segregation Report"

    dates     = report["dates"]
    n         = len(dates)
    TOT_START = LABEL_COLS + n * METRIC_N + 1
    LAST_COL  = TOT_START + METRIC_N - 1

    # ── ROW 1 ─────────────────────────────────────────────────────────────────
    _s(ws.cell(row=1, column=1, value=""), bg=_C_WHITE)
    _s(ws.cell(row=1, column=2, value="From Date"), bold=True, bg=_C_HDR_BLUE)
    _merge_write(ws, 1, 3, 6, _to_ist(report["from_dt"]), bold=True, bg=_C_HDR_BLUE)
    _s(ws.cell(row=1, column=7, value="To Date"), bold=True, bg=_C_WHITE)
    for c in range(8, 11):
        _s(ws.cell(row=1, column=c, value=""), bg=_C_WHITE)
    _merge_write(ws, 1, 11, 14, _to_ist(report["to_dt"]), bold=True, bg=_C_HDR_BLUE)
    note = "Note : Maximum 31 Days Date Range ( From to TO ) is allowed to Select"
    _merge_write(ws, 1, 15, max(LAST_COL, 15), note, bold=True, bg=_C_YELLOW_HDR)

    # ── ROW 2 ─────────────────────────────────────────────────────────────────
    _merge_write(ws, 2, 1, 2, "Select Date Range", bold=True, bg=_C_HDR_BLUE)
    for i, ds in enumerate(dates):
        label     = date.fromisoformat(ds).strftime("%d-%m-%y")
        bg        = _date_bg(i)
        col_start = LABEL_COLS + i * METRIC_N + 1
        _merge_write(ws, 2, col_start, col_start + METRIC_N - 1, label, bold=True, bg=bg)
    _merge_write(ws, 2, TOT_START, TOT_START + METRIC_N - 1, "Total", bold=True, bg=_C_YELLOW)

    # ── ROW 3 ─────────────────────────────────────────────────────────────────
    _s(ws.cell(row=3, column=1, value="Airline Name"), bold=True, bg=_C_GRAY, wrap=True)
    _s(ws.cell(row=3, column=2, value="Airline\nCode"), bold=True, bg=_C_GRAY, wrap=True)
    for i in range(n):
        bg        = _date_bg(i)
        col_start = LABEL_COLS + i * METRIC_N + 1
        for j, label in enumerate(_METRICS):
            _s(ws.cell(row=3, column=col_start + j, value=label), bold=True, bg=bg, wrap=True)
    for j, label in enumerate(_METRICS):
        _s(ws.cell(row=3, column=TOT_START + j, value=label), bold=True, bg=_C_YELLOW, wrap=True)

    # ── Row writers ───────────────────────────────────────────────────────────
    def _data_row(row_num, name, code, per_date, grand_total, bg_label=None, bold=False):
        _s(ws.cell(row=row_num, column=1, value=name), bold=bold, bg=bg_label, h_align="left")
        _s(ws.cell(row=row_num, column=2, value=code), bold=bold, bg=bg_label)
        for i, ds in enumerate(dates):
            col = LABEL_COLS + i * METRIC_N + 1
            _write_metrics(ws, row_num, col, per_date.get(ds, _empty_m()), _date_bg(i), bold, show_zeros)
        _write_metrics(ws, row_num, TOT_START, grand_total, _C_YELLOW, bold, show_zeros)

    def _group_header(row_num, name, code, per_date, grand_total, bg):
        _s(ws.cell(row=row_num, column=1, value=name), bold=True, bg=bg, h_align="left")
        _s(ws.cell(row=row_num, column=2, value=code), bold=True, bg=bg)
        for i, ds in enumerate(dates):
            col = LABEL_COLS + i * METRIC_N + 1
            _write_metrics(ws, row_num, col, per_date.get(ds, _empty_m()), bg, True, show_zeros)
        _write_metrics(ws, row_num, TOT_START, grand_total, _C_YELLOW, True, show_zeros)

    def _airline_block(cur, airline):
        """Write an airline row; if it's the AI parent, also write its children."""
        is_parent = airline.get("is_parent") and airline.get("children")
        bg = _C_AI_PARENT if is_parent else None
        _data_row(cur, airline["airline_name"], airline["airline_code"],
                  airline["per_date"], airline["grand_total"],
                  bg_label=bg, bold=bool(is_parent))
        cur += 1
        if is_parent:
            for child in airline["children"]:
                _data_row(cur, "    " + child["airline_name"], child["airline_code"],
                          child["per_date"], child["grand_total"],
                          bg_label=_C_AI_CHILD, bold=False)
                cur += 1
        return cur

    # ── PAX ───────────────────────────────────────────────────────────────────
    cur = 4
    pax = report["pax"]
    _group_header(cur, "Passenger Flights", "PAX", pax["per_date"], pax["grand_total"], _C_GREEN)
    cur += 1
    for airline in pax["airlines"]:
        cur = _airline_block(cur, airline)

    for col in range(1, LAST_COL + 1):
        _s(ws.cell(row=cur, column=col, value=""), bg=_C_WHITE)
    cur += 1

    # ── CAO ───────────────────────────────────────────────────────────────────
    cao = report["cao"]
    _group_header(cur, "Freighters", "CAO", cao["per_date"], cao["grand_total"], _C_PEACH)
    cur += 1
    for airline in cao["airlines"]:
        cur = _airline_block(cur, airline)

    for col in range(1, LAST_COL + 1):
        _s(ws.cell(row=cur, column=col, value=""), bg=_C_WHITE)
    cur += 1

    # ── Grand Total ─────────────────────────────────────────────────────────
    _s(ws.cell(row=cur, column=1, value="Total"), bold=True, bg=_C_YELLOW, h_align="left")
    _s(ws.cell(row=cur, column=2, value=""), bold=True, bg=_C_YELLOW)
    for i, ds in enumerate(dates):
        col = LABEL_COLS + i * METRIC_N + 1
        _write_metrics(ws, cur, col, report["per_date"].get(ds, _empty_m()), _C_YELLOW, True, show_zeros)
    _write_metrics(ws, cur, TOT_START, report["grand_total"], _C_YELLOW, True, show_zeros)

    # ── Widths / heights / freeze ────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    for col in range(3, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.row_dimensions[1].height = 20
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 38
    ws.freeze_panes = "C4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_csv(report: dict, show_zeros: bool = True) -> bytes:
    import csv
    buf = io.StringIO()
    w   = csv.writer(buf)
    w.writerow(["airline_type", "airline_code", "airline_name", "date",
                "flight_count", "mawb_count", "hawb_count", "pcs",
                "gross_wgt_mt", "chg_wgt_mt"])

    def cell(v):
        return "" if (not show_zeros and (v == 0 or v is None)) else v

    def _emit(air):
        for ds, m in air["per_date"].items():
            w.writerow([air["airline_type"], air["airline_code"], air["airline_name"], ds,
                        cell(m["flight_count"]), cell(m["mawb_count"]), cell(m["hawb_count"]),
                        cell(m["pcs"]), cell(m["gross_wgt_mt"]), cell(m["chg_wgt_mt"])])

    for gk in ("pax", "cao"):
        for air in report[gk]["airlines"]:
            _emit(air)
            for child in air.get("children", []):
                _emit(child)
    return buf.getvalue().encode("utf-8-sig")


# ═════════════════════════════════════════════════════════════════════════════
# DETAILED FORMAT — group → airline (AI: parent → children) → flights
# ═════════════════════════════════════════════════════════════════════════════

_C_AIRLINE_ROW = "FFD9E1F2"   # light blue  — airline rows
_C_FLIGHT_ROW  = "FFE2EFDA"   # light green — individual flight rows


def build_excel_detailed(report: dict, show_zeros: bool = True) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Segregation Detailed"

    dates = report["dates"]
    n     = len(dates)

    LBL = 3   # SN | Airline Name | Code
    TOT_START = LBL + n * METRIC_N + 1
    LAST_COL  = TOT_START + METRIC_N - 1

    # ── ROW 1 ─────────────────────────────────────────────────────────────────
    _s(ws.cell(row=1, column=1, value=""), bg=_C_WHITE)
    _s(ws.cell(row=1, column=2, value="From Date"), bold=True, bg=_C_HDR_BLUE)
    _merge_write(ws, 1, 3, 6, _to_ist(report["from_dt"]), bold=True, bg=_C_HDR_BLUE)
    _s(ws.cell(row=1, column=7, value="To Date"), bold=True, bg=_C_WHITE)
    for c in range(8, 11):
        _s(ws.cell(row=1, column=c, value=""), bg=_C_WHITE)
    _merge_write(ws, 1, 11, 14, _to_ist(report["to_dt"]), bold=True, bg=_C_HDR_BLUE)
    note = "Note : Maximum 31 Days Date Range ( From to TO ) is allowed to Select"
    _merge_write(ws, 1, 15, max(LAST_COL, 15), note, bold=True, bg=_C_YELLOW_HDR)

    # ── ROW 2 ─────────────────────────────────────────────────────────────────
    _s(ws.cell(row=2, column=1, value="SN"), bold=True, bg=_C_HDR_BLUE)
    _merge_write(ws, 2, 2, 3, "Select Date Range", bold=True, bg=_C_HDR_BLUE)
    for i, ds in enumerate(dates):
        bg = _date_bg(i)
        col_start = LBL + i * METRIC_N + 1
        _merge_write(ws, 2, col_start, col_start + METRIC_N - 1,
                     date.fromisoformat(ds).strftime("%d-%m-%y"), bold=True, bg=bg)
    _merge_write(ws, 2, TOT_START, TOT_START + METRIC_N - 1, "Total", bold=True, bg=_C_YELLOW)

    # ── ROW 3 ─────────────────────────────────────────────────────────────────
    _s(ws.cell(row=3, column=1, value="SN"), bold=True, bg=_C_GRAY, wrap=True)
    _s(ws.cell(row=3, column=2, value="Airline Name"), bold=True, bg=_C_GRAY, wrap=True)
    _s(ws.cell(row=3, column=3, value="Airline\nCode"), bold=True, bg=_C_GRAY, wrap=True)
    for i in range(n):
        bg = _date_bg(i)
        col_start = LBL + i * METRIC_N + 1
        for j, label in enumerate(_METRICS):
            _s(ws.cell(row=3, column=col_start + j, value=label), bold=True, bg=bg, wrap=True)
    for j, label in enumerate(_METRICS):
        _s(ws.cell(row=3, column=TOT_START + j, value=label), bold=True, bg=_C_YELLOW, wrap=True)

    def _row(row_num, sn, name, code, per_date, grand_total, bg, bold=False, name_align="center"):
        _s(ws.cell(row=row_num, column=1, value=sn),   bold=bold, bg=bg)
        _s(ws.cell(row=row_num, column=2, value=name), bold=bold, bg=bg, h_align=name_align)
        _s(ws.cell(row=row_num, column=3, value=code), bold=bold, bg=bg)
        for i, ds in enumerate(dates):
            col = LBL + i * METRIC_N + 1
            _write_metrics(ws, row_num, col, per_date.get(ds, _empty_m()), bg, bold, show_zeros)
        _write_metrics(ws, row_num, TOT_START, grand_total, _C_YELLOW, bold, show_zeros)

    cur = 4

    def _write_airline(airline, sn_airline):
        """Airline row (+ its flights). Returns nothing; advances cur."""
        nonlocal cur
        _row(cur, sn_airline, airline["airline_name"], airline["airline_code"],
             airline["per_date"], airline["grand_total"], _C_AIRLINE_ROW,
             bold=True, name_align="left")
        cur += 1
        for fi, flight in enumerate(airline.get("flights", [])):
            sn_flight = f"{sn_airline}({chr(97 + fi)})"
            _row(cur, sn_flight, flight["flight_no"], "",
                 flight["per_date"], flight["grand_total"], _C_FLIGHT_ROW, bold=False)
            cur += 1

    def _write_section(group_label, group_code, group_key, sn_group):
        nonlocal cur
        grp = report[group_key]
        _row(cur, sn_group, group_label, group_code,
             grp["per_date"], grp["grand_total"], _C_PEACH, bold=True, name_align="left")
        cur += 1

        for ai, airline in enumerate(grp["airlines"], start=1):
            sn_airline = f"{sn_group}.{ai}"
            if airline.get("is_parent") and airline.get("children"):
                # AI parent summary row, then children (each with its flights)
                _row(cur, sn_airline, airline["airline_name"], airline["airline_code"],
                     airline["per_date"], airline["grand_total"], _C_AI_PARENT,
                     bold=True, name_align="left")
                cur += 1
                for ci, child in enumerate(airline["children"], start=1):
                    _write_airline(child, f"{sn_airline}.{ci}")
            else:
                _write_airline(airline, sn_airline)

    _write_section("Passenger Flights", "PAX", "pax", 1)
    _write_section("Freighters",        "CAO", "cao", 2)

    _row(cur, "", "Total", "", report["per_date"], report["grand_total"],
         _C_PEACH, bold=True, name_align="left")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 8
    for col in range(4, LAST_COL + 1):
        ws.column_dimensions[get_column_letter(col)].width = 10
    ws.row_dimensions[3].height = 38
    ws.freeze_panes = "D4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def build_csv_detailed(report: dict, show_zeros: bool = True) -> bytes:
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["airline_type", "airline_code", "airline_name", "flight_no",
                "date", "flight_count", "mawb_count", "hawb_count", "pcs",
                "gross_wgt_mt", "chg_wgt_mt"])

    def cell(v):
        return "" if (not show_zeros and (v == 0 or v is None)) else v

    def _emit_airline(air):
        for flight in air.get("flights", []):
            for ds, m in flight["per_date"].items():
                w.writerow([air["airline_type"], air["airline_code"], air["airline_name"],
                            flight["flight_no"], ds,
                            cell(m["flight_count"]), cell(m["mawb_count"]), cell(m["hawb_count"]),
                            cell(m["pcs"]), cell(m["gross_wgt_mt"]), cell(m["chg_wgt_mt"])])

    for gk in ("pax", "cao"):
        for air in report[gk]["airlines"]:
            if air.get("is_parent") and air.get("children"):
                for child in air["children"]:
                    _emit_airline(child)
            else:
                _emit_airline(air)
    return buf.getvalue().encode("utf-8-sig")