# """
# Cleaning / parsing for the Import Pick Order (Examination) report.

# clean_pick_order_report(file_bytes) -> list[dict]

# Reads the COSYS "PICK ORDER REPORT" .xlsx and returns clean row dicts ready to
# insert. It does NOT touch the DB and does NOT know about report_date — the
# service layer adds report_date and persists. Keeping parsing pure makes it easy
# to unit-test and reuse.

# Report layout (0-based openpyxl is 1-based here):
#     Row 3  : title
#     Row 5  : FROM DATE / TO DATE
#     Row 7  : header  -> AWB No | HWB No | Pcs for Examination |
#                         RFE d&t | FFE d&t | POE start d&t | POE end d&t
#     Row 8+ : data

# Datetimes in the sheet are IST wall-clock; we convert to UTC for storage.
# """

# from __future__ import annotations

# import io
# import re
# from datetime import datetime, timezone, timedelta
# from typing import Optional

# import openpyxl

# IST = timezone(timedelta(hours=5, minutes=30))

# # Column positions (1-based) on the data rows, from the row-7 header.
# _COL_AWB = 1        # A  AWB No
# _COL_HWB = 2        # B  HWB No
# _COL_PCS = 3        # C  Pcs for Examination
# _COL_RFE = 4        # D  RFE date & Time
# _COL_FFE = 5        # E  FFE date & Time
# _COL_POE_START = 6  # F  POE start date & Time
# _COL_POE_END = 7    # G  POE end date & Time

# _HEADER_HINT = "awb no"     # used to locate the header row robustly
# _DATA_START_FALLBACK = 8    # if header detection fails, start here


# def _normalize_awb(raw) -> Optional[str]:
#     """Strip non-digits and keep the last 11 digits (matches other reports)."""
#     if raw is None:
#         return None
#     digits = re.sub(r"\D", "", str(raw))
#     if not digits:
#         return None
#     return digits[-11:] if len(digits) >= 11 else digits


# # def _to_utc(value) -> Optional[datetime]:
# #     """Interpret an Excel datetime as IST wall-clock, return a UTC datetime."""
# #     if value is None or not isinstance(value, datetime):
# #         return None
# #     # Excel datetimes are naive; treat as IST then convert to UTC.
# #     return value.replace(tzinfo=IST).astimezone(timezone.utc)

# def _to_utc(value):
#     if value is None or not isinstance(value, datetime):
#         return None
#     if value.tzinfo is None:
#         value = value.replace(tzinfo=IST)      # naive -> IST (no shift)
#     return value.astimezone(timezone.utc)      # -> real UTC instant


# def _to_int(value) -> Optional[int]:
#     if value is None or value == "":
#         return None
#     try:
#         return int(float(value))
#     except (TypeError, ValueError):
#         return None


# def _clean_str(value) -> Optional[str]:
#     if value is None:
#         return None
#     s = str(value).strip()
#     return s or None


# def _find_header_row(ws) -> int:
#     """Locate the header row by scanning for the 'AWB No' cell (defensive)."""
#     for r in range(1, min(ws.max_row, 20) + 1):
#         a = ws.cell(row=r, column=_COL_AWB).value
#         if a and _HEADER_HINT in str(a).strip().lower():
#             return r
#     return _DATA_START_FALLBACK - 1  # header just above the fallback data start


# def clean_pick_order_report_data_for_digital_reports(file_bytes: bytes) -> list[dict]:
#     """
#     Parse the Pick Order .xlsx bytes into clean row dicts:

#         {
#           "awb_no", "hawb_no", "pcs_for_examination",
#           "rfe_datetime", "ffe_datetime",
#           "poe_start_datetime", "poe_end_datetime",
#         }

#     Rows with no AWB and no HWB are skipped as empty/footer noise.
#     """
#     wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
#     ws = wb.worksheets[0]

#     header_row = _find_header_row(ws)
#     data_start = header_row + 1

#     rows: list[dict] = []
#     for r in range(data_start, ws.max_row + 1):
#         awb_raw = ws.cell(row=r, column=_COL_AWB).value
#         hwb_raw = ws.cell(row=r, column=_COL_HWB).value

#         awb = _normalize_awb(awb_raw)
#         hawb = _clean_str(hwb_raw)

#         # Skip completely empty rows (footers, blank separators).
#         if awb is None and hawb is None:
#             continue

#         rows.append({
#             "awb_no": awb,
#             "hawb_no": hawb,
#             "pcs_for_examination": _to_int(ws.cell(row=r, column=_COL_PCS).value),
#             "rfe_datetime": _to_utc(ws.cell(row=r, column=_COL_RFE).value),
#             "ffe_datetime": _to_utc(ws.cell(row=r, column=_COL_FFE).value),
#             "poe_start_datetime": _to_utc(ws.cell(row=r, column=_COL_POE_START).value),
#             "poe_end_datetime": _to_utc(ws.cell(row=r, column=_COL_POE_END).value),
#         })

#     return rows
























"""
Cleaning / parsing for the Import Pick Order (Examination) report.

clean_pick_order_report(file_bytes, filename) -> list[dict]

Reads the COSYS "PICK ORDER REPORT" as .xlsx OR .csv and returns clean row
dicts ready to insert. Pure (no DB, no report_date) so it's easy to test/reuse.

Report layout (both formats share it):
    Title row     : 'PICK ORDER REPORT:'
    FROM/TO row   : FROM DATE / TO DATE
    Header row    : AWB No | HWB No | Pcs for Examination |
                    RFE d&t | FFE d&t | POE start d&t | POE end d&t
    Data rows     : below the header

Datetimes:
    - .xlsx: usually real datetime objects (naive IST).
    - .csv : day-first STRINGS like '07-07-26 8:43'.
    Both are treated as IST wall-clock and converted to UTC for storage.
"""

from __future__ import annotations

import csv
import io
import math
import re
from datetime import datetime, timezone, timedelta
from typing import Optional

import openpyxl

IST = timezone(timedelta(hours=5, minutes=30))
UTC = timezone.utc

# Column positions (0-based) on the data rows, from the header.
_COL_AWB = 0        # AWB No
_COL_HWB = 1        # HWB No
_COL_PCS = 2        # Pcs for Examination
_COL_RFE = 3        # RFE date & Time
_COL_FFE = 4        # FFE date & Time
_COL_POE_START = 5  # POE start date & Time
_COL_POE_END = 6    # POE end date & Time

_HEADER_HINT = "awb no"

# Day-first string date formats (Indian source). strptime does not guess, so we
# try each in order. ISO is included last as an unambiguous fallback.
_DT_FORMATS = (

    "%d-%b-%y %I:%M:%S %p",   # 13-JUL-26 11:48:53 AM
    "%d-%b-%Y %I:%M:%S %p",   # 13-JUL-2026 11:48:53 AM

    "%d-%m-%y %H:%M:%S",   # 07-07-26 20:35:10
    "%d-%m-%y %H:%M",      # 07-07-26 8:43
    "%d-%m-%y",            # 07-07-26
    "%d-%b-%Y %H:%M:%S",   # 07-Jul-2026 20:35:10
    "%d-%b-%Y %H:%M",      # 07-Jul-2026 20:35
    "%d-%b-%Y",            # 07-Jul-2026
    "%d-%m-%Y %H:%M:%S",   # 07-07-2026 20:35:10
    "%d-%m-%Y %H:%M",      # 07-07-2026 20:35
    "%d-%m-%Y",            # 07-07-2026
    "%d/%m/%Y %H:%M:%S",   # 07/07/2026 20:35:10
    "%d/%m/%Y %H:%M",      # 07/07/2026 20:35
    "%d/%m/%Y",            # 07/07/2026
    "%Y-%m-%d %H:%M:%S",   # 2026-07-07 20:35:10 (ISO)
    "%Y-%m-%d %H:%M",      # 2026-07-07 20:35
    "%Y-%m-%d",            # 2026-07-07
    "%d%b%Y",              # 07JUL2026  (FROM/TO date cell format)
    "%d-%b-%y",            # 07-Jul-26
)


def _to_date(value):
    """Parse a FROM/TO date cell (string like '07JUL2026' or a datetime) -> date."""
    from datetime import date as _date
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, _date):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        pass
    for fmt in _DT_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


class PickOrderDateMismatch(ValueError):
    """Raised when the report's FROM/TO dates don't both equal report_date."""

    def __init__(self, from_date, to_date, report_date):
        self.from_date = from_date
        self.to_date = to_date
        self.report_date = report_date
        super().__init__(
            f"Report date range does not match the selected date. "
            f"File FROM={from_date}, TO={to_date}; selected report_date={report_date}."
        )


def extract_pick_order_range(raw_rows: list[list]) -> tuple[Optional["date"], Optional["date"]]:
    """
    Find the 'FROM DATE :' / 'TO DATE :' row and return (from_date, to_date).

    The value sits in the cell immediately AFTER each label. Robust to column
    position by scanning for the label text rather than a fixed column.
    """
    from_date = to_date = None
    for cells in raw_rows:
        row_texts = [str(c).strip().lower() if c is not None else "" for c in cells]
        for idx, text in enumerate(row_texts):
            if text.startswith("from date"):
                # value is the next non-empty cell after the label
                from_date = _next_value_date(cells, idx)
            elif text.startswith("to date"):
                to_date = _next_value_date(cells, idx)
        if from_date is not None or to_date is not None:
            break
    return from_date, to_date


def _next_value_date(cells: list, label_idx: int):
    """Return the parsed date from the first non-empty cell after label_idx."""
    for j in range(label_idx + 1, len(cells)):
        v = cells[j]
        if v is not None and str(v).strip() != "":
            d = _to_date(v)
            if d is not None:
                return d
    return None


def _normalize_awb(raw) -> Optional[str]:
    """Strip non-digits and keep the last 11 digits (matches other reports)."""
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if not digits:
        return None
    return digits[-11:] if len(digits) >= 11 else digits


def _to_utc(value) -> Optional[datetime]:
    """
    Interpret a value as IST wall-clock and return a UTC datetime.

    Accepts:
      - real datetime objects (naive -> assumed IST; aware -> converted)
      - day-first / ISO strings, parsed via _DT_FORMATS
    Returns None for blanks / unparseable values.
    """
    if value is None or (isinstance(value, float) and math.isnan(value)):
        return None

    # strings: try ISO first, then the day-first formats
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return None
        parsed = None
        try:
            parsed = datetime.fromisoformat(s)
        except ValueError:
            for fmt in _DT_FORMATS:
                try:
                    parsed = datetime.strptime(s, fmt)
                    break
                except ValueError:
                    continue
        if parsed is None:
            return None
        value = parsed

    if isinstance(value, datetime):
        if value.tzinfo is None:
            value = value.replace(tzinfo=IST)   # naive -> IST (no clock shift)
        return value.astimezone(UTC)            # -> real UTC instant

    return None


def _to_int(value) -> Optional[int]:
    if value is None or value == "" or (isinstance(value, float) and math.isnan(value)):
        return None
    try:
        return int(float(str(value)))
    except (TypeError, ValueError):
        return None


def _clean_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _row_to_dict(cells: list) -> Optional[dict]:
    """
    Turn a raw cell list (>=7 wide) into a clean dict, or None if the row is
    empty/footer noise (no AWB and no HWB).
    """
    def cell(i):
        return cells[i] if i < len(cells) else None

    awb = _normalize_awb(cell(_COL_AWB))
    hawb = _clean_str(cell(_COL_HWB))
    if awb is None and hawb is None:
        return None

    return {
        "awb_no": awb,
        "hawb_no": hawb,
        "pcs_for_examination": _to_int(cell(_COL_PCS)),
        "rfe_datetime": _to_utc(cell(_COL_RFE)),
        "ffe_datetime": _to_utc(cell(_COL_FFE)),
        "poe_start_datetime": _to_utc(cell(_COL_POE_START)),
        "poe_end_datetime": _to_utc(cell(_COL_POE_END)),
    }


def _is_header(cells: list) -> bool:
    first = cells[_COL_AWB] if len(cells) > _COL_AWB else None
    return bool(first) and _HEADER_HINT in str(first).strip().lower()


def _rows_from_xlsx(file_bytes: bytes) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]
    return [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]


def _rows_from_csv(file_bytes: bytes) -> list[list]:
    # decode tolerantly (COSYS exports are usually utf-8/latin-1)
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def clean_pick_order_report_data_for_digital_reports(
    file_bytes: bytes,
    filename: str = "",
    report_date=None,
) -> list[dict]:
    """
    Parse Pick Order .xlsx or .csv bytes into clean row dicts. File type is
    chosen by extension, with a CSV fallback if the bytes aren't a valid xlsx.

    If `report_date` is given, the report's FROM DATE and TO DATE are checked:
    both must equal report_date, else PickOrderDateMismatch is raised. This
    guards against uploading the wrong day's file.
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        raw_rows = _rows_from_csv(file_bytes)
    else:
        try:
            raw_rows = _rows_from_xlsx(file_bytes)
        except Exception:
            # Not a real xlsx (e.g. a .csv mislabeled) — try CSV.
            raw_rows = _rows_from_csv(file_bytes)

    # Optional guard: the file's FROM/TO date range must match report_date.
    if report_date is not None:
        from_date, to_date = extract_pick_order_range(raw_rows)
        if from_date != report_date or to_date != report_date:
            raise PickOrderDateMismatch(from_date, to_date, report_date)

    # Find the header row, then clean everything after it.
    header_idx = None
    for i, cells in enumerate(raw_rows):
        if _is_header(cells):
            header_idx = i
            break

    data_rows = raw_rows[header_idx + 1:] if header_idx is not None else raw_rows

    cleaned: list[dict] = []
    for cells in data_rows:
        row = _row_to_dict(cells)
        if row is not None:
            cleaned.append(row)
    return cleaned