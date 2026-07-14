
# # ----------------------------------------
# """
# Cleaning / parsing for the shift-worker Roster report.

# clean_roster_report(file_bytes, filename) -> list[dict]

# Reads the roster .xlsx (or .csv) and returns clean row dicts. Pure — no DB.

# File layout:
#     Row 1 : header  -> SN | Emp. Code | Emp. Name | Desg | Department |
#                        Date | Shift | Present
#     Row 2+: data

# Each clean row dict:
#     {
#       "emp_code":  str,
#       "emp_name":  str | None,
#       "desg":      str | None,
#       "department":str | None,
#       "date":      datetime.date,
#       "shift":     str,
#       "present_status": "P" | "A" | None,
#     }

# Present/Absent mapping (case-insensitive):
#     'Present' -> "P"
#     'Absent'  -> "A"
#     anything else / blank -> None   (preserves the "unknown" state, e.g. an
#                                       ex-employee still listed with no marking)

# Rows with no emp_code, or no parseable date, or no shift are skipped — those
# are the identifying fields an attendance record cannot exist without.
# """

# from __future__ import annotations

# import csv
# import io
# from datetime import date, datetime
# from typing import Optional

# import openpyxl

# # Column positions (0-based) from the row-1 header.
# _COL_SN = 0
# _COL_EMP_CODE = 1
# _COL_EMP_NAME = 2
# _COL_DESG = 3
# _COL_DEPARTMENT = 4
# _COL_DATE = 5
# _COL_SHIFT = 6
# _COL_PRESENT = 7

# _HEADER_HINT = "emp"   # header row's Emp. Code cell contains 'Emp'

# # Day-first date formats for string dates (xlsx usually gives real datetimes).
# _DATE_FORMATS = (
#     "%d-%m-%Y", "%d-%m-%y", "%d-%b-%Y", "%d-%b-%y",
#     "%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d",
#     "%d-%m-%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S",
# )


# def _clean_str(value) -> Optional[str]:
#     if value is None:
#         return None
#     s = str(value).strip()
#     return s or None


# def _emp_code(value) -> Optional[str]:
#     """Emp code as a clean string. Excel may give it as a float (523612.0)."""
#     if value is None:
#         return None
#     if isinstance(value, float) and value.is_integer():
#         return str(int(value))
#     s = str(value).strip()
#     # strip a trailing '.0' if it slipped through as text
#     if s.endswith(".0") and s[:-2].isdigit():
#         s = s[:-2]
#     return s or None


# def _to_date(value) -> Optional[date]:
#     if value is None:
#         return None
#     if isinstance(value, datetime):
#         return value.date()
#     if isinstance(value, date):
#         return value
#     s = str(value).strip()
#     if not s:
#         return None
#     try:
#         return datetime.fromisoformat(s).date()
#     except ValueError:
#         pass
#     for fmt in _DATE_FORMATS:
#         try:
#             return datetime.strptime(s, fmt).date()
#         except ValueError:
#             continue
#     return None


# def _present_status(value) -> Optional[str]:
#     """
#     Map the Present column to a stored status code (case-insensitive):
#         'Present' -> 'P'
#         'Absent'  -> 'A'
#         'LWP'     -> 'LWP'   (Leave Without Pay — excluded from WHA calcs)
#         anything else / blank -> None
#     """
#     if value is None:
#         return None
#     s = str(value).strip().lower()
#     if s == "present":
#         return "P"
#     if s == "absent":
#         return "A"
#     if s == "lwp":
#         return "LWP"
#     return None


# def _is_header(cells: list) -> bool:
#     c = cells[_COL_EMP_CODE] if len(cells) > _COL_EMP_CODE else None
#     return bool(c) and _HEADER_HINT in str(c).strip().lower()


# def _rows_from_xlsx(file_bytes: bytes) -> list[list]:
#     wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
#     ws = wb.worksheets[0]
#     return [
#         [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
#         for r in range(1, ws.max_row + 1)
#     ]


# def _rows_from_csv(file_bytes: bytes) -> list[list]:
#     text = file_bytes.decode("utf-8-sig", errors="replace")
#     return list(csv.reader(io.StringIO(text)))


# def clean_import_roster_report(file_bytes: bytes, filename: str = "") -> list[dict]:
#     name = (filename or "").lower()
#     if name.endswith(".csv"):
#         raw_rows = _rows_from_csv(file_bytes)
#     else:
#         try:
#             raw_rows = _rows_from_xlsx(file_bytes)
#         except Exception:
#             raw_rows = _rows_from_csv(file_bytes)

#     # Locate the header row, default to row 0 being header if not found.
#     header_idx = 0
#     for i, cells in enumerate(raw_rows):
#         if _is_header(cells):
#             header_idx = i
#             break

#     cleaned: list[dict] = []
#     for cells in raw_rows[header_idx + 1:]:
#         def cell(i):
#             return cells[i] if i < len(cells) else None

#         emp_code = _emp_code(cell(_COL_EMP_CODE))
#         d = _to_date(cell(_COL_DATE))
#         shift = _clean_str(cell(_COL_SHIFT))

#         # An attendance record needs at least: who, when (date), which shift.
#         if emp_code is None or d is None or shift is None:
#             continue

#         cleaned.append({
#             "emp_code": emp_code,
#             "emp_name": _clean_str(cell(_COL_EMP_NAME)),
#             "desg": _clean_str(cell(_COL_DESG)),
#             "department": _clean_str(cell(_COL_DEPARTMENT)),
#             "date": d,
#             "shift": shift,
#             "present_status": _present_status(cell(_COL_PRESENT)),
#         })

#     return cleaned







# ======================================================================================
# ============================= WITH NEW EMPLOYEE SHEETS  ==========================



"""
Cleaning / parsing for the Import Manpower Attendance Sheet (WIDE format).

clean_import_roster_report(file_bytes, filename) -> list[dict]

FILE SHAPE (new format)
──────────────────────
A monthly attendance MATRIX: one row per employee, every DATE its own column.

    Row 1 (header):  SN | Emp. Code | Emp. Name | Desg | Department | Shift |
                     2026-07-01 | 2026-07-02 | ... | 2026-07-31
    Row 2+ (data):   1  | 512175    | YAM ...   | WHA  | Documentation | Morning |
                     Present     | Absent     | ... | Present

ONE spreadsheet row holds MANY attendance records — one per date column. This
cleaner UNPIVOTS (melts) that matrix into the long rows the DB expects, the same
shape the previous (long-format) file produced:

    {emp_code, emp_name, desg, department, date, shift, present_status}

Nothing downstream changes: the model, upsert service and route already store
one row per (emp_code, date, shift).

BLANK CELLS
───────────
The sheet is filled in day by day, so future dates are blank. A blank cell means
"not recorded yet" — we SKIP it (no DB row). Re-uploading later, once more days
are filled, simply adds those days: the service upserts on
(emp_code, date, shift), so existing rows update and new ones append.

STATUS MAPPING (case-insensitive)
─────────────────────────────────
    'Present' -> 'P'
    'Absent'  -> 'A'
    'LWP'     -> 'LWP'   (Leave Without Pay — excluded from WHA counts)
    blank     -> skipped entirely (not stored)
    other text-> None (stored with unknown status)
"""

from __future__ import annotations

import csv
import io
from datetime import date, datetime
from typing import Optional

import openpyxl

# Fixed employee-info columns (0-based), left of the date columns.
_COL_SN = 0
_COL_EMP_CODE = 1
_COL_EMP_NAME = 2
_COL_DESG = 3
_COL_DEPARTMENT = 4
_COL_SHIFT = 5
# Date columns start here and run to the end of the sheet.
_FIRST_DATE_COL = 6

_HEADER_HINT = "emp"

_DATE_FORMATS = (
    "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%d-%b-%Y", "%d-%b-%y",
    "%d/%m/%Y", "%d/%m/%y",
    "%Y-%m-%d %H:%M:%S", "%d-%m-%Y %H:%M:%S",
)


def _clean_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


def _emp_code(value) -> Optional[str]:
    """Emp code as a clean string. Excel may hand it over as a float."""
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    s = str(value).strip()
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s or None


def _to_date(value) -> Optional[date]:
    """Parse a date-column header into a plain date."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except ValueError:
        pass
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _present_status(value) -> Optional[str]:
    """'Present'->'P', 'Absent'->'A', 'LWP'->'LWP', else None (case-insensitive)."""
    if value is None:
        return None
    s = str(value).strip().lower()
    if s == "present":
        return "P"
    if s == "absent":
        return "A"
    if s == "lwp":
        return "LWP"
    return None


def _is_header(cells: list) -> bool:
    c = cells[_COL_EMP_CODE] if len(cells) > _COL_EMP_CODE else None
    return bool(c) and _HEADER_HINT in str(c).strip().lower()


def _rows_from_xlsx(file_bytes: bytes) -> list[list]:
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.worksheets[0]          # the month sheet (e.g. 'July26')
    return [
        [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        for r in range(1, ws.max_row + 1)
    ]


def _rows_from_csv(file_bytes: bytes) -> list[list]:
    text = file_bytes.decode("utf-8-sig", errors="replace")
    return list(csv.reader(io.StringIO(text)))


def clean_new_import_roster_report(file_bytes: bytes, filename: str = "") -> list[dict]:
    """
    Unpivot the wide monthly attendance matrix into long attendance rows.

    Returns one dict per (employee, date) that has a status filled in:
        {emp_code, emp_name, desg, department, date, shift, present_status}
    """
    name = (filename or "").lower()
    if name.endswith(".csv"):
        raw_rows = _rows_from_csv(file_bytes)
    else:
        try:
            raw_rows = _rows_from_xlsx(file_bytes)
        except Exception:
            raw_rows = _rows_from_csv(file_bytes)

    if not raw_rows:
        return []

    # ── locate the header row ──────────────────────────────────────────────
    header_idx = 0
    for i, cells in enumerate(raw_rows[:10]):
        if _is_header(cells):
            header_idx = i
            break
    header = raw_rows[header_idx]

    # ── map each date column index -> its date ────────────────────────────
    # Columns from _FIRST_DATE_COL onward are dates. Any column whose header
    # doesn't parse as a date is ignored (stray notes / totals columns).
    date_cols: list[tuple[int, date]] = []
    for c in range(_FIRST_DATE_COL, len(header)):
        d = _to_date(header[c])
        if d is not None:
            date_cols.append((c, d))

    if not date_cols:
        raise ValueError(
            "No date columns found in the attendance sheet. Expected dates "
            "from column G onward in the header row."
        )

    # ── melt: one output row per (employee, date-with-a-status) ───────────
    cleaned: list[dict] = []
    for cells in raw_rows[header_idx + 1:]:

        def cell(i):
            return cells[i] if i < len(cells) else None

        emp_code = _emp_code(cell(_COL_EMP_CODE))
        if emp_code is None:
            continue                      # blank / footer row

        emp_name = _clean_str(cell(_COL_EMP_NAME))
        desg = _clean_str(cell(_COL_DESG))
        department = _clean_str(cell(_COL_DEPARTMENT))
        shift = _clean_str(cell(_COL_SHIFT))

        # Shift is part of the DB key (emp_code, date, shift) — a row without it
        # can't be stored. Real files always carry it; skip if genuinely absent.
        if shift is None:
            continue

        for col_idx, d in date_cols:
            raw_status = cell(col_idx)

            # BLANK cell = that day isn't filled in yet -> no record at all.
            # (Re-uploading later, once filled, adds it via the upsert.)
            if raw_status is None or str(raw_status).strip() == "":
                continue

            cleaned.append({
                "emp_code": emp_code,
                "emp_name": emp_name,
                "desg": desg,
                "department": department,
                "date": d,
                "shift": shift,
                "present_status": _present_status(raw_status),
            })


    # print(cleaned)        
    # print("Hello")        

    return cleaned


