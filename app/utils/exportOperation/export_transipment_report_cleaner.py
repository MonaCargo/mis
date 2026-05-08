# import re
# import pytz
# import numpy as np
# import pandas as pd
# from io import BytesIO
# from datetime import datetime, timezone, date, time as dt_time

# # ── Constants ────────────────────────────────────────────────────────────────

# BILLING_SHC_FILTER = {"TRM", "TPV"}

# REQUIRED_COLUMNS = [
#     "AWB No",
#     "PCS",
#     "Gross wgt",
#     "Rec_PCS",
#     "Received wgt",
#     "Received_Chg_Wgt",
#     "SHC",
#     "Billing SHC",
#     "Commodity",
#     "ORG",
#     "DES",
#     "DOC DATE & TIME",           # normalised from ' DOC DATE & TIME.'
#     "EXP TP SEG FLIGHT No.",
#     "EXP TP FLIGHT DATE",
#     "EXP TP SEG No DATE AND TIME",
#     "TRM NO",
#     "TRM DATE",                  # normalised from 'TRM  DATE'
#     "X-Ray DATE",
#     "X-Ray TIME",
#     "RAMP TRANSFER DATE/TIME",
#     "RAMP TRANSFER REMARK",
#     "RAMP TRANSFER USER",
#     "AIRLINE CD",
#     "FLIGHT NO",
#     "FLIGHT DATE",
#     "ULD LOAD",
#     "DEPARTURE DATE & TIME",
# ]

# NUMERIC_COLS = [
#     "PCS", "Gross wgt", "Rec_PCS",
#     "Received wgt", "Received_Chg_Wgt",
#     "TRM NO",
# ]

# # Columns that need IST → UTC datetime parsing
# DATETIME_COLS = [
#     "DOC DATE & TIME",
#     "EXP TP FLIGHT DATE",
#     "TRM DATE",
#     "RAMP TRANSFER DATE/TIME",
#     "DEPARTURE DATE & TIME",
#     "ULD LOAD",
# ]

# # X-Ray DATE and X-Ray TIME are merged into one UTC datetime column
# # FLIGHT DATE and EXP TP FLIGHT DATE are date-only → midnight IST → UTC

# _IST = pytz.timezone("Asia/Kolkata")


# # ── Helpers ──────────────────────────────────────────────────────────────────

# def _clean_awb(val) -> str | None:
#     if not val or str(val).strip().lower() in ("", "nan", "none"):
#         return None
#     s = re.sub(r"\s+", "", str(val).strip())
#     s = re.sub(r"-+", "-", s).replace("-", "")
#     s = re.sub(r"\D", "", s)
#     if s and len(s) == 10:
#         s = s.zfill(11)
#     return s or None


# def _localize_to_utc(dt_naive) -> pd.Timestamp:
#     """Take a naive Python datetime (or pd.Timestamp) → IST → UTC pd.Timestamp."""
#     if dt_naive is None or (isinstance(dt_naive, float) and pd.isna(dt_naive)):
#         return pd.NaT
#     try:
#         if pd.isnull(dt_naive):
#             return pd.NaT
#     except (TypeError, ValueError):
#         pass
#     # pd.Timestamp.localize() is the safe path (avoids pytz Naive time error)
#     if isinstance(dt_naive, pd.Timestamp):
#         if dt_naive.tzinfo is not None:
#             return dt_naive.tz_convert("UTC")
#         return dt_naive.tz_localize("Asia/Kolkata").tz_convert("UTC")
#     # Plain datetime
#     local_dt = _IST.localize(dt_naive.replace(tzinfo=None), is_dst=None)
#     return pd.Timestamp(local_dt.astimezone(pytz.utc))


# def _to_utc_str(val, formats: list) -> pd.Timestamp:
#     """Parse string datetime (assumed IST) → UTC pd.Timestamp."""
#     if not val or str(val).strip().lower() in ("", "nan", "none", "nat"):
#         return pd.NaT
#     s = str(val).strip()
#     for fmt in formats:
#         try:
#             tmp = pd.to_datetime(s, format=fmt, errors="coerce")
#             if pd.notna(tmp):
#                 return _localize_to_utc(tmp.to_pydatetime())
#         except Exception:
#             continue
#     try:
#         tmp = pd.to_datetime(s, errors="coerce", dayfirst=True)
#         if pd.notna(tmp):
#             return _localize_to_utc(tmp.to_pydatetime())
#     except Exception:
#         pass
#     return pd.NaT


# def _parse_date_to_utc(val, formats: list) -> pd.Timestamp:
#     """Parse date-only value → midnight IST → UTC."""
#     if val is None:
#         return pd.NaT
#     # Already a Python datetime from openpyxl
#     if isinstance(val, datetime):
#         naive = val.replace(hour=0, minute=0, second=0, microsecond=0)
#         return _localize_to_utc(naive)
#     if isinstance(val, date) and not isinstance(val, datetime):
#         naive = datetime(val.year, val.month, val.day)
#         return _localize_to_utc(naive)
#     s = str(val).strip()
#     if s.lower() in ("", "nan", "none", "nat"):
#         return pd.NaT
#     for fmt in formats:
#         try:
#             tmp = pd.to_datetime(s, format=fmt, errors="coerce")
#             if pd.notna(tmp):
#                 naive = tmp.to_pydatetime().replace(
#                     hour=0, minute=0, second=0, microsecond=0
#                 )
#                 return _localize_to_utc(naive)
#         except Exception:
#             continue
#     return pd.NaT


# def _combine_xray_datetime(date_val, time_val) -> pd.Timestamp:
#     """
#     Merge X-Ray DATE (str '01-05-2026') + X-Ray TIME (datetime.time or str)
#     into a single UTC Timestamp.
#     """
#     if date_val is None or str(date_val).strip().lower() in ("", "nan", "none"):
#         return pd.NaT

#     date_str = str(date_val).strip()
#     date_fmts = ["%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"]
#     parsed_date = None
#     for fmt in date_fmts:
#         try:
#             tmp = pd.to_datetime(date_str, format=fmt, errors="coerce")
#             if pd.notna(tmp):
#                 parsed_date = tmp.to_pydatetime().date()
#                 break
#         except Exception:
#             continue
#     if parsed_date is None:
#         return pd.NaT

#     # Parse time
#     if isinstance(time_val, dt_time):
#         t = time_val
#     elif time_val is None or str(time_val).strip().lower() in ("", "nan", "none"):
#         t = dt_time(0, 0, 0)
#     else:
#         ts = str(time_val).strip()
#         for fmt in ["%H:%M:%S", "%H:%M"]:
#             try:
#                 tmp = datetime.strptime(ts, fmt)
#                 t = tmp.time()
#                 break
#             except Exception:
#                 t = dt_time(0, 0, 0)

#     naive_dt = datetime.combine(parsed_date, t)
#     return _localize_to_utc(naive_dt)


# def _parse_datetime_cell(val, formats: list) -> pd.Timestamp:
#     """
#     Handle cells that may be Python datetime objects (from openpyxl)
#     OR strings. Returns UTC pd.Timestamp.
#     """
#     if val is None:
#         return pd.NaT
#     if isinstance(val, datetime):
#         return _localize_to_utc(val)
#     if isinstance(val, date) and not isinstance(val, datetime):
#         return _localize_to_utc(datetime(val.year, val.month, val.day))
#     return _to_utc_str(val, formats)


# # ── Metadata ─────────────────────────────────────────────────────────────────

# def _extract_metadata(ws) -> dict:
#     """
#     Read metadata directly from openpyxl worksheet rows.
#     Handles two patterns:
#       - "FROM DATE :" in one cell, datetime object in the next cell
#       - "FROM DATE :06MAY2026" all embedded in one string cell
#     """
#     metadata = {}
#     for row in ws.iter_rows(max_row=12, values_only=True):
#         # Work with all cells (including None gaps) to preserve positional next-cell logic
#         row_all = list(row)
#         row_vals = [v for v in row_all if v is not None]
#         for i, val in enumerate(row_vals):
#             s = str(val).strip().upper()
#             if "FROM DATE" in s:
#                 # Pattern A: next non-None cell is the value (datetime or string)
#                 if i + 1 < len(row_vals):
#                     nv = row_vals[i + 1]
#                     # Only use next cell if it looks like a date value, not another label
#                     if isinstance(nv, datetime) or (
#                         isinstance(nv, str) and "DATE" not in nv.upper()
#                     ):
#                         metadata["from_date"] = nv
#                 # Pattern B: value embedded in same string
#                 if "from_date" not in metadata:
#                     m = re.search(r"FROM\s*DATE\s*[:\s]+([\w/\-]+)", s)
#                     if m:
#                         metadata["from_date"] = m.group(1).strip(": ")
#             if "TO DATE" in s:
#                 if i + 1 < len(row_vals):
#                     nv = row_vals[i + 1]
#                     if isinstance(nv, datetime) or (
#                         isinstance(nv, str) and "DATE" not in nv.upper()
#                     ):
#                         metadata["to_date"] = nv
#                 if "to_date" not in metadata:
#                     m = re.search(r"TO\s*DATE\s*[:\s]+([\w/\-]+)", s)
#                     if m:
#                         metadata["to_date"] = m.group(1).strip(": ")
#             if "CARRIER" in s and "REPORT" not in s:
#                 if i + 1 < len(row_vals):
#                     metadata["carrier"] = str(row_vals[i + 1]).strip()
#     return metadata


# def _parse_meta_date(val) -> pd.Timestamp:
#     if isinstance(val, datetime):
#         return pd.Timestamp(val)
#     try:
#         return pd.to_datetime(str(val), errors="raise")
#     except Exception:
#         return pd.NaT


# def validate_same_month(metadata: dict):
#     from_dt = _parse_meta_date(metadata.get("from_date"))
#     to_dt   = _parse_meta_date(metadata.get("to_date"))
#     if pd.isna(from_dt) or pd.isna(to_dt):
#         raise ValueError("FROM DATE or TO DATE missing in report metadata.")
#     if from_dt.year != to_dt.year or from_dt.month != to_dt.month:
#         raise ValueError(
#             f"Invalid report period: FROM DATE ({from_dt.date()}) and "
#             f"TO DATE ({to_dt.date()}) must belong to same month."
#         )


# def _parse_month_uploaded(metadata: dict) -> str:
#     from_dt = _parse_meta_date(metadata.get("from_date"))
#     if pd.isna(from_dt):
#         return pd.Timestamp.now().strftime("%Y-%m")
#     return from_dt.strftime("%Y-%m")


# def _find_data_section(ws) -> tuple[int, int]:
#     """
#     Find the row index of the main header row ('AWB No' column present)
#     and the last data row index (stop at blank or second section header).
#     Returns (header_row_idx, last_data_row_idx) — 0-based.
#     """
#     all_rows = list(ws.iter_rows(values_only=True))
#     header_idx = None
#     for i, row in enumerate(all_rows):
#         vals = [str(v).strip().upper() for v in row if v is not None]
#         if "AWB NO" in vals and "BILLING SHC" in vals:
#             header_idx = i
#             break
#     if header_idx is None:
#         raise ValueError("Could not find header row with 'AWB No' and 'Billing SHC'.")

#     # Find end: stop at first completely blank row or a second section marker
#     last_data_idx = header_idx
#     for i in range(header_idx + 1, len(all_rows)):
#         row = all_rows[i]
#         non_none = [v for v in row if v is not None]
#         if not non_none:
#             # Blank row → data ends here
#             last_data_idx = i - 1
#             break
#         # Second section marker (e.g. "EXPORT TRANSFERED SHIPMENTS")
#         first_str = str(non_none[0]).strip().upper()
#         if "EXPORT TRANSFER" in first_str or "EXPORT TRANSHIP" in first_str:
#             last_data_idx = i - 1
#             break
#         last_data_idx = i

#     return header_idx, last_data_idx, all_rows


# # ── Main cleaner ─────────────────────────────────────────────────────────────

# def clean_export_transhipment_report(
#     file_bytes: BytesIO,
#     file_type: str = "excel",
# ) -> tuple[pd.DataFrame, dict]:
#     """
#     Parse and clean the Export Transhipment Report (.xlsx or .csv).
#     Returns (df, metadata) with all Billing SHC types included.
#     All datetimes are UTC. X-Ray DATE + TIME are merged into one column.
#     """

#     # ── 1. Load workbook ──────────────────────────────────────────────────────
#     if file_type == "excel":
#         import openpyxl
#         wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
#         ws = wb.active
#     else:
#         raise ValueError("Only 'excel' file_type is supported for this report.")

#     # ── 2. Extract + validate metadata ───────────────────────────────────────
#     metadata = _extract_metadata(ws)
#     validate_same_month(metadata)
#     month_uploaded = _parse_month_uploaded(metadata)
#     print(f"[meta] from={metadata.get('from_date')}  to={metadata.get('to_date')}  month={month_uploaded}")

#     # ── 3. Find header + data section ────────────────────────────────────────
#     header_idx, last_data_idx, all_rows = _find_data_section(ws)
#     print(f"[header] found at row index {header_idx}, data rows {header_idx+1}..{last_data_idx}")

#     # ── 4. Build DataFrame from raw rows ─────────────────────────────────────
#     header_vals = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
#     data_rows   = all_rows[header_idx + 1 : last_data_idx + 1]

#     df = pd.DataFrame(data_rows, columns=header_vals)

#     # ── 5. Drop junk columns (empty name, pure integer, UNNAMED) ─────────────
#     def _is_valid_col(name: str) -> bool:
#         n = name.strip()
#         if not n or n.lower() == "nan":
#             return False
#         if n.upper().startswith("UNNAMED"):
#             return False
#         try:
#             int(n)
#             return False
#         except ValueError:
#             return True

#     df = df.loc[:, [c for c in df.columns if _is_valid_col(c)]].copy()

#     # ── 6. Normalise column names ─────────────────────────────────────────────
#     rename_map = {
#         " DOC DATE & TIME." : "DOC DATE & TIME",
#         "DOC DATE & TIME."  : "DOC DATE & TIME",
#         "TRM  DATE"         : "TRM DATE",
#         " SL No"            : "SL No",
#     }
#     df = df.rename(columns=rename_map)

#     # ── 7. Drop rows with no SL No / AWB (footer/blank rows) ─────────────────
#     df = df[df["AWB No"].apply(lambda x: x is not None and str(x).strip().lower() not in ("", "nan", "none"))]
#     df = df.copy()

#     # ── 8. Validate required columns ─────────────────────────────────────────
#     missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
#     if missing:
#         raise ValueError(f"Missing required columns: {missing}")

#     df = df[REQUIRED_COLUMNS].copy()

#     # ── 9. Filter: Billing SHC must be TRM or TPV ───────────────────────────
#     df["Billing SHC"] = df["Billing SHC"].apply(
#         lambda x: str(x).strip().upper() if x is not None else None
#     )
#     df = df[df["Billing SHC"].isin(BILLING_SHC_FILTER)].copy()
#     print(f"[filter] rows after Billing SHC filter (TRM/TPV): {len(df)}")

#     if df.empty:
#         return df, metadata

#     # ── 18. Strip whitespace strings ──────────────────────────────────────────
#     for col in df.columns:
#         df[col] = df[col].apply(
#             lambda x: str(x).strip() if isinstance(x, str) else x
#         )

#     # ── 10. Clean AWB ─────────────────────────────────────────────────────────
#     df["AWB No"] = df["AWB No"].apply(_clean_awb)

#     # ── 11. Cast numeric columns ──────────────────────────────────────────────
#     for col in NUMERIC_COLS:
#         df[col] = pd.to_numeric(df[col], errors="coerce")

#     # ── 12. Parse datetime columns ────────────────────────────────────────────
#     # Common string formats seen in this file
#     dt_fmts = [
#         "%d-%m-%Y %H:%M",
#         "%d-%m-%Y %H:%M:%S",
#         "%d/%m/%Y %H:%M",
#         "%d/%m/%Y %H:%M:%S",
#         "%d-%b-%Y %H:%M",
#         "%d-%m-%y %H:%M",
#         "%Y-%m-%d %H:%M:%S",
#     ]
#     date_fmts = [
#         "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y",
#         "%d-%m-%y", "%d/%m/%y",
#     ]

#     for col in DATETIME_COLS:
#         df[col] = df[col].apply(lambda x: _parse_datetime_cell(x, dt_fmts))
#         df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

#     # Flight Date (date-only → midnight IST → UTC)
#     df["FLIGHT DATE"] = df["FLIGHT DATE"].apply(
#         lambda x: _parse_date_to_utc(x, date_fmts)
#     )
#     df["FLIGHT DATE"] = pd.to_datetime(df["FLIGHT DATE"], errors="coerce", utc=True)

#     # X-Ray DATE + X-Ray TIME → combined UTC datetime → new col, drop originals
#     df["XRAY DATETIME"] = df.apply(
#         lambda r: _combine_xray_datetime(r["X-Ray DATE"], r["X-Ray TIME"]),
#         axis=1,
#     )
#     df["XRAY DATETIME"] = pd.to_datetime(df["XRAY DATETIME"], errors="coerce", utc=True)
#     df = df.drop(columns=["X-Ray DATE", "X-Ray TIME"])

#     # EXP TP SEG No DATE AND TIME is a mixed string like "/30APR2026 16:13"
#     # or "15651/04MAY2026 19:37" — strip the prefix number/slash and parse
#     def _parse_exp_tp_seg(val):
#         if not val or str(val).strip().lower() in ("", "nan", "none"):
#             return pd.NaT
#         s = str(val).strip()
#         # Remove leading digits and slash (e.g. "15618/" or "/")
#         s = re.sub(r"^\d*/", "", s).strip()
#         seg_fmts = ["%d%b%Y %H:%M", "%d%b%Y %H:%M:%S"]
#         return _to_utc_str(s, seg_fmts)

#     df["EXP TP SEG No DATE AND TIME"] = df["EXP TP SEG No DATE AND TIME"].apply(_parse_exp_tp_seg)
#     df["EXP TP SEG No DATE AND TIME"] = pd.to_datetime(
#         df["EXP TP SEG No DATE AND TIME"], errors="coerce", utc=True
#     )

#     # ── 13. Standardise text columns ─────────────────────────────────────────
#     all_dt_cols = DATETIME_COLS + ["FLIGHT DATE", "XRAY DATETIME", "EXP TP SEG No DATE AND TIME"]
#     text_cols = [c for c in df.columns if c not in NUMERIC_COLS + all_dt_cols + ["AWB No"]]
#     for col in text_cols:
#         s = df[col].apply(lambda x: str(x).strip() if x is not None else "")
#         df[col] = s.where(s.str.lower() != "nan", other=None)

#     # ── 14. Drop rows with no valid AWB ──────────────────────────────────────
#     df = df.dropna(subset=["AWB No"])

#     # ── 15. Drop SL No (not stored) ──────────────────────────────────────────
#     df = df.drop(columns=["SL No"], errors="ignore")

#     # ── 16. Attach metadata ───────────────────────────────────────────────────
#     df["month_uploaded"] = month_uploaded
#     df["uploaded_at"]    = datetime.now(tz=timezone.utc)

#     # ── 17. Final dtype enforcement + cleanup ─────────────────────────────────
#     final_dt_cols = all_dt_cols + ["uploaded_at"]
#     for col in final_dt_cols:
#         if col in df.columns:
#             df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

#     df = df.where(df.notna(), other=None)
#     df = df.reset_index(drop=True)

#     return df, metadata


# # ── Quick inspection ──────────────────────────────────────────────────────────

# if __name__ == "__main__":
#     import sys, os

#     path = sys.argv[1] if len(sys.argv) > 1 else None

#     if path is None:
#         script_dir = os.path.dirname(os.path.abspath(__file__))
#         for fname in os.listdir(script_dir):
#             if fname.upper().endswith(".XLSX"):
#                 path = os.path.join(script_dir, fname)
#                 break

#     if path is None:
#         print("Usage: python export_transhipment_cleaning.py <path/to/file.xlsx>")
#         sys.exit(1)

#     print(f"[input] {path}")
#     with open(path, "rb") as f:
#         raw = f.read()

#     df, meta = clean_export_transhipment_report(BytesIO(raw), file_type="excel")

#     pd.set_option("display.max_columns", None)
#     pd.set_option("display.width", 240)
#     pd.set_option("display.max_colwidth", 28)

#     print(f"\n{'='*60}")
#     print(f"Metadata  : {meta}")
#     print(f"Rows      : {len(df)}")
#     print(f"Columns   : {list(df.columns)}")

#     print(f"\n--- dtypes ---")
#     print(df.dtypes.to_string())

#     print(f"\n--- Sample rows (first 5) ---")
#     print(df.head(5).to_string())

#     print(f"\n--- Billing SHC value counts ---")
#     print(df["Billing SHC"].value_counts())

#     print(f"\n--- FLIGHT DATE samples ---")
#     print(df["FLIGHT DATE"].dropna().head(5).to_string())

#     print(f"\n--- DOC DATE & TIME samples ---")
#     print(df["DOC DATE & TIME"].dropna().head(5).to_string())

#     print(f"\n--- XRAY DATETIME samples ---")
#     print(df["XRAY DATETIME"].dropna().head(5).to_string())

#     print(f"\n--- EXP TP SEG No DATE AND TIME samples ---")
#     print(df["EXP TP SEG No DATE AND TIME"].dropna().head(5).to_string())

#     print(f"\n--- DEPARTURE DATE & TIME samples ---")
#     print(df["DEPARTURE DATE & TIME"].dropna().head(5).to_string())

#     print(f"\n--- AWB No samples ---")
#     print(df["AWB No"].dropna().head(10).tolist())

#     print(f"\n--- Null counts ---")
#     print(df.isnull().sum().to_string())



































import re
import pytz
import numpy as np
import pandas as pd
from io import BytesIO
from datetime import datetime, timezone, date, time as dt_time

# ── Constants ────────────────────────────────────────────────────────────────

BILLING_SHC_FILTER = {"TRM", "TPV"}

REQUIRED_COLUMNS = [
    "AWB No",
    "PCS",
    "Gross wgt",
    "Rec_PCS",
    "Received wgt",
    "Received_Chg_Wgt",
    "SHC",
    "Billing SHC",
    "Commodity",
    "ORG",
    "DES",
    "DOC DATE & TIME",           # normalised from ' DOC DATE & TIME.'
    "EXP TP SEG FLIGHT No.",
    "EXP TP FLIGHT DATE",
    "EXP TP SEG No DATE AND TIME",
    "TRM NO",
    "TRM DATE",                  # normalised from 'TRM  DATE'
    "X-Ray DATE",
    "X-Ray TIME",
    "RAMP TRANSFER DATE/TIME",
    "RAMP TRANSFER REMARK",
    "RAMP TRANSFER USER",
    "AIRLINE CD",
    "FLIGHT NO",
    "FLIGHT DATE",
    "ULD LOAD",
    "DEPARTURE DATE & TIME",
]

NUMERIC_COLS = [
    "PCS", "Gross wgt", "Rec_PCS",
    "Received wgt", "Received_Chg_Wgt",
    "TRM NO",
]

# Columns that need IST → UTC datetime parsing
DATETIME_COLS = [
    "DOC DATE & TIME",
    "EXP TP FLIGHT DATE",
    "TRM DATE",
    "RAMP TRANSFER DATE/TIME",
    "DEPARTURE DATE & TIME",
    "ULD LOAD",
]

# X-Ray DATE and X-Ray TIME are merged into one UTC datetime column
# FLIGHT DATE and EXP TP FLIGHT DATE are date-only → midnight IST → UTC

_IST = pytz.timezone("Asia/Kolkata")


# ── Helpers ──────────────────────────────────────────────────────────────────

def _clean_awb(val) -> str | None:
    if not val or str(val).strip().lower() in ("", "nan", "none"):
        return None
    s = re.sub(r"\s+", "", str(val).strip())
    s = re.sub(r"-+", "-", s).replace("-", "")
    s = re.sub(r"\D", "", s)
    if s and len(s) == 10:
        s = s.zfill(11)
    return s or None


def _localize_to_utc(dt_naive) -> pd.Timestamp:
    """Take a naive Python datetime (or pd.Timestamp) → IST → UTC pd.Timestamp."""
    if dt_naive is None or (isinstance(dt_naive, float) and pd.isna(dt_naive)):
        return pd.NaT
    try:
        if pd.isnull(dt_naive):
            return pd.NaT
    except (TypeError, ValueError):
        pass
    # pd.Timestamp.localize() is the safe path (avoids pytz Naive time error)
    if isinstance(dt_naive, pd.Timestamp):
        if dt_naive.tzinfo is not None:
            return dt_naive.tz_convert("UTC")
        return dt_naive.tz_localize("Asia/Kolkata").tz_convert("UTC")
    # Plain datetime
    local_dt = _IST.localize(dt_naive.replace(tzinfo=None), is_dst=None)
    return pd.Timestamp(local_dt.astimezone(pytz.utc))


def _to_utc_str(val, formats: list) -> pd.Timestamp:
    """Parse string datetime (assumed IST) → UTC pd.Timestamp."""
    if not val or str(val).strip().lower() in ("", "nan", "none", "nat"):
        return pd.NaT
    s = str(val).strip()
    for fmt in formats:
        try:
            tmp = pd.to_datetime(s, format=fmt, errors="coerce")
            if pd.notna(tmp):
                return _localize_to_utc(tmp.to_pydatetime())
        except Exception:
            continue
    try:
        tmp = pd.to_datetime(s, errors="coerce", dayfirst=True)
        if pd.notna(tmp):
            return _localize_to_utc(tmp.to_pydatetime())
    except Exception:
        pass
    return pd.NaT


def _parse_date_to_utc(val, formats: list) -> pd.Timestamp:
    """Parse date-only value → midnight IST → UTC."""
    if val is None:
        return pd.NaT
    # Already a Python datetime from openpyxl
    if isinstance(val, datetime):
        naive = val.replace(hour=0, minute=0, second=0, microsecond=0)
        return _localize_to_utc(naive)
    if isinstance(val, date) and not isinstance(val, datetime):
        naive = datetime(val.year, val.month, val.day)
        return _localize_to_utc(naive)
    s = str(val).strip()
    if s.lower() in ("", "nan", "none", "nat"):
        return pd.NaT
    for fmt in formats:
        try:
            tmp = pd.to_datetime(s, format=fmt, errors="coerce")
            if pd.notna(tmp):
                naive = tmp.to_pydatetime().replace(
                    hour=0, minute=0, second=0, microsecond=0
                )
                return _localize_to_utc(naive)
        except Exception:
            continue
    return pd.NaT


def _combine_xray_datetime(date_val, time_val) -> pd.Timestamp:
    """
    Merge X-Ray DATE (str '01-05-2026') + X-Ray TIME (datetime.time or str)
    into a single UTC Timestamp.
    """
    if date_val is None or str(date_val).strip().lower() in ("", "nan", "none"):
        return pd.NaT

    date_str = str(date_val).strip()
    date_fmts = ["%d-%m-%Y", "%d-%m-%y", "%d/%m/%Y", "%d/%m/%y"]
    parsed_date = None
    for fmt in date_fmts:
        try:
            tmp = pd.to_datetime(date_str, format=fmt, errors="coerce")
            if pd.notna(tmp):
                parsed_date = tmp.to_pydatetime().date()
                break
        except Exception:
            continue
    if parsed_date is None:
        return pd.NaT

    # Parse time
    if isinstance(time_val, dt_time):
        t = time_val
    elif time_val is None or str(time_val).strip().lower() in ("", "nan", "none"):
        t = dt_time(0, 0, 0)
    else:
        ts = str(time_val).strip()
        for fmt in ["%H:%M:%S", "%H:%M"]:
            try:
                tmp = datetime.strptime(ts, fmt)
                t = tmp.time()
                break
            except Exception:
                t = dt_time(0, 0, 0)

    naive_dt = datetime.combine(parsed_date, t)
    return _localize_to_utc(naive_dt)


def _parse_datetime_cell(val, formats: list) -> pd.Timestamp:
    """
    Handle cells that may be Python datetime objects (from openpyxl)
    OR strings. Returns UTC pd.Timestamp.
    """
    if val is None:
        return pd.NaT
    if isinstance(val, datetime):
        return _localize_to_utc(val)
    if isinstance(val, date) and not isinstance(val, datetime):
        return _localize_to_utc(datetime(val.year, val.month, val.day))
    return _to_utc_str(val, formats)


# ── Metadata ─────────────────────────────────────────────────────────────────

def _extract_metadata(ws) -> dict:
    """
    Read metadata directly from openpyxl worksheet rows.
    Handles two patterns:
      - "FROM DATE :" in one cell, datetime object in the next cell
      - "FROM DATE :06MAY2026" all embedded in one string cell
    """
    metadata = {}
    for row in ws.iter_rows(max_row=12, values_only=True):
        # Work with all cells (including None gaps) to preserve positional next-cell logic
        row_all = list(row)
        row_vals = [v for v in row_all if v is not None]
        for i, val in enumerate(row_vals):
            s = str(val).strip().upper()
            if "FROM DATE" in s:
                # Pattern A: next non-None cell is the value (datetime or string)
                if i + 1 < len(row_vals):
                    nv = row_vals[i + 1]
                    # Only use next cell if it looks like a date value, not another label
                    if isinstance(nv, datetime) or (
                        isinstance(nv, str) and "DATE" not in nv.upper()
                    ):
                        metadata["from_date"] = nv
                # Pattern B: value embedded in same string
                if "from_date" not in metadata:
                    m = re.search(r"FROM\s*DATE\s*[:\s]+([\w/\-]+)", s)
                    if m:
                        metadata["from_date"] = m.group(1).strip(": ")
            if "TO DATE" in s:
                if i + 1 < len(row_vals):
                    nv = row_vals[i + 1]
                    if isinstance(nv, datetime) or (
                        isinstance(nv, str) and "DATE" not in nv.upper()
                    ):
                        metadata["to_date"] = nv
                if "to_date" not in metadata:
                    m = re.search(r"TO\s*DATE\s*[:\s]+([\w/\-]+)", s)
                    if m:
                        metadata["to_date"] = m.group(1).strip(": ")
            if "CARRIER" in s and "REPORT" not in s:
                if i + 1 < len(row_vals):
                    metadata["carrier"] = str(row_vals[i + 1]).strip()
    return metadata


def _parse_meta_date(val) -> pd.Timestamp:
    if isinstance(val, datetime):
        return pd.Timestamp(val)
    try:
        return pd.to_datetime(str(val), errors="raise")
    except Exception:
        return pd.NaT


def validate_same_month(metadata: dict):
    from_dt = _parse_meta_date(metadata.get("from_date"))
    to_dt   = _parse_meta_date(metadata.get("to_date"))
    if pd.isna(from_dt) or pd.isna(to_dt):
        raise ValueError("FROM DATE or TO DATE missing in report metadata.")
    if from_dt.year != to_dt.year or from_dt.month != to_dt.month:
        raise ValueError(
            f"Invalid report period: FROM DATE ({from_dt.date()}) and "
            f"TO DATE ({to_dt.date()}) must belong to same month."
        )


def _parse_month_uploaded(metadata: dict) -> str:
    from_dt = _parse_meta_date(metadata.get("from_date"))
    if pd.isna(from_dt):
        return pd.Timestamp.now().strftime("%Y-%m")
    return from_dt.strftime("%Y-%m")


def _find_data_section(ws) -> tuple[int, int]:
    """
    Find the row index of the main header row ('AWB No' column present)
    and the last data row index (stop at blank or second section header).
    Returns (header_row_idx, last_data_row_idx) — 0-based.
    """
    all_rows = list(ws.iter_rows(values_only=True))
    header_idx = None
    for i, row in enumerate(all_rows):
        vals = [str(v).strip().upper() for v in row if v is not None]
        if "AWB NO" in vals and "BILLING SHC" in vals:
            header_idx = i
            break
    if header_idx is None:
        raise ValueError("Could not find header row with 'AWB No' and 'Billing SHC'.")

    # Find end: stop at first completely blank row or a second section marker
    last_data_idx = header_idx
    for i in range(header_idx + 1, len(all_rows)):
        row = all_rows[i]
        non_none = [v for v in row if v is not None]
        if not non_none:
            # Blank row → data ends here
            last_data_idx = i - 1
            break
        # Second section marker (e.g. "EXPORT TRANSFERED SHIPMENTS")
        first_str = str(non_none[0]).strip().upper()
        if "EXPORT TRANSFER" in first_str or "EXPORT TRANSHIP" in first_str:
            last_data_idx = i - 1
            break
        last_data_idx = i

    return header_idx, last_data_idx, all_rows



# ── CSV-specific helpers ──────────────────────────────────────────────────────

def _read_transhipment_csv(file_bytes: BytesIO) -> pd.DataFrame:
    """Read CSV with variable column counts into uniform DataFrame."""
    raw_text = file_bytes.read().decode("utf-8", errors="replace")
    file_bytes.seek(0)
    lines    = raw_text.splitlines()
    max_cols = max((line.count(",") for line in lines), default=0) + 1
    return pd.read_csv(
        file_bytes, header=None, names=range(max_cols),
        dtype=str, keep_default_na=False,
        engine="c", on_bad_lines="skip",
    )


def _extract_metadata_csv(df_raw: pd.DataFrame) -> dict:
    """
    Extract FROM DATE / TO DATE from CSV header rows (first 12 only).
    Format seen: "FROM DATE :,01-Apr-26,,TO DATE :,30-Apr-26"
    """
    metadata = {}
    for _, row in df_raw.head(12).iterrows():
        row_vals = row.tolist()
        row_str  = " ".join(str(v).strip() for v in row_vals if str(v).strip()).upper()
        if "FROM DATE" not in row_str and "TO DATE" not in row_str:
            continue
        for i, val in enumerate(row_vals):
            s = str(val).strip().upper()
            # Embedded: "FROM DATE :01APR2026"
            m = re.search(r"FROM\s*DATE\s*[:\s]*([\w\-/]+)", s)
            if m and m.group(1):
                metadata["from_date"] = m.group(1).strip(": ")
            # Next cell: "FROM DATE :" then value in next cell
            if "FROM DATE" in s and "from_date" not in metadata:
                for j in range(i + 1, min(i + 4, len(row_vals))):
                    nv = str(row_vals[j]).strip()
                    if nv and nv.upper() not in ("", "NAN"):
                        metadata["from_date"] = nv
                        break
            m = re.search(r"TO\s*DATE\s*[:\s]*([\w\-/]+)", s)
            if m and m.group(1):
                metadata["to_date"] = m.group(1).strip(": ")
            if "TO DATE" in s and "to_date" not in metadata:
                for j in range(i + 1, min(i + 4, len(row_vals))):
                    nv = str(row_vals[j]).strip()
                    if nv and nv.upper() not in ("", "NAN") and "DATE" not in nv.upper():
                        metadata["to_date"] = nv
                        break
    return metadata


def _find_header_row_csv(df_raw: pd.DataFrame) -> int:
    """Find the header row index by locating 'AWB No' and 'Billing SHC'."""
    for i, row in df_raw.iterrows():
        vals = [str(v).strip().upper() for v in row.tolist() if str(v).strip()]
        if "AWB NO" in vals and "BILLING SHC" in vals:
            return i
    raise ValueError("Could not find header row with 'AWB No' and 'Billing SHC'.")


def _clean_transhipment_csv(file_bytes: BytesIO) -> tuple[pd.DataFrame, dict]:
    """
    Full cleaning pipeline for CSV variant of the Export Transhipment Report.
    Differences vs Excel:
      - metadata extracted from raw text rows (no openpyxl datetime objects)
      - X-Ray TIME is a plain string like "7:49", not datetime.time
      - All datetime columns are strings — same _parse_datetime_cell handles them
      - Second section ("EXPORT TRANSFERED SHIPMENTS") detected by string match
    """
    # 1. Read raw
    df_raw = _read_transhipment_csv(file_bytes)

    # 2. Metadata + validate
    metadata = _extract_metadata_csv(df_raw)
    validate_same_month(metadata)
    month_uploaded = _parse_month_uploaded(metadata)
    print(f"[meta] from={metadata.get('from_date')}  to={metadata.get('to_date')}  month={month_uploaded}")

    # 3. Find header row
    header_idx = _find_header_row_csv(df_raw)
    print(f"[header] found at row index {header_idx}")

    # 4. Build df from slice
    header_vals = [str(v).strip() for v in df_raw.iloc[header_idx].tolist()]
    data        = df_raw.iloc[header_idx + 1:].copy()
    data.columns = header_vals
    df = data.reset_index(drop=True)

    # 5. Drop junk columns
    def _is_valid_col(name: str) -> bool:
        n = name.strip()
        if not n or n.lower() == "nan" or n.upper().startswith("UNNAMED"):
            return False
        try: int(n); return False
        except ValueError: return True

    df = df.loc[:, [c for c in df.columns if _is_valid_col(c)]].copy()

    # 6. Normalise column names
    rename_map = {
        " DOC DATE & TIME." : "DOC DATE & TIME",
        "DOC DATE & TIME."  : "DOC DATE & TIME",
        "TRM  DATE"         : "TRM DATE",
        " SL No"            : "SL No",
    }
    df = df.rename(columns=rename_map)

    # 7. Drop second section ("EXPORT TRANSFERED SHIPMENTS") and below
    first_col = df.iloc[:, 0].astype(str).str.strip().str.upper()
    second_col = df.iloc[:, 1].astype(str).str.strip().str.upper() if df.shape[1] > 1 else first_col
    # Check first few cols for section marker
    section_mask = df.apply(
        lambda r: any(
            "EXPORT TRANSFER" in str(v).upper() or "EXPORT TRANSHIP" in str(v).upper()
            for v in r.tolist()[:4]
        ),
        axis=1,
    )
    if section_mask.any():
        df = df.iloc[: section_mask.values.argmax()].copy()

    # Drop subtotal rows: col index 2 == "Total"
    col2 = df.iloc[:, 2].astype(str).str.strip().str.upper()
    df = df[col2 != "TOTAL"].copy()

    # Drop grand total
    gc = df.iloc[:, 0].fillna("").astype(str).str.upper()
    df = df[~gc.str.contains("GRAND TOTAL", na=False)].copy()

    # 8. Basic cleanup
    df = df.replace(r"^\s*$", np.nan, regex=True).dropna(how="all")

    # 9. Keep rows where AWB No is non-empty
    if "AWB No" in df.columns:
        df = df[df["AWB No"].astype(str).str.strip().str.len() > 3]

    # 10. Validate required columns
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    df = df[REQUIRED_COLUMNS].copy()

    # 11. Filter Billing SHC
    df["Billing SHC"] = df["Billing SHC"].astype(str).str.strip().str.upper()
    df = df[df["Billing SHC"].isin(BILLING_SHC_FILTER)].copy()
    print(f"[filter] rows after Billing SHC filter (TRM/TPV): {len(df)}")

    if df.empty:
        return df, metadata

    # 12. Strip whitespace
    for col in df.columns:
        df[col] = df[col].apply(lambda x: str(x).strip() if isinstance(x, str) else x)

    # 13. Clean AWB
    df["AWB No"] = df["AWB No"].apply(_clean_awb)

    # 14. Cast numeric
    for col in NUMERIC_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # 15. Parse datetime columns (all strings in CSV)
    dt_fmts = [
        "%Y-%m-%d %H:%M",       # 2026-04-21 6:39  ← most common in this CSV
        "%Y-%m-%d %H:%M:%S",
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%b-%Y %H:%M",
        "%d-%m-%y %H:%M",
    ]
    date_fmts = [
        "%d-%b-%y",    # 20-Apr-26  ← EXP TP FLIGHT DATE format in CSV
        "%d-%b-%Y",
        "%d-%m-%Y", "%d/%m/%Y",
        "%d-%m-%y", "%d/%m/%y",
        "%Y-%m-%d",
    ]

    for col in DATETIME_COLS:
        df[col] = df[col].apply(lambda x: _parse_datetime_cell(x, dt_fmts))
        df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

    # Flight Date (date-only → midnight IST → UTC)
    df["FLIGHT DATE"] = df["FLIGHT DATE"].apply(
        lambda x: _parse_date_to_utc(x, date_fmts)
    )
    df["FLIGHT DATE"] = pd.to_datetime(df["FLIGHT DATE"], errors="coerce", utc=True)

    # X-Ray: DATE is string "21-04-2026", TIME is string "7:49"
    # _combine_xray_datetime handles both string time and datetime.time
    df["XRAY DATETIME"] = df.apply(
        lambda r: _combine_xray_datetime(r["X-Ray DATE"], r["X-Ray TIME"]),
        axis=1,
    )
    df["XRAY DATETIME"] = pd.to_datetime(df["XRAY DATETIME"], errors="coerce", utc=True)
    df = df.drop(columns=["X-Ray DATE", "X-Ray TIME"])

    # EXP TP SEG No DATE AND TIME — same mixed format as Excel
    def _parse_exp_tp_seg(val):
        if not val or str(val).strip().lower() in ("", "nan", "none"):
            return pd.NaT
        s = re.sub(r"^\d*/", "", str(val).strip()).strip()
        return _to_utc_str(s, ["%d%b%Y %H:%M", "%d%b%Y %H:%M:%S"])

    df["EXP TP SEG No DATE AND TIME"] = df["EXP TP SEG No DATE AND TIME"].apply(_parse_exp_tp_seg)
    df["EXP TP SEG No DATE AND TIME"] = pd.to_datetime(
        df["EXP TP SEG No DATE AND TIME"], errors="coerce", utc=True
    )

    # 16. Standardise text columns
    all_dt_cols = DATETIME_COLS + ["FLIGHT DATE", "XRAY DATETIME", "EXP TP SEG No DATE AND TIME"]
    text_cols = [c for c in df.columns if c not in NUMERIC_COLS + all_dt_cols + ["AWB No"]]
    for col in text_cols:
        s = df[col].apply(lambda x: str(x).strip() if x is not None else "")
        df[col] = s.where(s.str.lower() != "nan", other=None)

    # 17. Drop rows with no valid AWB
    df = df.dropna(subset=["AWB No"])

    # 18. Drop SL No
    df = df.drop(columns=["SL No"], errors="ignore")

    # 19. Attach metadata
    df["month_uploaded"] = month_uploaded
    df["uploaded_at"]    = datetime.now(tz=timezone.utc)

    # 20. Final dtype enforcement
    final_dt_cols = all_dt_cols + ["uploaded_at"]
    for col in final_dt_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

    df = df.where(df.notna(), other=None)
    df = df.reset_index(drop=True)

    return df, metadata

# ── Main cleaner ─────────────────────────────────────────────────────────────

def clean_export_transhipment_report(
    file_bytes: BytesIO,
    file_type: str = "excel",
) -> tuple[pd.DataFrame, dict]:
    """
    Parse and clean the Export Transhipment Report (.xlsx or .csv).
    Returns (df, metadata) with all Billing SHC types included.
    All datetimes are UTC. X-Ray DATE + TIME are merged into one column.
    """

    # ── 1. Load file ─────────────────────────────────────────────────────────
    if file_type == "excel":
        import openpyxl
        wb = openpyxl.load_workbook(file_bytes, read_only=True, data_only=True)
        ws = wb.active
    elif file_type == "csv":
        return _clean_transhipment_csv(file_bytes)
    else:
        raise ValueError("Unsupported file_type. Use 'excel' or 'csv'.")

    # ── 2. Extract + validate metadata ───────────────────────────────────────
    metadata = _extract_metadata(ws)
    validate_same_month(metadata)
    month_uploaded = _parse_month_uploaded(metadata)
    print(f"[meta] from={metadata.get('from_date')}  to={metadata.get('to_date')}  month={month_uploaded}")

    # ── 3. Find header + data section ────────────────────────────────────────
    header_idx, last_data_idx, all_rows = _find_data_section(ws)
    print(f"[header] found at row index {header_idx}, data rows {header_idx+1}..{last_data_idx}")

    # ── 4. Build DataFrame from raw rows ─────────────────────────────────────
    header_vals = [str(v).strip() if v is not None else "" for v in all_rows[header_idx]]
    data_rows   = all_rows[header_idx + 1 : last_data_idx + 1]

    df = pd.DataFrame(data_rows, columns=header_vals)

    # ── 5. Drop junk columns (empty name, pure integer, UNNAMED) ─────────────
    def _is_valid_col(name: str) -> bool:
        n = name.strip()
        if not n or n.lower() == "nan":
            return False
        if n.upper().startswith("UNNAMED"):
            return False
        try:
            int(n)
            return False
        except ValueError:
            return True

    df = df.loc[:, [c for c in df.columns if _is_valid_col(c)]].copy()

    # ── 6. Normalise column names ─────────────────────────────────────────────
    rename_map = {
        " DOC DATE & TIME." : "DOC DATE & TIME",
        "DOC DATE & TIME."  : "DOC DATE & TIME",
        "TRM  DATE"         : "TRM DATE",
        " SL No"            : "SL No",
    }
    df = df.rename(columns=rename_map)

    # ── 7. Drop rows with no SL No / AWB (footer/blank rows) ─────────────────
    df = df[df["AWB No"].apply(lambda x: x is not None and str(x).strip().lower() not in ("", "nan", "none"))]
    df = df.copy()

    # ── 8. Validate required columns ─────────────────────────────────────────
    missing = [c for c in REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"Missing required columns: {missing}")

    df = df[REQUIRED_COLUMNS].copy()

    # ── 9. Filter: Billing SHC must be TRM or TPV ───────────────────────────
    df["Billing SHC"] = df["Billing SHC"].apply(
        lambda x: str(x).strip().upper() if x is not None else None
    )
    df = df[df["Billing SHC"].isin(BILLING_SHC_FILTER)].copy()
    print(f"[filter] rows after Billing SHC filter (TRM/TPV): {len(df)}")

    if df.empty:
        return df, metadata

    # ── 18. Strip whitespace strings ──────────────────────────────────────────
    for col in df.columns:
        df[col] = df[col].apply(
            lambda x: str(x).strip() if isinstance(x, str) else x
        )

    # ── 10. Clean AWB ─────────────────────────────────────────────────────────
    df["AWB No"] = df["AWB No"].apply(_clean_awb)

    # ── 11. Cast numeric columns ──────────────────────────────────────────────
    for col in NUMERIC_COLS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    # ── 12. Parse datetime columns ────────────────────────────────────────────
    # Common string formats seen in this file
    dt_fmts = [
        "%d-%m-%Y %H:%M",
        "%d-%m-%Y %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
        "%d-%b-%Y %H:%M",
        "%d-%m-%y %H:%M",
        "%Y-%m-%d %H:%M:%S",
    ]
    date_fmts = [
        "%d-%m-%Y", "%d/%m/%Y", "%d-%b-%Y",
        "%d-%m-%y", "%d/%m/%y",
    ]

    for col in DATETIME_COLS:
        df[col] = df[col].apply(lambda x: _parse_datetime_cell(x, dt_fmts))
        df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

    # Flight Date (date-only → midnight IST → UTC)
    df["FLIGHT DATE"] = df["FLIGHT DATE"].apply(
        lambda x: _parse_date_to_utc(x, date_fmts)
    )
    df["FLIGHT DATE"] = pd.to_datetime(df["FLIGHT DATE"], errors="coerce", utc=True)

    # X-Ray DATE + X-Ray TIME → combined UTC datetime → new col, drop originals
    df["XRAY DATETIME"] = df.apply(
        lambda r: _combine_xray_datetime(r["X-Ray DATE"], r["X-Ray TIME"]),
        axis=1,
    )
    df["XRAY DATETIME"] = pd.to_datetime(df["XRAY DATETIME"], errors="coerce", utc=True)
    df = df.drop(columns=["X-Ray DATE", "X-Ray TIME"])

    # EXP TP SEG No DATE AND TIME is a mixed string like "/30APR2026 16:13"
    # or "15651/04MAY2026 19:37" — strip the prefix number/slash and parse
    def _parse_exp_tp_seg(val):
        if not val or str(val).strip().lower() in ("", "nan", "none"):
            return pd.NaT
        s = str(val).strip()
        # Remove leading digits and slash (e.g. "15618/" or "/")
        s = re.sub(r"^\d*/", "", s).strip()
        seg_fmts = ["%d%b%Y %H:%M", "%d%b%Y %H:%M:%S"]
        return _to_utc_str(s, seg_fmts)

    df["EXP TP SEG No DATE AND TIME"] = df["EXP TP SEG No DATE AND TIME"].apply(_parse_exp_tp_seg)
    df["EXP TP SEG No DATE AND TIME"] = pd.to_datetime(
        df["EXP TP SEG No DATE AND TIME"], errors="coerce", utc=True
    )

    # ── 13. Standardise text columns ─────────────────────────────────────────
    all_dt_cols = DATETIME_COLS + ["FLIGHT DATE", "XRAY DATETIME", "EXP TP SEG No DATE AND TIME"]
    text_cols = [c for c in df.columns if c not in NUMERIC_COLS + all_dt_cols + ["AWB No"]]
    for col in text_cols:
        s = df[col].apply(lambda x: str(x).strip() if x is not None else "")
        df[col] = s.where(s.str.lower() != "nan", other=None)

    # ── 14. Drop rows with no valid AWB ──────────────────────────────────────
    df = df.dropna(subset=["AWB No"])

    # ── 15. Drop SL No (not stored) ──────────────────────────────────────────
    df = df.drop(columns=["SL No"], errors="ignore")

    # ── 16. Attach metadata ───────────────────────────────────────────────────
    df["month_uploaded"] = month_uploaded
    df["uploaded_at"]    = datetime.now(tz=timezone.utc)

    # ── 17. Final dtype enforcement + cleanup ─────────────────────────────────
    final_dt_cols = all_dt_cols + ["uploaded_at"]
    for col in final_dt_cols:
        if col in df.columns:
            df[col] = pd.to_datetime(df[col], errors="coerce", utc=True)

    df = df.where(df.notna(), other=None)
    df = df.reset_index(drop=True)

    return df, metadata


# ── Quick inspection ──────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys, os

    path = sys.argv[1] if len(sys.argv) > 1 else None

    if path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        for fname in os.listdir(script_dir):
            if fname.upper().endswith(".XLSX"):
                path = os.path.join(script_dir, fname)
                break

    if path is None:
        print("Usage: python export_transhipment_cleaning.py <path/to/file.xlsx>")
        sys.exit(1)

    print(f"[input] {path}")
    with open(path, "rb") as f:
        raw = f.read()

    file_type = "csv" if path.lower().endswith(".csv") else "excel"
    df, meta = clean_export_transhipment_report(BytesIO(raw), file_type=file_type)

    pd.set_option("display.max_columns", None)
    pd.set_option("display.width", 240)
    pd.set_option("display.max_colwidth", 28)

    print(f"\n{'='*60}")
    print(f"Metadata  : {meta}")
    print(f"Rows      : {len(df)}")
    print(f"Columns   : {list(df.columns)}")

    print(f"\n--- dtypes ---")
    print(df.dtypes.to_string())

    print(f"\n--- Sample rows (first 5) ---")
    print(df.head(5).to_string())

    print(f"\n--- Billing SHC value counts ---")
    print(df["Billing SHC"].value_counts())

    print(f"\n--- FLIGHT DATE samples ---")
    print(df["FLIGHT DATE"].dropna().head(5).to_string())

    print(f"\n--- DOC DATE & TIME samples ---")
    print(df["DOC DATE & TIME"].dropna().head(5).to_string())

    print(f"\n--- XRAY DATETIME samples ---")
    print(df["XRAY DATETIME"].dropna().head(5).to_string())

    print(f"\n--- EXP TP SEG No DATE AND TIME samples ---")
    print(df["EXP TP SEG No DATE AND TIME"].dropna().head(5).to_string())

    print(f"\n--- DEPARTURE DATE & TIME samples ---")
    print(df["DEPARTURE DATE & TIME"].dropna().head(5).to_string())

    print(f"\n--- AWB No samples ---")
    print(df["AWB No"].dropna().head(10).tolist())

    print(f"\n--- Null counts ---")
    print(df.isnull().sum().to_string())