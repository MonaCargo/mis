"""
pdf_read_test.py
================
Extract Import Warehouse Inventory data from PDF → Excel.

Usage:
    python pdf_read_test.py getjobid5529279.pdf

Requirements:
    pip install pdfplumber openpyxl pandas
"""

import sys
import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Exact column x-boundaries (calibrated from this PDF's character positions) ─
# Each character is ~6px wide. Boundaries derived from coordinate analysis.
#
#  Column     x_start   x_end    Evidence from PDF
#  AWB_NO        9        82     "0","9","8","-","4","9"... at x=9,15,21,27,33,39...
#  HWB_NO      100       175     HWB chars at x=104,110,116,122,128,134,140,146,152,158,164,170
#  MH          215       235     "M" or "H" at x=222
#  STATUS      302       345     "R","C","F" / "D","L","V" at x=306,312,318
#  PCS         365       390     digits at x=369,375,381
#  WGT_CHG     438       458     digits at x=443,449 (sparse - only when value present)
#  SHC         460       500     "H","E","A","S","P","X"... at x=464,470,476,482,488,494
#  AGENT       510       540     "S","F","S" / "A","V","C" at x=520,526,532
#  FLT_NO      545       585     "A","I","2","3","3","3" at x=549,555,561,567,573,579
#  FLT_DATE    595       660     "2","5","-","F","E","B","-","2","6" at x=599,605,611,617,623,629,635,641,647

COL_BOUNDS = [
    ("AWB_NO",    9,    82),
    ("HWB_NO",   100,  175),
    ("MH",       215,  235),
    ("STATUS",   302,  345),
    ("PCS",      365,  390),
    ("WGT_CHG",  438,  458),
    ("SHC",      460,  500),
    ("AGENT",    510,  540),
    ("FLT_NO",   545,  585),
    ("FLT_DATE", 595,  660),
]

COLUMNS = [c[0] for c in COL_BOUNDS]

# Lines to skip
SKIP_PHRASES = [
    "import warehouse inventory",
    "from date", "to date",
    "delhi", "mumbai", "chennai", "kolkata", "bangalore", "hyderabad",
    "awb no", "hwb no", "m/h", "status", "wgt_chg", "flt no", "flt date",
    "agent", "shc", "grand total", "day total",
]


# ── Helpers ───────────────────────────────────────────────────────────────────

def is_valid_awb(text: str) -> bool:
    """AWB: 3 digits hyphen 7-8 digits  e.g. 098-31662105"""
    return bool(re.match(r"^\d{3}-\d{7,8}$", (text or "").strip()))


def is_skip_line(text: str) -> bool:
    t = text.strip()
    if not t:
        return True
    tl = t.lower()
    if re.search(r"\d{2}-[a-zA-Z]{3}-\d{2,4}\s+\d{2}:\d{2}", t):
        return True
    for phrase in SKIP_PHRASES:
        if phrase in tl:
            return True
    return False


def chars_to_text(chars: list) -> str:
    """Join individual character dicts (sorted by x) into a string."""
    return "".join(c["text"] for c in sorted(chars, key=lambda c: c["x0"])).strip()


def extract_cell_from_chars(chars: list, x0: float, x1: float) -> str:
    """Extract text from characters whose x-position falls within [x0, x1]."""
    in_range = [c for c in chars if c["x0"] >= x0 - 3 and c["x0"] <= x1 + 3]
    return chars_to_text(in_range)


# ── Core extraction ───────────────────────────────────────────────────────────

def extract_inventory_pdf(pdf_path: str) -> pd.DataFrame:
    rows = []

    with pdfplumber.open(pdf_path) as pdf:
        total = len(pdf.pages)

        for page_num, page in enumerate(pdf.pages, start=1):
            print(f"   Page {page_num}/{total}...", end="\r")

            # Extract individual characters with position
            chars = page.chars  # list of dicts: text, x0, x1, top, bottom, ...

            if not chars:
                continue

            # Group chars by line (y / top position, bucketed by 3pt)
            line_map: dict = {}
            for ch in chars:
                if not ch["text"].strip():
                    continue
                y_key = round(ch["top"] / 3) * 3
                line_map.setdefault(y_key, []).append(ch)

            # Process lines top → bottom
            for y_key in sorted(line_map):
                line_chars = line_map[y_key]
                line_text  = chars_to_text(line_chars)

                if is_skip_line(line_text):
                    continue

                # Extract each column by x-boundary
                row = {
                    col: extract_cell_from_chars(line_chars, x0, x1)
                    for col, x0, x1 in COL_BOUNDS
                }

                # Only keep rows with a valid AWB number
                if not is_valid_awb(row["AWB_NO"]):
                    continue

                rows.append(row)

    print()  # newline after progress
    df = pd.DataFrame(rows, columns=COLUMNS)

    # ── Type cleanup ──────────────────────────────────────────────────────────
    df.replace("", pd.NA, inplace=True)

    df["AWB_NO"]  = df["AWB_NO"].str.strip()
    df["HWB_NO"]  = df["HWB_NO"].str.strip()
    df["MH"]      = df["MH"].str.upper().str.strip()
    df["STATUS"]  = df["STATUS"].str.upper().str.strip()
    df["PCS"]     = pd.to_numeric(df["PCS"],     errors="coerce")
    df["WGT_CHG"] = pd.to_numeric(df["WGT_CHG"], errors="coerce")
    df["FLT_DATE"]= pd.to_datetime(df["FLT_DATE"], format="%d-%b-%y", errors="coerce")

    df.reset_index(drop=True, inplace=True)
    return df










